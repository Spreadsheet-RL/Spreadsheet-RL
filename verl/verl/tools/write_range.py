# Copyright 2026 Bytedance Ltd. and/or its affiliates
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
from __future__ import annotations

import asyncio
import json
import math
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .formula_fill import (
    _EXCEL_MAX_COLS,
    _EXCEL_MAX_ROWS,
    _commit_recalc_workbook_under_lock_sync,
    _copy_workbook_to_temp_under_lock_sync,
    _format_a1_address,
    _get_max_response_chars,
    _normalize_sheet_name,
    _quote_sheet_name_for_a1,
    _read_cells_formula_and_value_from_path,
    _resolve_target_worksheet,
    _resolve_workspace_file,
    _sample_indices,
    _sanitize_relpath,
    _scan_zip_metadata,
    _to_jsonable_excel_value,
    _truncate_payload_to_max_chars,
    _truncate_str,
)
from .schemas import OpenAIFunctionToolSchema, ToolResponse
from .workbook_formula_cache import (
    _collect_formula_cached_values_from_xlsx,
    _restore_formula_cached_values_in_xlsx,
)

_A1_CELL_RE = r"[A-Z]{1,3}[0-9]{1,7}"
_A1_RECT_RANGE_RE = re.compile(rf"^{_A1_CELL_RE}(?::{_A1_CELL_RE})?$")
_INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")


@dataclass(frozen=True)
class _RangeSpec:
    sheet_name: Optional[str]
    start_col: int
    start_row: int
    end_col: int
    end_row: int
    is_anchor: bool
    label: str


@dataclass(frozen=True)
class _CellPayload:
    kind: str
    value: Any = None
    force_text: bool = False


@dataclass(frozen=True)
class _WriteResult:
    resolved_sheet_name: str
    written_cells: int
    value_cells: int
    cleared_cells: int
    skipped_merged_cells: int
    overwritten_non_empty: int
    overwritten_samples: list[dict[str, Any]]
    created_sheet: bool
    formula_cache_preserved: bool = True


class _WriteRangeError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def _validation_response(code: str, message: str) -> tuple[ToolResponse, float, dict[str, Any]]:
    return ToolResponse(text=f"Error: {message}"), 0.0, {"status": "error", "error": code}


def _snippet(value: Any, *, max_chars: int = 160) -> str:
    try:
        text = json.dumps(value, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        text = repr(value)
    return _truncate_str(text, max_chars)


def _format_shape(rows: int, cols: int) -> str:
    row_label = "row" if rows == 1 else "rows"
    col_label = "col" if cols == 1 else "cols"
    return f"{rows} {row_label} x {cols} {col_label}"


def _get_positive_int(config: dict[str, Any], key: str, default: int) -> Optional[int]:
    raw = config.get(key, default)
    if raw is None:
        return None
    try:
        value = int(raw)
    except (TypeError, ValueError):
        value = default
    return value if value > 0 else None


def _is_valid_excel_sheet_name(name: str) -> bool:
    return bool(name) and len(name) <= 31 and _INVALID_SHEET_CHARS_RE.search(name) is None


def _split_sheet_range(value: str) -> tuple[Optional[str], str]:
    token = value.strip(' \t\n\r\f\v"')
    if not token:
        raise _WriteRangeError(
            "invalid_range",
            "range received an empty string; pass an A1 target like B2 or Sheet1!B2:D10.",
        )
    if "!" not in token:
        return None, token

    sheet_part, _, range_part = token.rpartition("!")
    sheet_name = _normalize_sheet_name(sheet_part)
    if not sheet_name:
        raise _WriteRangeError(
            "invalid_sheet_name",
            f"range received {value!r} with an empty sheet name; pass a valid sheet name before '!'.",
        )
    if not range_part.strip():
        raise _WriteRangeError(
            "invalid_range",
            f"range received {value!r} with an empty cell reference; pass a target like {sheet_name}!B2:D10.",
        )
    return sheet_name, range_part


def _parse_range_spec(value: Any) -> _RangeSpec:
    if not isinstance(value, str) or not value.strip():
        raise _WriteRangeError(
            "invalid_range",
            f"range received {_snippet(value)}; pass a non-empty A1 target like B2 or Sheet1!B2:D10.",
        )

    sheet_name, range_part = _split_sheet_range(value)
    cell_range = range_part.strip(" \t\n\r\f\v\"'").replace(" ", "").replace("$", "").upper()
    if not _A1_RECT_RANGE_RE.fullmatch(cell_range):
        raise _WriteRangeError(
            "invalid_range",
            f"range received {value!r}; pass a finite A1 cell or rectangle like B2 or B2:D10.",
        )

    try:
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    is_anchor = ":" not in cell_range
    parse_range = f"{cell_range}:{cell_range}" if is_anchor else cell_range
    try:
        min_col, min_row, max_col, max_row = range_boundaries(parse_range)
    except Exception as exc:
        raise _WriteRangeError(
            "invalid_range",
            f"range received {value!r} could not be parsed ({exc}); pass a target like B2:D10.",
        ) from None

    if None in (min_col, min_row, max_col, max_row):
        raise _WriteRangeError(
            "invalid_range",
            f"range received {value!r}; pass a finite A1 cell rectangle, not a whole row or column.",
        )

    start_col = int(min(min_col, max_col))
    end_col = int(max(min_col, max_col))
    start_row = int(min(min_row, max_row))
    end_row = int(max(min_row, max_row))
    if start_row < 1 or end_row > _EXCEL_MAX_ROWS or start_col < 1 or end_col > _EXCEL_MAX_COLS:
        raise _WriteRangeError(
            "invalid_range",
            f"range received {value!r} is outside Excel bounds; pass rows 1-{_EXCEL_MAX_ROWS} "
            f"and columns 1-{_EXCEL_MAX_COLS}.",
        )

    start_addr = f"{get_column_letter(start_col)}{start_row}"
    end_addr = f"{get_column_letter(end_col)}{end_row}"
    label = start_addr if start_addr == end_addr else f"{start_addr}:{end_addr}"
    return _RangeSpec(sheet_name, start_col, start_row, end_col, end_row, is_anchor, label)


def _decode_json_string(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if not value.lstrip().startswith("["):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def _raw_grid_from_input(
    value: Any,
    *,
    field_name: str,
    target_rows: int,
    target_cols: int,
    target_is_anchor: bool,
) -> list[list[Any]]:
    decoded = _decode_json_string(value)
    if isinstance(decoded, list):
        if not decoded:
            raise _WriteRangeError(
                "invalid_data",
                f"{field_name} received an empty array; pass a scalar, a 1D row, or a non-empty 2D array.",
            )

        if not any(isinstance(item, list) for item in decoded):
            if not target_is_anchor and target_cols == 1 and target_rows == len(decoded):
                return [[item] for item in decoded]
            return [decoded]

        expected_len: Optional[int] = None
        rows: list[list[Any]] = []
        for row_idx, row in enumerate(decoded, start=1):
            if not isinstance(row, list):
                raise _WriteRangeError(
                    "invalid_data",
                    f"{field_name}[{row_idx}] received {_snippet(row)}; pass a rectangular 2D array.",
                )
            if expected_len is None:
                expected_len = len(row)
                if expected_len == 0:
                    raise _WriteRangeError(
                        "invalid_data",
                        f"{field_name}[{row_idx}] received length 0; pass rows with at least one cell.",
                    )
            elif len(row) != expected_len:
                raise _WriteRangeError(
                    "invalid_data",
                    f"{field_name}[{row_idx}] received length {len(row)} but expected {expected_len}; "
                    "pass a rectangular 2D array.",
                )
            rows.append(row)
        return rows

    return [[decoded]]


def _check_text_length(*, field_name: str, row: int, col: int, value: str, max_cell_chars: Optional[int]) -> None:
    if max_cell_chars is not None and len(value) > max_cell_chars:
        raise _WriteRangeError(
            "cell_too_long",
            f"{field_name}[{row}][{col}] received string length {len(value)} > {max_cell_chars}; "
            "shorten the cell payload.",
        )


def _static_values_only_message(
    *,
    field_name: str,
    row: Optional[int] = None,
    col: Optional[int] = None,
    value: Any,
) -> str:
    location = field_name
    if row is not None and col is not None:
        location = f"{field_name}[{row}][{col}]"
    return (
        f"{location} received {_snippet(value)}; write_range writes static values only -- use fill_formula for "
        "formulas, or prefix with ' to write the text literally."
    )


def _cell_payload(
    value: Any,
    *,
    field_name: str,
    row: int,
    col: int,
    max_cell_chars: Optional[int],
) -> _CellPayload:
    if value is None:
        return _CellPayload("skip")
    if isinstance(value, bool):
        return _CellPayload("value", value)
    if isinstance(value, int):
        return _CellPayload("value", value)
    if isinstance(value, float):
        if math.isfinite(value):
            return _CellPayload("value", value)
        raise _WriteRangeError(
            "invalid_data",
            f"{field_name}[{row}][{col}] received non-finite number {value!r}; pass a finite JSON number.",
        )
    if isinstance(value, str):
        if value == "":
            return _CellPayload("clear")
        if value.startswith("'"):
            literal = value[1:]
            _check_text_length(field_name=field_name, row=row, col=col, value=literal, max_cell_chars=max_cell_chars)
            return _CellPayload("value", literal, force_text=True)
        if value.startswith("="):
            raise _WriteRangeError(
                "invalid_data",
                _static_values_only_message(field_name=field_name, row=row, col=col, value=value),
            )
        _check_text_length(field_name=field_name, row=row, col=col, value=value, max_cell_chars=max_cell_chars)
        return _CellPayload("value", value)

    raise _WriteRangeError(
        "invalid_data",
        f"{field_name}[{row}][{col}] received {_snippet(value)}; pass a JSON scalar, null, or a static string.",
    )


def _coerce_data_grid(
    raw_grid: list[list[Any]],
    *,
    field_name: str,
    max_cell_chars: Optional[int],
) -> list[list[_CellPayload]]:
    out: list[list[_CellPayload]] = []
    for row_idx, row in enumerate(raw_grid, start=1):
        out_row: list[_CellPayload] = []
        for col_idx, value in enumerate(row, start=1):
            out_row.append(
                _cell_payload(
                    value,
                    field_name=field_name,
                    row=row_idx,
                    col=col_idx,
                    max_cell_chars=max_cell_chars,
                )
            )
        out.append(out_row)
    return out


def _merge_legacy_grids(
    *,
    values: Any,
    formulas: Any,
    range_spec: _RangeSpec,
    max_cell_chars: Optional[int],
) -> list[list[_CellPayload]]:
    target_rows = range_spec.end_row - range_spec.start_row + 1
    target_cols = range_spec.end_col - range_spec.start_col + 1
    if formulas is not None:
        raise _WriteRangeError(
            "invalid_data",
            _static_values_only_message(field_name="formulas", value=formulas),
        )
    values_grid = None
    if values is not None:
        values_grid = _raw_grid_from_input(
            values,
            field_name="values",
            target_rows=target_rows,
            target_cols=target_cols,
            target_is_anchor=range_spec.is_anchor,
        )
    if values_grid is None:
        raise _WriteRangeError(
            "invalid_data",
            "data received missing; pass data as a scalar, 1D row, or rectangular 2D array.",
        )
    return _coerce_data_grid(values_grid, field_name="values", max_cell_chars=max_cell_chars)


def _coerce_input_grid(
    *,
    parameters: dict[str, Any],
    range_spec: _RangeSpec,
    max_cell_chars: Optional[int],
) -> list[list[_CellPayload]]:
    target_rows = range_spec.end_row - range_spec.start_row + 1
    target_cols = range_spec.end_col - range_spec.start_col + 1
    if parameters.get("formulas") is not None:
        raise _WriteRangeError(
            "invalid_data",
            _static_values_only_message(field_name="formulas", value=parameters.get("formulas")),
        )
    if parameters.get("data") is not None:
        raw_grid = _raw_grid_from_input(
            parameters.get("data"),
            field_name="data",
            target_rows=target_rows,
            target_cols=target_cols,
            target_is_anchor=range_spec.is_anchor,
        )
        return _coerce_data_grid(raw_grid, field_name="data", max_cell_chars=max_cell_chars)

    return _merge_legacy_grids(
        values=parameters.get("values") if "values" in parameters else None,
        formulas=parameters.get("formulas") if "formulas" in parameters else None,
        range_spec=range_spec,
        max_cell_chars=max_cell_chars,
    )


def _expand_grid_for_target(
    *,
    grid: list[list[_CellPayload]],
    range_spec: _RangeSpec,
) -> tuple[list[list[_CellPayload]], int, int]:
    data_rows = len(grid)
    data_cols = len(grid[0])
    if range_spec.is_anchor:
        return grid, data_rows, data_cols

    target_rows = range_spec.end_row - range_spec.start_row + 1
    target_cols = range_spec.end_col - range_spec.start_col + 1

    if data_rows == target_rows and data_cols == target_cols:
        return grid, target_rows, target_cols

    expanded: list[list[_CellPayload]] = []
    if data_rows == 1 and data_cols == 1:
        source = grid[0][0]
        for _ in range(target_rows):
            expanded.append([source for _ in range(target_cols)])
        return expanded, target_rows, target_cols

    if data_rows == 1 and data_cols == target_cols:
        for _ in range(target_rows):
            expanded.append(list(grid[0]))
        return expanded, target_rows, target_cols

    if data_cols == 1 and data_rows == target_rows:
        for source_row in grid:
            source = source_row[0]
            expanded.append([source for _ in range(target_cols)])
        return expanded, target_rows, target_cols

    hint = f"range='{range_spec.label.split(':', 1)[0]}'"
    raise _WriteRangeError(
        "shape_mismatch",
        f"data is {_format_shape(data_rows, data_cols)} but range {range_spec.label} is "
        f"{_format_shape(target_rows, target_cols)}; make data match the range, or pass only the top-left cell "
        f"({hint}) to size the write from data.",
    )


def _sample_coords(*, start_row: int, start_col: int, total_rows: int, total_cols: int) -> list[tuple[int, int]]:
    if total_rows <= 1 and total_cols <= 1:
        return [(start_row, start_col)]
    if total_rows == 1:
        sample_cols = [start_col + idx for idx in _sample_indices(total_cols, head=3, tail=3)]
        return [(start_row, col_num) for col_num in sample_cols]
    if total_cols == 1:
        sample_rows = [start_row + idx for idx in _sample_indices(total_rows, head=3, tail=3)]
        return [(row_num, start_col) for row_num in sample_rows]

    sample_rows = [start_row + idx for idx in _sample_indices(total_rows, head=2, tail=2)]
    sample_cols = [start_col + idx for idx in _sample_indices(total_cols, head=2, tail=2)]
    return [(row_num, col_num) for row_num in sample_rows for col_num in sample_cols]


def _resolved_range_text(*, sheet_name: str, start_col: int, start_row: int, total_rows: int, total_cols: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required to format ranges; install openpyxl>=3.1.5") from None

    end_row = start_row + total_rows - 1
    end_col = start_col + total_cols - 1
    start_addr = f"{get_column_letter(start_col)}{start_row}"
    end_addr = f"{get_column_letter(end_col)}{end_row}"
    cell_range = start_addr if start_addr == end_addr else f"{start_addr}:{end_addr}"
    return f"{_quote_sheet_name_for_a1(sheet_name)}!{cell_range}"


def _resolve_or_create_worksheet(wb, requested_name: Optional[str]):
    worksheets = getattr(wb, "worksheets", None) or []
    if requested_name is None:
        if worksheets:
            return worksheets[0], False
        return wb.create_sheet(title="Sheet1"), True

    try:
        return _resolve_target_worksheet(wb, requested_name), False
    except RuntimeError as exc:
        message = str(exc)
        if not message.startswith("sheet not found"):
            raise _WriteRangeError(
                "invalid_sheet_name",
                f"sheet_name received {requested_name!r}: {exc}; pass the exact sheet title.",
            ) from None
        if not _is_valid_excel_sheet_name(requested_name):
            raise _WriteRangeError(
                "invalid_sheet_name",
                f"sheet_name received {requested_name!r}; pass a valid Excel sheet name (1-31 chars, no []:*?/\\).",
            ) from None
        return wb.create_sheet(title=requested_name), True


def _old_cell_value(cell: Any) -> Any:
    value = getattr(cell, "value", None)
    if value is None or value == "":
        return None
    if getattr(cell, "data_type", None) == "f" and isinstance(value, str):
        formula = value if value.startswith("=") else f"={value}"
        return _truncate_str(formula, 200)
    return _to_jsonable_excel_value(value)


def _write_range_in_workbook_unlocked_sync(
    *,
    file_path: Path,
    sheet_name: Optional[str],
    start_row: int,
    start_col: int,
    grid: list[list[_CellPayload]],
) -> _WriteResult:
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    wb = None
    try:
        cached_formula_values = _collect_formula_cached_values_from_xlsx(file_path)
        try:
            wb = load_workbook(filename=str(file_path), keep_vba=False, data_only=False, keep_links=True)
        except Exception as exc:
            raise RuntimeError(f"failed to load workbook: {exc}") from None

        ws, created_sheet = _resolve_or_create_worksheet(wb, sheet_name)
        resolved_sheet_name = getattr(ws, "title", sheet_name) or sheet_name or ""

        written_cells = 0
        value_cells = 0
        cleared_cells = 0
        skipped_merged_cells = 0
        overwritten_non_empty = 0
        overwritten_samples: list[dict[str, Any]] = []

        for row_offset, row in enumerate(grid):
            row_num = start_row + row_offset
            for col_offset, payload in enumerate(row):
                if payload.kind == "skip":
                    continue
                col_num = start_col + col_offset
                cell = ws.cell(row=row_num, column=col_num)
                if isinstance(cell, MergedCell):
                    skipped_merged_cells += 1
                    continue

                old_value = _old_cell_value(cell)
                if old_value is not None:
                    overwritten_non_empty += 1
                    if len(overwritten_samples) < 3:
                        overwritten_samples.append(
                            {
                                "address": _format_a1_address(
                                    sheet_name=resolved_sheet_name,
                                    row=row_num,
                                    col=col_num,
                                ),
                                "old": old_value,
                            }
                        )

                if payload.kind == "clear":
                    cell.value = None
                    cleared_cells += 1
                else:
                    cell.value = payload.value
                    if payload.force_text:
                        cell.data_type = "s"
                    value_cells += 1
                written_cells += 1

        if written_cells == 0:
            raise _WriteRangeError(
                "no_cells_to_write",
                "data contains no cells to write (all null/merged-skipped).",
            )

        wb.save(str(file_path))
        formula_cache_preserved = True
        try:
            _restore_formula_cached_values_in_xlsx(file_path, cached_formula_values)
        except Exception:
            formula_cache_preserved = False
        return _WriteResult(
            resolved_sheet_name=resolved_sheet_name,
            written_cells=written_cells,
            value_cells=value_cells,
            cleared_cells=cleared_cells,
            skipped_merged_cells=skipped_merged_cells,
            overwritten_non_empty=overwritten_non_empty,
            overwritten_samples=overwritten_samples,
            created_sheet=created_sheet,
            formula_cache_preserved=formula_cache_preserved,
        )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


class WriteRangeTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            self.lock_timeout_s = max(0.0, float(lock_timeout_s_raw))
        except (TypeError, ValueError):
            self.lock_timeout_s = 30.0

        self.max_write_cells = _get_positive_int(config, "max_write_cells", 50_000)
        self.max_cell_chars = _get_positive_int(config, "max_cell_chars", 8192)
        self.max_response_chars = _get_max_response_chars(config)

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return _validation_response(
                    "invalid_path",
                    f"path received {_snippet(raw_path)}; pass a relative .xlsx path such as data.xlsx.",
                )
        if relpath.suffix.lower() != ".xlsx":
            return _validation_response(
                "invalid_path",
                f"path received {str(relpath)!r}; pass a relative .xlsx workbook path such as data.xlsx.",
            )

        range_raw = parameters.get("range")
        if range_raw is None and parameters.get("start_cell") is not None:
            range_raw = parameters.get("start_cell")
        try:
            range_spec = _parse_range_spec(range_raw)
        except _WriteRangeError as exc:
            return _validation_response(exc.code, str(exc))

        sheet_name_raw = parameters.get("sheet_name")
        sheet_name = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return _validation_response(
                    "invalid_sheet_name",
                    f"sheet_name received {_snippet(sheet_name_raw)}; pass a non-empty worksheet name.",
                )
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not _is_valid_excel_sheet_name(sheet_name):
                return _validation_response(
                    "invalid_sheet_name",
                    f"sheet_name received {sheet_name_raw!r}; pass a valid Excel sheet name (1-31 chars, no []:*?/\\).",
                )
            if range_spec.sheet_name is not None and sheet_name.casefold() != range_spec.sheet_name.casefold():
                return _validation_response(
                    "sheet_mismatch",
                    f"sheet_name {sheet_name!r} does not match range sheet {range_spec.sheet_name!r}; "
                    "make them the same or put the sheet name in only one parameter.",
                )
        elif range_spec.sheet_name is not None:
            sheet_name = range_spec.sheet_name
            if not _is_valid_excel_sheet_name(sheet_name):
                return _validation_response(
                    "invalid_sheet_name",
                    f"range sheet received {sheet_name!r}; pass a valid Excel sheet name (1-31 chars, no []:*?/\\).",
                )

        try:
            grid = _coerce_input_grid(
                parameters=parameters,
                range_spec=range_spec,
                max_cell_chars=self.max_cell_chars,
            )
            grid, total_rows, total_cols = _expand_grid_for_target(grid=grid, range_spec=range_spec)
        except _WriteRangeError as exc:
            return _validation_response(exc.code, str(exc))

        inferred_end_row = range_spec.start_row + total_rows - 1
        inferred_end_col = range_spec.start_col + total_cols - 1
        if inferred_end_row > _EXCEL_MAX_ROWS or inferred_end_col > _EXCEL_MAX_COLS:
            return _validation_response(
                "invalid_range",
                f"range received {range_spec.label!r} with data shape {_format_shape(total_rows, total_cols)} "
                "extends beyond Excel bounds; choose a top-left cell that keeps the write inside the worksheet.",
            )

        target_cells = total_rows * total_cols
        if self.max_write_cells is not None and target_cells > self.max_write_cells:
            return _validation_response(
                "range_too_large",
                f"range received {range_spec.label!r} resolves to {target_cells} cells > {self.max_write_cells}; "
                "split the write into smaller ranges.",
            )

        skipped_null_cells = sum(1 for row in grid for payload in row if payload.kind == "skip")
        if skipped_null_cells == target_cells:
            return _validation_response(
                "no_cells_to_write",
                "data contains no cells to write (all null/merged-skipped).",
            )

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return _validation_response(
                "missing_workspace_id",
                "workspace_id received missing/invalid; pass a valid workspace_id for the workbook.",
            )
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return _validation_response(
                "file_not_found",
                f"path received {str(relpath)!r} but no workbook was found; create it in the workspace or fix path.",
            )
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return _validation_response(
                    "file_not_found",
                    f"path received {str(relpath)!r} could not be stat'ed ({exc}); verify the workbook exists.",
                )
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return _validation_response(
                    "file_too_large",
                    f"path received {str(relpath)!r} is {actual_mb:.1f}MB > {max_mb}MB; use a smaller workbook.",
                )

        coords = _sample_coords(
            start_row=range_spec.start_row,
            start_col=range_spec.start_col,
            total_rows=total_rows,
            total_cols=total_cols,
        )

        tmp_fill_path: Optional[Path] = None
        expected_sig: Optional[tuple[int, int, int, int]] = None
        result: Optional[_WriteResult] = None
        samples: list[dict[str, Any]] = []
        samples_source = "none"
        writeback_error: Optional[str] = None
        error_code: Optional[str] = None

        try:
            try:
                tmp_fill_path, expected_sig = await asyncio.to_thread(
                    _copy_workbook_to_temp_under_lock_sync,
                    file_path=file_path,
                    lock_timeout_s=self.lock_timeout_s,
                )
                result = await asyncio.to_thread(
                    _write_range_in_workbook_unlocked_sync,
                    file_path=tmp_fill_path,
                    sheet_name=sheet_name,
                    start_row=range_spec.start_row,
                    start_col=range_spec.start_col,
                    grid=grid,
                )
            except asyncio.CancelledError:
                raise
            except _WriteRangeError as exc:
                return _validation_response(exc.code, str(exc))
            except Exception as exc:
                payload = {
                    "status": "error",
                    "error": "write_failed",
                    "file": str(relpath),
                    "target_cells": target_cells,
                    "rolled_back": True,
                    "write_error": _to_jsonable_excel_value(str(exc)),
                }
                response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
                return (
                    ToolResponse(text=response_text),
                    0.0,
                    {"status": "error", "error": "write_failed", "file": str(relpath), "rolled_back": True},
                )

            zip_error = await asyncio.to_thread(
                _scan_zip_metadata,
                tmp_fill_path,
                max_members=50_000,
                max_total_uncompressed_bytes=512 * 1024 * 1024,
                max_member_uncompressed_bytes=128 * 1024 * 1024,
                max_ratio=200.0,
            )
            if zip_error:
                writeback_error = f"filled workbook rejected by zip safety checks: {zip_error}"
                error_code = "unsafe_workbook"
            else:
                try:
                    filled_content = await asyncio.to_thread(tmp_fill_path.read_bytes)
                except asyncio.CancelledError:
                    raise
                except Exception as exc:
                    writeback_error = f"filled workbook could not be read before commit: {exc}"
                    error_code = "writeback_failed"
                else:
                    assert expected_sig is not None
                    try:
                        committed, writeback_error = await asyncio.to_thread(
                            _commit_recalc_workbook_under_lock_sync,
                            file_path=file_path,
                            expected_sig=expected_sig,
                            recalc_content=filled_content,
                            lock_timeout_s=self.lock_timeout_s,
                        )
                    except asyncio.CancelledError:
                        raise
                    except Exception as exc:
                        committed = False
                        writeback_error = f"writeback failed: {exc}"
                    if not committed and writeback_error is None:
                        writeback_error = "writeback failed"
                    if not committed:
                        error_code = "writeback_failed"

            if writeback_error is not None and tmp_fill_path is not None and result is not None:
                try:
                    samples = await asyncio.to_thread(
                        _read_cells_formula_and_value_from_path,
                        file_path=tmp_fill_path,
                        sheet_name=result.resolved_sheet_name,
                        coords=coords,
                    )
                    samples_source = "filled_workbook"
                except asyncio.CancelledError:
                    raise
                except Exception as exc:
                    samples = []
                    samples_source = "none"
                    writeback_error = f"{writeback_error}; failed to read filled workbook samples: {exc}"
        finally:
            if tmp_fill_path is not None:
                try:
                    tmp_fill_path.unlink(missing_ok=True)
                except Exception:
                    pass

        if result is None:
            return _validation_response(
                "write_failed",
                "data received could not be written; retry with a valid workbook and target range.",
            )

        resolved_range = _resolved_range_text(
            sheet_name=result.resolved_sheet_name,
            start_col=range_spec.start_col,
            start_row=range_spec.start_row,
            total_rows=total_rows,
            total_cols=total_cols,
        )

        if writeback_error is not None:
            payload: dict[str, Any] = {
                "status": "error",
                "error": error_code or "writeback_failed",
                "file": str(relpath),
                "sheet": result.resolved_sheet_name,
                "range": resolved_range,
                "target_cells": target_cells,
                "written_cells": result.written_cells,
                "value_cells": result.value_cells,
                "cleared_cells": result.cleared_cells,
                "skipped_null_cells": skipped_null_cells,
                "skipped_merged_cells": result.skipped_merged_cells,
                "overwritten_non_empty": result.overwritten_non_empty,
                "recalculated": False,
                "samples_source": samples_source,
                "rolled_back": True,
                "writeback_error": _to_jsonable_excel_value(writeback_error),
                "samples": samples,
            }
            if result.created_sheet:
                payload["created_sheet"] = True
            if result.overwritten_samples:
                payload["overwritten_samples"] = result.overwritten_samples
            response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
            samples_out = payload.get("samples")
            sample_count = len(samples_out) if isinstance(samples_out, list) else 0
            metrics = {
                "status": "error",
                "error": error_code or "writeback_failed",
                "file": str(relpath),
                "target_cells": target_cells,
                "written_cells": result.written_cells,
                "skipped_null_cells": skipped_null_cells,
                "skipped_merged_cells": result.skipped_merged_cells,
                "recalculated": False,
                "samples_source": samples_source,
                "sample_cells": sample_count,
            }
            return ToolResponse(text=response_text), 0.0, metrics

        try:
            samples = await asyncio.to_thread(
                _read_cells_formula_and_value_from_path,
                file_path=file_path,
                sheet_name=result.resolved_sheet_name,
                coords=coords,
            )
            samples_source = "disk"
        except asyncio.CancelledError:
            raise
        except Exception:
            samples = []
            samples_source = "none"

        status = "success" if result.skipped_merged_cells == 0 else "partial_success"
        payload = {
            "status": status,
            "file": str(relpath),
            "sheet": result.resolved_sheet_name,
            "range": resolved_range,
            "target_cells": target_cells,
            "written_cells": result.written_cells,
            "value_cells": result.value_cells,
            "cleared_cells": result.cleared_cells,
            "skipped_null_cells": skipped_null_cells,
            "skipped_merged_cells": result.skipped_merged_cells,
            "overwritten_non_empty": result.overwritten_non_empty,
            "recalculated": False,
            "samples_source": samples_source,
            "samples": samples,
        }
        if result.created_sheet:
            payload["created_sheet"] = True
        if result.overwritten_samples:
            payload["overwritten_samples"] = result.overwritten_samples
        if not result.formula_cache_preserved:
            payload["formula_cache_preserved"] = False

        response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
        samples_out = payload.get("samples")
        sample_count = len(samples_out) if isinstance(samples_out, list) else 0
        metrics = {
            "status": status,
            "file": str(relpath),
            "target_cells": target_cells,
            "written_cells": result.written_cells,
            "value_cells": result.value_cells,
            "cleared_cells": result.cleared_cells,
            "skipped_null_cells": skipped_null_cells,
            "skipped_merged_cells": result.skipped_merged_cells,
            "overwritten_non_empty": result.overwritten_non_empty,
            "recalculated": False,
            "samples_source": samples_source,
            "sample_cells": sample_count,
        }
        if result.created_sheet:
            metrics["created_sheet"] = True
        if not result.formula_cache_preserved:
            metrics["formula_cache_preserved"] = False
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
