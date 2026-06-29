from __future__ import annotations

import asyncio
import datetime as dt
import errno
import io
import json
import math
import mimetypes
import os
import re
import stat
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import get_sheet_arena_data_root, normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .schemas import OpenAIFunctionToolSchema, ToolResponse


def _is_success_status(status: int) -> bool:
    return 200 <= status < 300


_A1_CELL_RE = r"[A-Z]{1,3}[0-9]{1,7}"
_A1_COL_RE = r"[A-Z]{1,3}"
_A1_ROW_RE = r"[0-9]{1,7}"
_A1_RANGE_RE = re.compile(
    rf"^(?:{_A1_CELL_RE}(?::{_A1_CELL_RE})?|{_A1_COL_RE}:{_A1_COL_RE}|{_A1_ROW_RE}:{_A1_ROW_RE})$"
)

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z0-9_]+$")
DEFAULT_MAX_RESPONSE_CHARS = 4096
MAX_RESPONSE_CHARS = 8192


def _sanitize_relpath(value: Any) -> Optional[Path]:
    if not isinstance(value, str):
        return None
    raw = value.strip()
    if not raw:
        return None
    if "\0" in raw:
        return None
    p = Path(raw)
    if p.is_absolute():
        return None
    if any(part == ".." for part in p.parts):
        return None
    return p


def _resolve_workspace_file(
    *,
    workspace_id: Optional[str],
    relpath: Path,
) -> Optional[Path]:
    if not workspace_id:
        return None

    workspace_base = get_sheet_arena_data_root()
    workspaces_base = workspace_base / "_workspaces"
    try:
        if workspace_base.is_symlink():
            return None
        workspaces_base_resolved = workspaces_base.resolve(strict=True)
    except (OSError, RuntimeError):
        return None

    workspace_root = workspaces_base / workspace_id
    try:
        if workspace_root.is_symlink():
            return None
        workspace_root_resolved = workspace_root.resolve(strict=True)
    except (OSError, RuntimeError):
        return None
    if not workspace_root_resolved.is_relative_to(workspaces_base_resolved):
        return None

    rel_candidates = [relpath]
    if relpath.parts and relpath.parts[0] == "data" and len(relpath.parts) > 1:
        rel_candidates.append(Path(*relpath.parts[1:]))

    for rel_candidate in rel_candidates:
        candidate = workspace_root / rel_candidate
        try:
            resolved = candidate.resolve(strict=True)
        except (OSError, RuntimeError):
            continue
        if not resolved.is_relative_to(workspace_root_resolved):
            continue

        if candidate.is_symlink():
            continue

        cursor = workspace_root
        safe = True
        for part in rel_candidate.parts:
            cursor = cursor / part
            try:
                if cursor.is_symlink():
                    safe = False
                    break
            except OSError:
                safe = False
                break
        if not safe:
            continue

        if resolved.is_file():
            return resolved
    return None


def _add_query_params(url: str, params: dict[str, str]) -> str:
    parts = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qs(parts.query, keep_blank_values=True)
    for key, value in params.items():
        query[key] = [value]
    new_query = urllib.parse.urlencode(query, doseq=True)
    return urllib.parse.urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))


def _get_poll_wait_s() -> float:
    value = os.environ.get("SHEET_ARENA_RECALC_POLL_WAIT_S", "10")
    try:
        poll_wait_s = float(value)
        return min(max(0.0, poll_wait_s), 25.0)
    except ValueError:
        return 10.0


def _get_submit_retry_interval_s() -> float:
    value = os.environ.get("SHEET_ARENA_RECALC_SUBMIT_RETRY_INTERVAL_S", "20")
    try:
        retry_s = float(value)
        return max(0.1, retry_s)
    except ValueError:
        return 20.0


def _get_max_read_cells() -> int:
    value = os.environ.get("SHEET_ARENA_RECALC_MAX_READ_CELLS", "100")
    try:
        n = int(value)
        return min(max(1, n), 10_000)
    except ValueError:
        return 100


def _get_max_response_chars(config: Optional[dict[str, Any]] = None) -> int:
    config_value = None
    if isinstance(config, dict):
        config_value = config.get("max_response_chars")
    for raw in (config_value, os.environ.get("SHEET_ARENA_TOOL_MAX_RESPONSE_CHARS", "").strip()):
        if raw is None:
            continue
        try:
            n = int(raw)
            return min(max(128, n), MAX_RESPONSE_CHARS)
        except (TypeError, ValueError):
            continue
    return DEFAULT_MAX_RESPONSE_CHARS


def _get_max_string_chars() -> int:
    raw = os.environ.get("SHEET_ARENA_TOOL_MAX_STRING_CHARS", "200").strip()
    try:
        n = int(raw)
        return min(max(16, n), 100_000)
    except ValueError:
        return 200


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _truncate_str(value: str, max_chars: int) -> str:
    if max_chars <= 0:
        return ""
    if len(value) <= max_chars:
        return value
    if max_chars <= 3:
        return value[:max_chars]
    return value[: max_chars - 3] + "..."


def _scan_zip_metadata_bytes(
    content: bytes,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={getattr(info, 'filename', '?')!r}, bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={getattr(info, 'filename', '?')!r}, ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        f"zip contents too large (total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read zip metadata: {exc}"

    return None


def _scan_xlsx_container_bytes(content: bytes) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = {str(name).replace("\\", "/") for name in zf.namelist()}
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read xlsx container: {exc}"

    required = ("[Content_Types].xml", "xl/workbook.xml")
    for member in required:
        if member not in names:
            return f"missing required xlsx member: {member}"
    if not any(name.startswith("xl/worksheets/") and name.endswith(".xml") for name in names):
        return "missing worksheet parts"
    return None


def _scan_xlsx_loadable_bytes(content: bytes) -> Optional[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = None
    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=False, read_only=True, keep_links=False)
        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            return "workbook has no worksheets"
    except Exception as exc:
        return f"failed to load workbook: {exc}"
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return None


def _scan_zip_metadata(
    file_path: Path,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(file_path) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={getattr(info, 'filename', '?')!r}, bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={getattr(info, 'filename', '?')!r}, ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        f"zip contents too large (total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read zip metadata: {exc}"

    return None


def _quote_sheet_name_for_a1(sheet_name: str) -> str:
    if _SAFE_SHEET_NAME_RE.fullmatch(sheet_name):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _format_a1_address(*, sheet_name: str, row: int, col: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        col_letter = str(col)
    else:
        col_letter = get_column_letter(col)
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    return f"{sheet_ref}!{col_letter}{row}"


def _sample_linear_indices(total: int, *, head: int, tail: int) -> list[int]:
    if total <= 0:
        return []
    head_n = max(0, head)
    tail_n = max(0, tail)
    if total <= head_n + tail_n:
        return list(range(total))
    indices = list(range(head_n))
    indices.extend(range(total - tail_n, total))
    return indices


def _normalize_sheet_name(value: str) -> str:
    name = value.strip(' \t\n\r\f\v"')
    if len(name) >= 2 and name[0] == "'" and name[-1] == "'":
        name = name[1:-1].replace("''", "'")
        return name.strip()

    if name.startswith("'") and not name.endswith("'"):
        name = name[1:]
    elif name.endswith("'") and not name.startswith("'"):
        name = name[:-1]
    return name.strip()


def _normalize_cell_range(value: str) -> str:
    rng = value.strip(" \t\n\r\f\v\"'")
    rng = rng.replace(" ", "").replace("$", "")
    return rng.upper()


def _parse_sheet_cell_range(token: str, *, default_sheet_name: str) -> tuple[str, str]:
    token = token.strip(' \t\n\r\f\v"')
    if not token:
        raise ValueError("empty cell range")

    if "!" in token:
        sheet_part, _, range_part = token.rpartition("!")
        sheet_name = _normalize_sheet_name(sheet_part)
        if not sheet_name:
            raise ValueError(f"empty sheet name: {token!r}")
    else:
        sheet_name = default_sheet_name
        range_part = token

    cell_range = _normalize_cell_range(range_part)
    if not _A1_RANGE_RE.fullmatch(cell_range):
        raise ValueError(f"invalid A1 cell range: {range_part!r}")
    return sheet_name, cell_range


def _to_jsonable_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, str)):
        if isinstance(value, str):
            return _truncate_str(value, _get_max_string_chars())
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return str(value)


def _read_data_only_values(
    *, file_path: Optional[Path] = None, content: Optional[bytes] = None, cell_ranges: list[str]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required to read cell values; install openpyxl>=3.1.5") from None

    if (file_path is None) == (content is None):
        raise ValueError("provide exactly one of file_path or content")

    try:
        if content is not None:
            zip_error = _scan_zip_metadata_bytes(
                content,
                max_members=50_000,
                max_total_uncompressed_bytes=512 * 1024 * 1024,
                max_member_uncompressed_bytes=128 * 1024 * 1024,
                max_ratio=200.0,
            )
            if zip_error:
                raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")
            wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True, keep_links=False)
        else:
            wb = load_workbook(filename=str(file_path), data_only=True, read_only=True, keep_links=False)
    except Exception as exc:
        raise RuntimeError(f"failed to load workbook: {exc}") from None
    try:
        if not wb.sheetnames:
            raise RuntimeError("workbook has no sheets")
        default_sheet_name = wb.sheetnames[0]

        results: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        max_return_cells = _get_max_read_cells()
        returned_cells = 0
        sheet_map: dict[str, str | None] = {}
        for sheet in wb.sheetnames:
            key = sheet.lower()
            if key not in sheet_map:
                sheet_map[key] = sheet
            elif sheet_map[key] != sheet:
                sheet_map[key] = None

        for token in cell_ranges:
            try:
                sheet_name, cell_range = _parse_sheet_cell_range(token, default_sheet_name=default_sheet_name)
            except ValueError as exc:
                errors.append({"range": token, "error": str(exc)})
                continue

            sheet_key = sheet_name.lower()
            if sheet_key not in sheet_map:
                errors.append({"range": token, "error": f"sheet not found: {sheet_name}"})
                continue
            actual_sheet = sheet_map[sheet_key]
            if actual_sheet is None:
                errors.append({"range": token, "error": f"ambiguous sheet name: {sheet_name}"})
                continue
            sheet_name = actual_sheet

            try:
                boundaries = range_boundaries(cell_range if ":" in cell_range else f"{cell_range}:{cell_range}")
                min_col, min_row, max_col, max_row = boundaries
                if min_col is None:
                    min_col = 1
                if max_col is None:
                    max_col = _EXCEL_MAX_COLS
                if min_row is None:
                    min_row = 1
                if max_row is None:
                    max_row = _EXCEL_MAX_ROWS
                min_col = int(min_col)
                max_col = int(max_col)
                min_row = int(min_row)
                max_row = int(max_row)
                if not (
                    1 <= min_col <= _EXCEL_MAX_COLS
                    and 1 <= max_col <= _EXCEL_MAX_COLS
                    and 1 <= min_row <= _EXCEL_MAX_ROWS
                    and 1 <= max_row <= _EXCEL_MAX_ROWS
                ):
                    errors.append({"range": token, "error": "range out of bounds"})
                    continue
                range_cells = (max_col - min_col + 1) * (max_row - min_row + 1)
            except Exception as exc:
                errors.append({"range": token, "error": f"failed to parse range boundaries: {exc}"})
                continue

            if range_cells <= 0:
                errors.append({"range": token, "error": "invalid range boundaries"})
                continue

            ws = wb[sheet_name]
            rows = max_row - min_row + 1
            cols = max_col - min_col + 1
            remaining_budget = max(0, max_return_cells - returned_cells)
            want_full = range_cells <= remaining_budget

            if want_full:
                try:
                    values: list[list[Any]] = []
                    for r_idx, row in enumerate(
                        ws.iter_rows(
                            min_row=min_row,
                            max_row=max_row,
                            min_col=min_col,
                            max_col=max_col,
                        ),
                        start=0,
                    ):
                        values_row = [_to_jsonable_excel_value(getattr(cell, "value", None)) for cell in row]
                        if len(values_row) < cols:
                            values_row.extend([None] * (cols - len(values_row)))
                        elif len(values_row) > cols:
                            values_row = values_row[:cols]
                        values.append(values_row)

                    if len(values) < rows:
                        for r_idx in range(len(values), rows):
                            values.append([None] * cols)
                    elif len(values) > rows:
                        values = values[:rows]
                except Exception as exc:
                    errors.append({"range": token, "error": f"failed to read range: {exc}"})
                    continue

                returned_cells += range_cells
                sheet_ref = _quote_sheet_name_for_a1(sheet_name)
                addresses = [
                    [
                        _format_a1_address(sheet_name=sheet_name, row=min_row + r_off, col=min_col + c_off)
                        for c_off in range(cols)
                    ]
                    for r_off in range(rows)
                ]
                results.append(
                    {
                        "range": f"{sheet_ref}!{cell_range}",
                        "shape": {
                            "rows": rows,
                            "cols": cols,
                            "cells": range_cells,
                            "min_row": min_row,
                            "min_col": min_col,
                            "max_row": max_row,
                            "max_col": max_col,
                        },
                        "truncated": False,
                        "top_left": _format_a1_address(sheet_name=sheet_name, row=min_row, col=min_col),
                        "addresses": addresses,
                        "values": values,
                    }
                )
                continue

            try:
                sample_indices = _sample_linear_indices(range_cells, head=5, tail=5)
                sample_cells: list[dict[str, Any]] = []
                for idx in sample_indices:
                    r_off = idx // cols
                    c_off = idx - r_off * cols
                    row_num = min_row + r_off
                    col_num = min_col + c_off
                    try:
                        cell_obj = ws.cell(row=row_num, column=col_num)
                        value_obj = getattr(cell_obj, "value", None)
                    except Exception:
                        value_obj = None
                    sample_cells.append(
                        {
                            "address": _format_a1_address(sheet_name=sheet_name, row=row_num, col=col_num),
                            "value": _to_jsonable_excel_value(value_obj),
                        }
                    )
            except Exception as exc:
                errors.append({"range": token, "error": f"failed to read sample cells: {exc}"})
                continue

            results.append(
                {
                    "range": f"{_quote_sheet_name_for_a1(sheet_name)}!{cell_range}",
                    "shape": {
                        "rows": rows,
                        "cols": cols,
                        "cells": range_cells,
                        "min_row": min_row,
                        "min_col": min_col,
                        "max_row": max_row,
                        "max_col": max_col,
                    },
                    "truncated": True,
                    "cells": sample_cells,
                }
            )
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return results, errors


def _file_signature(file_path: Path) -> tuple[int, int, int, int]:
    st = file_path.stat()
    ino = int(getattr(st, "st_ino", 0) or 0)
    ctime_ns = getattr(st, "st_ctime_ns", None)
    if not isinstance(ctime_ns, int):
        ctime_ns = int(st.st_ctime * 1e9)
    mtime_ns = getattr(st, "st_mtime_ns", None)
    if not isinstance(mtime_ns, int):
        mtime_ns = int(st.st_mtime * 1e9)
    return ino, int(ctime_ns), int(mtime_ns), int(st.st_size)


def _write_bytes_atomically(file_path: Path, content: bytes, *, mode: int) -> None:
    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{file_path.name}.",
        suffix=file_path.suffix,
        dir=str(file_path.parent),
    )
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "wb") as f:
            fd = -1
            f.write(content)
        try:
            os.chmod(tmp_path, mode)
        except OSError:
            pass
        os.replace(tmp_path, file_path)
        tmp_path = None
    finally:
        if fd != -1:
            try:
                os.close(fd)
            except OSError:
                pass
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass


def _acquire_lockfile(lock_path: Path, timeout_s: float):
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    flags = os.O_RDWR | os.O_CREAT
    nofollow = getattr(os, "O_NOFOLLOW", 0)
    if nofollow:
        flags |= nofollow
    else:
        try:
            if lock_path.is_symlink():
                raise RuntimeError(f"lock file path is a symlink: {lock_path}")
        except OSError:
            pass

    try:
        fd = os.open(lock_path, flags, 0o600)
    except OSError as exc:
        if exc.errno == errno.ELOOP:
            raise RuntimeError(f"lock file path is a symlink: {lock_path}") from None
        raise RuntimeError(f"failed to open lock file: {exc}") from None

    try:
        st = os.fstat(fd)
    except OSError as exc:
        os.close(fd)
        raise RuntimeError(f"failed to stat lock file: {exc}") from None
    if not stat.S_ISREG(st.st_mode):
        os.close(fd)
        raise RuntimeError(f"lock file is not a regular file: {lock_path}") from None

    lock_file = os.fdopen(fd, "r+", encoding="utf-8", errors="replace")
    try:
        import fcntl
    except ImportError:
        lock_file.close()
        raise RuntimeError("file locking requires fcntl (Unix-only)") from None

    try:
        deadline = time.monotonic() + max(0.0, timeout_s)
        while True:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                return lock_file
            except BlockingIOError:
                if time.monotonic() >= deadline:
                    raise TimeoutError(f"timed out waiting for lock: {lock_path}") from None
                time.sleep(0.1)
    except BaseException:
        lock_file.close()
        raise


def _signature_from_stat(st: os.stat_result) -> tuple[int, int, int, int]:
    ino = int(getattr(st, "st_ino", 0) or 0)
    ctime_ns = getattr(st, "st_ctime_ns", None)
    if not isinstance(ctime_ns, int):
        ctime_ns = int(st.st_ctime * 1e9)
    mtime_ns = getattr(st, "st_mtime_ns", None)
    if not isinstance(mtime_ns, int):
        mtime_ns = int(st.st_mtime * 1e9)
    return ino, int(ctime_ns), int(mtime_ns), int(st.st_size)


def _read_bytes_and_signature_nofollow_sync(
    file_path: Path,
    *,
    max_bytes: Optional[int] = None,
) -> tuple[tuple[int, int, int, int], bytes]:
    flags = os.O_RDONLY
    nofollow = getattr(os, "O_NOFOLLOW", 0)
    if nofollow:
        flags |= nofollow
    else:
        try:
            if file_path.is_symlink():
                raise RuntimeError("workbook path is a symlink")
        except OSError:
            raise RuntimeError("failed to check workbook path") from None

    try:
        fd = os.open(file_path, flags)
    except OSError as exc:
        if exc.errno == errno.ELOOP:
            raise RuntimeError("workbook path is a symlink") from None
        raise RuntimeError(f"failed to open workbook: {exc}") from None

    try:
        st = os.fstat(fd)
        if not stat.S_ISREG(st.st_mode):
            raise RuntimeError("workbook is not a regular file")
        max_bytes_int: Optional[int]
        if max_bytes is None:
            max_bytes_int = None
        else:
            try:
                max_bytes_int = int(max_bytes)
            except (TypeError, ValueError):
                max_bytes_int = None
            if max_bytes_int is not None and max_bytes_int <= 0:
                max_bytes_int = None
        if max_bytes_int is not None:
            size = int(getattr(st, "st_size", 0) or 0)
            if size > max_bytes_int:
                raise RuntimeError(f"workbook is too large ({size} bytes > {max_bytes_int} bytes)")
        sig = _signature_from_stat(st)
        with os.fdopen(fd, "rb") as f:
            fd = -1
            if max_bytes_int is None:
                data = f.read()
            else:
                data = f.read(max_bytes_int + 1)
                if len(data) > max_bytes_int:
                    raise RuntimeError(f"workbook is too large (> {max_bytes_int} bytes)")
        return sig, data
    finally:
        if fd != -1:
            try:
                os.close(fd)
            except OSError:
                pass


def _read_workbook_bytes_under_lock_sync(
    *,
    file_path: Path,
    lock_timeout_s: float,
    max_bytes: Optional[int] = None,
) -> tuple[tuple[int, int, int, int], bytes]:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    try:
        return _read_bytes_and_signature_nofollow_sync(file_path, max_bytes=max_bytes)
    finally:
        try:
            lock_file.close()
        except Exception:
            pass


def _read_workbook_signature_under_lock_sync(*, file_path: Path, lock_timeout_s: float) -> tuple[int, int, int, int]:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    try:
        return _file_signature(file_path)
    finally:
        try:
            lock_file.close()
        except Exception:
            pass


def _write_recalc_under_lock_sync(
    *,
    file_path: Path,
    expected_sig: Optional[tuple[int, int, int, int]],
    recalc_content: bytes,
    lock_timeout_s: float,
) -> tuple[bool, Optional[str]]:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    try:
        if expected_sig is None:
            return False, "missing expected file signature"

        try:
            current_sig = _file_signature(file_path)
        except OSError as exc:
            return False, f"failed to stat workbook before recalc writeback: {exc}"
        if current_sig != expected_sig:
            return False, "workbook changed since recalc request; skipping writeback"

        try:
            mode = stat.S_IMODE(file_path.stat().st_mode)
        except OSError:
            mode = 0o644
        try:
            _write_bytes_atomically(file_path, recalc_content, mode=mode)
        except OSError as exc:
            return False, f"recalc writeback failed: {exc}"

        return True, None
    finally:
        try:
            lock_file.close()
        except Exception:
            pass


def _post_recalculate(
    *,
    url: str,
    timeout_s: float,
    file_path: Optional[Path] = None,
    file_bytes: Optional[bytes] = None,
    filename: Optional[str] = None,
    max_response_bytes: Optional[int] = None,
) -> bytes:
    if (file_path is None) == (file_bytes is None):
        raise ValueError("provide exactly one of file_path or file_bytes")

    if max_response_bytes is None:
        max_response_bytes = 200 * 1024 * 1024
    try:
        max_response_bytes = int(max_response_bytes)
    except (TypeError, ValueError):
        max_response_bytes = 200 * 1024 * 1024
    max_response_bytes = min(max(1, max_response_bytes), 512 * 1024 * 1024)
    error_read_limit = min(max_response_bytes, 1024 * 1024)

    if filename is None:
        if file_path is not None:
            filename = file_path.name
        else:
            filename = "workbook.xlsx"
    filename = filename.replace('"', "_").replace("\r", "_").replace("\n", "_")
    mime_type = mimetypes.guess_type(filename)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    boundary = uuid.uuid4().hex

    header = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        f"Content-Type: {mime_type}\r\n"
        "\r\n"
    ).encode()
    footer = f"\r\n--{boundary}--\r\n\r\n".encode()

    body = bytearray()
    body.extend(header)
    if file_bytes is not None:
        body.extend(file_bytes)
    else:
        if file_path is None:
            raise ValueError("file_path is required when file_bytes is not provided")
        flags = os.O_RDONLY
        nofollow = getattr(os, "O_NOFOLLOW", 0)
        if nofollow:
            flags |= nofollow
        try:
            fd = os.open(file_path, flags)
        except OSError as exc:
            if exc.errno == errno.ELOOP:
                raise RuntimeError("workbook path is a symlink") from None
            raise RuntimeError(f"failed to open workbook: {exc}") from None
        try:
            st = os.fstat(fd)
            if not stat.S_ISREG(st.st_mode):
                raise RuntimeError("workbook is not a regular file")
            size = int(getattr(st, "st_size", 0) or 0)
            if size > max_response_bytes:
                raise RuntimeError(f"workbook is too large ({size} bytes > {max_response_bytes} bytes)")
            with os.fdopen(fd, "rb") as f:
                fd = -1
                while True:
                    chunk = f.read(1024 * 1024)
                    if not chunk:
                        break
                    body.extend(chunk)
        finally:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
    body.extend(footer)

    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    req.add_header("Content-Length", str(len(body)))

    deadline_s = time.monotonic() + float(timeout_s)
    submit_retry_interval_s = _get_submit_retry_interval_s()

    job_id: str
    submit_parts = urllib.parse.urlsplit(url)
    if submit_parts.scheme not in {"http", "https"} or not submit_parts.netloc:
        raise RuntimeError("recalc_url must be an http(s) URL")

    def _origin_key(parts: urllib.parse.SplitResult) -> Optional[tuple[str, str, int]]:
        scheme = (parts.scheme or "").lower()
        if scheme not in {"http", "https"}:
            return None
        host = parts.hostname
        if not host:
            return None
        try:
            port = parts.port
        except ValueError:
            return None
        if port is None:
            port = 443 if scheme == "https" else 80
        return scheme, host.casefold(), port

    submit_origin = _origin_key(submit_parts)
    if submit_origin is None:
        raise RuntimeError("recalc_url must be an http(s) URL")

    class _SafeRedirectHandler(urllib.request.HTTPRedirectHandler):
        def redirect_request(self, req, fp, code, msg, headers, newurl):
            try:
                joined = urllib.parse.urljoin(req.full_url, newurl)
            except Exception:
                return None
            if _origin_key(urllib.parse.urlsplit(joined)) != submit_origin:
                return None
            return super().redirect_request(req, fp, code, msg, headers, joined)

    opener = urllib.request.build_opener(_SafeRedirectHandler())

    while True:
        remaining_s = deadline_s - time.monotonic()
        if remaining_s <= 0:
            raise TimeoutError("recalculate timed out during submit")

        submit_timeout_s = remaining_s
        try:
            if submit_parts.path.rstrip("/").endswith("/submit"):
                submit_timeout_s = min(30.0, remaining_s)
        except Exception:
            submit_timeout_s = remaining_s

        try:
            with opener.open(req, timeout=submit_timeout_s) as resp:
                if not _is_success_status(resp.status):
                    time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
                    continue
                content_type = (resp.headers.get("Content-Type") or "").lower()
                resp_bytes = resp.read(max_response_bytes + 1)
                if len(resp_bytes) > max_response_bytes:
                    raise RuntimeError("recalculate response too large")
        except urllib.error.HTTPError as exc:
            status = int(getattr(exc, "code", 0) or 0)
            if 300 <= status < 400:
                raise RuntimeError(f"recalculate submit returned redirect (HTTP {status})") from None
            if 400 <= status < 500 and status != 429:
                body = ""
                try:
                    raw_body = exc.read(error_read_limit + 1) or b""
                    if len(raw_body) > error_read_limit:
                        raw_body = raw_body[:error_read_limit]
                    body = raw_body.decode("utf-8", errors="replace").strip()
                except Exception:
                    body = ""
                if body:
                    body = body[:2000]
                    raise RuntimeError(f"recalculate submit failed with HTTP {status}: {body}") from None
                raise RuntimeError(f"recalculate submit failed with HTTP {status}") from None
            time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
            continue
        except (urllib.error.URLError, OSError, TimeoutError):
            time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
            continue

        if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content_type:
            return resp_bytes
        if resp_bytes.startswith(b"PK\x03\x04"):
            return resp_bytes

        try:
            payload = json.loads(resp_bytes.decode("utf-8", errors="replace"))
        except ValueError:
            time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
            continue
        if not isinstance(payload, dict):
            time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
            continue

        if payload.get("status") == "busy":
            time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
            continue

        job_id_raw = payload.get("job_id")
        if not isinstance(job_id_raw, str) or not job_id_raw:
            time.sleep(min(submit_retry_interval_s, max(0.0, remaining_s)))
            continue
        job_id = job_id_raw
        break

    result_path = payload.get("result_path")
    fallback_result_path = f"/recalculate/result/{job_id}"
    raw_result_path = result_path if isinstance(result_path, str) and result_path else fallback_result_path
    parsed_result_path = urllib.parse.urlsplit(raw_result_path)
    if parsed_result_path.scheme or parsed_result_path.netloc or raw_result_path.startswith("//"):
        result_url = urllib.parse.urljoin(url, raw_result_path)
        result_parts = urllib.parse.urlsplit(result_url)
        if _origin_key(result_parts) != submit_origin:
            raise RuntimeError("unsafe result URL host")
    else:
        raw_path = parsed_result_path.path or fallback_result_path
        path_parts = [part for part in submit_parts.path.split("/") if part]
        prefix_parts: list[str] = []
        for idx, part in enumerate(path_parts):
            if part.lower() == "recalculate":
                prefix_parts = path_parts[:idx]
                break
        prefix = f"/{'/'.join(prefix_parts)}" if prefix_parts else ""
        if raw_path.startswith("/"):
            if prefix and raw_path.startswith(f"{prefix}/"):
                full_path = raw_path
            elif prefix:
                full_path = f"{prefix}{raw_path}"
            else:
                full_path = raw_path
        elif prefix:
            full_path = f"{prefix}/{raw_path}"
        else:
            full_path = f"/{raw_path}"
        result_url = urllib.parse.urlunsplit(
            (
                submit_parts.scheme,
                submit_parts.netloc,
                full_path,
                parsed_result_path.query,
                parsed_result_path.fragment,
            )
        )

    poll_wait_s = _get_poll_wait_s()
    backoff_s = 0.5
    while True:
        remaining_s = deadline_s - time.monotonic()
        if remaining_s <= 0:
            raise TimeoutError("recalculate timed out while polling")

        wait_s = min(poll_wait_s, max(0.0, remaining_s))
        poll_url = _add_query_params(result_url, {"wait_s": f"{wait_s:.3f}"})
        poll_timeout_s = min(remaining_s, max(5.0, wait_s + 15.0))
        poll_parts = urllib.parse.urlsplit(poll_url)
        if _origin_key(poll_parts) != submit_origin:
            raise RuntimeError("unsafe poll URL")
        try:
            t0 = time.monotonic()
            with opener.open(poll_url, timeout=poll_timeout_s) as resp:
                if not _is_success_status(resp.status):
                    raise RuntimeError(f"recalculate poll failed with HTTP {resp.status}")
                content_type = (resp.headers.get("Content-Type") or "").lower()
                resp_bytes = resp.read(max_response_bytes + 1)
                if len(resp_bytes) > max_response_bytes:
                    raise RuntimeError("recalculate response too large")
            poll_elapsed_s = time.monotonic() - t0
        except urllib.error.HTTPError as exc:
            status = int(getattr(exc, "code", 0) or 0)
            if 300 <= status < 400:
                raise RuntimeError(f"recalculate poll returned redirect (HTTP {status})") from None
            if 400 <= status < 500 and status != 429:
                body = ""
                try:
                    raw_body = exc.read(error_read_limit + 1) or b""
                    if len(raw_body) > error_read_limit:
                        raw_body = raw_body[:error_read_limit]
                    body = raw_body.decode("utf-8", errors="replace").strip()
                except Exception:
                    body = ""
                if body:
                    body = body[:2000]
                    raise RuntimeError(f"recalculate poll failed with HTTP {status}: {body}") from None
                raise RuntimeError(f"recalculate poll failed with HTTP {status}") from None
            sleep_s = min(backoff_s, max(0.0, remaining_s))
            time.sleep(sleep_s)
            backoff_s = min(backoff_s * 2, 5.0)
            continue
        except (urllib.error.URLError, OSError, TimeoutError):
            sleep_s = min(backoff_s, max(0.0, remaining_s))
            time.sleep(sleep_s)
            backoff_s = min(backoff_s * 2, 5.0)
            continue

        if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content_type:
            return resp_bytes
        if resp_bytes.startswith(b"PK\x03\x04"):
            return resp_bytes

        try:
            poll_payload = json.loads(resp_bytes.decode("utf-8", errors="replace"))
        except ValueError:
            sleep_s = min(backoff_s, max(0.0, remaining_s))
            time.sleep(sleep_s)
            backoff_s = min(backoff_s * 2, 5.0)
            continue
        if not isinstance(poll_payload, dict):
            sleep_s = min(backoff_s, max(0.0, remaining_s))
            time.sleep(sleep_s)
            backoff_s = min(backoff_s * 2, 5.0)
            continue

        status = poll_payload.get("status")
        if status == "error":
            msg = poll_payload.get("msg", "")
            raise RuntimeError(str(msg or "recalculate failed"))
        if status == "done":
            raise RuntimeError("recalculate poll returned JSON done without a file")

        if poll_wait_s <= 0 or poll_elapsed_s < min(0.5, wait_s):
            sleep_s = min(backoff_s, max(0.0, remaining_s))
            time.sleep(sleep_s)
            backoff_s = min(backoff_s * 2, 5.0)


def _recalculate_workbook_for_commit(
    *,
    url: str,
    file_path: Path,
    filename: Optional[str] = None,
    timeout_s: float,
    max_response_bytes: Optional[int] = None,
) -> bytes:
    recalc_content = _post_recalculate(
        url=url,
        file_path=file_path,
        filename=filename,
        timeout_s=timeout_s,
        max_response_bytes=max_response_bytes,
    )
    zip_error = _scan_zip_metadata_bytes(
        recalc_content,
        max_members=50_000,
        max_total_uncompressed_bytes=512 * 1024 * 1024,
        max_member_uncompressed_bytes=128 * 1024 * 1024,
        max_ratio=200.0,
    )
    if zip_error:
        raise RuntimeError(f"recalculated workbook rejected by zip safety checks: {zip_error}")
    xlsx_error = _scan_xlsx_container_bytes(recalc_content)
    if xlsx_error:
        raise RuntimeError(f"recalculated workbook rejected by xlsx safety checks: {xlsx_error}")
    load_error = _scan_xlsx_loadable_bytes(recalc_content)
    if load_error:
        raise RuntimeError(f"recalculated workbook rejected by xlsx load checks: {load_error}")
    return recalc_content


class RecalculateAndReadTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self.recalc_url = (
            str(config.get("recalc_url") or "").strip()
            or os.environ.get("SHEET_ARENA_RECALC_URL", "").strip()
            or "http://127.0.0.1:5000/recalculate"
        )
        timeout_raw = config.get("timeout_s", 180)
        try:
            timeout_s = float(timeout_raw)
        except (TypeError, ValueError):
            timeout_s = 180.0
        if not math.isfinite(timeout_s):
            timeout_s = 180.0
        self.timeout_s = max(0.0, timeout_s)

        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)
        self.max_response_chars = _get_max_response_chars(config)
        self._instance_dict: dict[str, dict[str, Any]] = {}

    def get_openai_tool_schema(self) -> OpenAIFunctionToolSchema:
        return self.tool_schema

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            rel = Path("data.xlsx")
        else:
            rel = _sanitize_relpath(raw_path)
            if rel is None:
                return ToolResponse(text="Error: invalid path."), 0.0, {"status": "error", "error": "invalid_path"}
        if rel.suffix.lower() != ".xlsx":
            return (
                ToolResponse(text="Error: only .xlsx workbooks are supported."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_path",
                },
            )

        cell_ranges_raw = parameters.get("cell_ranges")
        if not isinstance(cell_ranges_raw, list):
            return (
                ToolResponse(text="Error: 'cell_ranges' must be a list of A1 ranges (e.g. ['A1', 'B2:C3'])."),
                0.0,
                {"status": "error", "error": "invalid_cell_ranges"},
            )
        if any(not isinstance(item, str) or not item.strip() for item in cell_ranges_raw):
            return (
                ToolResponse(text="Error: 'cell_ranges' must contain non-empty strings."),
                0.0,
                {"status": "error", "error": "invalid_cell_ranges"},
            )
        cell_ranges: list[str] = [item.strip() for item in cell_ranges_raw]

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return (
                ToolResponse(text="Error: workspace_id is missing/invalid."),
                0.0,
                {
                    "status": "error",
                    "error": "missing_workspace_id",
                },
            )
        target_file = _resolve_workspace_file(
            workspace_id=workspace_id,
            relpath=rel,
        )
        if target_file is None:
            return (
                ToolResponse(text=f"Error: file not found: {rel}"),
                0.0,
                {"status": "error", "error": "file_not_found"},
            )

        if self.max_file_size_bytes is not None:
            try:
                file_size = target_file.stat().st_size
            except OSError as exc:
                return (
                    ToolResponse(text=f"Error: failed to stat file: {exc}"),
                    0.0,
                    {
                        "status": "error",
                        "error": "stat_failed",
                    },
                )
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return (
                    ToolResponse(text=f"Error: workbook is too large ({actual_mb:.1f}MB > {max_mb}MB)."),
                    0.0,
                    {
                        "status": "error",
                        "error": "file_too_large",
                    },
                )

        if not self.recalc_url:
            return (
                ToolResponse(text="Error: recalc_url is not configured."),
                0.0,
                {
                    "status": "error",
                    "error": "missing_url",
                },
            )

        try:
            expected_sig, request_bytes = await asyncio.to_thread(
                _read_workbook_bytes_under_lock_sync,
                file_path=target_file,
                lock_timeout_s=self.lock_timeout_s,
                max_bytes=self.max_file_size_bytes,
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return (
                ToolResponse(text=f"Error: failed to read workbook: {exc}"),
                0.0,
                {
                    "status": "error",
                    "error": "stat_failed",
                },
            )

        zip_error = await asyncio.to_thread(
            _scan_zip_metadata_bytes,
            request_bytes,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            return (
                ToolResponse(text=f"Error: workbook rejected by zip safety checks: {zip_error}"),
                0.0,
                {
                    "status": "error",
                    "error": "zip_limits_exceeded",
                },
            )

        try:
            content = await asyncio.to_thread(
                _post_recalculate,
                url=self.recalc_url,
                file_bytes=request_bytes,
                filename=target_file.name,
                timeout_s=self.timeout_s,
                max_response_bytes=self.max_file_size_bytes,
            )
        except (urllib.error.URLError, urllib.error.HTTPError, OSError, RuntimeError, TimeoutError, ValueError) as exc:
            return (
                ToolResponse(text=f"Error: recalculate request failed: {exc}"),
                0.0,
                {
                    "status": "error",
                    "error": "request_failed",
                },
            )

        try:
            read_values, read_errors = await asyncio.to_thread(
                _read_data_only_values,
                content=content,
                cell_ranges=cell_ranges,
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return (
                ToolResponse(text=f"Error: failed to read recalculated values: {exc}"),
                0.0,
                {
                    "status": "error",
                    "error": "read_failed",
                },
            )

        updated = False
        writeback_error = None
        try:
            updated, writeback_error = await asyncio.to_thread(
                _write_recalc_under_lock_sync,
                file_path=target_file,
                expected_sig=expected_sig,
                recalc_content=content,
                lock_timeout_s=self.lock_timeout_s,
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            updated = False
            writeback_error = f"failed to write updated file: {exc}"

        values_source = "recalc_response"
        if (
            writeback_error is not None
            and not updated
            and "workbook changed since recalc request" in str(writeback_error)
        ):
            try:
                _, live_bytes = await asyncio.to_thread(
                    _read_workbook_bytes_under_lock_sync,
                    file_path=target_file,
                    lock_timeout_s=self.lock_timeout_s,
                    max_bytes=self.max_file_size_bytes,
                )
                live_values, live_errors = await asyncio.to_thread(
                    _read_data_only_values,
                    content=live_bytes,
                    cell_ranges=cell_ranges,
                )
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                writeback_error = f"{writeback_error}; failed to read live workbook: {exc}"
            else:
                read_values = live_values
                read_errors = live_errors
                values_source = "live_workbook"

        status = "success" if not read_errors else ("partial_success" if read_values else "error")
        if writeback_error is not None and status == "success":
            status = "partial_success"
        payload = {
            "status": status,
            "file": str(rel),
            "updated": bool(updated),
            "values_source": values_source,
            "truncated": False,
            "values": read_values,
            "errors": read_errors,
        }
        if writeback_error is not None:
            payload["writeback_error"] = _to_jsonable_excel_value(writeback_error)
        response_text = _json_dumps_compact(payload)
        if len(response_text) > self.max_response_chars:
            payload["truncated"] = True
            trimmed_values: list[dict[str, Any]] = []
            for entry in read_values:
                if not isinstance(entry, dict) or entry.get("truncated") is True:
                    trimmed_values.append(entry)
                    continue
                shape = entry.get("shape") if isinstance(entry.get("shape"), dict) else {}
                sheet_name = entry.get("sheet")
                if not isinstance(sheet_name, str) or not sheet_name:
                    sheet_name = entry.get("range", "")
                    if isinstance(sheet_name, str) and "!" in sheet_name:
                        sheet_name = sheet_name.rsplit("!", 1)[0].strip()
                        if sheet_name.startswith("'") and sheet_name.endswith("'"):
                            sheet_name = sheet_name[1:-1].replace("''", "'")
                values = entry.get("values")
                if (
                    isinstance(sheet_name, str)
                    and isinstance(values, list)
                    and isinstance(shape.get("rows"), int)
                    and isinstance(shape.get("cols"), int)
                    and isinstance(shape.get("min_row"), int)
                    and isinstance(shape.get("min_col"), int)
                ):
                    rows = int(shape["rows"])
                    cols = int(shape["cols"])
                    min_row = int(shape["min_row"])
                    min_col = int(shape["min_col"])
                    sample_indices = _sample_linear_indices(rows * cols, head=3, tail=3)
                    sample_cells: list[dict[str, Any]] = []
                    for idx in sample_indices:
                        r_off = idx // cols
                        c_off = idx - r_off * cols
                        row_num = min_row + r_off
                        col_num = min_col + c_off
                        value_obj = None
                        try:
                            value_obj = values[r_off][c_off]
                        except Exception:
                            value_obj = None
                        sample_cells.append(
                            {
                                "address": _format_a1_address(sheet_name=sheet_name, row=row_num, col=col_num),
                                "value": value_obj,
                            }
                        )
                    trimmed_values.append(
                        {
                            "range": entry.get("range"),
                            "shape": shape,
                            "truncated": True,
                            "cells": sample_cells,
                        }
                    )
                else:
                    trimmed_values.append(
                        {
                            "range": entry.get("range"),
                            "shape": entry.get("shape"),
                            "truncated": True,
                            "cells": entry.get("cells", []),
                        }
                    )

            payload["values"] = trimmed_values
            response_text = _json_dumps_compact(payload)
            if len(response_text) > self.max_response_chars:
                payload.pop("errors", None)
                response_text = _json_dumps_compact(payload)
            if len(response_text) > self.max_response_chars:
                values_list = payload.get("values")
                if isinstance(values_list, list) and len(values_list) > 1:
                    for head, tail in ((2, 2), (1, 1)):
                        if len(values_list) <= head + tail:
                            continue
                        values_list = values_list[:head] + values_list[-tail:]
                        payload["values"] = values_list
                        response_text = _json_dumps_compact(payload)
                        if len(response_text) <= self.max_response_chars:
                            break
                if len(response_text) > self.max_response_chars:
                    values_list = payload.get("values")
                    if isinstance(values_list, list) and len(values_list) > 1:
                        payload["values"] = values_list[:1]
                        response_text = _json_dumps_compact(payload)
                if len(response_text) > self.max_response_chars:
                    values_list = payload.get("values")
                    if isinstance(values_list, list):
                        for entry in values_list:
                            if not isinstance(entry, dict):
                                continue
                            cells = entry.get("cells")
                            if not isinstance(cells, list) or len(cells) <= 1:
                                continue
                            for head, tail in ((2, 2), (1, 1)):
                                if len(cells) <= head + tail:
                                    continue
                                entry["cells"] = cells[:head] + cells[-tail:]
                                cells = entry.get("cells")
                                response_text = _json_dumps_compact(payload)
                                if len(response_text) <= self.max_response_chars:
                                    break
                            if len(response_text) > self.max_response_chars:
                                cells = entry.get("cells")
                                if isinstance(cells, list) and len(cells) > 1:
                                    entry["cells"] = cells[:1]
                                    response_text = _json_dumps_compact(payload)
                            if len(response_text) <= self.max_response_chars:
                                break

                    if len(response_text) > self.max_response_chars and isinstance(values_list, list):
                        for entry in values_list:
                            if not isinstance(entry, dict):
                                continue
                            cells = entry.get("cells")
                            if not isinstance(cells, list):
                                continue
                            for cell in cells:
                                if isinstance(cell, dict):
                                    cell.pop("value", None)
                        response_text = _json_dumps_compact(payload)

                    if len(response_text) > self.max_response_chars and isinstance(values_list, list):
                        for entry in values_list:
                            if not isinstance(entry, dict):
                                continue
                            cells = entry.get("cells")
                            if not isinstance(cells, list) or not cells:
                                continue
                            addr = cells[0].get("address") if isinstance(cells[0], dict) else None
                            entry["cells"] = [{"address": addr}] if isinstance(addr, str) and addr else []
                        response_text = _json_dumps_compact(payload)
        metrics = {
            "status": status,
            "file": str(rel),
            "ranges": len(cell_ranges),
            "values_source": values_source,
            "values_read": len(read_values),
            "errors": len(read_errors),
            "updated": bool(updated),
            "truncated": bool(payload.get("truncated")),
        }
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
