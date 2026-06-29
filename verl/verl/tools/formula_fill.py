from __future__ import annotations

import asyncio
import datetime as dt
import errno
import io
import json
import math
import os
import re
import shutil
import stat
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import get_sheet_arena_data_root, normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .formula_fill_core import fill_formula, normalize_formula_for_excel
from .recalculate import _post_recalculate, _scan_zip_metadata_bytes
from .schemas import OpenAIFunctionToolSchema, ToolResponse

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384

_A1_CELL_RE = re.compile(r"^([A-Z]{1,3})([0-9]{1,7})$")
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z0-9_]+$")
DEFAULT_MAX_RESPONSE_CHARS = 4096
MAX_RESPONSE_CHARS = 8192


def _col_letter_to_index(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def _parse_a1_cell(value: str) -> tuple[str, int]:
    cell = value.strip().upper().replace("$", "")
    match = _A1_CELL_RE.fullmatch(cell)
    if match is None:
        raise ValueError(f"invalid A1 cell: {value!r}")
    col = match.group(1)
    row = int(match.group(2))
    if not 1 <= row <= _EXCEL_MAX_ROWS:
        raise ValueError(f"row out of bounds: {row}")
    col_idx = _col_letter_to_index(col)
    if not 1 <= col_idx <= _EXCEL_MAX_COLS:
        raise ValueError(f"column out of bounds: {col}")
    return col, row


def _normalize_sheet_name(value: str) -> str:
    name = value.strip(' \t\n\r\f\v"')
    if len(name) >= 2 and name[0] == "'" and name[-1] == "'":
        name = name[1:-1].replace("''", "'")
        return name.strip()

    if name.startswith("'") and not name.endswith("'"):
        name = name[1:]
    elif name.endswith("'") and not name.startswith("'"):
        name = name[:-1]
    return name.strip()


def _quote_sheet_name_for_a1(sheet_name: str) -> str:
    if _SAFE_SHEET_NAME_RE.fullmatch(sheet_name):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _split_sheet_cell(value: str) -> tuple[Optional[str], str]:
    token = value.strip(' \t\n\r\f\v"')
    if not token:
        raise ValueError("empty cell")
    if "!" not in token:
        return None, token
    sheet_part, _, cell_part = token.rpartition("!")
    sheet_name = _normalize_sheet_name(sheet_part)
    if not sheet_name:
        raise ValueError(f"empty sheet name: {value!r}")
    if not cell_part.strip():
        raise ValueError(f"empty cell reference: {value!r}")
    return sheet_name, cell_part


def _sanitize_relpath(value: Any) -> Optional[Path]:
    if not isinstance(value, str):
        return None
    raw = value.strip()
    if not raw:
        return None
    if "\0" in raw:
        return None
    p = Path(raw)
    if p.is_absolute():
        return None
    if any(part == ".." for part in p.parts):
        return None
    return p


def _resolve_workspace_file(*, workspace_id: Optional[str], relpath: Path) -> Optional[Path]:
    if not workspace_id:
        return None

    workspace_base = get_sheet_arena_data_root()
    workspaces_base = workspace_base / "_workspaces"
    try:
        if workspace_base.is_symlink():
            return None
        workspaces_base_resolved = workspaces_base.resolve(strict=True)
    except (OSError, RuntimeError):
        return None

    workspace_root = workspaces_base / workspace_id
    try:
        if workspace_root.is_symlink():
            return None
        workspace_root_resolved = workspace_root.resolve(strict=True)
    except (OSError, RuntimeError):
        return None
    if not workspace_root_resolved.is_relative_to(workspaces_base_resolved):
        return None

    rel_candidates = [relpath]
    if relpath.parts and relpath.parts[0] == "data" and len(relpath.parts) > 1:
        rel_candidates.append(Path(*relpath.parts[1:]))

    for rel_candidate in rel_candidates:
        candidate = workspace_root / rel_candidate
        try:
            resolved = candidate.resolve(strict=True)
        except (OSError, RuntimeError):
            continue
        if not resolved.is_relative_to(workspace_root_resolved):
            continue

        if candidate.is_symlink():
            continue

        cursor = workspace_root
        safe = True
        for part in rel_candidate.parts:
            cursor = cursor / part
            try:
                if cursor.is_symlink():
                    safe = False
                    break
            except OSError:
                safe = False
                break
        if not safe:
            continue

        if resolved.is_file():
            return resolved
    return None


def _resolve_target_worksheet(wb, requested_name: str):
    worksheets = getattr(wb, "worksheets", None) or []
    for ws in worksheets:
        if getattr(ws, "title", None) == requested_name:
            return ws

    requested_cf = requested_name.casefold()
    matches = [ws for ws in worksheets if getattr(ws, "title", "").casefold() == requested_cf]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        candidates = ", ".join(repr(getattr(ws, "title", "")) for ws in matches[:5])
        raise RuntimeError(f"ambiguous sheet name: {requested_name!r} matches {candidates}")

    sheetnames = getattr(wb, "sheetnames", None) or []
    if any(isinstance(name, str) and name.casefold() == requested_cf for name in sheetnames):
        raise RuntimeError(f"sheet is not a worksheet: {requested_name!r}")
    raise RuntimeError(f"sheet not found: {requested_name!r}")


def _get_max_string_chars() -> int:
    raw = os.environ.get("SHEET_ARENA_TOOL_MAX_STRING_CHARS", "200").strip()
    try:
        n = int(raw)
        return min(max(16, n), 100_000)
    except ValueError:
        return 200


def _get_max_response_chars(config: Optional[dict[str, Any]] = None) -> int:
    config_value = None
    if isinstance(config, dict):
        config_value = config.get("max_response_chars")
    for raw in (config_value, os.environ.get("SHEET_ARENA_TOOL_MAX_RESPONSE_CHARS", "").strip()):
        if raw is None:
            continue
        try:
            n = int(raw)
            return min(max(128, n), MAX_RESPONSE_CHARS)
        except (TypeError, ValueError):
            continue
    return DEFAULT_MAX_RESPONSE_CHARS


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _truncate_str(value: str, max_chars: int) -> str:
    if max_chars <= 0:
        return ""
    if len(value) <= max_chars:
        return value
    if max_chars <= 3:
        return value[:max_chars]
    return value[: max_chars - 3] + "..."


def _to_jsonable_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, str)):
        if isinstance(value, str):
            return _truncate_str(value, _get_max_string_chars())
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return str(value)


def _format_a1_address(*, sheet_name: str, row: int, col: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        col_letter = str(col)
    else:
        col_letter = get_column_letter(col)
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    return f"{sheet_ref}!{col_letter}{row}"


def _sample_indices(total: int, *, head: int, tail: int) -> list[int]:
    if total <= 0:
        return []
    head_n = max(0, head)
    tail_n = max(0, tail)
    if total <= head_n + tail_n:
        return list(range(total))
    indices = list(range(head_n))
    indices.extend(range(total - tail_n, total))
    return indices


def _normalize_formula_template_for_excel(formula: str) -> str:
    return normalize_formula_for_excel(formula)


def _file_signature(file_path: Path) -> tuple[int, int, int, int]:
    st = file_path.stat()
    ino = int(getattr(st, "st_ino", 0) or 0)
    ctime_ns = getattr(st, "st_ctime_ns", None)
    if not isinstance(ctime_ns, int):
        ctime_ns = int(st.st_ctime * 1e9)
    mtime_ns = getattr(st, "st_mtime_ns", None)
    if not isinstance(mtime_ns, int):
        mtime_ns = int(st.st_mtime * 1e9)
    return ino, int(ctime_ns), int(mtime_ns), int(st.st_size)


def _write_bytes_atomically(file_path: Path, content: bytes, *, mode: int) -> None:
    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{file_path.name}.",
        suffix=file_path.suffix,
        dir=str(file_path.parent),
    )
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "wb") as f:
            fd = -1
            f.write(content)
        try:
            os.chmod(tmp_path, mode)
        except OSError:
            pass
        os.replace(tmp_path, file_path)
        tmp_path = None
    finally:
        if fd != -1:
            try:
                os.close(fd)
            except OSError:
                pass
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass


def _read_cells_formula_and_value_from_bytes(
    *,
    content: bytes,
    sheet_name: str,
    coords: list[tuple[int, int]],
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required; install openpyxl>=3.1.5") from None

    wb_formula = None
    wb_values = None
    try:
        wb_formula = load_workbook(filename=io.BytesIO(content), data_only=False, read_only=True, keep_links=False)
        ws_formula = wb_formula[sheet_name]
        wb_values = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True, keep_links=False)
        ws_values = wb_values[sheet_name]

        out: list[dict[str, Any]] = []
        for row_num, col_num in coords:
            cell_formula = None
            try:
                cell = ws_formula.cell(row=row_num, column=col_num)
                if getattr(cell, "data_type", None) == "f":
                    raw = getattr(cell, "value", None)
                    if isinstance(raw, str) and raw:
                        cell_formula = raw if raw.startswith("=") else f"={raw}"
            except Exception:
                cell_formula = None
            if isinstance(cell_formula, str):
                cell_formula = _truncate_str(cell_formula, _get_max_string_chars())

            cell_value = None
            try:
                cell_value = ws_values.cell(row=row_num, column=col_num).value
            except Exception:
                cell_value = None

            out.append(
                {
                    "address": _format_a1_address(sheet_name=sheet_name, row=row_num, col=col_num),
                    "formula": cell_formula,
                    "value": _to_jsonable_excel_value(cell_value),
                }
            )
        return out
    finally:
        for wb in (wb_values, wb_formula):
            if wb is None:
                continue
            try:
                wb.close()
            except Exception:
                pass


def _read_cells_formula_and_value_from_path(
    *,
    file_path: Path,
    sheet_name: str,
    coords: list[tuple[int, int]],
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required; install openpyxl>=3.1.5") from None

    zip_error = _scan_zip_metadata(
        file_path,
        max_members=50_000,
        max_total_uncompressed_bytes=512 * 1024 * 1024,
        max_member_uncompressed_bytes=128 * 1024 * 1024,
        max_ratio=200.0,
    )
    if zip_error:
        raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

    wb_formula = None
    wb_values = None
    try:
        wb_formula = load_workbook(filename=str(file_path), data_only=False, read_only=True, keep_links=False)
        ws_formula = wb_formula[sheet_name]
        wb_values = load_workbook(filename=str(file_path), data_only=True, read_only=True, keep_links=False)
        ws_values = wb_values[sheet_name]

        out: list[dict[str, Any]] = []
        for row_num, col_num in coords:
            cell_formula = None
            try:
                cell = ws_formula.cell(row=row_num, column=col_num)
                if getattr(cell, "data_type", None) == "f":
                    raw = getattr(cell, "value", None)
                    if isinstance(raw, str) and raw:
                        cell_formula = raw if raw.startswith("=") else f"={raw}"
            except Exception:
                cell_formula = None
            if isinstance(cell_formula, str):
                cell_formula = _truncate_str(cell_formula, _get_max_string_chars())

            cell_value = None
            try:
                cell_value = ws_values.cell(row=row_num, column=col_num).value
            except Exception:
                cell_value = None

            out.append(
                {
                    "address": _format_a1_address(sheet_name=sheet_name, row=row_num, col=col_num),
                    "formula": cell_formula,
                    "value": _to_jsonable_excel_value(cell_value),
                }
            )
        return out
    finally:
        for wb in (wb_values, wb_formula):
            if wb is None:
                continue
            try:
                wb.close()
            except Exception:
                pass


def _scan_zip_metadata(
    file_path: Path,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(file_path) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={getattr(info, 'filename', '?')!r}, bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={getattr(info, 'filename', '?')!r}, ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        f"zip contents too large (total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read zip metadata: {exc}"

    return None


def _acquire_lockfile(lock_path: Path, timeout_s: float):
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    flags = os.O_RDWR | os.O_CREAT
    nofollow = getattr(os, "O_NOFOLLOW", 0)
    if nofollow:
        flags |= nofollow
    else:
        try:
            if lock_path.is_symlink():
                raise RuntimeError(f"lock file path is a symlink: {lock_path}")
        except OSError:
            pass

    try:
        fd = os.open(lock_path, flags, 0o600)
    except OSError as exc:
        if exc.errno == errno.ELOOP:
            raise RuntimeError(f"lock file path is a symlink: {lock_path}") from None
        raise RuntimeError(f"failed to open lock file: {exc}") from None

    try:
        st = os.fstat(fd)
    except OSError as exc:
        os.close(fd)
        raise RuntimeError(f"failed to stat lock file: {exc}") from None
    if not stat.S_ISREG(st.st_mode):
        os.close(fd)
        raise RuntimeError(f"lock file is not a regular file: {lock_path}") from None

    lock_file = os.fdopen(fd, "r+", encoding="utf-8", errors="replace")
    try:
        import fcntl
    except ImportError:
        lock_file.close()
        raise RuntimeError("file locking requires fcntl (Unix-only)") from None

    try:
        deadline = time.monotonic() + max(0.0, timeout_s)
        while True:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                return lock_file
            except BlockingIOError:
                if time.monotonic() >= deadline:
                    raise TimeoutError(f"timed out waiting for lock: {lock_path}") from None
                time.sleep(0.1)
    except BaseException:
        lock_file.close()
        raise


def _copy_workbook_to_temp_under_lock_sync(
    *,
    file_path: Path,
    lock_timeout_s: float,
) -> tuple[Path, tuple[int, int, int, int]]:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), lock_timeout_s)
    tmp_path: Optional[Path] = None
    try:
        zip_error = _scan_zip_metadata(
            file_path,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

        expected_sig = _file_signature(file_path)
        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{file_path.name}.",
            suffix=file_path.suffix,
            dir=str(file_path.parent),
        )
        tmp_path = Path(tmp_name)
        try:
            with os.fdopen(fd, "wb") as dst:
                fd = -1
                with file_path.open("rb") as src:
                    shutil.copyfileobj(src, dst)
        except Exception as exc:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
            raise RuntimeError(f"failed to copy workbook: {exc}") from None
        copied_path = tmp_path
        tmp_path = None
        return copied_path, expected_sig
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
        try:
            lock_file.close()
        except Exception:
            pass


def _fill_formula_in_workbook_unlocked_sync(
    *,
    file_path: Path,
    sheet_name: Optional[str],
    start_cell: str,
    end_row: int,
    end_col: str,
    formula_template: str,
) -> tuple[str, int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    wb = None
    try:
        try:
            wb = load_workbook(filename=str(file_path), keep_vba=False, data_only=False, keep_links=True)
        except Exception as exc:
            raise RuntimeError(f"failed to load workbook: {exc}") from None

        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")
        if sheet_name is None:
            ws = worksheets[0]
        else:
            ws = _resolve_target_worksheet(wb, sheet_name)
        resolved_sheet_name = getattr(ws, "title", sheet_name) or sheet_name or ""

        filled_cells, skipped_merged_cells = fill_formula(
            ws,
            start_cell=start_cell,
            end_row=end_row,
            end_col=end_col,
            formula_template=formula_template,
        )

        wb.save(str(file_path))
        return resolved_sheet_name, filled_cells, skipped_merged_cells
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _commit_recalc_workbook_under_lock_sync(
    *,
    file_path: Path,
    expected_sig: tuple[int, int, int, int],
    recalc_content: bytes,
    lock_timeout_s: float,
) -> tuple[bool, Optional[str]]:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), lock_timeout_s)
    try:
        try:
            current_sig = _file_signature(file_path)
        except OSError as exc:
            return False, f"failed to stat workbook before recalc writeback: {exc}"
        if current_sig != expected_sig:
            return False, "workbook changed since fill request; aborting writeback"

        try:
            mode = stat.S_IMODE(file_path.stat().st_mode)
        except OSError:
            mode = 0o644
        try:
            _write_bytes_atomically(file_path, recalc_content, mode=mode)
        except OSError as exc:
            return False, f"recalc writeback failed: {exc}"
        return True, None
    finally:
        try:
            lock_file.close()
        except Exception:
            pass


def _truncate_payload_to_max_chars(payload: dict[str, Any], max_chars: int) -> str:
    response_text = _json_dumps_compact(payload)
    if len(response_text) <= max_chars:
        return response_text

    payload["truncated"] = True
    if "recalc_samples" in payload:
        payload.pop("recalc_samples", None)
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text
    samples_out = payload.get("samples")
    if not isinstance(samples_out, list):
        return _json_dumps_compact(payload)

    for head, tail in ((2, 2), (1, 1)):
        if len(samples_out) > head + tail:
            payload["samples"] = samples_out[:head] + samples_out[-tail:]
            samples_out = payload.get("samples")
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    samples_out = payload.get("samples")
    if len(response_text) > max_chars and isinstance(samples_out, list) and len(samples_out) > 1:
        payload["samples"] = samples_out[:1]
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    samples_out = payload.get("samples")
    if len(response_text) > max_chars and isinstance(samples_out, list):
        for sample in samples_out:
            if not isinstance(sample, dict):
                continue
            for key in ("formula", "value"):
                val = sample.get(key)
                if isinstance(val, str):
                    sample[key] = _truncate_str(val, 80)
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    samples_out = payload.get("samples")
    if len(response_text) > max_chars and isinstance(samples_out, list):
        for sample in samples_out:
            if isinstance(sample, dict):
                sample["formula"] = None
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    samples_out = payload.get("samples")
    if len(response_text) > max_chars and isinstance(samples_out, list):
        for sample in samples_out:
            if isinstance(sample, dict):
                sample["value"] = None
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    if len(response_text) > max_chars:
        payload.pop("recalc_error", None)
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    samples_out = payload.get("samples")
    if len(response_text) > max_chars and isinstance(samples_out, list) and samples_out:
        addr = samples_out[0].get("address") if isinstance(samples_out[0], dict) else None
        payload["samples"] = (
            [{"address": addr, "formula": None, "value": None}] if isinstance(addr, str) and addr else []
        )
        response_text = _json_dumps_compact(payload)

    return response_text


class FillFormulaTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)

        self.recalc_url = (
            str(config.get("recalc_url") or "").strip()
            or os.environ.get("SHEET_ARENA_RECALC_URL", "").strip()
            or "http://127.0.0.1:5000/recalculate"
        )
        recalc_timeout_s_raw = config.get("recalc_timeout_s", 180)
        try:
            self.recalc_timeout_s = max(0.0, float(recalc_timeout_s_raw))
        except (TypeError, ValueError):
            self.recalc_timeout_s = 180.0
        self.max_response_chars = _get_max_response_chars(config)

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return ToolResponse(text="Error: invalid path."), 0.0, {"status": "error", "error": "invalid_path"}
        if relpath.suffix.lower() != ".xlsx":
            return (
                ToolResponse(text="Error: only .xlsx workbooks are supported."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_path",
                },
            )

        start_cell_raw = parameters.get("start_cell")
        if not isinstance(start_cell_raw, str) or not start_cell_raw.strip():
            return (
                ToolResponse(text="Error: start_cell must be a non-empty A1 cell (e.g. D2 or Sheet1!D2)."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_start_cell",
                },
            )

        sheet_from_cell = None
        try:
            sheet_from_cell, start_cell_token = _split_sheet_cell(start_cell_raw)
        except ValueError as exc:
            return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "invalid_start_cell"}

        sheet_name_raw = parameters.get("sheet_name")
        sheet_name = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return (
                    ToolResponse(text="Error: sheet_name must be a non-empty string."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_sheet_name",
                    },
                )
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return (
                    ToolResponse(text="Error: sheet_name is empty after normalization."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_sheet_name",
                    },
                )
            if sheet_from_cell is not None and sheet_name.casefold() != sheet_from_cell.casefold():
                return (
                    ToolResponse(
                        text=f"Error: sheet_name {sheet_name!r} does not match start_cell sheet {sheet_from_cell!r}."
                    ),
                    0.0,
                    {"status": "error", "error": "sheet_mismatch"},
                )
        elif sheet_from_cell is not None:
            sheet_name = sheet_from_cell

        end_row_raw = parameters.get("end_row")
        end_col_raw = parameters.get("end_col")

        formula_template_raw = parameters.get("formula_template")
        if not isinstance(formula_template_raw, str) or not formula_template_raw.strip():
            return (
                ToolResponse(text="Error: formula_template must be a non-empty string."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_formula_template",
                },
            )
        formula_template = formula_template_raw.strip()
        if not formula_template.startswith("="):
            formula_template = "=" + formula_template
        formula_template = _normalize_formula_template_for_excel(formula_template)
        max_formula_chars_raw = self.config.get("max_formula_template_chars", 8192)
        max_formula_chars: Optional[int] = None
        if max_formula_chars_raw is not None:
            try:
                max_formula_chars = int(max_formula_chars_raw)
            except (TypeError, ValueError):
                max_formula_chars = 8192
            if max_formula_chars <= 0:
                max_formula_chars = None
        if max_formula_chars is not None and len(formula_template) > max_formula_chars:
            return (
                ToolResponse(
                    text=f"Error: formula_template too long (len={len(formula_template)} > {max_formula_chars})."
                ),
                0.0,
                {"status": "error", "error": "formula_template_too_long"},
            )

        normalized_start_cell = start_cell_token.strip().upper().replace("$", "")
        try:
            start_col_letter, start_row = _parse_a1_cell(normalized_start_cell)
        except ValueError as exc:
            return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "invalid_start_cell"}

        if end_row_raw is None:
            end_row = start_row
        else:
            try:
                end_row = int(end_row_raw)
            except (TypeError, ValueError):
                return (
                    ToolResponse(text="Error: end_row must be an integer."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_end_row",
                    },
                )
            if not 1 <= end_row <= _EXCEL_MAX_ROWS:
                return (
                    ToolResponse(text=f"Error: end_row must be between 1 and {_EXCEL_MAX_ROWS}."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_end_row",
                    },
                )
            if end_row < start_row:
                return (
                    ToolResponse(text=f"Error: end_row ({end_row}) must be >= start_cell row ({start_row})."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_end_row",
                    },
                )

        if end_col_raw is None:
            end_col = start_col_letter
            end_col_idx = _col_letter_to_index(end_col)
        else:
            if not isinstance(end_col_raw, str) or not end_col_raw.strip():
                return (
                    ToolResponse(text="Error: end_col must be a non-empty column letter (e.g. H)."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_end_col",
                    },
                )
            end_col = end_col_raw.strip().upper().replace("$", "")
            if not re.fullmatch(r"[A-Z]{1,3}", end_col):
                return (
                    ToolResponse(text="Error: end_col must be a column letter (e.g. H or AB)."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_end_col",
                    },
                )
            end_col_idx = _col_letter_to_index(end_col)
            if not 1 <= end_col_idx <= _EXCEL_MAX_COLS:
                return (
                    ToolResponse(text=f"Error: end_col out of bounds: {end_col}."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_end_col",
                    },
                )

        start_col_idx = _col_letter_to_index(start_col_letter)
        if end_col_idx < start_col_idx:
            return (
                ToolResponse(text=f"Error: end_col ({end_col}) must be >= start_cell col ({start_col_letter})."),
                0.0,
                {"status": "error", "error": "invalid_end_col"},
            )

        total_rows = end_row - start_row + 1
        total_cols = end_col_idx - start_col_idx + 1
        total_cells = total_rows * total_cols
        max_fill_cells_raw = self.config.get("max_fill_cells", 300_000)
        max_fill_cells = None
        if max_fill_cells_raw is not None:
            try:
                max_fill_cells = int(max_fill_cells_raw)
            except (TypeError, ValueError):
                max_fill_cells = 300_000
            if max_fill_cells <= 0:
                max_fill_cells = None
        if max_fill_cells is not None and total_cells > max_fill_cells:
            return (
                ToolResponse(text=f"Error: requested range is too large (cells={total_cells} > {max_fill_cells})."),
                0.0,
                {"status": "error", "error": "range_too_large"},
            )

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return (
                ToolResponse(text="Error: workspace_id is missing/invalid."),
                0.0,
                {
                    "status": "error",
                    "error": "missing_workspace_id",
                },
            )
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return (
                ToolResponse(text=f"Error: file not found: {relpath}"),
                0.0,
                {
                    "status": "error",
                    "error": "file_not_found",
                },
            )
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return (
                    ToolResponse(text=f"Error: failed to stat file: {exc}"),
                    0.0,
                    {
                        "status": "error",
                        "error": "stat_failed",
                    },
                )
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return (
                    ToolResponse(text=f"Error: workbook is too large ({actual_mb:.1f}MB > {max_mb}MB)."),
                    0.0,
                    {
                        "status": "error",
                        "error": "file_too_large",
                    },
                )

        if total_rows <= 1 and total_cols <= 1:
            coords = [(start_row, start_col_idx)]
        elif total_rows == 1:
            sample_cols = [start_col_idx + idx for idx in _sample_indices(total_cols, head=3, tail=3)]
            coords = [(start_row, col_num) for col_num in sample_cols]
        elif total_cols == 1:
            sample_rows = [start_row + idx for idx in _sample_indices(total_rows, head=3, tail=3)]
            coords = [(row_num, start_col_idx) for row_num in sample_rows]
        else:
            sample_rows = [start_row + idx for idx in _sample_indices(total_rows, head=2, tail=2)]
            sample_cols = [start_col_idx + idx for idx in _sample_indices(total_cols, head=2, tail=2)]
            coords = [(row_num, col_num) for row_num in sample_rows for col_num in sample_cols]

        tmp_fill_path: Optional[Path] = None
        expected_sig: Optional[tuple[int, int, int, int]] = None
        resolved_sheet_name: Optional[str] = None
        filled_cells = 0
        skipped_merged_cells = 0
        recalc_written = False
        samples: list[dict[str, Any]] = []
        recalc_samples: Optional[list[dict[str, Any]]] = None
        recalc_error: Optional[str] = None
        writeback_error: Optional[str] = None
        error_code: Optional[str] = None
        samples_source = "none"

        try:
            try:
                tmp_fill_path, expected_sig = await asyncio.to_thread(
                    _copy_workbook_to_temp_under_lock_sync,
                    file_path=file_path,
                    lock_timeout_s=self.lock_timeout_s,
                )
                resolved_sheet_name, filled_cells, skipped_merged_cells = await asyncio.to_thread(
                    _fill_formula_in_workbook_unlocked_sync,
                    file_path=tmp_fill_path,
                    sheet_name=sheet_name,
                    start_cell=normalized_start_cell,
                    end_row=end_row,
                    end_col=end_col,
                    formula_template=formula_template,
                )
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                return (
                    ToolResponse(text=f"Error: failed to fill formulas: {exc}"),
                    0.0,
                    {
                        "status": "error",
                        "error": "fill_failed",
                    },
                )

            zip_error = await asyncio.to_thread(
                _scan_zip_metadata,
                tmp_fill_path,
                max_members=50_000,
                max_total_uncompressed_bytes=512 * 1024 * 1024,
                max_member_uncompressed_bytes=128 * 1024 * 1024,
                max_ratio=200.0,
            )
            if zip_error:
                recalc_error = f"filled workbook rejected by zip safety checks: {zip_error}"
                error_code = "unsafe_workbook"
            else:
                try:
                    recalc_content = await asyncio.to_thread(
                        _post_recalculate,
                        url=self.recalc_url,
                        file_path=tmp_fill_path,
                        filename=file_path.name,
                        timeout_s=self.recalc_timeout_s,
                        max_response_bytes=self.max_file_size_bytes,
                    )
                except asyncio.CancelledError:
                    raise
                except Exception as exc:
                    recalc_error = str(exc)
                    error_code = "recalc_failed"
                else:
                    zip_error = _scan_zip_metadata_bytes(
                        recalc_content,
                        max_members=50_000,
                        max_total_uncompressed_bytes=512 * 1024 * 1024,
                        max_member_uncompressed_bytes=128 * 1024 * 1024,
                        max_ratio=200.0,
                    )
                    if zip_error:
                        recalc_error = f"recalc workbook rejected by zip safety checks: {zip_error}"
                        error_code = "recalc_failed"
                    else:
                        assert resolved_sheet_name is not None
                        try:
                            recalc_samples = await asyncio.to_thread(
                                _read_cells_formula_and_value_from_bytes,
                                content=recalc_content,
                                sheet_name=resolved_sheet_name,
                                coords=coords,
                            )
                        except asyncio.CancelledError:
                            raise
                        except Exception as exc:
                            recalc_error = f"failed to read samples from recalc response: {exc}"
                            error_code = "recalc_failed"
                        else:
                            assert expected_sig is not None
                            try:
                                recalc_written, writeback_error = await asyncio.to_thread(
                                    _commit_recalc_workbook_under_lock_sync,
                                    file_path=file_path,
                                    expected_sig=expected_sig,
                                    recalc_content=recalc_content,
                                    lock_timeout_s=self.lock_timeout_s,
                                )
                            except asyncio.CancelledError:
                                raise
                            except Exception as exc:
                                recalc_written = False
                                writeback_error = f"recalc writeback failed: {exc}"

            if (
                recalc_error is not None
                and error_code != "unsafe_workbook"
                and tmp_fill_path is not None
                and resolved_sheet_name is not None
            ):
                try:
                    samples = await asyncio.to_thread(
                        _read_cells_formula_and_value_from_path,
                        file_path=tmp_fill_path,
                        sheet_name=resolved_sheet_name,
                        coords=coords,
                    )
                    samples_source = "filled_workbook"
                except asyncio.CancelledError:
                    raise
                except Exception as exc:
                    samples = []
                    samples_source = "none"
                    recalc_error = f"{recalc_error}; failed to read filled workbook samples: {exc}"
        finally:
            if tmp_fill_path is not None:
                try:
                    tmp_fill_path.unlink(missing_ok=True)
                except Exception:
                    pass

        if resolved_sheet_name is None:
            return (
                ToolResponse(text="Error: failed to resolve sheet name."),
                0.0,
                {
                    "status": "error",
                    "error": "fill_failed",
                },
            )

        start_addr = f"{start_col_letter}{start_row}"
        range_end = f"{end_col}{end_row}"
        sheet_ref = _quote_sheet_name_for_a1(resolved_sheet_name)

        if recalc_error is not None:
            payload: dict[str, Any] = {
                "status": "error",
                "error": error_code or "recalc_failed",
                "file": str(relpath),
                "sheet": resolved_sheet_name,
                "range": f"{sheet_ref}!{start_addr}:{range_end}",
                "target_cells": total_cells,
                "filled_cells": filled_cells,
                "skipped_merged_cells": skipped_merged_cells,
                "samples_source": samples_source,
                "recalc_written": False,
                "rolled_back": True,
                "recalc_error": _to_jsonable_excel_value(recalc_error),
                "samples": samples,
            }
            response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
            samples_out = payload.get("samples")
            sample_count = len(samples_out) if isinstance(samples_out, list) else 0
            metrics = {
                "status": "error",
                "error": error_code or "recalc_failed",
                "file": str(relpath),
                "target_cells": total_cells,
                "filled_cells": filled_cells,
                "skipped_merged_cells": skipped_merged_cells,
                "recalc_written": False,
                "samples_source": samples_source,
                "sample_cells": sample_count,
            }
            return ToolResponse(text=response_text), 0.0, metrics

        if writeback_error is not None:
            samples_source = "disk"
            recalc_written = False
            zip_error = await asyncio.to_thread(
                _scan_zip_metadata,
                file_path,
                max_members=50_000,
                max_total_uncompressed_bytes=512 * 1024 * 1024,
                max_member_uncompressed_bytes=128 * 1024 * 1024,
                max_ratio=200.0,
            )
            if zip_error:
                samples = recalc_samples or []
                samples_source = "recalc"
                writeback_error = f"{writeback_error}; workbook rejected by zip safety checks: {zip_error}"
            else:
                try:
                    samples = await asyncio.to_thread(
                        _read_cells_formula_and_value_from_path,
                        file_path=file_path,
                        sheet_name=resolved_sheet_name,
                        coords=coords,
                    )
                except Exception as exc:
                    samples = recalc_samples or []
                    samples_source = "recalc"
                    writeback_error = f"{writeback_error}; failed to read live workbook samples: {exc}"

            payload = {
                "status": "error",
                "error": "writeback_failed",
                "file": str(relpath),
                "sheet": resolved_sheet_name,
                "range": f"{sheet_ref}!{start_addr}:{range_end}",
                "target_cells": total_cells,
                "filled_cells": filled_cells,
                "skipped_merged_cells": skipped_merged_cells,
                "samples_source": samples_source,
                "recalc_written": False,
                "rolled_back": True,
                "writeback_error": _to_jsonable_excel_value(writeback_error),
                "samples": samples,
            }
            if recalc_samples:
                payload["recalc_samples"] = recalc_samples
            response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
            samples_out = payload.get("samples")
            sample_count = len(samples_out) if isinstance(samples_out, list) else 0
            metrics = {
                "status": "error",
                "error": "writeback_failed",
                "file": str(relpath),
                "target_cells": total_cells,
                "filled_cells": filled_cells,
                "skipped_merged_cells": skipped_merged_cells,
                "recalc_written": False,
                "samples_source": samples_source,
                "sample_cells": sample_count,
            }
            return ToolResponse(text=response_text), 0.0, metrics

        samples = recalc_samples or []
        samples_source = "recalc"
        recalc_written = True

        status = "success" if skipped_merged_cells == 0 else "partial_success"
        payload = {
            "status": status,
            "file": str(relpath),
            "sheet": resolved_sheet_name,
            "range": f"{sheet_ref}!{start_addr}:{range_end}",
            "target_cells": total_cells,
            "filled_cells": filled_cells,
            "skipped_merged_cells": skipped_merged_cells,
            "samples_source": samples_source,
            "recalc_written": recalc_written,
            "samples": samples,
        }

        response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
        samples_out = payload.get("samples")
        sample_count = len(samples_out) if isinstance(samples_out, list) else 0
        metrics = {
            "status": status,
            "file": str(relpath),
            "target_cells": total_cells,
            "filled_cells": filled_cells,
            "skipped_merged_cells": skipped_merged_cells,
            "recalc_written": recalc_written,
            "samples_source": samples_source,
            "sample_cells": sample_count,
        }
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
