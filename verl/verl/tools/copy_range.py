from __future__ import annotations

import asyncio
import csv
import io
import os
import shutil
import stat
import tempfile
import uuid
from copy import copy as copy_obj
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .clear_range import (
    _acquire_lockfile,
    _format_a1_address,
    _get_max_response_chars,
    _get_max_string_chars,
    _json_dumps_compact,
    _normalize_sheet_name,
    _quote_sheet_name_for_a1,
    _resolve_target_worksheet,
    _resolve_workspace_file,
    _sample_indices,
    _sanitize_relpath,
    _scan_zip_metadata,
    _split_sheet_cell_range,
    _to_jsonable_excel_value,
    _truncate_str,
)
from .recalculate import _file_signature
from .schemas import OpenAIFunctionToolSchema, ToolResponse

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384


@dataclass(frozen=True)
class _RangeRef:
    sheet_name: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def rows(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def cols(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def cells(self) -> int:
        return self.rows * self.cols


@dataclass(frozen=True)
class _CellSnapshot:
    row: int
    col: int
    value: Any
    data_type: str
    style: Any
    comment: Any
    hyperlink: Any


def _normalize_optional_sheet_name(value: Any, *, field_name: str) -> Optional[str]:
    if value is None:
        return None
    if not isinstance(value, str) or not value.strip():
        raise RuntimeError(f"{field_name} must be a non-empty string")
    sheet_name = _normalize_sheet_name(value)
    if not sheet_name:
        raise RuntimeError(f"{field_name} is empty after normalization")
    return sheet_name


def _parse_cell_range_bounds(range_part: str) -> tuple[int, int, int, int, bool]:
    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required to parse ranges; install openpyxl>=3.1.5") from None

    normalized = range_part.strip(" \t\n\r\f\v\"'").replace(" ", "").replace("$", "").upper()
    if not normalized:
        raise RuntimeError("range is empty")
    if "," in normalized:
        raise RuntimeError("copy_range only accepts one rectangular A1 range")

    token_had_range = ":" in normalized
    parse_token = normalized if token_had_range else f"{normalized}:{normalized}"
    try:
        min_col, min_row, max_col, max_row = range_boundaries(parse_token)
    except Exception as exc:
        raise RuntimeError(f"invalid A1 cell range: {range_part!r}: {exc}") from None

    if None in (min_col, min_row, max_col, max_row):
        raise RuntimeError("copy_range requires cell ranges such as A1:C3, not whole rows or columns")
    min_col = int(min_col)
    min_row = int(min_row)
    max_col = int(max_col)
    max_row = int(max_row)
    if min_col > max_col:
        min_col, max_col = max_col, min_col
    if min_row > max_row:
        min_row, max_row = max_row, min_row
    if min_col < 1 or min_row < 1 or max_col > _EXCEL_MAX_COLS or max_row > _EXCEL_MAX_ROWS:
        raise RuntimeError("range is outside Excel worksheet bounds")
    return min_row, min_col, max_row, max_col, token_had_range


def _resolve_source_ref(
    *,
    wb,
    source_range: str,
    sheet_name: Optional[str],
    source_sheet_name: Optional[str],
) -> tuple[_RangeRef, Any]:
    worksheets = getattr(wb, "worksheets", None) or []
    default_sheet = getattr(worksheets[0], "title", "") if worksheets else ""
    token_sheet, range_part = _split_sheet_cell_range(source_range)
    if (
        token_sheet is not None
        and source_sheet_name is not None
        and token_sheet.casefold() != source_sheet_name.casefold()
    ):
        raise RuntimeError(f"source sheet mismatch: {source_sheet_name!r} vs {token_sheet!r}")

    resolved_name = token_sheet or source_sheet_name or sheet_name or default_sheet
    if not resolved_name:
        raise RuntimeError("failed to resolve source sheet")
    ws = _resolve_target_worksheet(wb, resolved_name)
    actual_name = getattr(ws, "title", resolved_name) or resolved_name
    min_row, min_col, max_row, max_col, _ = _parse_cell_range_bounds(range_part)
    return (
        _RangeRef(
            sheet_name=actual_name,
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
        ),
        ws,
    )


def _resolve_destination_ref(
    *,
    wb,
    destination_range: str,
    sheet_name: Optional[str],
    destination_sheet_name: Optional[str],
    source_ref: _RangeRef,
) -> tuple[_RangeRef, Any]:
    token_sheet, range_part = _split_sheet_cell_range(destination_range)
    if (
        token_sheet is not None
        and destination_sheet_name is not None
        and token_sheet.casefold() != destination_sheet_name.casefold()
    ):
        raise RuntimeError(f"destination sheet mismatch: {destination_sheet_name!r} vs {token_sheet!r}")

    resolved_name = token_sheet or destination_sheet_name or sheet_name or source_ref.sheet_name
    if not resolved_name:
        raise RuntimeError("failed to resolve destination sheet")
    ws = _resolve_target_worksheet(wb, resolved_name)
    actual_name = getattr(ws, "title", resolved_name) or resolved_name
    min_row, min_col, max_row, max_col, token_had_range = _parse_cell_range_bounds(range_part)

    if token_had_range:
        dest_rows = max_row - min_row + 1
        dest_cols = max_col - min_col + 1
        if dest_rows != source_ref.rows or dest_cols != source_ref.cols:
            raise RuntimeError(
                "destination range must be a single top-left cell or match source dimensions "
                f"(source={source_ref.rows}x{source_ref.cols}, destination={dest_rows}x{dest_cols})"
            )
    else:
        max_row = min_row + source_ref.rows - 1
        max_col = min_col + source_ref.cols - 1
        if max_row > _EXCEL_MAX_ROWS or max_col > _EXCEL_MAX_COLS:
            raise RuntimeError("destination range extends outside Excel worksheet bounds")

    return (
        _RangeRef(
            sheet_name=actual_name,
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
        ),
        ws,
    )


def _format_range(ref: _RangeRef) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        start = f"{ref.min_col}{ref.min_row}"
        end = f"{ref.max_col}{ref.max_row}"
    else:
        start = f"{get_column_letter(ref.min_col)}{ref.min_row}"
        end = f"{get_column_letter(ref.max_col)}{ref.max_row}"
    sheet_ref = _quote_sheet_name_for_a1(ref.sheet_name)
    if start == end:
        return f"{sheet_ref}!{start}"
    return f"{sheet_ref}!{start}:{end}"


def _sample_coords(ref: _RangeRef) -> list[tuple[int, int]]:
    if ref.rows <= 1 and ref.cols <= 1:
        return [(ref.min_row, ref.min_col)]
    if ref.rows == 1:
        sample_cols = [ref.min_col + idx for idx in _sample_indices(ref.cols, head=3, tail=3)]
        return [(ref.min_row, col_num) for col_num in sample_cols]
    if ref.cols == 1:
        sample_rows = [ref.min_row + idx for idx in _sample_indices(ref.rows, head=3, tail=3)]
        return [(row_num, ref.min_col) for row_num in sample_rows]

    sample_rows = [ref.min_row + idx for idx in _sample_indices(ref.rows, head=2, tail=2)]
    sample_cols = [ref.min_col + idx for idx in _sample_indices(ref.cols, head=2, tail=2)]
    return [(row_num, col_num) for row_num in sample_rows for col_num in sample_cols]


def _snapshot_source_cells(source_ws, source_ref: _RangeRef) -> list[_CellSnapshot]:
    snapshots: list[_CellSnapshot] = []
    for row_num in range(source_ref.min_row, source_ref.max_row + 1):
        for col_num in range(source_ref.min_col, source_ref.max_col + 1):
            cell = source_ws.cell(row=row_num, column=col_num)
            snapshots.append(
                _CellSnapshot(
                    row=row_num,
                    col=col_num,
                    value=getattr(cell, "value", None),
                    data_type=str(getattr(cell, "data_type", "") or ""),
                    style=copy_obj(getattr(cell, "_style", None)),
                    comment=copy_obj(getattr(cell, "comment", None)),
                    hyperlink=copy_obj(getattr(cell, "hyperlink", None)),
                )
            )
    return snapshots


def _apply_cell_snapshot(*, snapshot: _CellSnapshot, dest_cell, dest_row: int, dest_col: int) -> None:
    try:
        from openpyxl.formula.translate import Translator
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required to translate formulas; install openpyxl>=3.1.5") from None

    value = snapshot.value
    if snapshot.data_type == "f" and isinstance(value, str) and value:
        formula = value if value.startswith("=") else f"={value}"
        row_delta = dest_row - snapshot.row
        col_delta = dest_col - snapshot.col
        origin = f"{get_column_letter(snapshot.col)}{snapshot.row}"
        try:
            value = Translator(formula, origin=origin).translate_formula(row_delta=row_delta, col_delta=col_delta)
        except Exception:
            value = formula

    if snapshot.style is not None:
        dest_cell._style = copy_obj(snapshot.style)
    dest_cell.value = value
    if snapshot.data_type != "f" and isinstance(value, str) and value.startswith("="):
        dest_cell.data_type = "s"
    dest_cell.comment = copy_obj(snapshot.comment) if snapshot.comment is not None else None
    hyperlink = copy_obj(snapshot.hyperlink) if snapshot.hyperlink is not None else None
    if hyperlink is not None:
        try:
            hyperlink.ref = dest_cell.coordinate
        except Exception:
            pass
    dest_cell.hyperlink = hyperlink


def _sample_destination_csv(dest_ws, dest_ref: _RangeRef) -> str:
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(["address", "value", "formula"])
    for row_num, col_num in _sample_coords(dest_ref):
        cell = dest_ws.cell(row=row_num, column=col_num)
        raw = getattr(cell, "value", None)
        formula = None
        if getattr(cell, "data_type", None) == "f" and isinstance(raw, str) and raw:
            formula = raw if raw.startswith("=") else f"={raw}"
        if isinstance(formula, str):
            formula = _truncate_str(formula, _get_max_string_chars())
        value = "" if formula is not None else _to_jsonable_excel_value(raw)
        if value is None:
            value = ""
        writer.writerow(
            [
                _format_a1_address(sheet_name=dest_ref.sheet_name, row=row_num, col=col_num),
                value,
                formula or "",
            ]
        )
    return out.getvalue().rstrip("\n")


def _validate_saved_workbook(file_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to validate workbooks; install openpyxl>=3.1.5") from None

    zip_error = _scan_zip_metadata(
        file_path,
        max_members=50_000,
        max_total_uncompressed_bytes=512 * 1024 * 1024,
        max_member_uncompressed_bytes=128 * 1024 * 1024,
        max_ratio=200.0,
    )
    if zip_error:
        raise RuntimeError(f"saved workbook rejected by zip safety checks: {zip_error}")

    wb = None
    try:
        wb = load_workbook(filename=str(file_path), data_only=False, read_only=True, keep_links=False)
        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("saved workbook has no worksheets")
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError(f"saved workbook failed load validation: {exc}") from None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _copy_range_in_workbook(
    *,
    file_path: Path,
    sheet_name: Optional[str],
    source_sheet_name: Optional[str],
    destination_sheet_name: Optional[str],
    source_range: str,
    destination_range: str,
    lock_timeout_s: float,
    max_copy_cells: int,
) -> tuple[str, str, int, int, str]:
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    wb = None
    tmp_path: Optional[Path] = None
    try:
        zip_error = _scan_zip_metadata(
            file_path,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

        try:
            orig_mode = stat.S_IMODE(file_path.stat().st_mode)
            expected_sig = _file_signature(file_path)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook: {exc}") from None

        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{file_path.name}.",
            suffix=file_path.suffix,
            dir=str(file_path.parent),
        )
        tmp_path = Path(tmp_name)
        try:
            with os.fdopen(fd, "wb") as dst:
                fd = -1
                with file_path.open("rb") as src:
                    shutil.copyfileobj(src, dst)
        except Exception as exc:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
            raise RuntimeError(f"failed to copy workbook to temp file: {exc}") from None
        try:
            os.chmod(tmp_path, orig_mode)
        except OSError:
            pass

        try:
            wb = load_workbook(filename=str(tmp_path), keep_vba=False, data_only=False, keep_links=True)
        except Exception as exc:
            raise RuntimeError(f"failed to load workbook: {exc}") from None

        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")

        source_ref, source_ws = _resolve_source_ref(
            wb=wb,
            source_range=source_range,
            sheet_name=sheet_name,
            source_sheet_name=source_sheet_name,
        )
        dest_ref, dest_ws = _resolve_destination_ref(
            wb=wb,
            destination_range=destination_range,
            sheet_name=sheet_name,
            destination_sheet_name=destination_sheet_name,
            source_ref=source_ref,
        )
        if max_copy_cells > 0 and source_ref.cells > max_copy_cells:
            raise RuntimeError(f"source range is too large (cells={source_ref.cells} > {max_copy_cells})")
        snapshots = _snapshot_source_cells(source_ws, source_ref)

        copied_cells = 0
        skipped_merged_cells = 0
        idx = 0
        for row_offset in range(source_ref.rows):
            for col_offset in range(source_ref.cols):
                snapshot = snapshots[idx]
                idx += 1
                dest_row = dest_ref.min_row + row_offset
                dest_col = dest_ref.min_col + col_offset
                dest_cell = dest_ws.cell(row=dest_row, column=dest_col)
                if isinstance(dest_cell, MergedCell):
                    skipped_merged_cells += 1
                    continue
                _apply_cell_snapshot(snapshot=snapshot, dest_cell=dest_cell, dest_row=dest_row, dest_col=dest_col)
                copied_cells += 1

        sample_csv = _sample_destination_csv(dest_ws, dest_ref)

        try:
            wb.save(str(tmp_path))
        except Exception as exc:
            raise RuntimeError(f"failed to save workbook: {exc}") from None
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
            wb = None

        _validate_saved_workbook(tmp_path)
        try:
            current_sig = _file_signature(file_path)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook before writeback: {exc}") from None
        if current_sig != expected_sig:
            raise RuntimeError("workbook changed since copy request; aborting writeback")

        os.replace(tmp_path, file_path)
        tmp_path = None
        return _format_range(source_ref), _format_range(dest_ref), copied_cells, skipped_merged_cells, sample_csv
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        try:
            lock_file.close()
        except Exception:
            pass


def _truncate_payload_to_max_chars(payload: dict[str, Any], max_chars: int) -> str:
    response_text = _json_dumps_compact(payload)
    if len(response_text) <= max_chars:
        return response_text

    payload["truncated"] = True
    sample_csv = payload.get("sample_csv")
    if isinstance(sample_csv, str) and sample_csv:
        lines = sample_csv.splitlines()
        if len(lines) > 4:
            payload["sample_csv"] = "\n".join(lines[:4])
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text
        payload.pop("sample_csv", None)
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    minimal_payload = {
        "status": payload.get("status"),
        "file": payload.get("file"),
        "source_range": payload.get("source_range"),
        "destination_range": payload.get("destination_range"),
        "copied_cells": payload.get("copied_cells"),
        "truncated": True,
    }
    return _json_dumps_compact(minimal_payload)


def _json_error(error: str, message: str, *, file: Optional[str] = None) -> ToolResponse:
    payload: dict[str, Any] = {
        "status": "error",
        "error": error,
        "message": message,
        "truncated": False,
    }
    if file is not None:
        payload["file"] = file
    return ToolResponse(text=_json_dumps_compact(payload))


class CopyRangeTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        max_copy_cells_raw = config.get("max_copy_cells", os.environ.get("SHEET_ARENA_COPY_RANGE_MAX", "300000"))
        try:
            self.max_copy_cells = max(0, int(max_copy_cells_raw))
        except (TypeError, ValueError):
            self.max_copy_cells = 300_000

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)
        self.max_response_chars = _get_max_response_chars(config)

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return _json_error("invalid_path", "invalid path."), 0.0, {
                    "status": "error",
                    "error": "invalid_path",
                }
        if relpath.suffix.lower() != ".xlsx":
            return _json_error("invalid_path", "only .xlsx workbooks are supported."), 0.0, {
                "status": "error",
                "error": "invalid_path",
            }

        source_range_raw = parameters.get("source_range")
        if not isinstance(source_range_raw, str) or not source_range_raw.strip():
            return _json_error(
                "invalid_source_range",
                "source_range must be a non-empty A1 range such as A1:C3 or Sheet1!A1:C3.",
                file=str(relpath),
            ), 0.0, {"status": "error", "error": "invalid_source_range"}
        source_range = source_range_raw.strip()

        destination_range_raw = parameters.get("destination_range")
        if not isinstance(destination_range_raw, str) or not destination_range_raw.strip():
            return _json_error(
                "invalid_destination_range",
                "destination_range must be a non-empty A1 cell/range such as E1 or Sheet2!E1:G3.",
                file=str(relpath),
            ), 0.0, {"status": "error", "error": "invalid_destination_range"}
        destination_range = destination_range_raw.strip()

        try:
            sheet_name = _normalize_optional_sheet_name(parameters.get("sheet_name"), field_name="sheet_name")
            source_sheet_name = _normalize_optional_sheet_name(
                parameters.get("source_sheet_name"),
                field_name="source_sheet_name",
            )
            destination_sheet_name = _normalize_optional_sheet_name(
                parameters.get("destination_sheet_name"),
                field_name="destination_sheet_name",
            )
        except RuntimeError as exc:
            return _json_error("invalid_sheet_name", str(exc), file=str(relpath)), 0.0, {
                "status": "error",
                "error": "invalid_sheet_name",
            }

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return _json_error("missing_workspace_id", "workspace_id is missing/invalid.", file=str(relpath)), 0.0, {
                "status": "error",
                "error": "missing_workspace_id",
            }
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return _json_error("file_not_found", f"file not found: {relpath}", file=str(relpath)), 0.0, {
                "status": "error",
                "error": "file_not_found",
            }
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return _json_error("stat_failed", f"failed to stat file: {exc}", file=str(relpath)), 0.0, {
                    "status": "error",
                    "error": "stat_failed",
                }
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return _json_error(
                    "file_too_large",
                    f"workbook is too large ({actual_mb:.1f}MB > {max_mb}MB).",
                    file=str(relpath),
                ), 0.0, {
                    "status": "error",
                    "error": "file_too_large",
                }

        try:
            source_range_out, destination_range_out, copied_cells, skipped_merged_cells, sample_csv = (
                await asyncio.to_thread(
                    _copy_range_in_workbook,
                    file_path=file_path,
                    sheet_name=sheet_name,
                    source_sheet_name=source_sheet_name,
                    destination_sheet_name=destination_sheet_name,
                    source_range=source_range,
                    destination_range=destination_range,
                    lock_timeout_s=self.lock_timeout_s,
                    max_copy_cells=self.max_copy_cells,
                )
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return _json_error("copy_failed", f"failed to copy range: {exc}", file=str(relpath)), 0.0, {
                "status": "error",
                "error": "copy_failed",
            }

        status = "success" if skipped_merged_cells == 0 else "partial_success"
        payload: dict[str, Any] = {
            "status": status,
            "file": str(relpath),
            "source_range": source_range_out,
            "destination_range": destination_range_out,
            "copied_cells": copied_cells,
            "truncated": False,
            "sample_csv": sample_csv,
        }
        if skipped_merged_cells:
            payload["skipped_merged_cells"] = skipped_merged_cells
        response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
        metrics = {
            "status": status,
            "file": str(relpath),
            "source_range": source_range_out,
            "destination_range": destination_range_out,
            "copied_cells": copied_cells,
            "skipped_merged_cells": skipped_merged_cells,
            "sample_cells": max(0, len(sample_csv.splitlines()) - 1) if sample_csv else 0,
        }
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
