from __future__ import annotations

import asyncio
import datetime as dt
import json
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional
from uuid import uuid4

from verl.utils.paths import get_spreadsheet_rl_data_root, normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .schemas import OpenAIFunctionToolSchema, ToolResponse
from .worksheet_resolution import WorksheetResolutionError, resolve_worksheet

_A1_CELL_RE = r"[A-Z]{1,3}[0-9]{1,7}"
_A1_COL_RE = r"[A-Z]{1,3}"
_A1_ROW_RE = r"[0-9]{1,7}"
_A1_RANGE_RE = re.compile(
    rf"^(?:{_A1_CELL_RE}(?::{_A1_CELL_RE})?|{_A1_COL_RE}:{_A1_COL_RE}|{_A1_ROW_RE}:{_A1_ROW_RE})$"
)
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z0-9_]+$")

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384


def _normalize_primary_ext(primary_ext: Any) -> Optional[str]:
    if not isinstance(primary_ext, str):
        return None
    primary_ext = primary_ext.strip()
    if not primary_ext:
        return None
    if len(primary_ext) > 16 or not primary_ext.startswith("."):
        return None
    if not re.fullmatch(r"\.[A-Za-z0-9]+", primary_ext):
        return None
    if primary_ext.lower() != ".xlsx":
        return None
    return primary_ext


def _normalize_sheet_name(value: str) -> str:
    name = value.strip('\t\n\r\f\v"')
    quoted = name.strip()
    if len(quoted) >= 2 and quoted[0] == "'" and quoted[-1] == "'":
        return quoted[1:-1].replace("''", "'")
    if len(quoted) >= 2 and quoted[0] == '"' and quoted[-1] == '"':
        return quoted[1:-1]

    if quoted.startswith("'") and not quoted.endswith("'"):
        return quoted[1:]
    if quoted.endswith("'") and not quoted.startswith("'"):
        return quoted[:-1]
    return name.strip("\t\n\r\f\v")


def _normalize_cell_range(value: str) -> str:
    rng = value.strip(" \t\n\r\f\v\"'")
    rng = rng.replace(" ", "").replace("$", "")
    return rng.upper()


def _parse_sheet_cell_range(token: str, *, default_sheet_name: str) -> tuple[str, str]:
    token = token.strip(' \t\n\r\f\v"')
    if not token:
        raise ValueError("empty cell range")

    if "!" in token:
        sheet_part, _, range_part = token.rpartition("!")
        sheet_name = _normalize_sheet_name(sheet_part)
        if not sheet_name:
            raise ValueError(f"empty sheet name: {token!r}")
    else:
        sheet_name = default_sheet_name
        range_part = token

    cell_range = _normalize_cell_range(range_part)
    if not _A1_RANGE_RE.fullmatch(cell_range):
        raise ValueError(f"invalid A1 cell range: {range_part!r}")
    return sheet_name, cell_range


class _TooManyRangesError(ValueError):
    pass


class _SummaryRangeTooLargeError(ValueError):
    pass


class _SheetMismatchError(ValueError):
    def __init__(self, *, expected_sheet: str, requested_sheet: str) -> None:
        super().__init__(f"sheet_name {expected_sheet!r} does not match range sheet {requested_sheet!r}.")
        self.expected_sheet = expected_sheet
        self.requested_sheet = requested_sheet


@dataclass
class _RangeGroup:
    tokens: list[str]
    separators: list[str]


def _split_disjoint_range_group(value: str, *, max_tokens: Optional[int] = None) -> _RangeGroup:
    tokens: list[str] = []
    separators: list[str] = []
    current: list[str] = []
    in_sheet_quote = False
    i = 0
    while i < len(value):
        ch = value[i]
        if ch == "'":
            current.append(ch)
            if in_sheet_quote and i + 1 < len(value) and value[i + 1] == "'":
                current.append(value[i + 1])
                i += 2
                continue
            in_sheet_quote = not in_sheet_quote
            i += 1
            continue
        if ch == "," and not in_sheet_quote:
            token = "".join(current).strip()
            if not token:
                raise ValueError("empty range in comma-separated range list")
            tokens.append(token)
            if max_tokens is not None and len(tokens) > max_tokens:
                raise _TooManyRangesError
            j = i + 1
            while j < len(value) and value[j].isspace():
                j += 1
            separators.append(value[i:j])
            current = []
            i = j
            continue
        current.append(ch)
        i += 1
    token = "".join(current).strip()
    if not token:
        raise ValueError("empty range in comma-separated range list")
    tokens.append(token)
    if max_tokens is not None and len(tokens) > max_tokens:
        raise _TooManyRangesError
    return _RangeGroup(tokens=tokens, separators=separators)


def _split_disjoint_range_tokens(value: str, *, max_tokens: Optional[int] = None) -> list[str]:
    return _split_disjoint_range_group(value, max_tokens=max_tokens).tokens


def _to_jsonable_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool | int | str):
        if isinstance(value, str):
            return _truncate_str(value, _get_max_string_chars())
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, dt.datetime | dt.date | dt.time):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return str(value)


def _sanitize_relpath(value: Any) -> Optional[Path]:
    if not isinstance(value, str):
        return None
    raw = value.strip()
    if not raw:
        return None
    if "\0" in raw:
        return None
    p = Path(raw)
    if p.is_absolute():
        return None
    if any(part == ".." for part in p.parts):
        return None
    return p


def _resolve_safe_file(root: Path, relpath: Path, *, allow_root_symlink: bool = False) -> Optional[Path]:
    if not allow_root_symlink and root.is_symlink():
        return None
    candidate = root / relpath
    try:
        root_resolved = root.resolve(strict=True)
        resolved = candidate.resolve(strict=True)
    except (OSError, RuntimeError):
        return None
    if not resolved.is_relative_to(root_resolved):
        return None

    if candidate.is_symlink():
        return None

    cursor = root
    for part in relpath.parts:
        cursor = cursor / part
        try:
            if cursor.is_symlink():
                return None
        except OSError:
            return None

    if resolved.is_file():
        return resolved
    return None


def _get_max_cells(config: dict[str, Any]) -> int:
    env_value = os.environ.get("SHEET_ARENA_INSPECT_MAX_CELLS", "").strip()
    if env_value:
        try:
            n = int(env_value)
            return min(max(1, n), 100_000)
        except ValueError:
            pass

    try:
        n = int(config.get("max_cells", 400))
        return min(max(1, n), 100_000)
    except Exception:
        return 400


def _get_max_response_chars(config: dict[str, Any]) -> int:
    for raw in (config.get("max_response_chars"), os.environ.get("SHEET_ARENA_TOOL_MAX_RESPONSE_CHARS", "").strip()):
        if raw is None:
            continue
        try:
            n = int(raw)
            return min(max(128, n), 10240)
        except (TypeError, ValueError):
            continue
    return 900


def _get_max_ranges(config: dict[str, Any]) -> int:
    env_value = os.environ.get("SHEET_ARENA_INSPECT_MAX_RANGES", "").strip()
    if env_value:
        try:
            n = int(env_value)
            return min(max(1, n), 1_000)
        except ValueError:
            pass
    try:
        n = int(config.get("max_ranges", 20))
        return min(max(1, n), 1_000)
    except Exception:
        return 20


def _get_max_summary_cells(config: dict[str, Any]) -> int:
    try:
        n = int(config.get("max_summary_cells", 100_000))
        return min(max(1, n), 1_000_000)
    except Exception:
        return 100_000


def _coerce_mode(value: Any) -> str:
    if value is None:
        return "cells"
    if not isinstance(value, str):
        raise ValueError("mode must be 'cells' or 'summary'.")
    mode = value.strip().casefold()
    if mode not in {"cells", "summary"}:
        raise ValueError("mode must be 'cells' or 'summary'.")
    return mode


def _get_max_string_chars() -> int:
    raw = os.environ.get("SHEET_ARENA_TOOL_MAX_STRING_CHARS", "200").strip()
    try:
        n = int(raw)
        return min(max(16, n), 100_000)
    except ValueError:
        return 200


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _coerce_bool_parameter(value: Any, *, parameter: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        token = value.strip().casefold()
        if token == "true":
            return True
        if token == "false":
            return False
    raise ValueError(f"{parameter} received {value!r}; pass true or false.")


def _coerce_ranges_parameter(value: Any) -> list[Any]:
    if isinstance(value, str):
        token = value.strip()
        if not token:
            raise ValueError("ranges received empty; pass a JSON list of A1 range strings.")
        try:
            value = json.loads(token)
        except json.JSONDecodeError as exc:
            raise ValueError('ranges received invalid JSON string; pass a JSON list like ["A1", "B2:C3"].') from exc
    if not isinstance(value, list):
        raise ValueError(f"ranges received {type(value).__name__}; pass a list of A1 range strings.")
    return value


def _truncate_str(value: str, max_chars: int) -> str:
    if max_chars <= 0:
        return ""
    if len(value) <= max_chars:
        return value
    if max_chars <= 3:
        return value[:max_chars]
    return value[: max_chars - 3] + "..."


def _get_max_cf_rules(config: dict[str, Any]) -> int:
    env_value = os.environ.get("SHEET_ARENA_INSPECT_MAX_CF_RULES", "").strip()
    if env_value:
        try:
            n = int(env_value)
            return min(max(0, n), 100_000)
        except ValueError:
            pass

    try:
        n = int(config.get("max_cf_rules", 2000))
        return min(max(0, n), 100_000)
    except Exception:
        return 2000


def _get_max_file_mb(config: dict[str, Any]) -> int:
    env_value = os.environ.get("SHEET_ARENA_INSPECT_MAX_FILE_MB", "").strip()
    if env_value:
        try:
            n = int(env_value)
            return min(max(0, n), 10_000)
        except ValueError:
            pass

    try:
        n = int(config.get("max_file_mb", 100))
        return min(max(0, n), 10_000)
    except Exception:
        return 100


def _scan_zip_metadata(
    file_path: Path,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(file_path) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={getattr(info, 'filename', '?')!r}, "
                        f"bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={getattr(info, 'filename', '?')!r}, "
                        f"ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        f"zip contents too large (total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to scan zip metadata: {exc}"

    return None


def _quote_sheet_name_for_a1(sheet_name: str) -> str:
    if _SAFE_SHEET_NAME_RE.fullmatch(sheet_name):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _merge_unquoted_comma_sheet_tokens(
    range_tokens: list[str],
    sheet_names: list[str],
    *,
    default_sheet: Optional[str] = None,
    separators: Optional[list[str]] = None,
) -> list[str]:
    sheet_lookup = {sheet.casefold(): sheet for sheet in sheet_names}
    default_sheet_key = default_sheet.casefold() if default_sheet else None
    first_sheet_key = sheet_names[0].casefold() if sheet_names else None
    carried_aliases: dict[str, str] = {}
    merged: list[str] = []
    i = 0
    while i < len(range_tokens):
        token = range_tokens[i]
        if "!" in token:
            sheet_part, _, cell_part = token.rpartition("!")
            alias_target = carried_aliases.get(_normalize_sheet_name(sheet_part).casefold())
            if alias_target is None:
                merged.append(token)
            else:
                merged.append(f"{_quote_sheet_name_for_a1(alias_target)}!{cell_part}")
            i += 1
            continue

        prefix_parts = [token]
        merged_token: str | None = None
        consumed = i + 1
        for j in range(i + 1, len(range_tokens)):
            next_token = range_tokens[j]
            if "!" not in next_token:
                prefix_parts.append(next_token)
                continue

            sheet_part, _, cell_part = next_token.rpartition("!")
            next_sheet = _normalize_sheet_name(sheet_part)
            next_sheet_key = next_sheet.casefold()
            next_sheet_exists = sheet_lookup.get(next_sheet_key) is not None
            prefix_can_stand_alone = all(_A1_RANGE_RE.fullmatch(_normalize_cell_range(part)) for part in prefix_parts)
            if separators is None:
                candidates = (
                    ", ".join([*prefix_parts, next_sheet]),
                    ",".join([*prefix_parts, next_sheet]),
                )
            else:
                exact_candidate = range_tokens[i]
                for sep_idx in range(i, j):
                    exact_candidate += separators[sep_idx]
                    exact_candidate += next_sheet if sep_idx == j - 1 else range_tokens[sep_idx + 1]
                candidates = (exact_candidate,)
            for candidate in candidates:
                candidate_key = candidate.casefold()
                actual_sheet = sheet_lookup.get(candidate_key)
                if actual_sheet is not None:
                    if (
                        next_sheet_exists
                        and candidate_key != default_sheet_key
                        and candidate_key != first_sheet_key
                        and prefix_can_stand_alone
                    ):
                        continue
                    merged_token = f"{_quote_sheet_name_for_a1(actual_sheet)}!{cell_part}"
                    if not next_sheet_exists:
                        carried_aliases[next_sheet.casefold()] = actual_sheet
                    consumed = j + 1
                    break
            break

        if merged_token is None:
            merged.append(token)
            i += 1
        else:
            merged.append(merged_token)
            i = consumed
    return merged


def _range_groups_may_need_unquoted_comma_merge(range_groups: list[_RangeGroup]) -> bool:
    return any(len(group.tokens) > 1 and any("!" in token for token in group.tokens[1:]) for group in range_groups)


def _validate_range_groups(
    requested_range_groups: list[_RangeGroup],
    *,
    sheet_name: Optional[str],
    range_boundaries: Any,
    sheet_names: Optional[list[str]] = None,
) -> tuple[list[str], int]:
    range_tokens: list[str] = []
    total_cell_count = 0
    if sheet_names is None:
        range_groups = [group.tokens for group in requested_range_groups]
    else:
        range_groups = [
            _merge_unquoted_comma_sheet_tokens(
                group.tokens,
                sheet_names,
                default_sheet=sheet_name,
                separators=group.separators,
            )
            for group in requested_range_groups
        ]

    for raw_range_group in range_groups:
        default_sheet_for_disjoint = sheet_name
        for raw_range_token in raw_range_group:
            has_sheet_qualified_range = "!" in raw_range_token
            requested_sheet, cell_range = _parse_sheet_cell_range(
                raw_range_token,
                default_sheet_name=default_sheet_for_disjoint or "Sheet1",
            )
            if sheet_name is not None and has_sheet_qualified_range:
                if requested_sheet.casefold() != sheet_name.casefold():
                    raise _SheetMismatchError(expected_sheet=sheet_name, requested_sheet=requested_sheet)
            if has_sheet_qualified_range:
                default_sheet_for_disjoint = requested_sheet
            if not has_sheet_qualified_range and default_sheet_for_disjoint is not None:
                range_tokens.append(f"{_quote_sheet_name_for_a1(default_sheet_for_disjoint)}!{cell_range}")
            else:
                range_tokens.append(raw_range_token)
            boundaries = range_boundaries(cell_range if ":" in cell_range else f"{cell_range}:{cell_range}")
            filled = _fill_missing_boundaries(boundaries)
            min_col, min_row, max_col, max_row = filled
            if min_col < 1 or min_row < 1 or max_col > _EXCEL_MAX_COLS or max_row > _EXCEL_MAX_ROWS:
                raise ValueError("range out of bounds")
            total_cell_count += (max_col - min_col + 1) * (max_row - min_row + 1)

    if total_cell_count <= 0:
        raise ValueError("invalid range size")
    return range_tokens, total_cell_count


def _validate_range_groups_with_workbook(
    *,
    file_path: Path,
    requested_range_groups: list[_RangeGroup],
    sheet_name: Optional[str],
    range_boundaries: Any,
) -> tuple[list[str], int]:
    wb = _load_workbook(file_path, data_only=False, read_only=True)
    try:
        return _validate_range_groups(
            requested_range_groups,
            sheet_name=sheet_name,
            range_boundaries=range_boundaries,
            sheet_names=list(wb.sheetnames),
        )
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _get_data_only_worksheet(
    *,
    file_path: Path,
    sheet: str,
    shared_state: dict[str, Any],
) -> Any:
    wb_values = shared_state.get("wb_values")
    if wb_values is None:
        wb_values = _load_workbook(file_path, data_only=True, read_only=True)
        shared_state["wb_values"] = wb_values

    ws_values_by_sheet = shared_state.setdefault("ws_values_by_sheet", {})
    ws_values = ws_values_by_sheet.get(sheet)
    if ws_values is None:
        ws_values = wb_values[sheet]
        ws_values_by_sheet[sheet] = ws_values
    return ws_values


def _format_a1_address(*, sheet_name: str, row: int, col: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        col_letter = str(col)
    else:
        col_letter = get_column_letter(col)
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    return f"{sheet_ref}!{col_letter}{row}"


def _sample_linear_indices(total: int, *, head: int, tail: int) -> list[int]:
    if total <= 0:
        return []
    head_n = max(0, head)
    tail_n = max(0, tail)
    if total <= head_n + tail_n:
        return list(range(total))
    indices = list(range(head_n))
    indices.extend(range(total - tail_n, total))
    return indices


def _address_for_linear_index(*, sheet: str, min_row: int, min_col: int, cols: int, index: int) -> str:
    row_num = min_row + index // cols
    col_num = min_col + index % cols
    return _format_a1_address(sheet_name=sheet, row=row_num, col=col_num)


def _omitted_from_address(cells: list[Any], returned_count: int, fallback: Optional[str]) -> Optional[str]:
    if returned_count < len(cells):
        cell = cells[returned_count]
        if isinstance(cell, dict):
            address = cell.get("address")
            if isinstance(address, str) and address:
                return address
    return fallback


def _set_prefix_cell_window(
    payload: dict[str, Any],
    metrics: Optional[dict[str, Any]],
    *,
    cells: list[Any],
    returned_count: int,
    total_cells: int,
    omitted_from_fallback: Optional[str] = None,
    force_truncated: bool = False,
) -> None:
    returned = max(0, min(returned_count, len(cells)))
    total = max(total_cells, returned)
    payload["cells"] = cells[:returned]
    payload["returned_cells"] = returned
    if metrics is not None:
        metrics["returned_cells"] = returned

    is_truncated = force_truncated or returned < total
    payload["truncated"] = is_truncated
    if metrics is not None:
        metrics["truncated"] = is_truncated

    if returned < total:
        payload["omitted_cells"] = total - returned
        omitted_from = _omitted_from_address(cells, returned, omitted_from_fallback)
        if omitted_from:
            payload["omitted_from"] = omitted_from
        else:
            payload.pop("omitted_from", None)
    else:
        payload.pop("omitted_cells", None)
        payload.pop("omitted_from", None)


def _fit_prefix_cells_to_budget(
    payload: dict[str, Any],
    metrics: Optional[dict[str, Any]],
    *,
    cells: list[Any],
    total_cells: int,
    max_response_chars: int,
    omitted_from_fallback: Optional[str] = None,
    force_truncated: bool = False,
) -> bool:
    keys = ("cells", "returned_cells", "omitted_cells", "omitted_from", "truncated")
    sentinel = object()
    saved_payload = {key: payload.get(key, sentinel) for key in keys}
    saved_metrics = {}
    if metrics is not None:
        saved_metrics = {key: metrics.get(key, sentinel) for key in ("returned_cells", "truncated")}

    low = 0
    high = len(cells)
    best = -1
    while low <= high:
        mid = (low + high) // 2
        _set_prefix_cell_window(
            payload,
            metrics,
            cells=cells,
            returned_count=mid,
            total_cells=total_cells,
            omitted_from_fallback=omitted_from_fallback,
            force_truncated=force_truncated,
        )
        if len(_json_dumps_compact(payload)) <= max_response_chars:
            best = mid
            low = mid + 1
        else:
            high = mid - 1

    if best >= 1:
        _set_prefix_cell_window(
            payload,
            metrics,
            cells=cells,
            returned_count=best,
            total_cells=total_cells,
            omitted_from_fallback=omitted_from_fallback,
            force_truncated=force_truncated,
        )
        return True

    for key, value in saved_payload.items():
        if value is sentinel:
            payload.pop(key, None)
        else:
            payload[key] = value
    if metrics is not None:
        for key, value in saved_metrics.items():
            if value is sentinel:
                metrics.pop(key, None)
            else:
                metrics[key] = value
    return False


def _fill_missing_boundaries(
    boundaries: tuple[Optional[int], Optional[int], Optional[int], Optional[int]],
) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = boundaries
    if min_col is None:
        min_col = 1
    if max_col is None:
        max_col = _EXCEL_MAX_COLS
    if min_row is None:
        min_row = 1
    if max_row is None:
        max_row = _EXCEL_MAX_ROWS
    return int(min_col), int(min_row), int(max_col), int(max_row)


def _boundaries_intersect(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = a
    b_min_col, b_min_row, b_max_col, b_max_row = b
    if a_max_col < b_min_col or b_max_col < a_min_col:
        return False
    if a_max_row < b_min_row or b_max_row < a_min_row:
        return False
    return True


def _split_sqref(value: Any) -> list[str]:
    if value is None:
        return []
    try:
        if hasattr(value, "ranges"):
            return [str(rng) for rng in list(value.ranges)]
    except Exception:
        pass
    text = str(value).strip()
    if not text:
        return []
    return [part for part in text.split() if part]


def _color_to_json(color: Any) -> Any:
    if color is None:
        return None
    data: dict[str, Any] = {}
    for key in ("type", "rgb", "indexed", "theme", "tint"):
        try:
            value = getattr(color, key)
        except Exception:
            continue
        if key in {"indexed", "theme"} and not isinstance(value, int):
            continue
        if key == "tint" and not isinstance(value, int | float):
            continue
        if key in {"type", "rgb"} and not isinstance(value, str):
            continue
        if value is None or value == "":
            continue
        if isinstance(value, float):
            value = value if math.isfinite(value) else str(value)
        data[key] = value
    return data or None


def _side_to_json(side: Any) -> Any:
    if side is None:
        return None
    try:
        style = getattr(side, "style", None)
    except Exception:
        style = None
    try:
        color = getattr(side, "color", None)
    except Exception:
        color = None
    if style is None and color is None:
        return None
    return {"style": style, "color": _color_to_json(color)}


def _style_to_json(cell: Any) -> Any:
    try:
        if not getattr(cell, "has_style", False):
            return None
    except Exception:
        return None

    style: dict[str, Any] = {}
    try:
        style["style"] = getattr(cell, "style", None)
    except Exception:
        pass
    try:
        style_id = getattr(cell, "style_id", None)
        if style_id is not None:
            style["style_id"] = int(style_id)
    except Exception:
        pass

    try:
        font = getattr(cell, "font", None)
    except Exception:
        font = None
    if font is not None:
        font_data: dict[str, Any] = {}
        for key in ("name", "sz", "b", "i", "u", "strike", "outline", "shadow", "vertAlign"):
            try:
                value = getattr(font, key)
            except Exception:
                continue
            if value is None or value is False:
                continue
            font_data[key] = value
        try:
            font_data["color"] = _color_to_json(getattr(font, "color", None))
        except Exception:
            pass
        style["font"] = font_data or None

    try:
        fill = getattr(cell, "fill", None)
    except Exception:
        fill = None
    if fill is not None:
        fill_data: dict[str, Any] = {}
        for key in ("patternType", "fill_type"):
            try:
                value = getattr(fill, key)
            except Exception:
                continue
            if value:
                fill_data["patternType"] = value
                break
        try:
            fill_data["fgColor"] = _color_to_json(getattr(fill, "fgColor", None))
        except Exception:
            pass
        try:
            fill_data["bgColor"] = _color_to_json(getattr(fill, "bgColor", None))
        except Exception:
            pass
        style["fill"] = fill_data or None

    try:
        border = getattr(cell, "border", None)
    except Exception:
        border = None
    if border is not None:
        border_data: dict[str, Any] = {}
        for key, attr in (
            ("left", "left"),
            ("right", "right"),
            ("top", "top"),
            ("bottom", "bottom"),
            ("diagonal", "diagonal"),
        ):
            try:
                side_data = _side_to_json(getattr(border, attr, None))
            except Exception:
                side_data = None
            if side_data is not None:
                border_data[key] = side_data
        style["border"] = border_data or None

    try:
        alignment = getattr(cell, "alignment", None)
    except Exception:
        alignment = None
    if alignment is not None:
        alignment_data: dict[str, Any] = {}
        for key in (
            "horizontal",
            "vertical",
            "wrap_text",
            "shrink_to_fit",
            "text_rotation",
            "indent",
        ):
            try:
                value = getattr(alignment, key)
            except Exception:
                continue
            if value is None or value is False:
                continue
            alignment_data[key] = value
        style["alignment"] = alignment_data or None

    try:
        protection = getattr(cell, "protection", None)
    except Exception:
        protection = None
    if protection is not None:
        protection_data: dict[str, Any] = {}
        for key in ("locked", "hidden"):
            try:
                value = getattr(protection, key)
            except Exception:
                continue
            if value is None or value is False:
                continue
            protection_data[key] = value
        style["protection"] = protection_data or None

    if all(v is None for v in style.values()):
        return None
    return style


def _serialisable_to_json(obj: Any, *, depth: int = 3) -> Any:
    if obj is None:
        return None
    if obj.__class__.__name__ == "Color":
        return _color_to_json(obj) or str(obj)
    if depth <= 0:
        return str(obj)
    if isinstance(obj, str | bool | int):
        return obj
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else str(obj)
    if isinstance(obj, dt.datetime | dt.date | dt.time):
        return obj.isoformat()
    if isinstance(obj, dt.timedelta):
        return obj.total_seconds()
    if isinstance(obj, Path):
        return str(obj)
    if isinstance(obj, list | tuple | set):
        items = list(obj)
        return [_serialisable_to_json(item, depth=depth - 1) for item in items[:200]]
    if isinstance(obj, dict):
        out: dict[str, Any] = {}
        for idx, (k, v) in enumerate(obj.items()):
            if idx >= 200:
                break
            out[str(k)] = _serialisable_to_json(v, depth=depth - 1)
        return out

    attrs = getattr(obj, "__attrs__", None)
    elems = getattr(obj, "__elements__", None)
    keys: list[str] = []
    if isinstance(attrs, tuple | list):
        keys.extend([str(item) for item in attrs])
    if isinstance(elems, tuple | list):
        keys.extend([str(item) for item in elems])
    if keys:
        out: dict[str, Any] = {"_type": obj.__class__.__name__}
        for key in keys[:200]:
            try:
                value = getattr(obj, key)
            except Exception:
                continue
            out[key] = _serialisable_to_json(value, depth=depth - 1)
        return out

    try:
        raw = vars(obj)
    except Exception:
        raw = None
    if isinstance(raw, dict) and raw:
        out = {"_type": obj.__class__.__name__}
        for idx, (k, v) in enumerate(raw.items()):
            if idx >= 200:
                break
            if str(k).startswith("_"):
                continue
            out[str(k)] = _serialisable_to_json(v, depth=depth - 1)
        if len(out) > 1:
            return out

    return str(obj)


def _has_detail_payload(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value)
    if isinstance(value, dict):
        return any(_has_detail_payload(item) for item in value.values())
    if isinstance(value, list | tuple | set):
        return any(_has_detail_payload(item) for item in value)
    return bool(value)


@dataclass(frozen=True)
class _RangeTarget:
    sheet: str
    cell_range: str
    boundaries: tuple[int, int, int, int]


def _load_workbook(path: Path, *, data_only: bool, read_only: bool):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required; install openpyxl==3.1.5") from None

    try:
        return load_workbook(
            filename=str(path),
            data_only=data_only,
            read_only=read_only,
            keep_links=False,
        )
    except Exception as exc:
        raise RuntimeError(f"failed to load workbook: {exc}") from None


def _resolve_sheet_name(wb: Any, requested_sheet: str) -> str:
    try:
        return str(resolve_worksheet(wb, requested_sheet).title)
    except WorksheetResolutionError as exc:
        raise ValueError(str(exc)) from None


def _build_target(wb: Any, token: str) -> _RangeTarget:
    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required; install openpyxl==3.1.5") from None

    if not wb.sheetnames:
        raise RuntimeError("workbook has no sheets")
    default_sheet = wb.sheetnames[0]

    sheet_name, cell_range = _parse_sheet_cell_range(token, default_sheet_name=default_sheet)
    sheet_name = _resolve_sheet_name(wb, sheet_name)

    try:
        boundaries = range_boundaries(cell_range if ":" in cell_range else f"{cell_range}:{cell_range}")
    except Exception as exc:
        raise ValueError(f"failed to parse range boundaries: {exc}") from None

    filled = _fill_missing_boundaries(boundaries)
    if filled[0] < 1 or filled[1] < 1 or filled[2] > _EXCEL_MAX_COLS or filled[3] > _EXCEL_MAX_ROWS:
        raise ValueError("invalid range boundaries")
    return _RangeTarget(sheet=sheet_name, cell_range=cell_range, boundaries=filled)


def _iter_overlapping_ranges(sqref: Any, *, requested_boundaries: tuple[int, int, int, int]) -> list[str]:
    tokens = _split_sqref(sqref)
    if not tokens:
        return []

    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        return []

    matches: list[str] = []
    for token in tokens:
        token_norm = _normalize_cell_range(token)
        if not _A1_RANGE_RE.fullmatch(token_norm):
            continue
        try:
            boundaries = range_boundaries(token_norm if ":" in token_norm else f"{token_norm}:{token_norm}")
        except Exception:
            continue
        filled = _fill_missing_boundaries(boundaries)
        if _boundaries_intersect(requested_boundaries, filled):
            matches.append(token_norm)
    return matches


def _extract_auto_filter(ws: Any, *, requested_boundaries: tuple[int, int, int, int]) -> Any:
    auto_filter = getattr(ws, "auto_filter", None)
    if auto_filter is None:
        return None

    ref = getattr(auto_filter, "ref", None)
    applies: Optional[bool] = None
    if isinstance(ref, str) and ref.strip():
        try:
            from openpyxl.utils.cell import range_boundaries
        except ImportError:
            applies = None
        else:
            try:
                filled = _fill_missing_boundaries(range_boundaries(_normalize_cell_range(ref)))
                applies = _boundaries_intersect(requested_boundaries, filled)
            except Exception:
                applies = None

    filter_columns = getattr(auto_filter, "filterColumn", None)
    if not isinstance(filter_columns, list):
        filter_columns = []

    out: dict[str, Any] = {
        "ref": ref,
        "applies_to_requested_range": applies,
        "filter_columns": [_serialisable_to_json(col, depth=3) for col in filter_columns[:50]],
        "sort_state": _serialisable_to_json(getattr(auto_filter, "sortState", None), depth=3),
    }
    return out


def _extract_tables(ws: Any, *, requested_boundaries: tuple[int, int, int, int]) -> list[dict[str, Any]]:
    tables_obj = getattr(ws, "tables", None)
    if tables_obj is None:
        return []

    try:
        if hasattr(tables_obj, "values"):
            tables = list(tables_obj.values())
        else:
            tables = list(tables_obj)
    except Exception:
        return []

    out: list[dict[str, Any]] = []
    for table in tables[:200]:
        ref = getattr(table, "ref", None)
        if not isinstance(ref, str) or not ref.strip():
            continue
        try:
            from openpyxl.utils.cell import range_boundaries
        except ImportError:
            continue
        try:
            filled = _fill_missing_boundaries(range_boundaries(_normalize_cell_range(ref)))
        except Exception:
            continue
        if not _boundaries_intersect(requested_boundaries, filled):
            continue

        columns: list[str] = []
        try:
            cols = getattr(table, "tableColumns", None)
            if isinstance(cols, list):
                for col in cols[:200]:
                    col_name = getattr(col, "name", None)
                    if isinstance(col_name, str) and col_name:
                        columns.append(col_name)
        except Exception:
            columns = []

        out.append(
            {
                "name": getattr(table, "name", None),
                "displayName": getattr(table, "displayName", None),
                "ref": ref,
                "style": _serialisable_to_json(getattr(table, "tableStyleInfo", None), depth=2),
                "auto_filter": _serialisable_to_json(getattr(table, "autoFilter", None), depth=2),
                "columns": columns or None,
            }
        )
    return out


def _extract_named_ranges(
    wb: Any,
    *,
    ws: Any,
    requested_sheet: str,
    requested_boundaries: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    wb_defined_names = getattr(wb, "defined_names", None)
    ws_defined_names = getattr(ws, "defined_names", None)
    if wb_defined_names is None and ws_defined_names is None:
        return []

    sources: list[tuple[Any, Optional[str]]] = []
    if wb_defined_names is not None:
        entries: list[Any] = []
        for attr in ("definedName", "defined_names", "values"):
            try:
                maybe = getattr(wb_defined_names, attr)
            except Exception:
                continue
            if isinstance(maybe, list):
                entries = list(maybe)
                break

        if not entries:
            try:
                if hasattr(wb_defined_names, "definedName"):
                    entries = list(wb_defined_names.definedName or [])
            except Exception:
                entries = []

        if not entries:
            try:
                entries = list(wb_defined_names.values())
            except Exception:
                entries = []

        sources.extend([(entry, None) for entry in entries])

    if ws_defined_names is not None:
        try:
            ws_entries = list(ws_defined_names.values())
        except Exception:
            ws_entries = []
        sources.extend([(entry, requested_sheet) for entry in ws_entries])

    out: list[dict[str, Any]] = []
    seen: set[int] = set()
    for entry, fallback_scope in sources:
        if entry is None:
            continue
        entry_id = id(entry)
        if entry_id in seen:
            continue
        seen.add(entry_id)
        name = getattr(entry, "name", None)
        if not isinstance(name, str) or not name:
            continue

        scope = None
        try:
            local_id = getattr(entry, "localSheetId", None)
        except Exception:
            local_id = None
        if local_id is not None:
            try:
                scope = wb.sheetnames[int(local_id)]
            except Exception:
                scope = None
        elif fallback_scope:
            scope = fallback_scope

        destinations: list[dict[str, str]] = []
        try:
            dest_iter = getattr(entry, "destinations", None)
        except Exception:
            dest_iter = None
        if dest_iter is None:
            continue

        try:
            for sheet_name, rng in list(dest_iter):
                if sheet_name is None:
                    continue
                if not isinstance(sheet_name, str) or not isinstance(rng, str):
                    continue
                if sheet_name.lower() != requested_sheet.lower():
                    continue
                rng_norm = _normalize_cell_range(rng)
                if not _A1_RANGE_RE.fullmatch(rng_norm):
                    continue
                overlaps = _iter_overlapping_ranges(rng_norm, requested_boundaries=requested_boundaries)
                if not overlaps:
                    continue
                destinations.append({"sheet": sheet_name, "range": rng_norm})
        except Exception:
            destinations = []

        if not destinations:
            continue

        out.append(
            {
                "name": name,
                "scope": scope,
                "destinations": destinations,
                "value": getattr(entry, "attr_text", None),
            }
        )
    return out


def _extract_data_validations(ws: Any, *, requested_boundaries: tuple[int, int, int, int]) -> list[dict[str, Any]]:
    data_validations = getattr(ws, "data_validations", None)
    if data_validations is None:
        return []
    rules = getattr(data_validations, "dataValidation", None)
    if not isinstance(rules, list):
        return []

    out: list[dict[str, Any]] = []
    for dv in rules[:200]:
        sqref = getattr(dv, "sqref", None)
        overlaps = _iter_overlapping_ranges(sqref, requested_boundaries=requested_boundaries)
        if not overlaps:
            continue

        out.append(
            {
                "ranges": overlaps,
                "type": getattr(dv, "type", None),
                "operator": getattr(dv, "operator", None),
                "allow_blank": getattr(dv, "allowBlank", None),
                "show_drop_down": getattr(dv, "showDropDown", None),
                "show_error_message": getattr(dv, "showErrorMessage", None),
                "show_input_message": getattr(dv, "showInputMessage", None),
                "formula1": getattr(dv, "formula1", None),
                "formula2": getattr(dv, "formula2", None),
                "prompt_title": getattr(dv, "promptTitle", None),
                "prompt": getattr(dv, "prompt", None),
                "error_title": getattr(dv, "errorTitle", None),
                "error": getattr(dv, "error", None),
                "error_style": getattr(dv, "errorStyle", None),
            }
        )
    return out


def _extract_conditional_formatting(
    ws: Any,
    *,
    requested_boundaries: tuple[int, int, int, int],
    max_rules: int,
) -> list[dict[str, Any]]:
    cf = getattr(ws, "conditional_formatting", None)
    if cf is None:
        return []

    out: list[dict[str, Any]] = []
    if max_rules <= 0:
        return out

    for cf_item in list(cf)[:200]:
        sqref = getattr(cf_item, "sqref", None)
        if sqref is None:
            continue
        overlaps = _iter_overlapping_ranges(str(sqref), requested_boundaries=requested_boundaries)
        if not overlaps:
            continue

        rules = getattr(cf_item, "rules", None)
        if not isinstance(rules, list):
            continue
        for rule in rules[:200]:
            data: dict[str, Any] = {
                "ranges": overlaps,
                "type": getattr(rule, "type", None),
                "operator": getattr(rule, "operator", None),
                "priority": getattr(rule, "priority", None),
                "stop_if_true": getattr(rule, "stopIfTrue", None),
                "text": getattr(rule, "text", None),
                "time_period": getattr(rule, "timePeriod", None),
                "formula": getattr(rule, "formula", None),
                "dxf": _serialisable_to_json(getattr(rule, "dxf", None), depth=2),
            }
            out.append(data)
            if len(out) >= max_rules:
                return out

    return out


def _extract_merged_ranges(ws: Any, *, requested_boundaries: tuple[int, int, int, int]) -> list[str]:
    merged_cells = getattr(ws, "merged_cells", None)
    if merged_cells is None:
        return []

    ranges = getattr(merged_cells, "ranges", None)
    if ranges is None:
        return []

    out: list[str] = []
    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        return []

    for rng in list(ranges)[:500]:
        coord = getattr(rng, "coord", None)
        if not isinstance(coord, str) or not coord:
            coord = str(rng)
        coord_norm = _normalize_cell_range(coord)
        if not _A1_RANGE_RE.fullmatch(coord_norm):
            continue
        try:
            filled = _fill_missing_boundaries(range_boundaries(coord_norm))
        except Exception:
            continue
        if _boundaries_intersect(requested_boundaries, filled):
            out.append(coord_norm)
    return out


def _extract_hidden_dims(ws: Any, *, min_row: int, max_row: int, min_col: int, max_col: int) -> dict[str, Any]:
    try:
        from openpyxl.utils import column_index_from_string, get_column_letter
    except ImportError:
        return {"rows": [], "columns": []}

    hidden_rows: list[int] = []
    row_dimensions = getattr(ws, "row_dimensions", None)
    if row_dimensions is not None:
        for row in range(min_row, max_row + 1):
            dim = None
            try:
                dim = row_dimensions.get(row)
            except Exception:
                dim = None
            if dim is None:
                continue
            if getattr(dim, "hidden", None) is True:
                hidden_rows.append(row)

    hidden_col_indices: set[int] = set()
    col_dimensions = getattr(ws, "column_dimensions", None)
    if col_dimensions is not None:
        try:
            items = col_dimensions.items()
        except Exception:
            items = []
        for letter, dim in items:
            if getattr(dim, "hidden", None) is not True:
                continue

            span_min = getattr(dim, "min", None)
            span_max = getattr(dim, "max", None)
            if isinstance(span_min, int) and isinstance(span_max, int) and span_min <= span_max:
                idx_min = span_min
                idx_max = span_max
            elif not isinstance(letter, str) or not letter:
                continue
            elif ":" in letter:
                left, _, right = letter.partition(":")
                try:
                    idx_min = column_index_from_string(left)
                    idx_max = column_index_from_string(right)
                except Exception:
                    continue
                if idx_max < idx_min:
                    idx_min, idx_max = idx_max, idx_min
            else:
                try:
                    idx_min = idx_max = column_index_from_string(letter)
                except Exception:
                    continue

            if idx_max < min_col or idx_min > max_col:
                continue
            start = max(idx_min, min_col)
            end = min(idx_max, max_col)
            for idx in range(start, end + 1):
                hidden_col_indices.add(idx)

    hidden_cols = [get_column_letter(idx) for idx in sorted(hidden_col_indices)]
    return {"rows": hidden_rows, "columns": hidden_cols}


def _inspect_range_sync(
    *,
    file_path: Path,
    display_file: str | None = None,
    range_token: str,
    max_cells: int,
    max_summary_cells: int,
    max_cf_rules: int,
    include_details: bool,
    mode: str,
    max_response_chars: int,
    wb: Any | None = None,
    shared_state: Optional[dict[str, Any]] = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    close_wb = wb is None
    wb_values = None
    if wb is None:
        wb = _load_workbook(file_path, data_only=False, read_only=not include_details)
    try:
        target = _build_target(wb, range_token)
        sheet = target.sheet
        cell_range = target.cell_range
        min_col, min_row, max_col, max_row = target.boundaries
        cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
        if cell_count <= 0:
            raise ValueError("invalid range size")

        ws = wb[sheet]

        rows = max_row - min_row + 1
        cols = max_col - min_col + 1

        if mode == "summary":
            if cell_count > max_summary_cells:
                raise _SummaryRangeTooLargeError(
                    f"summary range is too large (cells={cell_count} > {max_summary_cells}); narrow the range."
                )
            non_empty_cells = 0
            formula_cells = 0
            error_cells = 0
            formula_positions: set[tuple[int, int]] = set()
            for row_num, row in enumerate(
                ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col),
                start=min_row,
            ):
                for col_num, cell in enumerate(row, start=min_col):
                    data_type = getattr(cell, "data_type", None)
                    value = getattr(cell, "value", None)
                    is_formula = data_type == "f"
                    if is_formula:
                        formula_cells += 1
                        formula_positions.add((row_num, col_num))
                    if is_formula or value is not None:
                        non_empty_cells += 1
                    if data_type == "e":
                        error_cells += 1

            if formula_positions:
                summary_wb_values = None
                if shared_state is None:
                    summary_wb_values = _load_workbook(file_path, data_only=True, read_only=True)
                    ws_summary_values = summary_wb_values[sheet]
                else:
                    ws_summary_values = _get_data_only_worksheet(
                        file_path=file_path,
                        sheet=sheet,
                        shared_state=shared_state,
                    )
                try:
                    for row_num, row in enumerate(
                        ws_summary_values.iter_rows(
                            min_row=min_row,
                            max_row=max_row,
                            min_col=min_col,
                            max_col=max_col,
                        ),
                        start=min_row,
                    ):
                        for col_num, cell in enumerate(row, start=min_col):
                            if (row_num, col_num) in formula_positions and getattr(cell, "data_type", None) == "e":
                                error_cells += 1
                finally:
                    if summary_wb_values is not None:
                        summary_wb_values.close()

            payload = {
                "status": "success",
                "file": display_file or str(file_path),
                "sheet": sheet,
                "range": f"{_quote_sheet_name_for_a1(sheet)}!{cell_range}",
                "shape": {
                    "rows": rows,
                    "cols": cols,
                    "cells": cell_count,
                    "min_col": min_col,
                    "min_row": min_row,
                    "max_col": max_col,
                    "max_row": max_row,
                },
                "summary": {
                    "non_empty_cells": non_empty_cells,
                    "empty_cells": cell_count - non_empty_cells,
                    "formula_cells": formula_cells,
                    "error_cells": error_cells,
                    "non_empty_fraction": non_empty_cells / cell_count,
                },
                "truncated": False,
            }
            metrics = {
                "status": "success",
                "file": display_file or str(file_path),
                "sheet": sheet,
                "requested_cells": cell_count,
                "scanned_cells": cell_count,
                "returned_cells": 0,
                "truncated": False,
                "include_details": False,
                "mode": "summary",
            }
            return payload, metrics

        truncated = cell_count > max_cells
        ws_values = None
        try:
            cells: list[dict[str, Any]] = []
            if not truncated:
                for r_idx, row in enumerate(
                    ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col),
                    start=0,
                ):
                    row_num = min_row + r_idx
                    for c_idx, cell in enumerate(row, start=0):
                        col_num = min_col + c_idx
                        is_formula = getattr(cell, "data_type", None) == "f"
                        formula = None
                        if is_formula:
                            raw = getattr(cell, "value", None)
                            if isinstance(raw, str) and raw:
                                formula = raw if raw.startswith("=") else f"={raw}"
                                formula = _truncate_str(formula, _get_max_string_chars())
                        if ws_values is None and is_formula:
                            if shared_state is None:
                                wb_values = _load_workbook(file_path, data_only=True, read_only=True)
                                ws_values = wb_values[sheet]
                            else:
                                ws_values = _get_data_only_worksheet(
                                    file_path=file_path,
                                    sheet=sheet,
                                    shared_state=shared_state,
                                )

                        if is_formula and ws_values is not None:
                            try:
                                val = ws_values.cell(row=row_num, column=col_num).value
                            except Exception:
                                val = None
                        else:
                            val = getattr(cell, "value", None)

                        entry: dict[str, Any] = {
                            "address": _format_a1_address(sheet_name=sheet, row=row_num, col=col_num),
                            "value": _to_jsonable_excel_value(val),
                            "formula": formula,
                        }
                        if include_details:
                            try:
                                entry["number_format"] = getattr(cell, "number_format", None)
                            except Exception:
                                entry["number_format"] = None
                            entry["style"] = _style_to_json(cell)
                        cells.append(entry)
            else:
                for idx in range(min(cell_count, max_cells)):
                    r_off = idx // cols
                    c_off = idx - r_off * cols
                    row_num = min_row + r_off
                    col_num = min_col + c_off
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                    except Exception:
                        cell = None
                    is_formula = getattr(cell, "data_type", None) == "f" if cell is not None else False
                    formula = None
                    if is_formula:
                        raw = getattr(cell, "value", None)
                        if isinstance(raw, str) and raw:
                            formula = raw if raw.startswith("=") else f"={raw}"
                            formula = _truncate_str(formula, _get_max_string_chars())
                    if ws_values is None and is_formula:
                        if shared_state is None:
                            wb_values = _load_workbook(file_path, data_only=True, read_only=True)
                            ws_values = wb_values[sheet]
                        else:
                            ws_values = _get_data_only_worksheet(
                                file_path=file_path,
                                sheet=sheet,
                                shared_state=shared_state,
                            )

                    if is_formula and ws_values is not None:
                        try:
                            val = ws_values.cell(row=row_num, column=col_num).value
                        except Exception:
                            val = None
                    else:
                        val = getattr(cell, "value", None) if cell is not None else None

                    entry = {
                        "address": _format_a1_address(sheet_name=sheet, row=row_num, col=col_num),
                        "value": _to_jsonable_excel_value(val),
                        "formula": formula,
                    }
                    if include_details and cell is not None:
                        try:
                            entry["number_format"] = getattr(cell, "number_format", None)
                        except Exception:
                            entry["number_format"] = None
                        entry["style"] = _style_to_json(cell)
                    cells.append(entry)

            payload: dict[str, Any] = {
                "status": "success",
                "file": display_file or str(file_path),
                "sheet": sheet,
                "range": f"{_quote_sheet_name_for_a1(sheet)}!{cell_range}",
                "shape": {
                    "rows": rows,
                    "cols": cols,
                    "cells": cell_count,
                    "min_col": min_col,
                    "min_row": min_row,
                    "max_col": max_col,
                    "max_row": max_row,
                },
                "truncated": truncated,
                "cells": cells,
            }
            if truncated:
                omitted_from = _address_for_linear_index(
                    sheet=sheet,
                    min_row=min_row,
                    min_col=min_col,
                    cols=cols,
                    index=len(cells),
                )
                _set_prefix_cell_window(
                    payload,
                    None,
                    cells=cells,
                    returned_count=len(cells),
                    total_cells=cell_count,
                    omitted_from_fallback=omitted_from,
                )

            if include_details and not truncated:
                merged_ranges = _extract_merged_ranges(ws, requested_boundaries=target.boundaries)
                hidden = _extract_hidden_dims(ws, min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
                tables = _extract_tables(ws, requested_boundaries=target.boundaries)
                named_ranges = _extract_named_ranges(
                    wb,
                    ws=ws,
                    requested_sheet=sheet,
                    requested_boundaries=target.boundaries,
                )
                auto_filter = _extract_auto_filter(ws, requested_boundaries=target.boundaries)
                data_validations = _extract_data_validations(ws, requested_boundaries=target.boundaries)
                conditional_formatting = _extract_conditional_formatting(
                    ws,
                    requested_boundaries=target.boundaries,
                    max_rules=max_cf_rules,
                )
                for key, value in (
                    ("merged_ranges", merged_ranges),
                    ("hidden", hidden),
                    ("tables", tables),
                    ("named_ranges", named_ranges),
                    ("auto_filter", auto_filter),
                    ("data_validations", data_validations),
                    ("conditional_formatting", conditional_formatting),
                ):
                    if _has_detail_payload(value):
                        payload[key] = value

            metrics = {
                "status": "success",
                "file": display_file or str(file_path),
                "sheet": sheet,
                "requested_cells": cell_count,
                "returned_cells": len(cells),
                "truncated": truncated,
                "include_details": include_details,
                "mode": "cells",
            }
            text = _json_dumps_compact(payload)
            if len(text) <= max_response_chars:
                return payload, metrics

            metadata_keys = (
                "merged_ranges",
                "hidden",
                "tables",
                "named_ranges",
                "auto_filter",
                "data_validations",
                "conditional_formatting",
            )
            removed_any = False
            for key in metadata_keys:
                if key in payload:
                    removed_any = True
                payload.pop(key, None)
            if removed_any:
                payload["truncated"] = True
                metrics["truncated"] = True

            text = _json_dumps_compact(payload)
            if len(text) <= max_response_chars:
                return payload, metrics

            original_cells = payload.get("cells")
            if not isinstance(original_cells, list) or not original_cells:
                return payload, metrics
            original_cells = list(original_cells)
            omitted_from_fallback = None
            if len(original_cells) < cell_count:
                omitted_from_fallback = _address_for_linear_index(
                    sheet=sheet,
                    min_row=min_row,
                    min_col=min_col,
                    cols=cols,
                    index=len(original_cells),
                )

            if _fit_prefix_cells_to_budget(
                payload,
                metrics,
                cells=original_cells,
                total_cells=cell_count,
                max_response_chars=max_response_chars,
                omitted_from_fallback=omitted_from_fallback,
                force_truncated=bool(payload.get("truncated")),
            ):
                return payload, metrics

            stripped_cells: list[Any] = []
            for cell in original_cells:
                if isinstance(cell, dict):
                    stripped = dict(cell)
                    stripped.pop("style", None)
                    stripped.pop("number_format", None)
                    stripped_cells.append(stripped)
                else:
                    stripped_cells.append(cell)

            if _fit_prefix_cells_to_budget(
                payload,
                metrics,
                cells=stripped_cells,
                total_cells=cell_count,
                max_response_chars=max_response_chars,
                omitted_from_fallback=omitted_from_fallback,
                force_truncated=bool(payload.get("truncated")),
            ):
                return payload, metrics

            first_omitted = _omitted_from_address(stripped_cells, 0, omitted_from_fallback)
            _set_prefix_cell_window(
                payload,
                metrics,
                cells=stripped_cells,
                returned_count=0,
                total_cells=cell_count,
                omitted_from_fallback=first_omitted,
                force_truncated=True,
            )
            text = _json_dumps_compact(payload)
            if len(text) <= max_response_chars:
                return payload, metrics

            minimal_payload = {
                "status": "success",
                "truncated": True,
                "returned_cells": 0,
            }
            for key in (
                "file",
                "sheet",
                "range",
                "shape",
            ):
                value = payload.get(key)
                if value is None:
                    continue
                candidate = {**minimal_payload, key: value}
                if len(_json_dumps_compact(candidate)) <= max_response_chars:
                    minimal_payload = candidate
            payload.clear()
            payload.update(minimal_payload)
            return payload, metrics
        finally:
            if shared_state is None and wb_values is not None:
                try:
                    wb_values.close()
                except Exception:
                    pass
    finally:
        if close_wb:
            try:
                wb.close()
            except Exception:
                pass


def _inspect_ranges_sync(
    *,
    file_path: Path,
    display_file: str | None = None,
    range_tokens: list[str],
    max_cells: int,
    max_summary_cells: int,
    max_cf_rules: int,
    include_details: bool,
    mode: str,
    max_response_chars: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = _load_workbook(file_path, data_only=False, read_only=not include_details)
    shared_state: dict[str, Any] = {}
    try:
        range_payloads = []
        range_metrics = []
        for token in range_tokens:
            item_payload, item_metrics = _inspect_range_sync(
                file_path=file_path,
                display_file=display_file,
                range_token=token,
                max_cells=max_cells,
                max_summary_cells=max_summary_cells,
                max_cf_rules=max_cf_rules,
                include_details=include_details,
                mode=mode,
                max_response_chars=max_response_chars,
                wb=wb,
                shared_state=shared_state,
            )
            range_payloads.append(item_payload)
            range_metrics.append(item_metrics)
        return range_payloads, range_metrics
    finally:
        wb_values = shared_state.get("wb_values")
        if wb_values is not None:
            try:
                wb_values.close()
            except Exception:
                pass
        try:
            wb.close()
        except Exception:
            pass


def _trim_multi_range_payload(
    payload: dict[str, Any],
    metrics: dict[str, Any],
    *,
    max_response_chars: int,
) -> None:
    if len(_json_dumps_compact(payload)) <= max_response_chars:
        return

    ranges = payload.get("ranges")
    if isinstance(ranges, list) and payload.get("file") is not None:
        for item in ranges:
            if isinstance(item, dict):
                item.pop("file", None)
        if len(_json_dumps_compact(payload)) <= max_response_chars:
            return

    metadata_keys = (
        "merged_ranges",
        "hidden",
        "tables",
        "named_ranges",
        "auto_filter",
        "data_validations",
        "conditional_formatting",
    )
    for item in payload.get("ranges", []):
        if not isinstance(item, dict):
            continue
        for key in metadata_keys:
            item.pop(key, None)
    payload["truncated"] = True
    metrics["truncated"] = True
    if len(_json_dumps_compact(payload)) <= max_response_chars:
        return

    ranges = payload.get("ranges")
    if (
        isinstance(ranges, list)
        and ranges
        and all(isinstance(item, dict) and isinstance(item.get("summary"), dict) for item in ranges)
    ):
        try:
            total_range_count = int(payload.get("range_count") or len(ranges))
        except (TypeError, ValueError):
            total_range_count = len(ranges)
        original_ranges = list(ranges)
        low = 0
        high = len(original_ranges)
        best_count = -1
        while low <= high:
            count = (low + high) // 2
            payload["ranges"] = original_ranges[:count]
            payload["omitted_ranges"] = max(total_range_count - count, 0)
            if len(_json_dumps_compact(payload)) <= max_response_chars:
                best_count = count
                low = count + 1
            else:
                high = count - 1
        if best_count >= 0:
            payload["ranges"] = original_ranges[:best_count]
            payload["omitted_ranges"] = max(total_range_count - best_count, 0)
            payload["returned_cells"] = 0
            metrics["returned_cells"] = 0
            return
        payload["ranges"] = []
        payload["omitted_ranges"] = total_range_count

    item_cells: list[tuple[dict[str, Any], list[Any], int, Optional[str], bool]] = []
    for item in payload.get("ranges", []):
        if not isinstance(item, dict):
            continue
        prior_truncated = bool(item.get("truncated"))
        cells = item.get("cells")
        if not isinstance(cells, list):
            continue
        shape = item.get("shape")
        total_cells = len(cells)
        if isinstance(shape, dict):
            try:
                total_cells = int(shape.get("cells") or total_cells)
            except (TypeError, ValueError):
                total_cells = len(cells)
        omitted_from = item.get("omitted_from")
        if not isinstance(omitted_from, str):
            omitted_from = None
        item_cells.append((item, list(cells), total_cells, omitted_from, prior_truncated))

    for per_range_cells in (6, 4, 2, 1, 0):
        returned_cells = 0
        for item, cells, total_cells, omitted_from, prior_truncated in item_cells:
            keep = max(0, min(per_range_cells, len(cells)))
            fallback = omitted_from
            if keep < len(cells):
                fallback = _omitted_from_address(cells, keep, omitted_from)
            _set_prefix_cell_window(
                item,
                None,
                cells=cells,
                returned_count=keep,
                total_cells=total_cells,
                omitted_from_fallback=fallback,
                force_truncated=prior_truncated,
            )
            returned_cells += len(item.get("cells") or [])
        payload["returned_cells"] = returned_cells
        metrics["returned_cells"] = returned_cells
        if len(_json_dumps_compact(payload)) <= max_response_chars:
            return

    ranges = payload.get("ranges")
    if isinstance(ranges, list):
        try:
            total_range_count = int(payload.get("range_count") or len(ranges))
        except (TypeError, ValueError):
            total_range_count = len(ranges)
        payload["ranges"] = [
            {
                "status": item.get("status"),
                "sheet": item.get("sheet"),
                "range": item.get("range"),
                "shape": item.get("shape"),
                "truncated": True,
            }
            for item in ranges
            if isinstance(item, dict)
        ]
        ranges = payload["ranges"]
        if len(_json_dumps_compact(payload)) > max_response_chars:
            original_ranges = list(ranges)
            low = 0
            high = len(original_ranges)
            best_count = 0
            while low <= high:
                mid = (low + high) // 2
                payload["ranges"] = original_ranges[:mid]
                payload["omitted_ranges"] = max(total_range_count - mid, 0)
                if len(_json_dumps_compact(payload)) <= max_response_chars:
                    best_count = mid
                    low = mid + 1
                else:
                    high = mid - 1
            payload["ranges"] = original_ranges[:best_count]
            payload["omitted_ranges"] = max(total_range_count - best_count, 0)
        if len(_json_dumps_compact(payload)) <= max_response_chars:
            return
        payload["ranges"] = []
        payload["omitted_ranges"] = total_range_count
    payload["returned_cells"] = 0
    metrics["returned_cells"] = 0
    if len(_json_dumps_compact(payload)) <= max_response_chars:
        return

    file_value = payload.get("file")
    range_count = payload.get("range_count", 0)
    omitted_ranges = payload.get("omitted_ranges", range_count)
    minimal_payload = {
        "status": "success",
        "range_count": range_count,
        "truncated": True,
        "omitted_ranges": omitted_ranges,
        "returned_cells": 0,
    }
    if file_value is not None:
        minimal_with_file = {**minimal_payload, "file": file_value}
        if len(_json_dumps_compact(minimal_with_file)) <= max_response_chars:
            minimal_payload = minimal_with_file
    payload.clear()
    payload.update(minimal_payload)


def _trim_single_range_payload(
    payload: dict[str, Any],
    metrics: dict[str, Any],
    *,
    max_response_chars: int,
) -> None:
    if len(_json_dumps_compact(payload)) <= max_response_chars:
        return

    payload["truncated"] = True
    metrics["truncated"] = True
    file_value = payload.pop("file", None)
    if len(_json_dumps_compact(payload)) <= max_response_chars:
        return

    if file_value is not None:
        payload["file"] = file_value
    cells = payload.get("cells")
    if isinstance(cells, list) and cells:
        shape = payload.get("shape")
        total_cells = len(cells)
        if isinstance(shape, dict):
            try:
                total_cells = int(shape.get("cells") or total_cells)
            except (TypeError, ValueError):
                total_cells = len(cells)
        omitted_from = payload.get("omitted_from")
        if not isinstance(omitted_from, str):
            omitted_from = None
        if _fit_prefix_cells_to_budget(
            payload,
            metrics,
            cells=list(cells),
            total_cells=total_cells,
            max_response_chars=max_response_chars,
            omitted_from_fallback=omitted_from,
            force_truncated=True,
        ):
            return
        _set_prefix_cell_window(
            payload,
            metrics,
            cells=list(cells),
            returned_count=0,
            total_cells=total_cells,
            omitted_from_fallback=_omitted_from_address(list(cells), 0, omitted_from),
            force_truncated=True,
        )
        if len(_json_dumps_compact(payload)) <= max_response_chars:
            return

    if file_value is not None:
        payload.pop("file", None)
        if len(_json_dumps_compact(payload)) <= max_response_chars:
            return

    minimal_payload = {
        "status": "success",
        "truncated": True,
        "returned_cells": 0,
    }
    for key in ("file", "sheet", "range", "shape"):
        value = file_value if key == "file" else payload.get(key)
        if value is None:
            continue
        candidate = {**minimal_payload, key: value}
        if len(_json_dumps_compact(candidate)) <= max_response_chars:
            minimal_payload = candidate
    metrics["returned_cells"] = 0
    payload.clear()
    payload.update(minimal_payload)


def _resolve_workbook_path(
    *,
    workspace_id: Optional[str],
    thread_dir: Optional[str],
    primary_ext: Optional[str],
    relpath: Optional[Path],
) -> tuple[Optional[Path], list[Path]]:
    candidates: list[Path] = []
    if relpath is not None:
        rel_candidates = [relpath]
        if relpath.parts and relpath.parts[0] == "data" and len(relpath.parts) > 1:
            rel_candidates.append(Path(*relpath.parts[1:]))
        for cand in rel_candidates:
            candidates.append(cand)

    ext = _normalize_primary_ext(primary_ext)
    if relpath is None:
        if thread_dir and ext:
            candidates.append(Path(thread_dir) / f"output{ext}")
        if thread_dir:
            candidates.append(Path(thread_dir) / "output.xlsx")
        if ext:
            candidates.append(Path(f"output{ext}"))
        candidates.append(Path("output.xlsx"))

    seen: set[Path] = set()
    unique: list[Path] = []
    for p in candidates:
        if p in seen:
            continue
        seen.add(p)
        unique.append(p)

    attempted: list[Path] = []
    if not workspace_id:
        return None, attempted

    workspace_base = get_spreadsheet_rl_data_root()
    workspaces_base = workspace_base / "_workspaces"
    try:
        if workspace_base.is_symlink():
            return None, attempted
        workspaces_base_resolved = workspaces_base.resolve(strict=True)
    except (OSError, RuntimeError):
        return None, attempted

    root = workspaces_base / workspace_id
    try:
        if root.is_symlink():
            return None, attempted
        root_resolved = root.resolve(strict=True)
    except (OSError, RuntimeError):
        return None, attempted
    if not root_resolved.is_relative_to(workspaces_base_resolved):
        return None, attempted
    for cand in unique:
        attempted.append(root / cand)
        resolved = _resolve_safe_file(root, cand)
        if resolved is not None:
            return resolved, attempted
    return None, attempted


class InspectRangeTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        self.max_cells = _get_max_cells(config)
        self.max_summary_cells = _get_max_summary_cells(config)
        self.max_cf_rules = _get_max_cf_rules(config)
        self.max_file_mb = _get_max_file_mb(config)
        self.max_response_chars = _get_max_response_chars(config)
        self.max_ranges = _get_max_ranges(config)

    def get_openai_tool_schema(self) -> OpenAIFunctionToolSchema:
        return self.tool_schema

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = str(uuid4())
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        requested_range_groups: list[_RangeGroup] = []
        range_token_raw = parameters.get("range")
        ranges_raw = parameters.get("ranges")
        if range_token_raw is not None and ranges_raw is not None:
            return (
                ToolResponse(text="Error: use either 'range' or 'ranges', not both."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_range",
                },
            )
        range_token_count = 0
        try:
            if isinstance(range_token_raw, str) and range_token_raw.strip():
                group = _split_disjoint_range_group(range_token_raw.strip())
                requested_range_groups.append(group)
                if not _range_groups_may_need_unquoted_comma_merge([group]):
                    range_token_count += len(group.tokens)
                    if range_token_count > self.max_ranges:
                        raise _TooManyRangesError
            elif range_token_raw is not None and not (isinstance(range_token_raw, str) and not range_token_raw.strip()):
                return (
                    ToolResponse(text="Error: 'range' must be a non-empty string when provided."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_range",
                    },
                )
            if ranges_raw is not None:
                try:
                    ranges_raw = _coerce_ranges_parameter(ranges_raw)
                except ValueError as exc:
                    return (
                        ToolResponse(text=f"Error: {exc}"),
                        0.0,
                        {
                            "status": "error",
                            "error": "invalid_range",
                        },
                    )
                for idx, item in enumerate(ranges_raw):
                    if not isinstance(item, str) or not item.strip():
                        return (
                            ToolResponse(text=f"Error: 'ranges[{idx}]' must be a non-empty string."),
                            0.0,
                            {
                                "status": "error",
                                "error": "invalid_range",
                            },
                        )
                    group = _split_disjoint_range_group(item.strip())
                    requested_range_groups.append(group)
                    if not _range_groups_may_need_unquoted_comma_merge([group]):
                        range_token_count += len(group.tokens)
                        if range_token_count > self.max_ranges:
                            raise _TooManyRangesError
        except _TooManyRangesError:
            return (
                ToolResponse(text=f"Error: too many ranges requested (maximum is {self.max_ranges})."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_range",
                },
            )
        except ValueError as exc:
            return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "invalid_range"}
        if not any(requested_range_groups):
            return (
                ToolResponse(text="Error: missing/invalid 'range' or 'ranges' (A1 notation)."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_range",
                },
            )
        sheet_name_raw = parameters.get("sheet_name")
        sheet_name: Optional[str] = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return (
                    ToolResponse(text="Error: sheet_name must be a non-empty string."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_sheet_name",
                    },
                )
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return (
                    ToolResponse(text="Error: sheet_name is empty after normalization."),
                    0.0,
                    {
                        "status": "error",
                        "error": "invalid_sheet_name",
                    },
                )
        try:
            include_details = _coerce_bool_parameter(
                parameters.get("include_details", False),
                parameter="include_details",
            )
        except ValueError as exc:
            return (
                ToolResponse(text=f"Error: {exc}"),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_include_details",
                },
            )
        try:
            mode = _coerce_mode(parameters.get("mode"))
        except ValueError as exc:
            return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "invalid_mode"}
        if mode == "summary" and include_details:
            return (
                ToolResponse(text="Error: include_details is unavailable in summary mode; use mode='cells'."),
                0.0,
                {"status": "error", "error": "invalid_mode"},
            )
        try:
            from openpyxl.utils.cell import range_boundaries
        except ImportError:
            return (
                ToolResponse(text="Error: openpyxl is required; install openpyxl==3.1.5."),
                0.0,
                {
                    "status": "error",
                    "error": "missing_dependency",
                },
            )

        range_tokens: list[str] = []
        needs_workbook_range_validation = _range_groups_may_need_unquoted_comma_merge(requested_range_groups)
        try:
            range_tokens, _total_cell_count = _validate_range_groups(
                requested_range_groups,
                sheet_name=sheet_name,
                range_boundaries=range_boundaries,
            )
        except _SheetMismatchError as exc:
            if _range_groups_may_need_unquoted_comma_merge(requested_range_groups):
                needs_workbook_range_validation = True
            else:
                return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "sheet_mismatch"}
        except Exception as exc:
            if _range_groups_may_need_unquoted_comma_merge(requested_range_groups):
                needs_workbook_range_validation = True
            else:
                return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "invalid_range"}

        if not needs_workbook_range_validation and len(range_tokens) > self.max_ranges:
            return (
                ToolResponse(text=f"Error: too many ranges requested (maximum is {self.max_ranges})."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_range",
                },
            )

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return (
                ToolResponse(text="Error: workspace_id is missing/invalid."),
                0.0,
                {
                    "status": "error",
                    "error": "missing_workspace_id",
                },
            )

        path_raw = parameters.get("path")
        if path_raw is None or (isinstance(path_raw, str) and not path_raw.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(path_raw)
            if relpath is None:
                return ToolResponse(text="Error: invalid 'path'."), 0.0, {"status": "error", "error": "invalid_path"}
        if relpath.suffix.lower() != ".xlsx":
            return (
                ToolResponse(text="Error: only .xlsx workbooks are supported."),
                0.0,
                {
                    "status": "error",
                    "error": "invalid_path",
                },
            )

        file_path, attempted = _resolve_workbook_path(
            workspace_id=workspace_id,
            thread_dir=None,
            primary_ext=None,
            relpath=relpath,
        )
        if file_path is None:
            attempted_display: list[str] = []
            root = get_spreadsheet_rl_data_root() / "_workspaces" / workspace_id
            for p in attempted:
                try:
                    attempted_display.append(str(p.relative_to(root)))
                except Exception:
                    attempted_display.append(str(p))
            attempted_text = "\n".join(attempted_display[:12])
            return (
                ToolResponse(
                    text=(
                        "Error: workbook not found.\n"
                        f"Tried:\n{attempted_text}\n"
                        "Save the workbook as data.xlsx in the workspace root, or pass 'path' explicitly."
                    )
                ),
                0.0,
                {"status": "error", "error": "workbook_not_found"},
            )

        try:
            if self.max_file_mb > 0:
                try:
                    file_size = file_path.stat().st_size
                except OSError:
                    file_size = None
                if file_size is not None and file_size > self.max_file_mb * 1024 * 1024:
                    return (
                        ToolResponse(
                            text=(
                                "Error: workbook too large to inspect.\n"
                                f"file: {relpath}\n"
                                f"size_bytes: {file_size}\n"
                                f"max_file_mb: {self.max_file_mb}"
                            )
                        ),
                        0.0,
                        {"status": "error", "error": "file_too_large", "size_bytes": int(file_size)},
                    )

            zip_error = await asyncio.to_thread(
                _scan_zip_metadata,
                file_path,
                max_members=50_000,
                max_total_uncompressed_bytes=512 * 1024 * 1024,
                max_member_uncompressed_bytes=128 * 1024 * 1024,
                max_ratio=200.0,
            )
            if zip_error:
                return (
                    ToolResponse(
                        text=(f"Error: workbook rejected by zip safety checks.\nfile: {relpath}\ndetail: {zip_error}")
                    ),
                    0.0,
                    {"status": "error", "error": "zip_limits_exceeded"},
                )

            if needs_workbook_range_validation:
                try:
                    range_tokens, _total_cell_count = await asyncio.to_thread(
                        _validate_range_groups_with_workbook,
                        file_path=file_path,
                        requested_range_groups=requested_range_groups,
                        sheet_name=sheet_name,
                        range_boundaries=range_boundaries,
                    )
                except _SheetMismatchError as exc:
                    return (
                        ToolResponse(text=f"Error: {exc}"),
                        0.0,
                        {
                            "status": "error",
                            "error": "sheet_mismatch",
                        },
                    )
                except Exception as exc:
                    return (
                        ToolResponse(text=f"Error: {exc}"),
                        0.0,
                        {
                            "status": "error",
                            "error": "invalid_range",
                        },
                    )
                if len(range_tokens) > self.max_ranges:
                    return (
                        ToolResponse(text=f"Error: too many ranges requested (maximum is {self.max_ranges})."),
                        0.0,
                        {
                            "status": "error",
                            "error": "invalid_range",
                        },
                    )

            if mode == "summary" and _total_cell_count > self.max_summary_cells:
                return (
                    ToolResponse(
                        text=(
                            "Error: summary request is too large "
                            f"(cells={_total_cell_count} > {self.max_summary_cells}); narrow the ranges."
                        )
                    ),
                    0.0,
                    {"status": "error", "error": "range_too_large"},
                )

            if len(range_tokens) == 1:
                payload, metrics = await asyncio.to_thread(
                    _inspect_range_sync,
                    file_path=file_path,
                    display_file=str(relpath),
                    range_token=range_tokens[0],
                    max_cells=self.max_cells,
                    max_summary_cells=self.max_summary_cells,
                    max_cf_rules=self.max_cf_rules,
                    include_details=include_details,
                    mode=mode,
                    max_response_chars=self.max_response_chars,
                )
            else:
                range_payloads, range_metrics = await asyncio.to_thread(
                    _inspect_ranges_sync,
                    file_path=file_path,
                    display_file=str(relpath),
                    range_tokens=range_tokens,
                    max_cells=self.max_cells,
                    max_summary_cells=self.max_summary_cells,
                    max_cf_rules=self.max_cf_rules,
                    include_details=include_details,
                    mode=mode,
                    max_response_chars=self.max_response_chars,
                )
                if len(range_payloads) == 1:
                    payload = range_payloads[0]
                    metrics = range_metrics[0]
                else:
                    payload = {
                        "status": "success",
                        "file": str(relpath),
                        "range_count": len(range_payloads),
                        "total_cells": int(sum(item.get("shape", {}).get("cells", 0) for item in range_payloads)),
                        "returned_cells": int(
                            sum(
                                len(item.get("cells", []))
                                for item in range_payloads
                                if isinstance(item.get("cells"), list)
                            )
                        ),
                        "truncated": any(bool(item.get("truncated")) for item in range_payloads),
                        "ranges": range_payloads,
                    }
                    metrics = {
                        "status": "success",
                        "file": str(relpath),
                        "range_count": len(range_payloads),
                        "requested_cells": int(sum(item.get("requested_cells", 0) for item in range_metrics)),
                        "returned_cells": int(sum(item.get("returned_cells", 0) for item in range_metrics)),
                        "truncated": any(bool(item.get("truncated")) for item in range_metrics),
                        "include_details": include_details,
                        "mode": mode,
                    }
                    _trim_multi_range_payload(payload, metrics, max_response_chars=self.max_response_chars)
        except _SummaryRangeTooLargeError as exc:
            return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "range_too_large"}
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return ToolResponse(text=f"Error: {exc}"), 0.0, {"status": "error", "error": "inspect_failed"}

        if int(metrics.get("range_count", 1) or 1) <= 1:
            metrics["file"] = str(relpath)
            _trim_single_range_payload(payload, metrics, max_response_chars=self.max_response_chars)
        return ToolResponse(text=_json_dumps_compact(payload)), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
