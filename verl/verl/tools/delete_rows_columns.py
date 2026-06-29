from __future__ import annotations

import asyncio
import errno
import json
import os
import re
import stat
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import get_sheet_arena_data_root, normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .recalculate import _file_signature, _recalculate_workbook_for_commit
from .schemas import OpenAIFunctionToolSchema, ToolResponse

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384

_A1_COL_RE = re.compile(r"^[A-Z]{1,3}$")
_A1_ROW_RE = re.compile(r"^[0-9]{1,7}$")
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z0-9_]+$")


def _col_letter_to_index(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def _col_index_to_letter(index: int) -> str:
    if index < 1:
        raise ValueError(f"invalid column index: {index}")
    letters: list[str] = []
    current = index
    while current:
        current, rem = divmod(current - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _coerce_token_list(value: Any, *, field_name: str) -> list[str]:
    if not isinstance(value, list) or not value:
        raise RuntimeError(f"{field_name} must be a non-empty list of non-empty strings")

    tokens: list[str] = []
    for idx, item in enumerate(value, start=1):
        if not isinstance(item, str):
            raise RuntimeError(f"{field_name}[{idx}] must be a string")
        token = item.strip()
        if not token:
            raise RuntimeError(f"{field_name}[{idx}] must be a non-empty string")
        tokens.append(token)
    return tokens


def _merge_index_ranges(ranges: list[tuple[int, int]]) -> list[tuple[int, int]]:
    if not ranges:
        return []
    ordered = sorted(ranges, key=lambda item: (item[0], item[1]))
    merged: list[tuple[int, int]] = [ordered[0]]
    for start, end in ordered[1:]:
        prev_start, prev_end = merged[-1]
        if start <= (prev_end + 1):
            merged[-1] = (prev_start, max(prev_end, end))
            continue
        merged.append((start, end))
    return merged


def _format_row_ranges_for_output(sheet_name: str, ranges: list[tuple[int, int]]) -> str:
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    out_parts: list[str] = []
    for start, end in ranges:
        if start == end:
            out_parts.append(f"{sheet_ref}!{start}")
        else:
            out_parts.append(f"{sheet_ref}!{start}:{end}")
    return ",".join(out_parts)


def _format_column_ranges_for_output(sheet_name: str, ranges: list[tuple[int, int]]) -> str:
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    out_parts: list[str] = []
    for start_idx, end_idx in ranges:
        start_col = _col_index_to_letter(start_idx)
        end_col = _col_index_to_letter(end_idx)
        if start_idx == end_idx:
            out_parts.append(f"{sheet_ref}!{start_col}")
        else:
            out_parts.append(f"{sheet_ref}!{start_col}:{end_col}")
    return ",".join(out_parts)


def _parse_row_range_token(token: str) -> tuple[Optional[str], tuple[int, int]]:
    token_sheet, rows_part = _split_sheet_token(token)
    rows_part = rows_part.replace("$", "")
    if ":" in rows_part:
        start_raw, end_raw = rows_part.split(":", 1)
    else:
        start_raw, end_raw = rows_part, rows_part
    start_raw = start_raw.strip()
    end_raw = end_raw.strip()
    if not _A1_ROW_RE.fullmatch(start_raw or "") or not _A1_ROW_RE.fullmatch(end_raw or ""):
        raise RuntimeError(f"invalid row range: {rows_part!r}")

    start_row = int(start_raw)
    end_row = int(end_raw)
    if not 1 <= start_row <= _EXCEL_MAX_ROWS or not 1 <= end_row <= _EXCEL_MAX_ROWS:
        raise RuntimeError("row out of bounds")
    if start_row > end_row:
        start_row, end_row = end_row, start_row
    return token_sheet, (start_row, end_row)


def _parse_column_range_token(token: str) -> tuple[Optional[str], tuple[int, int]]:
    token_sheet, cols_part = _split_sheet_token(token)
    cols_part = cols_part.upper().replace("$", "")
    if ":" in cols_part:
        start_raw, end_raw = cols_part.split(":", 1)
    else:
        start_raw, end_raw = cols_part, cols_part
    start_raw = start_raw.strip()
    end_raw = end_raw.strip()
    if not _A1_COL_RE.fullmatch(start_raw or "") or not _A1_COL_RE.fullmatch(end_raw or ""):
        raise RuntimeError(f"invalid column range: {cols_part!r}")

    start_idx = _col_letter_to_index(start_raw)
    end_idx = _col_letter_to_index(end_raw)
    if not 1 <= start_idx <= _EXCEL_MAX_COLS or not 1 <= end_idx <= _EXCEL_MAX_COLS:
        raise RuntimeError("column out of bounds")
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return token_sheet, (start_idx, end_idx)


def _sanitize_relpath(value: Any) -> Optional[Path]:
    if not isinstance(value, str):
        return None
    raw = value.strip()
    if not raw:
        return None
    if "\0" in raw:
        return None
    p = Path(raw)
    if p.is_absolute():
        return None
    if any(part == ".." for part in p.parts):
        return None
    return p


def _resolve_workspace_file(*, workspace_id: Optional[str], relpath: Path) -> Optional[Path]:
    if not workspace_id:
        return None

    workspace_base = get_sheet_arena_data_root()
    workspaces_base = workspace_base / "_workspaces"
    try:
        if workspace_base.is_symlink():
            return None
        workspaces_base_resolved = workspaces_base.resolve(strict=True)
    except (OSError, RuntimeError):
        return None

    workspace_root = workspaces_base / workspace_id
    try:
        if workspace_root.is_symlink():
            return None
        workspace_root_resolved = workspace_root.resolve(strict=True)
    except (OSError, RuntimeError):
        return None
    if not workspace_root_resolved.is_relative_to(workspaces_base_resolved):
        return None

    rel_candidates = [relpath]
    if relpath.parts and relpath.parts[0] == "data" and len(relpath.parts) > 1:
        rel_candidates.append(Path(*relpath.parts[1:]))

    for rel_candidate in rel_candidates:
        candidate = workspace_root / rel_candidate
        try:
            resolved = candidate.resolve(strict=True)
        except (OSError, RuntimeError):
            continue
        if not resolved.is_relative_to(workspace_root_resolved):
            continue

        if candidate.is_symlink():
            continue

        cursor = workspace_root
        safe = True
        for part in rel_candidate.parts:
            cursor = cursor / part
            try:
                if cursor.is_symlink():
                    safe = False
                    break
            except OSError:
                safe = False
                break
        if not safe:
            continue

        if resolved.is_file():
            return resolved
    return None


def _normalize_sheet_name(value: str) -> str:
    name = value.strip(" \t\n\r\f\v\"")
    if len(name) >= 2 and name[0] == "'" and name[-1] == "'":
        name = name[1:-1].replace("''", "'")
        return name.strip()

    if name.startswith("'") and not name.endswith("'"):
        name = name[1:]
    elif name.endswith("'") and not name.startswith("'"):
        name = name[:-1]
    return name.strip()


def _quote_sheet_name_for_a1(sheet_name: str) -> str:
    if _SAFE_SHEET_NAME_RE.fullmatch(sheet_name):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _split_sheet_token(token: str) -> tuple[Optional[str], str]:
    token = token.strip(" \t\n\r\f\v\"")
    if not token:
        raise ValueError("empty range")
    if "!" not in token:
        return None, token
    sheet_part, _, range_part = token.rpartition("!")
    sheet_name = _normalize_sheet_name(sheet_part)
    if not sheet_name:
        raise ValueError(f"empty sheet name: {token!r}")
    if not range_part.strip():
        raise ValueError(f"empty range: {token!r}")
    return sheet_name, range_part


def _scan_zip_metadata(
    file_path: Path,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(file_path) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={getattr(info, 'filename', '?')!r}, bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={getattr(info, 'filename', '?')!r}, ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        "zip contents too large "
                        f"(total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read zip metadata: {exc}"

    return None


def _acquire_lockfile(lock_path: Path, timeout_s: float):
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    flags = os.O_RDWR | os.O_CREAT
    nofollow = getattr(os, "O_NOFOLLOW", 0)
    if nofollow:
        flags |= nofollow
    else:
        try:
            if lock_path.is_symlink():
                raise RuntimeError(f"lock file path is a symlink: {lock_path}")
        except OSError:
            pass

    try:
        fd = os.open(lock_path, flags, 0o600)
    except OSError as exc:
        if exc.errno == errno.ELOOP:
            raise RuntimeError(f"lock file path is a symlink: {lock_path}") from None
        raise RuntimeError(f"failed to open lock file: {exc}") from None

    try:
        st = os.fstat(fd)
    except OSError as exc:
        os.close(fd)
        raise RuntimeError(f"failed to stat lock file: {exc}") from None
    if not stat.S_ISREG(st.st_mode):
        os.close(fd)
        raise RuntimeError(f"lock file is not a regular file: {lock_path}") from None

    lock_file = os.fdopen(fd, "r+", encoding="utf-8", errors="replace")
    try:
        import fcntl
    except ImportError:
        lock_file.close()
        raise RuntimeError("file locking requires fcntl (Unix-only)") from None

    try:
        deadline = time.monotonic() + max(0.0, timeout_s)
        while True:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                return lock_file
            except BlockingIOError:
                if time.monotonic() >= deadline:
                    raise TimeoutError(f"timed out waiting for lock: {lock_path}") from None
                time.sleep(0.1)
    except BaseException:
        lock_file.close()
        raise


def _resolve_target_worksheet(wb, requested_name: str):
    worksheets = getattr(wb, "worksheets", None) or []
    for ws in worksheets:
        if getattr(ws, "title", None) == requested_name:
            return ws

    requested_cf = requested_name.casefold()
    matches = [ws for ws in worksheets if getattr(ws, "title", "").casefold() == requested_cf]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        candidates = ", ".join(repr(getattr(ws, "title", "")) for ws in matches[:5])
        raise RuntimeError(f"ambiguous sheet name: {requested_name!r} matches {candidates}")

    sheetnames = getattr(wb, "sheetnames", None) or []
    if any(isinstance(name, str) and name.casefold() == requested_cf for name in sheetnames):
        raise RuntimeError(f"sheet is not a worksheet: {requested_name!r}")
    raise RuntimeError(f"sheet not found: {requested_name!r}")


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _delete_rows_in_workbook(
    *,
    file_path: Path,
    sheet_name: Optional[str],
    rows_token: list[str],
    lock_timeout_s: float,
    max_delete_rows: int,
    recalc_url: str,
    recalc_timeout_s: float,
    max_response_bytes: Optional[int],
) -> tuple[str, str, int, int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), lock_timeout_s)
    wb = None
    tmp_path: Optional[Path] = None
    try:
        zip_error = _scan_zip_metadata(
            file_path,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

        try:
            orig_mode = stat.S_IMODE(file_path.stat().st_mode)
            expected_sig = _file_signature(file_path)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook: {exc}") from None

        try:
            wb = load_workbook(filename=str(file_path), keep_vba=False, data_only=False, keep_links=True)
        except Exception as exc:
            raise RuntimeError(f"failed to load workbook: {exc}") from None

        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")

        row_tokens = _coerce_token_list(rows_token, field_name="rows")
        target_sheet_name = sheet_name
        row_ranges: list[tuple[int, int]] = []
        has_sheet_qualified_tokens = False
        has_unqualified_tokens = False
        for token in row_tokens:
            token_sheet, row_range = _parse_row_range_token(token)
            if token_sheet is not None:
                has_sheet_qualified_tokens = True
                if target_sheet_name is None:
                    target_sheet_name = token_sheet
                elif token_sheet.casefold() != target_sheet_name.casefold():
                    raise RuntimeError(f"sheet mismatch: {target_sheet_name!r} vs {token_sheet!r}")
            else:
                has_unqualified_tokens = True
            row_ranges.append(row_range)
        if sheet_name is None and has_sheet_qualified_tokens and has_unqualified_tokens:
            raise RuntimeError("mixed sheet-qualified and unqualified row ranges require sheet_name")

        merged_ranges = _merge_index_ranges(row_ranges)
        amount = sum((end - start + 1) for start, end in merged_ranges)
        if max_delete_rows > 0 and amount > max_delete_rows:
            raise RuntimeError(f"too many rows to delete (rows={amount}, max={max_delete_rows})")

        if target_sheet_name is None:
            ws = worksheets[0]
        else:
            ws = _resolve_target_worksheet(wb, target_sheet_name)
        resolved_sheet = getattr(ws, "title", target_sheet_name) or target_sheet_name or ""

        for start_row, end_row in reversed(merged_ranges):
            ws.delete_rows(start_row, end_row - start_row + 1)
        remaining_rows = int(getattr(ws, "max_row", 0) or 0)
        remaining_cols = int(getattr(ws, "max_column", 0) or 0)

        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{file_path.name}.",
            suffix=file_path.suffix,
            dir=str(file_path.parent),
        )
        tmp_path = Path(tmp_name)
        try:
            with os.fdopen(fd, "w+b") as f:
                fd = -1
                wb.save(f)
        finally:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
        try:
            os.chmod(tmp_path, orig_mode)
        except OSError:
            pass
        if recalc_url:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
                wb = None
            try:
                lock_file.close()
            except Exception:
                pass
            lock_file = None
            recalc_content = _recalculate_workbook_for_commit(
                url=recalc_url,
                file_path=tmp_path,
                filename=file_path.name,
                timeout_s=recalc_timeout_s,
                max_response_bytes=max_response_bytes,
            )
            tmp_path.write_bytes(recalc_content)
            try:
                os.chmod(tmp_path, orig_mode)
            except OSError:
                pass
            lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), lock_timeout_s)
            try:
                current_sig = _file_signature(file_path)
            except OSError as exc:
                raise RuntimeError(f"failed to stat workbook before recalc writeback: {exc}") from None
            if current_sig != expected_sig:
                raise RuntimeError("workbook changed since delete rows request; aborting writeback")
        os.replace(tmp_path, file_path)
        tmp_path = None

        out_range = _format_row_ranges_for_output(resolved_sheet, merged_ranges)
        return resolved_sheet, out_range, amount, remaining_rows, remaining_cols
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if lock_file is not None:
            try:
                lock_file.close()
            except Exception:
                pass


def _delete_columns_in_workbook(
    *,
    file_path: Path,
    sheet_name: Optional[str],
    columns_token: list[str],
    lock_timeout_s: float,
    max_delete_cols: int,
    recalc_url: str,
    recalc_timeout_s: float,
    max_response_bytes: Optional[int],
) -> tuple[str, str, int, int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), lock_timeout_s)
    wb = None
    tmp_path: Optional[Path] = None
    try:
        zip_error = _scan_zip_metadata(
            file_path,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

        try:
            orig_mode = stat.S_IMODE(file_path.stat().st_mode)
            expected_sig = _file_signature(file_path)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook: {exc}") from None

        try:
            wb = load_workbook(filename=str(file_path), keep_vba=False, data_only=False, keep_links=True)
        except Exception as exc:
            raise RuntimeError(f"failed to load workbook: {exc}") from None

        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")

        column_tokens = _coerce_token_list(columns_token, field_name="columns")
        target_sheet_name = sheet_name
        column_ranges: list[tuple[int, int]] = []
        has_sheet_qualified_tokens = False
        has_unqualified_tokens = False
        for token in column_tokens:
            token_sheet, col_range = _parse_column_range_token(token)
            if token_sheet is not None:
                has_sheet_qualified_tokens = True
                if target_sheet_name is None:
                    target_sheet_name = token_sheet
                elif token_sheet.casefold() != target_sheet_name.casefold():
                    raise RuntimeError(f"sheet mismatch: {target_sheet_name!r} vs {token_sheet!r}")
            else:
                has_unqualified_tokens = True
            column_ranges.append(col_range)
        if sheet_name is None and has_sheet_qualified_tokens and has_unqualified_tokens:
            raise RuntimeError("mixed sheet-qualified and unqualified column ranges require sheet_name")

        merged_ranges = _merge_index_ranges(column_ranges)
        amount = sum((end - start + 1) for start, end in merged_ranges)
        if max_delete_cols > 0 and amount > max_delete_cols:
            raise RuntimeError(f"too many columns to delete (cols={amount}, max={max_delete_cols})")

        if target_sheet_name is None:
            ws = worksheets[0]
        else:
            ws = _resolve_target_worksheet(wb, target_sheet_name)
        resolved_sheet = getattr(ws, "title", target_sheet_name) or target_sheet_name or ""

        for start_idx, end_idx in reversed(merged_ranges):
            ws.delete_cols(start_idx, end_idx - start_idx + 1)
        remaining_rows = int(getattr(ws, "max_row", 0) or 0)
        remaining_cols = int(getattr(ws, "max_column", 0) or 0)

        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{file_path.name}.",
            suffix=file_path.suffix,
            dir=str(file_path.parent),
        )
        tmp_path = Path(tmp_name)
        try:
            with os.fdopen(fd, "w+b") as f:
                fd = -1
                wb.save(f)
        finally:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
        try:
            os.chmod(tmp_path, orig_mode)
        except OSError:
            pass
        if recalc_url:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
                wb = None
            try:
                lock_file.close()
            except Exception:
                pass
            lock_file = None
            recalc_content = _recalculate_workbook_for_commit(
                url=recalc_url,
                file_path=tmp_path,
                filename=file_path.name,
                timeout_s=recalc_timeout_s,
                max_response_bytes=max_response_bytes,
            )
            tmp_path.write_bytes(recalc_content)
            try:
                os.chmod(tmp_path, orig_mode)
            except OSError:
                pass
            lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), lock_timeout_s)
            try:
                current_sig = _file_signature(file_path)
            except OSError as exc:
                raise RuntimeError(f"failed to stat workbook before recalc writeback: {exc}") from None
            if current_sig != expected_sig:
                raise RuntimeError("workbook changed since delete columns request; aborting writeback")
        os.replace(tmp_path, file_path)
        tmp_path = None

        out_range = _format_column_ranges_for_output(resolved_sheet, merged_ranges)
        return resolved_sheet, out_range, amount, remaining_rows, remaining_cols
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if lock_file is not None:
            try:
                lock_file.close()
            except Exception:
                pass


class DeleteRowsTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)

        max_delete_rows_raw = config.get("max_delete_rows", os.environ.get("SHEET_ARENA_DELETE_ROWS_MAX", "10000"))
        try:
            self.max_delete_rows = max(0, int(max_delete_rows_raw))
        except (TypeError, ValueError):
            self.max_delete_rows = 10_000

        self.recalc_url = (
            str(config.get("recalc_url") or "").strip()
            or os.environ.get("SHEET_ARENA_RECALC_URL", "").strip()
            or "http://127.0.0.1:5000/recalculate"
        )
        recalc_timeout_s_raw = config.get("recalc_timeout_s", 180)
        try:
            self.recalc_timeout_s = max(0.0, float(recalc_timeout_s_raw))
        except (TypeError, ValueError):
            self.recalc_timeout_s = 180.0

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return ToolResponse(text="Error: invalid path."), 0.0, {"status": "error", "error": "invalid_path"}
        if relpath.suffix.lower() != ".xlsx":
            return ToolResponse(text="Error: only .xlsx workbooks are supported."), 0.0, {
                "status": "error",
                "error": "invalid_path",
            }

        rows_raw = parameters.get("rows")
        try:
            rows_token = _coerce_token_list(rows_raw, field_name="rows")
        except RuntimeError:
            return ToolResponse(
                text=(
                    "Error: rows must be a non-empty list of row range strings "
                    "(e.g. ['2:3'] or ['3', '5', '7:9'])."
                )
            ), 0.0, {"status": "error", "error": "invalid_rows"}

        sheet_name_raw = parameters.get("sheet_name")
        sheet_name = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return ToolResponse(text="Error: sheet_name must be a non-empty string."), 0.0, {
                    "status": "error",
                    "error": "invalid_sheet_name",
                }
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return ToolResponse(text="Error: sheet_name is empty after normalization."), 0.0, {
                    "status": "error",
                    "error": "invalid_sheet_name",
                }

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return ToolResponse(text="Error: workspace_id is missing/invalid."), 0.0, {
                "status": "error",
                "error": "missing_workspace_id",
            }
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return ToolResponse(text=f"Error: file not found: {relpath}"), 0.0, {
                "status": "error",
                "error": "file_not_found",
            }
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return ToolResponse(text=f"Error: failed to stat file: {exc}"), 0.0, {
                    "status": "error",
                    "error": "stat_failed",
                }
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return ToolResponse(text=f"Error: workbook is too large ({actual_mb:.1f}MB > {max_mb}MB)."), 0.0, {
                    "status": "error",
                    "error": "file_too_large",
                }

        try:
            resolved_sheet, range_out, deleted_rows, max_row, max_col = await asyncio.to_thread(
                _delete_rows_in_workbook,
                file_path=file_path,
                sheet_name=sheet_name,
                rows_token=rows_token,
                lock_timeout_s=self.lock_timeout_s,
                max_delete_rows=self.max_delete_rows,
                recalc_url=self.recalc_url,
                recalc_timeout_s=self.recalc_timeout_s,
                max_response_bytes=self.max_file_size_bytes,
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return ToolResponse(text=f"Error: failed to delete rows: {exc}"), 0.0, {
                "status": "error",
                "error": "delete_failed",
            }

        payload = {
            "status": "success",
            "file": str(relpath),
            "sheet": resolved_sheet,
            "range": range_out,
            "deleted_rows": deleted_rows,
            "max_row": max_row,
            "max_col": max_col,
        }
        text = _json_dumps_compact(payload)
        metrics = {
            "status": "success",
            "file": str(relpath),
            "deleted_rows": deleted_rows,
        }
        return ToolResponse(text=text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)


class DeleteColumnsTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)

        max_delete_cols_raw = config.get("max_delete_cols", os.environ.get("SHEET_ARENA_DELETE_COLS_MAX", "1000"))
        try:
            self.max_delete_cols = max(0, int(max_delete_cols_raw))
        except (TypeError, ValueError):
            self.max_delete_cols = 1000

        self.recalc_url = (
            str(config.get("recalc_url") or "").strip()
            or os.environ.get("SHEET_ARENA_RECALC_URL", "").strip()
            or "http://127.0.0.1:5000/recalculate"
        )
        recalc_timeout_s_raw = config.get("recalc_timeout_s", 180)
        try:
            self.recalc_timeout_s = max(0.0, float(recalc_timeout_s_raw))
        except (TypeError, ValueError):
            self.recalc_timeout_s = 180.0

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return ToolResponse(text="Error: invalid path."), 0.0, {"status": "error", "error": "invalid_path"}
        if relpath.suffix.lower() != ".xlsx":
            return ToolResponse(text="Error: only .xlsx workbooks are supported."), 0.0, {
                "status": "error",
                "error": "invalid_path",
            }

        columns_raw = parameters.get("columns")
        try:
            columns_token = _coerce_token_list(columns_raw, field_name="columns")
        except RuntimeError:
            return ToolResponse(
                text=(
                    "Error: columns must be a non-empty list of column range strings "
                    "(e.g. ['B:D'] or ['B', 'D', 'F:G'])."
                )
            ), 0.0, {"status": "error", "error": "invalid_columns"}

        sheet_name_raw = parameters.get("sheet_name")
        sheet_name = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return ToolResponse(text="Error: sheet_name must be a non-empty string."), 0.0, {
                    "status": "error",
                    "error": "invalid_sheet_name",
                }
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return ToolResponse(text="Error: sheet_name is empty after normalization."), 0.0, {
                    "status": "error",
                    "error": "invalid_sheet_name",
                }

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return ToolResponse(text="Error: workspace_id is missing/invalid."), 0.0, {
                "status": "error",
                "error": "missing_workspace_id",
            }
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return ToolResponse(text=f"Error: file not found: {relpath}"), 0.0, {
                "status": "error",
                "error": "file_not_found",
            }
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return ToolResponse(text=f"Error: failed to stat file: {exc}"), 0.0, {
                    "status": "error",
                    "error": "stat_failed",
                }
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return ToolResponse(text=f"Error: workbook is too large ({actual_mb:.1f}MB > {max_mb}MB)."), 0.0, {
                    "status": "error",
                    "error": "file_too_large",
                }

        try:
            resolved_sheet, range_out, deleted_cols, max_row, max_col = await asyncio.to_thread(
                _delete_columns_in_workbook,
                file_path=file_path,
                sheet_name=sheet_name,
                columns_token=columns_token,
                lock_timeout_s=self.lock_timeout_s,
                max_delete_cols=self.max_delete_cols,
                recalc_url=self.recalc_url,
                recalc_timeout_s=self.recalc_timeout_s,
                max_response_bytes=self.max_file_size_bytes,
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return ToolResponse(text=f"Error: failed to delete columns: {exc}"), 0.0, {
                "status": "error",
                "error": "delete_failed",
            }

        payload = {
            "status": "success",
            "file": str(relpath),
            "sheet": resolved_sheet,
            "range": range_out,
            "deleted_columns": deleted_cols,
            "max_row": max_row,
            "max_col": max_col,
        }
        text = _json_dumps_compact(payload)
        metrics = {
            "status": "success",
            "file": str(relpath),
            "deleted_columns": deleted_cols,
        }
        return ToolResponse(text=text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
