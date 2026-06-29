from __future__ import annotations

import asyncio
import copy
import os
import re
import shutil
import stat
import tempfile
import uuid
import zipfile
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .clear_range import (
    _acquire_lockfile,
    _format_a1_address,
    _get_max_response_chars,
    _json_dumps_compact,
    _normalize_cell_range,
    _normalize_sheet_name,
    _quote_sheet_name_for_a1,
    _resolve_target_worksheet,
    _resolve_workspace_file,
    _sanitize_relpath,
    _scan_zip_metadata,
    _split_sheet_cell_range,
)
from .recalculate import _file_signature
from .schemas import OpenAIFunctionToolSchema, ToolResponse

_A1_CELL_RE = r"[A-Z]{1,3}[0-9]{1,7}"
_A1_RECT_RANGE_RE = re.compile(rf"^{_A1_CELL_RE}(?::{_A1_CELL_RE})?$")
_COLOR_RE = re.compile(r"^#?(?:[0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$")
_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384
_MAX_NUMBER_FORMAT_CHARS = 255

_HORIZONTAL_ALIGNMENTS = {
    "general",
    "left",
    "center",
    "right",
    "fill",
    "justify",
    "centerContinuous",
    "distributed",
}
_VERTICAL_ALIGNMENTS = {"top", "center", "bottom", "justify", "distributed"}
_BORDER_STYLES = {
    "dashDot",
    "dashDotDot",
    "dashed",
    "dotted",
    "double",
    "hair",
    "medium",
    "mediumDashDot",
    "mediumDashDotDot",
    "mediumDashed",
    "slantDashDot",
    "thick",
    "thin",
}


class _FormatOptionError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def _default_tool_schema() -> OpenAIFunctionToolSchema:
    return OpenAIFunctionToolSchema.model_validate(
        {
            "type": "function",
            "function": {
                "name": "format_range",
                "description": (
                    "Apply common formatting to a finite A1 cell range in an .xlsx workbook. Supports fill color, "
                    "font color, bold, italic, underline, number_format, alignment, border style/color, row height, "
                    "and column width. Does not recalculate formulas."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Optional relative path to the workbook file. Defaults to data.xlsx.",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Optional worksheet name. If range includes a sheet name, it must match.",
                        },
                        "range": {
                            "type": "string",
                            "description": "Finite A1 cell range to format, e.g. A1:C3 or Sheet1!B2:D10.",
                        },
                        "fill_color": {
                            "type": "string",
                            "description": "Cell fill/background color as #RRGGBB, RRGGBB, or AARRGGBB.",
                        },
                        "background_color": {
                            "type": "string",
                            "description": "Alias for fill_color.",
                        },
                        "font_color": {
                            "type": "string",
                            "description": "Font color as #RRGGBB, RRGGBB, or AARRGGBB.",
                        },
                        "bold": {"type": "boolean", "description": "Set font bold on or off."},
                        "italic": {"type": "boolean", "description": "Set font italic on or off."},
                        "underline": {"type": "boolean", "description": "Set single underline on or off."},
                        "number_format": {
                            "type": "string",
                            "description": "Excel number format code, e.g. $#,##0.00 or 0.0%.",
                        },
                        "alignment": {
                            "type": "object",
                            "description": (
                                "Optional alignment object with horizontal, vertical, wrap_text, shrink_to_fit, "
                                "or text_rotation."
                            ),
                        },
                        "horizontal_alignment": {
                            "type": "string",
                            "description": "Horizontal alignment, e.g. left, center, right, or general.",
                        },
                        "vertical_alignment": {
                            "type": "string",
                            "description": "Vertical alignment, e.g. top, center, bottom.",
                        },
                        "wrap_text": {"type": "boolean", "description": "Set text wrapping on or off."},
                        "text_rotation": {
                            "type": "integer",
                            "description": "Text rotation from 0 to 180, or 255 for stacked text.",
                        },
                        "border_style": {
                            "type": "string",
                            "description": "Border style for all sides, e.g. thin, medium, thick, dashed, dotted.",
                        },
                        "border_color": {
                            "type": "string",
                            "description": "Border color as #RRGGBB, RRGGBB, or AARRGGBB.",
                        },
                        "row_height": {
                            "type": "number",
                            "description": "Optional row height applied to rows touched by the range.",
                        },
                        "column_width": {
                            "type": "number",
                            "description": "Optional column width applied to columns touched by the range.",
                        },
                    },
                    "required": ["range"],
                },
            },
        }
    )


def _error_response(error: str, message: str, **extra: Any) -> tuple[ToolResponse, float, dict[str, Any]]:
    payload = {"status": "error", "error": error, "message": message, "truncated": False}
    payload.update(extra)
    metrics = {"status": "error", "error": error, "truncated": False}
    return ToolResponse(text=_json_dumps_compact(payload)), 0.0, metrics


def _normalize_color(value: Any, *, field_name: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise _FormatOptionError("invalid_color", f"{field_name} must be a hex color like #RRGGBB.")
    token = value.strip()
    if not _COLOR_RE.fullmatch(token):
        raise _FormatOptionError("invalid_color", f"{field_name} must be a hex color like #RRGGBB.")
    hex_value = token.lstrip("#").upper()
    if len(hex_value) == 6:
        return f"FF{hex_value}"
    return hex_value


def _parse_bool_option(parameters: dict[str, Any], key: str) -> Optional[bool]:
    if key not in parameters or parameters.get(key) is None:
        return None
    value = parameters.get(key)
    if not isinstance(value, bool):
        raise _FormatOptionError(f"invalid_{key}", f"{key} must be a boolean.")
    return value


def _parse_number_option(parameters: dict[str, Any], key: str, *, minimum: float, maximum: float) -> Optional[float]:
    if key not in parameters or parameters.get(key) is None:
        return None
    value = parameters.get(key)
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise _FormatOptionError(f"invalid_{key}", f"{key} must be a number.")
    number = float(value)
    if not minimum <= number <= maximum:
        raise _FormatOptionError(f"invalid_{key}", f"{key} must be between {minimum:g} and {maximum:g}.")
    return number


def _merge_alignment_value(
    alignment: dict[str, Any],
    *,
    key: str,
    value: Any,
    source: str,
) -> None:
    if value is None:
        return
    if key in alignment and alignment[key] != value:
        raise _FormatOptionError(
            "invalid_alignment",
            f"conflicting alignment option for {key}: alignment.{key} and {source}.",
        )
    alignment[key] = value


def _parse_alignment_options(parameters: dict[str, Any]) -> dict[str, Any]:
    alignment: dict[str, Any] = {}
    raw_alignment = parameters.get("alignment")
    if raw_alignment is not None:
        if not isinstance(raw_alignment, dict):
            raise _FormatOptionError("invalid_alignment", "alignment must be an object.")
        allowed = {"horizontal", "vertical", "wrap_text", "shrink_to_fit", "text_rotation"}
        unknown = sorted(str(key) for key in raw_alignment if key not in allowed)
        if unknown:
            raise _FormatOptionError("invalid_alignment", f"unsupported alignment keys: {', '.join(unknown)}.")
        for key in allowed:
            if key in raw_alignment:
                alignment[key] = raw_alignment[key]

    horizontal = parameters.get("horizontal_alignment")
    if horizontal is not None:
        _merge_alignment_value(alignment, key="horizontal", value=horizontal, source="horizontal_alignment")

    vertical = parameters.get("vertical_alignment")
    if vertical is not None:
        _merge_alignment_value(alignment, key="vertical", value=vertical, source="vertical_alignment")

    if "wrap_text" in parameters:
        wrap_text = _parse_bool_option(parameters, "wrap_text")
        _merge_alignment_value(alignment, key="wrap_text", value=wrap_text, source="wrap_text")

    if "text_rotation" in parameters:
        text_rotation_raw = parameters.get("text_rotation")
        if isinstance(text_rotation_raw, bool):
            raise _FormatOptionError("invalid_alignment", "text_rotation must be an integer.")
        try:
            text_rotation = int(text_rotation_raw)
        except (TypeError, ValueError):
            raise _FormatOptionError("invalid_alignment", "text_rotation must be an integer.") from None
        _merge_alignment_value(alignment, key="text_rotation", value=text_rotation, source="text_rotation")

    if "shrink_to_fit" in parameters:
        shrink_to_fit = _parse_bool_option(parameters, "shrink_to_fit")
        _merge_alignment_value(alignment, key="shrink_to_fit", value=shrink_to_fit, source="shrink_to_fit")

    if not alignment:
        return {}

    horizontal_value = alignment.get("horizontal")
    if horizontal_value is not None:
        if not isinstance(horizontal_value, str) or horizontal_value not in _HORIZONTAL_ALIGNMENTS:
            allowed = ", ".join(sorted(_HORIZONTAL_ALIGNMENTS))
            raise _FormatOptionError("invalid_alignment", f"horizontal alignment must be one of: {allowed}.")

    vertical_value = alignment.get("vertical")
    if vertical_value is not None:
        if not isinstance(vertical_value, str) or vertical_value not in _VERTICAL_ALIGNMENTS:
            allowed = ", ".join(sorted(_VERTICAL_ALIGNMENTS))
            raise _FormatOptionError("invalid_alignment", f"vertical alignment must be one of: {allowed}.")

    for key in ("wrap_text", "shrink_to_fit"):
        if key in alignment and not isinstance(alignment[key], bool):
            raise _FormatOptionError("invalid_alignment", f"alignment.{key} must be a boolean.")

    if "text_rotation" in alignment:
        text_rotation = alignment["text_rotation"]
        if isinstance(text_rotation, bool) or not isinstance(text_rotation, int):
            raise _FormatOptionError("invalid_alignment", "alignment.text_rotation must be an integer.")
        if text_rotation != 255 and not 0 <= text_rotation <= 180:
            raise _FormatOptionError("invalid_alignment", "alignment.text_rotation must be 0-180 or 255.")

    return alignment


def _parse_format_options(parameters: dict[str, Any]) -> dict[str, Any]:
    options: dict[str, Any] = {}

    fill_color_raw = parameters.get("fill_color")
    background_color_raw = parameters.get("background_color")
    if fill_color_raw is not None and background_color_raw is not None:
        fill_color = _normalize_color(fill_color_raw, field_name="fill_color")
        background_color = _normalize_color(background_color_raw, field_name="background_color")
        if fill_color != background_color:
            raise _FormatOptionError("invalid_color", "fill_color and background_color conflict.")
        options["fill_color"] = fill_color
    elif fill_color_raw is not None:
        options["fill_color"] = _normalize_color(fill_color_raw, field_name="fill_color")
    elif background_color_raw is not None:
        options["fill_color"] = _normalize_color(background_color_raw, field_name="background_color")

    if parameters.get("font_color") is not None:
        options["font_color"] = _normalize_color(parameters.get("font_color"), field_name="font_color")

    for key in ("bold", "italic", "underline"):
        value = _parse_bool_option(parameters, key)
        if value is not None:
            options[key] = value

    if parameters.get("number_format") is not None:
        number_format = parameters.get("number_format")
        if not isinstance(number_format, str) or not number_format.strip():
            raise _FormatOptionError("invalid_number_format", "number_format must be a non-empty string.")
        number_format = number_format.strip()
        if len(number_format) > _MAX_NUMBER_FORMAT_CHARS:
            raise _FormatOptionError(
                "invalid_number_format",
                f"number_format is too long (len={len(number_format)} > {_MAX_NUMBER_FORMAT_CHARS}).",
            )
        options["number_format"] = number_format

    alignment = _parse_alignment_options(parameters)
    if alignment:
        options["alignment"] = alignment

    border_style = parameters.get("border_style")
    border_color = parameters.get("border_color")
    if border_style is not None:
        if not isinstance(border_style, str) or border_style not in _BORDER_STYLES:
            allowed = ", ".join(sorted(_BORDER_STYLES))
            raise _FormatOptionError("invalid_border", f"border_style must be one of: {allowed}.")
        options["border_style"] = border_style
    if border_color is not None:
        options["border_color"] = _normalize_color(border_color, field_name="border_color")
        options.setdefault("border_style", "thin")

    row_height = _parse_number_option(parameters, "row_height", minimum=0.1, maximum=409)
    if row_height is not None:
        options["row_height"] = row_height

    column_width = _parse_number_option(parameters, "column_width", minimum=0.1, maximum=255)
    if column_width is not None:
        options["column_width"] = column_width

    if not options:
        raise _FormatOptionError("no_format_options", "provide at least one formatting option.")
    return options


def _validate_range_token(range_token: str, sheet_name: Optional[str]) -> tuple[Optional[str], str]:
    try:
        parsed_sheet, range_part = _split_sheet_cell_range(range_token)
    except ValueError as exc:
        raise _FormatOptionError("invalid_range", str(exc)) from None
    if parsed_sheet is not None and sheet_name is not None and parsed_sheet.casefold() != sheet_name.casefold():
        raise _FormatOptionError(
            "sheet_mismatch",
            f"sheet_name {sheet_name!r} does not match range sheet {parsed_sheet!r}.",
        )
    cell_range = _normalize_cell_range(range_part)
    if not _A1_RECT_RANGE_RE.fullmatch(cell_range):
        raise _FormatOptionError(
            "invalid_range",
            "range must be a finite A1 cell or rectangular cell range such as A1:C3.",
        )
    return parsed_sheet, cell_range


def _scan_xlsx_container_path(file_path: Path) -> Optional[str]:
    try:
        with zipfile.ZipFile(file_path) as zf:
            names = {str(name).replace("\\", "/") for name in zf.namelist()}
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read xlsx container: {exc}"

    for member in ("[Content_Types].xml", "xl/workbook.xml"):
        if member not in names:
            return f"missing required xlsx member: {member}"
    if not any(name.startswith("xl/worksheets/") and name.endswith(".xml") for name in names):
        return "missing worksheet parts"
    return None


def _validate_workbook_loadable(file_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    zip_error = _scan_zip_metadata(
        file_path,
        max_members=50_000,
        max_total_uncompressed_bytes=512 * 1024 * 1024,
        max_member_uncompressed_bytes=128 * 1024 * 1024,
        max_ratio=200.0,
    )
    if zip_error:
        raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

    xlsx_error = _scan_xlsx_container_path(file_path)
    if xlsx_error:
        raise RuntimeError(f"workbook rejected by xlsx safety checks: {xlsx_error}")

    wb = None
    try:
        wb = load_workbook(filename=str(file_path), data_only=False, read_only=True, keep_links=False)
        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError(f"failed to load workbook: {exc}") from None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _copy_workbook_to_temp_under_lock(
    *,
    file_path: Path,
    lock_timeout_s: float,
) -> tuple[Path, tuple[int, int, int, int], int]:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    tmp_path: Optional[Path] = None
    try:
        _validate_workbook_loadable(file_path)
        try:
            expected_sig = _file_signature(file_path)
            orig_mode = stat.S_IMODE(file_path.stat().st_mode)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook: {exc}") from None

        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{file_path.name}.",
            suffix=file_path.suffix,
            dir=str(file_path.parent),
        )
        tmp_path = Path(tmp_name)
        try:
            with os.fdopen(fd, "wb") as dst:
                fd = -1
                with file_path.open("rb") as src:
                    shutil.copyfileobj(src, dst)
        finally:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
        try:
            os.chmod(tmp_path, orig_mode)
        except OSError:
            pass
        copied_path = tmp_path
        tmp_path = None
        return copied_path, expected_sig, orig_mode
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
        try:
            lock_file.close()
        except Exception:
            pass


def _commit_temp_workbook_under_lock(
    *,
    file_path: Path,
    tmp_path: Path,
    expected_sig: tuple[int, int, int, int],
    orig_mode: int,
    lock_timeout_s: float,
) -> None:
    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    try:
        try:
            current_sig = _file_signature(file_path)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook before writeback: {exc}") from None
        if current_sig != expected_sig:
            raise RuntimeError("workbook changed since format request; aborting writeback")
        try:
            os.chmod(tmp_path, orig_mode)
        except OSError:
            pass
        os.replace(tmp_path, file_path)
    finally:
        try:
            lock_file.close()
        except Exception:
            pass


def _color_to_argb(color: Any) -> Optional[str]:
    if color is None:
        return None
    try:
        if getattr(color, "type", None) == "rgb":
            rgb = getattr(color, "rgb", None)
            return str(rgb).upper() if rgb else None
    except Exception:
        return None
    return None


def _format_sample(cell: Any) -> dict[str, Any]:
    return {
        "fill_color": _color_to_argb(getattr(getattr(cell, "fill", None), "fgColor", None)),
        "font_color": _color_to_argb(getattr(getattr(cell, "font", None), "color", None)),
        "bold": bool(getattr(getattr(cell, "font", None), "bold", False)),
        "italic": bool(getattr(getattr(cell, "font", None), "italic", False)),
        "underline": getattr(getattr(cell, "font", None), "underline", None),
        "number_format": getattr(cell, "number_format", None),
        "horizontal_alignment": getattr(getattr(cell, "alignment", None), "horizontal", None),
        "vertical_alignment": getattr(getattr(cell, "alignment", None), "vertical", None),
        "wrap_text": getattr(getattr(cell, "alignment", None), "wrap_text", None),
    }


def _sample_coords(min_row: int, max_row: int, min_col: int, max_col: int) -> list[tuple[int, int]]:
    candidates = [
        (min_row, min_col),
        (min_row, max_col),
        (max_row, min_col),
        (max_row, max_col),
    ]
    seen: set[tuple[int, int]] = set()
    out: list[tuple[int, int]] = []
    for coord in candidates:
        if coord not in seen:
            out.append(coord)
            seen.add(coord)
    return out


def _truncate_success_payload_to_max_chars(payload: dict[str, Any], max_chars: int) -> str:
    payload["truncated"] = False
    response_text = _json_dumps_compact(payload)
    if len(response_text) <= max_chars:
        return response_text

    payload["truncated"] = True
    samples = payload.get("sample_formats")
    if isinstance(samples, list) and len(samples) > 1:
        payload["sample_formats"] = samples[:1]
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    payload.pop("sample_formats", None)
    response_text = _json_dumps_compact(payload)
    if len(response_text) <= max_chars:
        return response_text

    payload.pop("applied_options", None)
    response_text = _json_dumps_compact(payload)
    if len(response_text) <= max_chars:
        return response_text

    compact_payload = {
        "status": payload.get("status", "success"),
        "file": payload.get("file"),
        "sheet": payload.get("sheet"),
        "formatted_range": payload.get("formatted_range"),
        "formatted_cells": payload.get("formatted_cells"),
        "formatted_rows": payload.get("formatted_rows"),
        "formatted_columns": payload.get("formatted_columns"),
        "skipped_merged_cells": payload.get("skipped_merged_cells"),
        "truncated": True,
    }
    response_text = _json_dumps_compact(compact_payload)
    if len(response_text) <= max_chars:
        payload.clear()
        payload.update(compact_payload)
        return response_text

    compact_payload = {
        "status": payload.get("status", "success"),
        "formatted_range": payload.get("formatted_range"),
        "formatted_cells": payload.get("formatted_cells"),
        "truncated": True,
    }
    response_text = _json_dumps_compact(compact_payload)
    if len(response_text) <= max_chars:
        payload.clear()
        payload.update(compact_payload)
        return response_text

    minimal_payload = {
        "status": payload.get("status", "success"),
        "formatted_cells": payload.get("formatted_cells"),
        "truncated": True,
    }
    response_text = _json_dumps_compact(minimal_payload)
    if len(response_text) <= max_chars:
        payload.clear()
        payload.update(minimal_payload)
        return response_text

    minimal_payload = {"status": payload.get("status", "success"), "truncated": True}
    payload.clear()
    payload.update(minimal_payload)
    return _json_dumps_compact(minimal_payload)


def _apply_cell_format(cell: Any, options: dict[str, Any]) -> None:
    from openpyxl.styles import Alignment, Border, PatternFill, Side

    if "fill_color" in options:
        cell.fill = PatternFill(fill_type="solid", fgColor=options["fill_color"])

    font_options = {"font_color", "bold", "italic", "underline"} & options.keys()
    if font_options:
        font = copy.copy(cell.font)
        if "font_color" in options:
            font.color = options["font_color"]
        if "bold" in options:
            font.bold = options["bold"]
        if "italic" in options:
            font.italic = options["italic"]
        if "underline" in options:
            font.underline = "single" if options["underline"] else None
        cell.font = font

    if "number_format" in options:
        cell.number_format = options["number_format"]

    if "alignment" in options:
        raw_alignment = options["alignment"]
        alignment_kwargs = {
            "horizontal": raw_alignment.get("horizontal"),
            "vertical": raw_alignment.get("vertical"),
            "wrap_text": raw_alignment.get("wrap_text"),
            "shrink_to_fit": raw_alignment.get("shrink_to_fit"),
            "text_rotation": raw_alignment.get("text_rotation"),
        }
        current = copy.copy(cell.alignment)
        alignment = Alignment(
            horizontal=(
                alignment_kwargs["horizontal"]
                if alignment_kwargs["horizontal"] is not None
                else current.horizontal
            ),
            vertical=alignment_kwargs["vertical"] if alignment_kwargs["vertical"] is not None else current.vertical,
            text_rotation=(
                alignment_kwargs["text_rotation"]
                if alignment_kwargs["text_rotation"] is not None
                else current.text_rotation
            ),
            wrap_text=alignment_kwargs["wrap_text"] if alignment_kwargs["wrap_text"] is not None else current.wrap_text,
            shrink_to_fit=(
                alignment_kwargs["shrink_to_fit"]
                if alignment_kwargs["shrink_to_fit"] is not None
                else current.shrink_to_fit
            ),
            indent=current.indent,
            relativeIndent=current.relativeIndent,
            justifyLastLine=current.justifyLastLine,
            readingOrder=current.readingOrder,
        )
        cell.alignment = alignment

    if "border_style" in options:
        side = Side(style=options["border_style"], color=options.get("border_color"))
        current_border = copy.copy(cell.border)
        cell.border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
            diagonal=current_border.diagonal,
            diagonal_direction=current_border.diagonal_direction,
            diagonalUp=current_border.diagonalUp,
            diagonalDown=current_border.diagonalDown,
            outline=current_border.outline,
            vertical=current_border.vertical,
            horizontal=current_border.horizontal,
        )


def _format_temp_workbook(
    *,
    tmp_path: Path,
    sheet_name: Optional[str],
    range_token: str,
    options: dict[str, Any],
    max_format_cells: int,
) -> tuple[str, str, int, list[dict[str, Any]], int, int, int]:
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    wb = None
    try:
        wb = load_workbook(filename=str(tmp_path), keep_vba=False, data_only=False, keep_links=True)
        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")

        default_sheet = getattr(worksheets[0], "title", "") or ""
        parsed_sheet, range_part = _split_sheet_cell_range(range_token)
        if parsed_sheet is None:
            requested_sheet = sheet_name or default_sheet
        else:
            if sheet_name is not None and parsed_sheet.casefold() != sheet_name.casefold():
                raise RuntimeError(f"sheet mismatch: {sheet_name!r} vs {parsed_sheet!r}")
            requested_sheet = parsed_sheet

        ws = _resolve_target_worksheet(wb, requested_sheet) if requested_sheet else worksheets[0]
        resolved_sheet_name = getattr(ws, "title", requested_sheet) or requested_sheet or ""

        cell_range = _normalize_cell_range(range_part)
        if not _A1_RECT_RANGE_RE.fullmatch(cell_range):
            raise RuntimeError("range must be a finite A1 cell or rectangular cell range")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(
                cell_range if ":" in cell_range else f"{cell_range}:{cell_range}"
            )
        except Exception as exc:
            raise RuntimeError(f"failed to parse range boundaries: {exc}") from None
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        if min_col > max_col:
            min_col, max_col = max_col, min_col
        if min_row > max_row:
            min_row, max_row = max_row, min_row
        if min_col < 1 or min_row < 1 or max_col > _EXCEL_MAX_COLS or max_row > _EXCEL_MAX_ROWS:
            raise RuntimeError("invalid range boundaries")

        total_rows = max_row - min_row + 1
        total_cols = max_col - min_col + 1
        formatted_cells = total_rows * total_cols
        if max_format_cells > 0 and formatted_cells > max_format_cells:
            raise RuntimeError(f"requested range is too large (cells={formatted_cells} > {max_format_cells})")

        has_cell_options = any(
            key in options
            for key in (
                "fill_color",
                "font_color",
                "bold",
                "italic",
                "underline",
                "number_format",
                "alignment",
                "border_style",
            )
        )
        if has_cell_options:
            formatted_cells = 0
            skipped_merged_cells = 0
            for row_num in range(min_row, max_row + 1):
                for col_num in range(min_col, max_col + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if isinstance(cell, MergedCell):
                        skipped_merged_cells += 1
                        continue
                    _apply_cell_format(cell, options)
                    formatted_cells += 1
        else:
            skipped_merged_cells = 0

        if "row_height" in options:
            for row_num in range(min_row, max_row + 1):
                ws.row_dimensions[row_num].height = options["row_height"]

        if "column_width" in options:
            for col_num in range(min_col, max_col + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = options["column_width"]

        normalized_range = (
            f"{get_column_letter(min_col)}{min_row}"
            if min_col == max_col and min_row == max_row
            else f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        )
        formatted_range = f"{_quote_sheet_name_for_a1(resolved_sheet_name)}!{normalized_range}"

        samples = []
        for row_num, col_num in _sample_coords(min_row, max_row, min_col, max_col):
            cell = ws.cell(row=row_num, column=col_num)
            if isinstance(cell, MergedCell):
                continue
            samples.append(
                {
                    "address": _format_a1_address(sheet_name=resolved_sheet_name, row=row_num, col=col_num),
                    "format": _format_sample(cell),
                }
            )

        wb.save(str(tmp_path))
        _validate_workbook_loadable(tmp_path)
        return (
            resolved_sheet_name,
            formatted_range,
            formatted_cells,
            samples,
            total_rows,
            total_cols,
            skipped_merged_cells,
        )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


class FormatRangeTool(BaseTool):
    def __init__(self, config: dict, tool_schema: Optional[OpenAIFunctionToolSchema]):
        super().__init__(config, tool_schema or _default_tool_schema())
        self._instance_dict: dict[str, dict[str, Any]] = {}

        max_file_size_mb_raw = config.get("max_file_size_mb", config.get("max_file_mb", 100))
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)

        max_format_cells_raw = config.get(
            "max_format_cells", os.environ.get("SHEET_ARENA_FORMAT_RANGE_MAX_CELLS", "100000")
        )
        try:
            self.max_format_cells = max(0, int(max_format_cells_raw))
        except (TypeError, ValueError):
            self.max_format_cells = 100_000

        self.max_response_chars = _get_max_response_chars(config)

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return _error_response("invalid_path", "path must be a relative workbook path.")
        if relpath.suffix.lower() != ".xlsx":
            return _error_response("invalid_path", "only .xlsx workbooks are supported.")

        sheet_name_raw = parameters.get("sheet_name")
        sheet_name = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return _error_response("invalid_sheet_name", "sheet_name must be a non-empty string.")
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return _error_response("invalid_sheet_name", "sheet_name is empty after normalization.")

        range_raw = parameters.get("range")
        if not isinstance(range_raw, str) or not range_raw.strip():
            return _error_response(
                "invalid_range",
                "range must be a non-empty finite A1 range such as A1:C3 or Sheet1!B2:D10.",
            )
        range_token = range_raw.strip()

        try:
            _validate_range_token(range_token, sheet_name)
            options = _parse_format_options(parameters)
        except _FormatOptionError as exc:
            return _error_response(exc.code, str(exc))

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return _error_response("missing_workspace_id", "workspace_id is missing/invalid.")

        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return _error_response("file_not_found", f"file not found: {relpath}", file=str(relpath))

        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return _error_response("stat_failed", f"failed to stat file: {exc}", file=str(relpath))
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return _error_response(
                    "file_too_large",
                    f"workbook is too large ({actual_mb:.1f}MB > {max_mb}MB).",
                    file=str(relpath),
                )

        tmp_path: Optional[Path] = None
        try:
            tmp_path, expected_sig, orig_mode = await asyncio.to_thread(
                _copy_workbook_to_temp_under_lock,
                file_path=file_path,
                lock_timeout_s=self.lock_timeout_s,
            )
            (
                resolved_sheet,
                formatted_range,
                formatted_cells,
                samples,
                formatted_rows,
                formatted_columns,
                skipped_merged_cells,
            ) = await asyncio.to_thread(
                _format_temp_workbook,
                tmp_path=tmp_path,
                sheet_name=sheet_name,
                range_token=range_token,
                options=options,
                max_format_cells=self.max_format_cells,
            )
            await asyncio.to_thread(
                _commit_temp_workbook_under_lock,
                file_path=file_path,
                tmp_path=tmp_path,
                expected_sig=expected_sig,
                orig_mode=orig_mode,
                lock_timeout_s=self.lock_timeout_s,
            )
            tmp_path = None
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return _error_response("format_failed", f"failed to format range: {exc}", file=str(relpath))
        finally:
            if tmp_path is not None:
                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass

        payload: dict[str, Any] = {
            "status": "partial_success" if skipped_merged_cells else "success",
            "file": str(relpath),
            "sheet": resolved_sheet,
            "formatted_range": formatted_range,
            "formatted_cells": formatted_cells,
            "formatted_rows": formatted_rows,
            "formatted_columns": formatted_columns,
            "skipped_merged_cells": skipped_merged_cells,
            "applied_options": options,
            "sample_formats": samples,
            "truncated": False,
        }
        response_text = _truncate_success_payload_to_max_chars(payload, self.max_response_chars)

        metrics = {
            "status": payload.get("status", "success"),
            "file": str(relpath),
            "formatted_cells": formatted_cells,
            "formatted_rows": formatted_rows,
            "formatted_columns": formatted_columns,
            "skipped_merged_cells": skipped_merged_cells,
            "truncated": bool(payload.get("truncated")),
        }
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
