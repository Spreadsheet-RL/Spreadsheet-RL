# Copyright 2026 Bytedance Ltd. and/or its affiliates
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
from __future__ import annotations

import asyncio
import datetime as dt
import itertools
import json
import math
import os
import re
import uuid
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import get_spreadsheet_rl_data_root, normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .response_format import records_to_csv
from .schemas import OpenAIFunctionToolSchema, ToolResponse
from .worksheet_resolution import resolve_worksheet as _resolve_target_worksheet

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z0-9_]+$")
DEFAULT_MAX_RESPONSE_CHARS = 4096
MAX_RESPONSE_CHARS = 8192


def _get_max_scan_cells() -> int:
    value = os.environ.get("SHEET_ARENA_FIND_MAX_SCAN_CELLS", "200000").strip()
    try:
        n = int(value)
        return min(max(1, n), 10_000_000)
    except ValueError:
        return 200_000


def _get_max_response_chars(config: Optional[dict[str, Any]] = None) -> int:
    config_value = None
    if isinstance(config, dict):
        config_value = config.get("max_response_chars")
    for raw in (config_value, os.environ.get("SHEET_ARENA_TOOL_MAX_RESPONSE_CHARS", "").strip()):
        if raw is None:
            continue
        try:
            n = int(raw)
            return min(max(128, n), MAX_RESPONSE_CHARS)
        except (TypeError, ValueError):
            continue
    return DEFAULT_MAX_RESPONSE_CHARS


def _get_max_string_chars() -> int:
    raw = os.environ.get("SHEET_ARENA_TOOL_MAX_STRING_CHARS", "200").strip()
    try:
        n = int(raw)
        return min(max(16, n), 100_000)
    except ValueError:
        return 200


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _error_payload(error: str, message: str) -> dict[str, Any]:
    return {
        "status": "error",
        "error": error,
        "message": message,
        "truncated": False,
    }


def _error_tool_response(error: str, message: str) -> tuple[ToolResponse, float, dict]:
    payload = _error_payload(error, message)
    return (
        ToolResponse(text=_json_dumps_compact(payload)),
        0.0,
        {
            "status": "error",
            "error": error,
        },
    )


def _coerce_bool_parameter(value: Any, *, parameter: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        token = value.strip().casefold()
        if token == "true":
            return True
        if token == "false":
            return False
    raise ValueError(f"{parameter} received {value!r}; pass true or false.")


def _coerce_int_parameter(value: Any, *, parameter: str, default: int) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        raise ValueError(f"{parameter} received {value!r}; pass an integer.")
    try:
        return int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{parameter} received {value!r}; pass an integer.") from None


def _truncate_str(value: str, max_chars: int) -> str:
    if max_chars <= 0:
        return ""
    if len(value) <= max_chars:
        return value
    if max_chars <= 3:
        return value[:max_chars]
    return value[: max_chars - 3] + "..."


def _bounds_to_a1(*, min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        left = f"{min_col}{min_row}"
        right = f"{max_col}{max_row}"
    else:
        left = f"{get_column_letter(min_col)}{min_row}"
        right = f"{get_column_letter(max_col)}{max_row}"
    return left if left == right else f"{left}:{right}"


def _sanitize_relpath(value: Any) -> Optional[Path]:
    if not isinstance(value, str):
        return None
    raw = value.strip()
    if not raw:
        return None
    if "\0" in raw:
        return None
    p = Path(raw)
    if p.is_absolute():
        return None
    if any(part == ".." for part in p.parts):
        return None
    return p


def _normalize_sheet_name(value: str) -> str:
    name = value.strip('\t\n\r\f\v"')
    quoted = name.strip()
    if len(quoted) >= 2 and quoted[0] == "'" and quoted[-1] == "'":
        return quoted[1:-1].replace("''", "'")
    if len(quoted) >= 2 and quoted[0] == '"' and quoted[-1] == '"':
        return quoted[1:-1]

    if quoted.startswith("'") and not quoted.endswith("'"):
        return quoted[1:]
    if quoted.endswith("'") and not quoted.startswith("'"):
        return quoted[:-1]
    return name.strip("\t\n\r\f\v")


def _normalize_cell_range(value: str) -> str:
    rng = value.strip(" \t\n\r\f\v\"'")
    rng = rng.replace(" ", "").replace("$", "")
    return rng.upper()


def _resolve_workspace_file(*, workspace_id: Optional[str], relpath: Path) -> Optional[Path]:
    if not workspace_id:
        return None

    workspace_base = get_spreadsheet_rl_data_root()
    workspaces_base = workspace_base / "_workspaces"
    try:
        if workspace_base.is_symlink():
            return None
        workspaces_base_resolved = workspaces_base.resolve(strict=True)
    except (OSError, RuntimeError):
        return None

    workspace_root = workspaces_base / workspace_id
    try:
        if workspace_root.is_symlink():
            return None
        workspace_root_resolved = workspace_root.resolve(strict=True)
    except (OSError, RuntimeError):
        return None
    if not workspace_root_resolved.is_relative_to(workspaces_base_resolved):
        return None

    rel_candidates = [relpath]
    if relpath.parts and relpath.parts[0] == "data" and len(relpath.parts) > 1:
        rel_candidates.append(Path(*relpath.parts[1:]))

    for rel_candidate in rel_candidates:
        candidate = workspace_root / rel_candidate
        try:
            resolved = candidate.resolve(strict=True)
        except (OSError, RuntimeError):
            continue
        if not resolved.is_relative_to(workspace_root_resolved):
            continue

        if candidate.is_symlink():
            continue

        cursor = workspace_root
        safe = True
        for part in rel_candidate.parts:
            cursor = cursor / part
            try:
                if cursor.is_symlink():
                    safe = False
                    break
            except OSError:
                safe = False
                break
        if not safe:
            continue

        if resolved.is_file():
            return resolved
    return None


def _quote_sheet_name_for_a1(sheet_name: str) -> str:
    if _SAFE_SHEET_NAME_RE.fullmatch(sheet_name):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _format_a1_address(*, sheet_name: str, row: int, col: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        col_letter = str(col)
    else:
        col_letter = get_column_letter(col)
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    return f"{sheet_ref}!{col_letter}{row}"


def _to_jsonable_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool | int | str):
        if isinstance(value, str):
            return _truncate_str(value, _get_max_string_chars())
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, dt.datetime | dt.date | dt.time):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return str(value)


def _match_columns(*, include_values: bool, search_in: str) -> list[str]:
    columns = ["address"]
    if include_values and search_in in {"values", "both"}:
        columns.append("value")
    if include_values and search_in in {"formulas", "both"}:
        columns.append("formula")
    return columns


def _set_matches_csv(
    payload: dict[str, Any],
    matches: list[dict[str, Any]],
    *,
    include_values: bool,
    search_in: str,
) -> None:
    payload["include_values"] = include_values
    payload["matches_csv"] = records_to_csv(
        matches,
        _match_columns(include_values=include_values, search_in=search_in),
    )
    payload["returned"] = len(matches)
    payload["returned_matches"] = len(matches)
    payload.pop("matches", None)


def _mark_values_omitted(payload: dict[str, Any]) -> None:
    payload["values_omitted"] = True
    _add_truncation_reason(payload, "values_omitted")


def _fits_response_budget(payload: dict[str, Any], max_response_chars: int) -> bool:
    return len(_json_dumps_compact(payload)) <= max_response_chars


def _fit_success_payload_to_max_chars(payload: dict[str, Any], max_response_chars: int) -> dict[str, Any]:
    if _fits_response_budget(payload, max_response_chars):
        return payload

    payload["truncated"] = True
    payload["response_truncated"] = True
    _add_truncation_reason(payload, "response_budget")

    for keys in (
        ("query", "requested_range", "ranges", "sheet_summaries", "searched_sheets", "range", "max_results"),
        ("file",),
        ("sheet_scope",),
        ("sheet",),
        ("result_truncated", "scan_truncated"),
        ("include_values",),
        ("truncation_reasons",),
    ):
        for key in keys:
            payload.pop(key, None)
        if _fits_response_budget(payload, max_response_chars):
            return payload

    total_matches = payload.get("total_matches", payload.get("returned_matches", payload.get("returned", 0)))
    compact: dict[str, Any] = {
        "status": payload.get("status", "success"),
        "total_matches": total_matches,
        "truncated": True,
        "response_truncated": True,
    }
    for key in ("matches_csv", "returned", "returned_matches", "values_omitted"):
        if key in payload:
            compact[key] = payload[key]
    if _fits_response_budget(compact, max_response_chars):
        payload.clear()
        payload.update(compact)
        return payload

    for key in ("matches_csv", "returned", "returned_matches"):
        compact.pop(key, None)
        if _fits_response_budget(compact, max_response_chars):
            payload.clear()
            payload.update(compact)
            return payload

    compact.pop("values_omitted", None)
    if _fits_response_budget(compact, max_response_chars):
        payload.clear()
        payload.update(compact)
        return payload

    payload.clear()
    payload.update({"status": "success", "truncated": True})
    return payload


def _add_truncation_reason(payload: dict[str, Any], reason: str) -> None:
    payload["truncated"] = True
    reasons = payload.setdefault("truncation_reasons", [])
    if isinstance(reasons, list) and reason not in reasons:
        reasons.append(reason)


def _fit_matches_csv_payload(
    payload: dict[str, Any],
    matches: list[dict[str, Any]],
    *,
    include_values: bool,
    search_in: str,
    max_response_chars: int,
) -> dict[str, Any]:
    matches_out = list(matches)
    _set_matches_csv(payload, matches_out, include_values=include_values, search_in=search_in)
    text = _json_dumps_compact(payload)
    if len(text) <= max_response_chars:
        return payload

    while len(text) > max_response_chars and len(matches_out) > 1:
        matches_out = matches_out[: max(1, len(matches_out) // 2)]
        _set_matches_csv(payload, matches_out, include_values=include_values, search_in=search_in)
        payload["response_truncated"] = True
        _add_truncation_reason(payload, "response_budget")
        text = _json_dumps_compact(payload)

    if len(text) > max_response_chars:
        for key in ("query", "ranges", "sheet_summaries", "searched_sheets"):
            payload.pop(key, None)
        payload["response_truncated"] = True
        _add_truncation_reason(payload, "response_budget")
        text = _json_dumps_compact(payload)

    if len(text) > max_response_chars and include_values:
        matches_out = [{"address": match.get("address")} for match in matches_out if match.get("address")]
        _set_matches_csv(payload, matches_out, include_values=False, search_in=search_in)
        payload["response_truncated"] = True
        _add_truncation_reason(payload, "response_budget")
        _mark_values_omitted(payload)
        text = _json_dumps_compact(payload)

    if len(text) <= max_response_chars:
        return payload

    first_match = matches_out[:1]
    minimal = {
        "status": "success",
        "file": payload.get("file"),
        "sheet_scope": payload.get("sheet_scope"),
        "total_matches": payload.get("total_matches", 0),
        "max_results": payload.get("max_results"),
        "truncated": True,
        "result_truncated": bool(payload.get("result_truncated")),
        "scan_truncated": bool(payload.get("scan_truncated")),
        "response_truncated": True,
        "truncation_reasons": ["response_budget"],
    }
    if include_values:
        minimal["values_omitted"] = True
        minimal["truncation_reasons"].append("values_omitted")
    _set_matches_csv(minimal, first_match, include_values=False, search_in=search_in)
    return _fit_success_payload_to_max_chars(minimal, max_response_chars)


def _scan_zip_metadata(
    file_path: Path,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(file_path) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)
                filename = getattr(info, "filename", "?")

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={filename!r}, bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={filename!r}, ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        f"zip contents too large (total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read zip metadata: {exc}"

    return None


def _normalize_match_mode(value: Any) -> str:
    if value is None:
        return "contains"
    if not isinstance(value, str):
        raise ValueError("match must be a string")
    mode = value.strip().lower()
    if mode in {"contains", "equals", "prefix", "regex"}:
        return mode
    raise ValueError("match must be one of: contains, equals, prefix, regex")


def _normalize_search_in(value: Any) -> str:
    if value is None:
        return "values"
    if not isinstance(value, str):
        raise ValueError("search_in must be a string")
    mode = value.strip().lower()
    if mode in {"values", "formulas", "both"}:
        return mode
    raise ValueError("search_in must be one of: values, formulas, both")


def _normalize_return_mode(value: Any) -> str:
    if value is None:
        return "first"
    if not isinstance(value, str):
        raise ValueError("return must be a string")
    mode = value.strip().lower()
    if mode in {"first", "all"}:
        return mode
    raise ValueError("return must be one of: first, all")


def _get_regex_haystack_chars() -> int:
    raw = os.environ.get("SHEET_ARENA_FIND_REGEX_HAYSTACK_CHARS", "2000").strip()
    try:
        n = int(raw)
        return min(max(64, n), 100_000)
    except ValueError:
        return 2000


def _get_max_regex_pattern_chars() -> int:
    raw = os.environ.get("SHEET_ARENA_FIND_REGEX_MAX_PATTERN_CHARS", "200").strip()
    try:
        n = int(raw)
        return min(max(16, n), 2000)
    except ValueError:
        return 200


def _validate_regex_pattern(pattern: str) -> Optional[str]:
    if not pattern:
        return "regex pattern is empty"

    in_class = False
    escaped = False
    for ch in pattern:
        if escaped:
            escaped = False
            continue
        if ch == "\\":
            escaped = True
            continue
        if ch == "[" and not in_class:
            in_class = True
            continue
        if ch == "]" and in_class:
            in_class = False
            continue
        if in_class:
            continue
        if ch in ("(", ")", "{", "}", "|", "+", "?"):
            return (
                "regex pattern uses unsupported constructs; supported: literal text, '.', '*', '^', '$', "
                "character classes '[...]', and escapes"
            )

    if escaped:
        return "regex pattern ends with a trailing backslash"
    if in_class:
        return "regex pattern has an unterminated character class"
    return None


def _find_cells_sync(
    *,
    file_path: Path,
    sheet_name: str,
    query: str,
    search_range: Optional[str],
    match_mode: str,
    search_in: str,
    case_sensitive: bool,
    return_mode: str,
    max_results: int,
    max_response_chars: int,
    include_values: bool,
) -> tuple[dict[str, Any], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required; install openpyxl>=3.1.5") from None

    if max_results < 1:
        max_results = 1

    wb_values = None
    wb_formula = None
    try:
        if search_in in {"values", "both"}:
            wb_values = load_workbook(filename=str(file_path), data_only=True, read_only=True, keep_links=False)
        if search_in in {"formulas", "both"}:
            wb_formula = load_workbook(filename=str(file_path), data_only=False, read_only=True, keep_links=False)

        wb_ref = wb_values or wb_formula
        if wb_ref is None:
            raise RuntimeError("invalid search_in mode")

        ws_ref = _resolve_target_worksheet(wb_ref, sheet_name)
        actual_sheet = getattr(ws_ref, "title", sheet_name) or sheet_name

        used_range = None
        try:
            used_range = ws_ref.calculate_dimension()
        except Exception:
            used_range = None
        if not isinstance(used_range, str) or not used_range.strip():
            try:
                dim = getattr(ws_ref, "dimensions", None)
            except Exception:
                dim = None
            used_range = dim.strip() if isinstance(dim, str) else ""
        if not used_range:
            try:
                max_row = int(getattr(ws_ref, "max_row", 1) or 1)
                max_col = int(getattr(ws_ref, "max_column", 1) or 1)
                max_row = min(max(1, max_row), _EXCEL_MAX_ROWS)
                max_col = min(max(1, max_col), _EXCEL_MAX_COLS)
                used_range = _bounds_to_a1(min_col=1, min_row=1, max_col=max_col, max_row=max_row)
            except Exception:
                used_range = "A1:A1"

        effective_range = search_range.strip() if isinstance(search_range, str) and search_range.strip() else used_range
        if "!" in effective_range:
            sheet_part, _, range_part = effective_range.rpartition("!")
            requested_sheet = _normalize_sheet_name(sheet_part)
            if not requested_sheet:
                raise RuntimeError(f"empty sheet name in range: {effective_range!r}")
            if requested_sheet.casefold() != actual_sheet.casefold():
                raise RuntimeError(f"range sheet {requested_sheet!r} does not match sheet_name {actual_sheet!r}")
            effective_range = range_part.strip()
        effective_range = _normalize_cell_range(effective_range)
        used_range = _normalize_cell_range(used_range)

        used_boundaries = range_boundaries(used_range if ":" in used_range else f"{used_range}:{used_range}")
        used_min_col, used_min_row, used_max_col, used_max_row = used_boundaries
        if used_min_col is None:
            used_min_col = 1
        if used_max_col is None:
            used_max_col = _EXCEL_MAX_COLS
        if used_min_row is None:
            used_min_row = 1
        if used_max_row is None:
            used_max_row = _EXCEL_MAX_ROWS
        used_min_col = int(used_min_col)
        used_min_row = int(used_min_row)
        used_max_col = int(used_max_col)
        used_max_row = int(used_max_row)

        req_boundaries = range_boundaries(
            effective_range if ":" in effective_range else f"{effective_range}:{effective_range}"
        )
        req_min_col, req_min_row, req_max_col, req_max_row = req_boundaries
        if req_min_col is None:
            req_min_col = 1
        if req_max_col is None:
            req_max_col = _EXCEL_MAX_COLS
        if req_min_row is None:
            req_min_row = 1
        if req_max_row is None:
            req_max_row = _EXCEL_MAX_ROWS
        req_min_col = int(req_min_col)
        req_min_row = int(req_min_row)
        req_max_col = int(req_max_col)
        req_max_row = int(req_max_row)
        requested_cell_count = (req_max_col - req_min_col + 1) * (req_max_row - req_min_row + 1)

        min_col = max(req_min_col, used_min_col)
        min_row = max(req_min_row, used_min_row)
        max_col = min(req_max_col, used_max_col)
        max_row = min(req_max_row, used_max_row)
        scanned_range = None
        if max_col >= min_col and max_row >= min_row:
            scanned_range = _bounds_to_a1(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            candidate_cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
        else:
            candidate_cell_count = 0

        max_scan_cells = _get_max_scan_cells()

        regex = None
        if match_mode == "regex":
            flags = 0 if case_sensitive else re.IGNORECASE
            regex = re.compile(query, flags=flags)
        q = query if case_sensitive else query.casefold()

        regex_haystack_chars = _get_regex_haystack_chars()
        if match_mode == "regex":
            max_scan_cells = min(max_scan_cells, 10_000)
            regex_haystack_chars = min(regex_haystack_chars, 200)

        def matches_text(text: str) -> bool:
            if match_mode == "regex":
                if regex is None:
                    return False
                return regex.search(text[:regex_haystack_chars]) is not None
            hay = text if case_sensitive else text.casefold()
            if match_mode == "contains":
                return q in hay
            if match_mode == "equals":
                return hay == q
            if match_mode == "prefix":
                return hay.startswith(q)
            return False

        ws_values = None
        ws_formula = None
        if wb_values is not None:
            ws_values = _resolve_target_worksheet(wb_values, actual_sheet)
        if wb_formula is not None:
            ws_formula = _resolve_target_worksheet(wb_formula, actual_sheet)

        results: list[dict[str, Any]] = []
        total_matches = 0
        truncated = False
        scan_truncated = False
        scanned = 0

        sheet_ref = _quote_sheet_name_for_a1(actual_sheet)
        range_out = scanned_range or effective_range
        requested_range_out = (
            effective_range if scanned_range is not None and scanned_range != effective_range else None
        )

        if candidate_cell_count <= 0:
            payload = {
                "status": "success",
                "file": str(file_path),
                "sheet": actual_sheet,
                "range": f"{sheet_ref}!{range_out}",
                "query": _to_jsonable_excel_value(query),
                "match": match_mode,
                "search_in": search_in,
                "case_sensitive": case_sensitive,
                "include_values": include_values,
                "total_matches": 0,
                "max_results": max_results,
                "truncated": False,
                "result_truncated": False,
                "scan_truncated": False,
            }
            _set_matches_csv(payload, [], include_values=include_values, search_in=search_in)
            if requested_range_out is not None:
                payload["requested_range"] = f"{sheet_ref}!{requested_range_out}"
            metrics = {
                "status": "success",
                "file": str(file_path),
                "sheet": actual_sheet,
                "scanned_cells": 0,
                "requested_cells": requested_cell_count,
                "candidate_cells": candidate_cell_count,
                "returned_matches": 0,
                "total_matches": 0,
                "include_values": include_values,
                "truncated": False,
                "result_truncated": False,
                "scan_truncated": False,
            }
            return payload, metrics

        def iter_rows_values():
            if ws_values is None:
                return []
            return ws_values.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)

        def iter_rows_formula():
            if ws_formula is None:
                return []
            return ws_formula.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)

        if search_in == "both":
            row_iter = itertools.zip_longest(iter_rows_values(), iter_rows_formula(), fillvalue=())
        elif search_in == "values":
            row_iter = ((row, None) for row in iter_rows_values())
        else:
            row_iter = ((None, row) for row in iter_rows_formula())

        for r_idx, (row_vals, row_forms) in enumerate(row_iter, start=0):
            row_num = min_row + r_idx
            if row_vals is None:
                row_vals = []
            if row_forms is None:
                row_forms = []
            max_len = max(len(row_vals), len(row_forms))
            for c_idx in range(max_len):
                col_num = min_col + c_idx
                if scanned >= max_scan_cells:
                    scan_truncated = True
                    break
                scanned += 1

                value_obj = None
                if search_in in {"values", "both"} and c_idx < len(row_vals):
                    try:
                        value_obj = getattr(row_vals[c_idx], "value", None)
                    except Exception:
                        value_obj = None

                formula_text = None
                if search_in in {"formulas", "both"} and c_idx < len(row_forms):
                    try:
                        cell_f = row_forms[c_idx]
                        if getattr(cell_f, "data_type", None) == "f":
                            raw = getattr(cell_f, "value", None)
                            if isinstance(raw, str) and raw:
                                formula_text = raw if raw.startswith("=") else f"={raw}"
                    except Exception:
                        formula_text = None

                haystacks: list[str] = []
                if search_in in {"values", "both"}:
                    if value_obj is not None:
                        haystacks.append(str(value_obj))
                if search_in in {"formulas", "both"}:
                    if formula_text is not None:
                        haystacks.append(formula_text)

                if not haystacks:
                    continue
                if not any(matches_text(text) for text in haystacks):
                    continue

                total_matches += 1
                if return_mode == "all" and len(results) >= max_results:
                    truncated = True
                    continue

                match_entry: dict[str, Any] = {
                    "address": _format_a1_address(sheet_name=actual_sheet, row=row_num, col=col_num),
                }
                if include_values and search_in in {"values", "both"}:
                    match_entry["value"] = _to_jsonable_excel_value(value_obj)
                if include_values and search_in in {"formulas", "both"}:
                    match_entry["formula"] = (
                        _truncate_str(formula_text, _get_max_string_chars()) if isinstance(formula_text, str) else None
                    )
                results.append(match_entry)

                if return_mode == "first":
                    truncated = False
                    payload = {
                        "status": "success",
                        "file": str(file_path),
                        "sheet": actual_sheet,
                        "range": f"{sheet_ref}!{range_out}",
                        "query": _to_jsonable_excel_value(query),
                        "match": match_mode,
                        "search_in": search_in,
                        "case_sensitive": case_sensitive,
                        "include_values": include_values,
                        "total_matches": len(results),
                        "max_results": max_results,
                        "truncated": False,
                        "result_truncated": False,
                        "scan_truncated": False,
                    }
                    _set_matches_csv(
                        payload,
                        results,
                        include_values=include_values,
                        search_in=search_in,
                    )
                    if requested_range_out is not None:
                        payload["requested_range"] = f"{sheet_ref}!{requested_range_out}"
                    metrics = {
                        "status": "success",
                        "file": str(file_path),
                        "sheet": actual_sheet,
                        "scanned_cells": scanned,
                        "requested_cells": requested_cell_count,
                        "candidate_cells": candidate_cell_count,
                        "returned_matches": len(results),
                        "total_matches": len(results),
                        "include_values": include_values,
                        "truncated": False,
                        "result_truncated": False,
                        "scan_truncated": False,
                    }
                    text = _json_dumps_compact(payload)
                    if len(text) <= max_response_chars:
                        return payload, metrics

                    matches_out = list(results)
                    if matches_out:
                        for match in matches_out:
                            match.pop("value", None)
                            match.pop("formula", None)
                        matches_out = matches_out[:1]
                        _set_matches_csv(
                            payload,
                            matches_out,
                            include_values=False,
                            search_in=search_in,
                        )
                        payload["truncated"] = True
                        payload["response_truncated"] = True
                        _add_truncation_reason(payload, "response_budget")
                        if include_values:
                            _mark_values_omitted(payload)
                        metrics["truncated"] = True
                        metrics["include_values"] = False
                        metrics["returned_matches"] = len(matches_out)
                        text = _json_dumps_compact(payload)

                    if len(text) > max_response_chars:
                        payload.pop("query", None)
                        payload.pop("requested_range", None)
                        text = _json_dumps_compact(payload)

                    if len(text) > max_response_chars:
                        addr = matches_out[0].get("address") if matches_out else None
                        payload = {
                            "status": "success",
                            "file": str(file_path),
                            "sheet": actual_sheet,
                            "range": f"{sheet_ref}!{range_out}",
                            "total_matches": len(results),
                            "max_results": max_results,
                            "truncated": True,
                            "result_truncated": False,
                            "response_truncated": True,
                            "scan_truncated": False,
                        }
                        _set_matches_csv(
                            payload,
                            [{"address": addr}] if isinstance(addr, str) and addr else [],
                            include_values=False,
                            search_in=search_in,
                        )
                        _add_truncation_reason(payload, "response_budget")
                        if include_values:
                            _mark_values_omitted(payload)
                        metrics["truncated"] = True
                        metrics["include_values"] = False
                        metrics["returned_matches"] = 1 if isinstance(addr, str) and addr else 0
                        payload = _fit_success_payload_to_max_chars(payload, max_response_chars)
                        metrics["returned_matches"] = payload.get(
                            "returned_matches",
                            payload.get("returned", metrics["returned_matches"]),
                        )
                    return payload, metrics
            if scan_truncated:
                break

        payload = {
            "status": "success",
            "file": str(file_path),
            "sheet": actual_sheet,
            "range": f"{sheet_ref}!{range_out}",
            "query": _to_jsonable_excel_value(query),
            "match": match_mode,
            "search_in": search_in,
            "case_sensitive": case_sensitive,
            "include_values": include_values,
            "total_matches": total_matches,
            "max_results": max_results,
            "truncated": truncated,
            "result_truncated": truncated,
            "scan_truncated": scan_truncated,
        }
        _set_matches_csv(payload, results, include_values=include_values, search_in=search_in)
        if truncated:
            _add_truncation_reason(payload, "max_results")
        if requested_range_out is not None:
            payload["requested_range"] = f"{sheet_ref}!{requested_range_out}"
        metrics = {
            "status": "success",
            "file": str(file_path),
            "sheet": actual_sheet,
            "scanned_cells": scanned,
            "requested_cells": requested_cell_count,
            "candidate_cells": candidate_cell_count,
            "returned_matches": len(results),
            "total_matches": total_matches,
            "include_values": include_values,
            "truncated": truncated,
            "result_truncated": truncated,
            "scan_truncated": scan_truncated,
        }
        text = _json_dumps_compact(payload)
        if len(text) <= max_response_chars:
            return payload, metrics

        matches = list(results)
        if matches:
            while len(text) > max_response_chars and len(matches) > 1:
                matches = matches[: max(1, len(matches) // 2)]
                _set_matches_csv(payload, matches, include_values=include_values, search_in=search_in)
                payload["truncated"] = True
                payload["response_truncated"] = True
                _add_truncation_reason(payload, "response_budget")
                metrics["truncated"] = True
                metrics["returned_matches"] = len(matches)
                text = _json_dumps_compact(payload)

            if len(text) > max_response_chars:
                for match in matches:
                    match.pop("value", None)
                    match.pop("formula", None)
                matches = matches[:1]
                _set_matches_csv(payload, matches, include_values=False, search_in=search_in)
                payload["truncated"] = True
                payload["response_truncated"] = True
                _add_truncation_reason(payload, "response_budget")
                if include_values:
                    _mark_values_omitted(payload)
                metrics["truncated"] = True
                metrics["include_values"] = False
                metrics["returned_matches"] = len(matches)
                text = _json_dumps_compact(payload)

        if len(text) > max_response_chars:
            payload.pop("query", None)
            payload.pop("requested_range", None)
            text = _json_dumps_compact(payload)

        if len(text) > max_response_chars:
            addr = None
            if matches:
                addr = matches[0].get("address")
            payload = {
                "status": "success",
                "file": str(file_path),
                "sheet": actual_sheet,
                "range": f"{sheet_ref}!{range_out}",
                "total_matches": total_matches,
                "max_results": max_results,
                "truncated": True,
                "result_truncated": truncated,
                "response_truncated": True,
                "scan_truncated": scan_truncated,
            }
            _set_matches_csv(
                payload,
                [{"address": addr}] if isinstance(addr, str) and addr else [],
                include_values=False,
                search_in=search_in,
            )
            _add_truncation_reason(payload, "response_budget")
            if include_values:
                _mark_values_omitted(payload)
            metrics["truncated"] = True
            metrics["include_values"] = False
            metrics["returned_matches"] = 1 if isinstance(addr, str) and addr else 0
            payload = _fit_success_payload_to_max_chars(payload, max_response_chars)
            metrics["returned_matches"] = payload.get(
                "returned_matches",
                payload.get("returned", metrics["returned_matches"]),
            )
        return payload, metrics
    finally:
        for wb in (wb_formula, wb_values):
            if wb is None:
                continue
            try:
                wb.close()
            except Exception:
                pass


def _find_cells_all_sheets_sync(
    *,
    file_path: Path,
    query: str,
    search_range: Optional[str],
    match_mode: str,
    search_in: str,
    case_sensitive: bool,
    return_mode: str,
    max_results: int,
    max_response_chars: int,
    include_values: bool,
) -> tuple[dict[str, Any], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required; install openpyxl>=3.1.5") from None

    if max_results < 1:
        max_results = 1

    wb_values = None
    wb_formula = None
    try:
        if search_in in {"values", "both"}:
            wb_values = load_workbook(filename=str(file_path), data_only=True, read_only=True, keep_links=False)
        if search_in in {"formulas", "both"}:
            wb_formula = load_workbook(filename=str(file_path), data_only=False, read_only=True, keep_links=False)

        wb_ref = wb_values or wb_formula
        if wb_ref is None:
            raise RuntimeError("invalid search_in mode")

        requested_range_sheet = None
        range_without_sheet = search_range
        if isinstance(search_range, str) and "!" in search_range:
            sheet_part, _, range_part = search_range.rpartition("!")
            requested_range_sheet = _normalize_sheet_name(sheet_part)
            if not requested_range_sheet:
                raise RuntimeError(f"empty sheet name in range: {search_range!r}")
            if not range_part.strip():
                raise RuntimeError(f"empty range in range: {search_range!r}")
            range_without_sheet = range_part.strip()

        if requested_range_sheet:
            ws_ref = _resolve_target_worksheet(wb_ref, requested_range_sheet)
            sheet_names = [getattr(ws_ref, "title", requested_range_sheet) or requested_range_sheet]
            sheet_scope = sheet_names[0]
        else:
            worksheets = getattr(wb_ref, "worksheets", None) or []
            sheet_names = [getattr(ws, "title", "") for ws in worksheets if getattr(ws, "title", "")]
            sheet_scope = "all"
        if not sheet_names:
            raise RuntimeError("workbook has no worksheets")

        max_scan_cells = _get_max_scan_cells()
        regex = None
        if match_mode == "regex":
            flags = 0 if case_sensitive else re.IGNORECASE
            regex = re.compile(query, flags=flags)
            max_scan_cells = min(max_scan_cells, 10_000)
        q = query if case_sensitive else query.casefold()
        regex_haystack_chars = _get_regex_haystack_chars()
        if match_mode == "regex":
            regex_haystack_chars = min(regex_haystack_chars, 200)

        def matches_text(text: str) -> bool:
            if match_mode == "regex":
                if regex is None:
                    return False
                return regex.search(text[:regex_haystack_chars]) is not None
            hay = text if case_sensitive else text.casefold()
            if match_mode == "contains":
                return q in hay
            if match_mode == "equals":
                return hay == q
            if match_mode == "prefix":
                return hay.startswith(q)
            return False

        results: list[dict[str, Any]] = []
        ranges: list[str] = []
        searched_sheets: list[str] = []
        sheet_summaries: list[dict[str, Any]] = []
        scanned = 0
        requested_cells = 0
        candidate_cells = 0
        total_matches = 0
        result_truncated = False
        scan_truncated = False
        stopped_after_first = False

        for actual_sheet in sheet_names:
            if scanned >= max_scan_cells:
                scan_truncated = True
                break

            ws_ref = _resolve_target_worksheet(wb_ref, actual_sheet)
            actual_sheet = getattr(ws_ref, "title", actual_sheet) or actual_sheet
            searched_sheets.append(actual_sheet)

            used_range = None
            try:
                used_range = ws_ref.calculate_dimension()
            except Exception:
                used_range = None
            if not isinstance(used_range, str) or not used_range.strip():
                try:
                    dim = getattr(ws_ref, "dimensions", None)
                except Exception:
                    dim = None
                used_range = dim.strip() if isinstance(dim, str) else ""
            if not used_range:
                try:
                    max_row = int(getattr(ws_ref, "max_row", 1) or 1)
                    max_col = int(getattr(ws_ref, "max_column", 1) or 1)
                    max_row = min(max(1, max_row), _EXCEL_MAX_ROWS)
                    max_col = min(max(1, max_col), _EXCEL_MAX_COLS)
                    used_range = _bounds_to_a1(min_col=1, min_row=1, max_col=max_col, max_row=max_row)
                except Exception:
                    used_range = "A1:A1"

            effective_range = (
                range_without_sheet.strip()
                if isinstance(range_without_sheet, str) and range_without_sheet.strip()
                else used_range
            )
            effective_range = _normalize_cell_range(effective_range)
            used_range = _normalize_cell_range(used_range)

            used_boundaries = range_boundaries(used_range if ":" in used_range else f"{used_range}:{used_range}")
            used_min_col, used_min_row, used_max_col, used_max_row = used_boundaries
            if used_min_col is None:
                used_min_col = 1
            if used_max_col is None:
                used_max_col = _EXCEL_MAX_COLS
            if used_min_row is None:
                used_min_row = 1
            if used_max_row is None:
                used_max_row = _EXCEL_MAX_ROWS
            used_min_col = int(used_min_col)
            used_min_row = int(used_min_row)
            used_max_col = int(used_max_col)
            used_max_row = int(used_max_row)

            req_boundaries = range_boundaries(
                effective_range if ":" in effective_range else f"{effective_range}:{effective_range}"
            )
            req_min_col, req_min_row, req_max_col, req_max_row = req_boundaries
            if req_min_col is None:
                req_min_col = 1
            if req_max_col is None:
                req_max_col = _EXCEL_MAX_COLS
            if req_min_row is None:
                req_min_row = 1
            if req_max_row is None:
                req_max_row = _EXCEL_MAX_ROWS
            req_min_col = int(req_min_col)
            req_min_row = int(req_min_row)
            req_max_col = int(req_max_col)
            req_max_row = int(req_max_row)
            requested_cell_count = (req_max_col - req_min_col + 1) * (req_max_row - req_min_row + 1)
            requested_cells += requested_cell_count

            min_col = max(req_min_col, used_min_col)
            min_row = max(req_min_row, used_min_row)
            max_col = min(req_max_col, used_max_col)
            max_row = min(req_max_row, used_max_row)
            if max_col >= min_col and max_row >= min_row:
                scanned_range = _bounds_to_a1(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                candidate_cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
            else:
                scanned_range = None
                candidate_cell_count = 0
            candidate_cells += candidate_cell_count

            sheet_ref = _quote_sheet_name_for_a1(actual_sheet)
            range_out = scanned_range or effective_range
            range_label = f"{sheet_ref}!{range_out}"
            ranges.append(range_label)
            sheet_scanned = 0
            sheet_matches = 0

            if candidate_cell_count <= 0:
                sheet_summaries.append(
                    {
                        "sheet": actual_sheet,
                        "range": range_label,
                        "scanned_cells": 0,
                        "matches": 0,
                    }
                )
                continue

            ws_values = _resolve_target_worksheet(wb_values, actual_sheet) if wb_values is not None else None
            ws_formula = _resolve_target_worksheet(wb_formula, actual_sheet) if wb_formula is not None else None

            def iter_rows_values(
                ws_values=ws_values,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                if ws_values is None:
                    return []
                return ws_values.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)

            def iter_rows_formula(
                ws_formula=ws_formula,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                if ws_formula is None:
                    return []
                return ws_formula.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)

            if search_in == "both":
                row_iter = itertools.zip_longest(iter_rows_values(), iter_rows_formula(), fillvalue=())
            elif search_in == "values":
                row_iter = ((row, None) for row in iter_rows_values())
            else:
                row_iter = ((None, row) for row in iter_rows_formula())

            for r_idx, (row_vals, row_forms) in enumerate(row_iter, start=0):
                row_num = min_row + r_idx
                if row_vals is None:
                    row_vals = []
                if row_forms is None:
                    row_forms = []
                max_len = max(len(row_vals), len(row_forms))
                for c_idx in range(max_len):
                    col_num = min_col + c_idx
                    if scanned >= max_scan_cells:
                        scan_truncated = True
                        break
                    scanned += 1
                    sheet_scanned += 1

                    value_obj = None
                    if search_in in {"values", "both"} and c_idx < len(row_vals):
                        try:
                            value_obj = getattr(row_vals[c_idx], "value", None)
                        except Exception:
                            value_obj = None

                    formula_text = None
                    if search_in in {"formulas", "both"} and c_idx < len(row_forms):
                        try:
                            cell_f = row_forms[c_idx]
                            if getattr(cell_f, "data_type", None) == "f":
                                raw = getattr(cell_f, "value", None)
                                if isinstance(raw, str) and raw:
                                    formula_text = raw if raw.startswith("=") else f"={raw}"
                        except Exception:
                            formula_text = None

                    haystacks: list[str] = []
                    if search_in in {"values", "both"} and value_obj is not None:
                        haystacks.append(str(value_obj))
                    if search_in in {"formulas", "both"} and formula_text is not None:
                        haystacks.append(formula_text)
                    if not haystacks or not any(matches_text(text) for text in haystacks):
                        continue

                    sheet_matches += 1
                    total_matches += 1
                    if return_mode == "all" and len(results) >= max_results:
                        result_truncated = True
                        continue

                    match_entry: dict[str, Any] = {
                        "address": _format_a1_address(sheet_name=actual_sheet, row=row_num, col=col_num),
                    }
                    if include_values and search_in in {"values", "both"}:
                        match_entry["value"] = _to_jsonable_excel_value(value_obj)
                    if include_values and search_in in {"formulas", "both"}:
                        match_entry["formula"] = (
                            _truncate_str(formula_text, _get_max_string_chars())
                            if isinstance(formula_text, str)
                            else None
                        )
                    results.append(match_entry)

                    if return_mode == "first":
                        stopped_after_first = True
                        break
                if scan_truncated or stopped_after_first:
                    break

            sheet_summaries.append(
                {
                    "sheet": actual_sheet,
                    "range": range_label,
                    "scanned_cells": sheet_scanned,
                    "matches": sheet_matches,
                }
            )
            if scan_truncated or stopped_after_first:
                break

        payload = {
            "status": "success",
            "file": str(file_path),
            "sheet_scope": sheet_scope,
            "searched_sheets": searched_sheets,
            "ranges": ranges,
            "query": _to_jsonable_excel_value(query),
            "match": match_mode,
            "search_in": search_in,
            "case_sensitive": case_sensitive,
            "include_values": include_values,
            "total_matches": total_matches,
            "max_results": max_results,
            "truncated": result_truncated or scan_truncated,
            "result_truncated": result_truncated,
            "scan_truncated": scan_truncated,
            "sheet_summaries": sheet_summaries,
        }
        if sheet_scope != "all" and searched_sheets:
            payload["sheet"] = searched_sheets[0]
            payload["range"] = ranges[0] if ranges else None
        if result_truncated:
            _add_truncation_reason(payload, "max_results")
        if scan_truncated:
            _add_truncation_reason(payload, "scan_limit")
        payload = _fit_matches_csv_payload(
            payload,
            results,
            include_values=include_values,
            search_in=search_in,
            max_response_chars=max_response_chars,
        )

        metrics = {
            "status": "success",
            "file": str(file_path),
            "sheet_scope": sheet_scope,
            "searched_sheets": len(searched_sheets),
            "scanned_cells": scanned,
            "requested_cells": requested_cells,
            "candidate_cells": candidate_cells,
            "total_matches": total_matches,
            "returned_matches": payload.get("returned_matches", payload.get("returned", len(results))),
            "include_values": bool(payload.get("include_values")),
            "truncated": bool(payload.get("truncated")),
            "result_truncated": bool(payload.get("result_truncated")),
            "scan_truncated": scan_truncated,
        }
        if sheet_scope != "all" and searched_sheets:
            metrics["sheet"] = searched_sheets[0]
        return payload, metrics
    finally:
        for wb in (wb_formula, wb_values):
            if wb is None:
                continue
            try:
                wb.close()
            except Exception:
                pass


class FindCellsTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None
        self.max_response_chars = _get_max_response_chars(config)

    def get_openai_tool_schema(self) -> OpenAIFunctionToolSchema:
        return self.tool_schema

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return _error_tool_response("invalid_path", "invalid path.")
        if relpath.suffix.lower() != ".xlsx":
            return _error_tool_response("invalid_path", "only .xlsx workbooks are supported.")

        sheet_name_raw = parameters.get("sheet_name")
        if sheet_name_raw is None or (isinstance(sheet_name_raw, str) and not sheet_name_raw.strip()):
            sheet_name = None
        elif not isinstance(sheet_name_raw, str):
            return _error_tool_response("invalid_sheet_name", "sheet_name must be a string when provided.")
        else:
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return _error_tool_response("invalid_sheet_name", "sheet_name is empty after normalization.")

        query_raw = parameters.get("query")
        if not isinstance(query_raw, str) or not query_raw.strip():
            return _error_tool_response(
                "invalid_query",
                "query received empty; pass the text to search for (e.g. a header name).",
            )
        query = query_raw.strip()

        try:
            match_mode = _normalize_match_mode(parameters.get("match"))
            search_in = _normalize_search_in(parameters.get("search_in"))
            return_mode = _normalize_return_mode(parameters.get("return"))
        except ValueError as exc:
            return _error_tool_response("invalid_parameters", str(exc))

        if match_mode == "regex":
            max_pat = _get_max_regex_pattern_chars()
            if len(query) > max_pat:
                return _error_tool_response(
                    "invalid_query",
                    f"regex pattern too long (len={len(query)}, max={max_pat}).",
                )
            regex_error = _validate_regex_pattern(query)
            if regex_error:
                return _error_tool_response("invalid_query", regex_error)

        try:
            case_sensitive = _coerce_bool_parameter(
                parameters.get("case_sensitive", False),
                parameter="case_sensitive",
            )
        except ValueError as exc:
            return _error_tool_response("invalid_case_sensitive", str(exc))

        try:
            include_values = _coerce_bool_parameter(
                parameters.get("include_values", False),
                parameter="include_values",
            )
        except ValueError as exc:
            return _error_tool_response("invalid_include_values", str(exc))

        search_range = parameters.get("range")
        if search_range is not None and (not isinstance(search_range, str) or not search_range.strip()):
            return _error_tool_response("invalid_range", "range must be a non-empty string if provided.")

        try:
            max_results = _coerce_int_parameter(parameters.get("max_results"), parameter="max_results", default=20)
        except ValueError as exc:
            return _error_tool_response("invalid_max_results", str(exc))
        max_results = min(max(1, max_results), 1000)

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return _error_tool_response("missing_workspace_id", "workspace_id is missing/invalid.")
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return _error_tool_response("file_not_found", f"file not found: {relpath}")
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return _error_tool_response("stat_failed", f"failed to stat file: {exc}")
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return _error_tool_response(
                    "file_too_large",
                    f"workbook is too large ({actual_mb:.1f}MB > {max_mb}MB).",
                )
        zip_error = await asyncio.to_thread(
            _scan_zip_metadata,
            file_path,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            return _error_tool_response("zip_limits_exceeded", f"workbook rejected by zip safety checks: {zip_error}")

        try:
            if sheet_name is None:
                payload, metrics = await asyncio.to_thread(
                    _find_cells_all_sheets_sync,
                    file_path=file_path,
                    query=query,
                    search_range=search_range,
                    match_mode=match_mode,
                    search_in=search_in,
                    case_sensitive=case_sensitive,
                    return_mode=return_mode,
                    max_results=max_results,
                    max_response_chars=self.max_response_chars,
                    include_values=include_values,
                )
            else:
                payload, metrics = await asyncio.to_thread(
                    _find_cells_sync,
                    file_path=file_path,
                    sheet_name=sheet_name,
                    query=query,
                    search_range=search_range,
                    match_mode=match_mode,
                    search_in=search_in,
                    case_sensitive=case_sensitive,
                    return_mode=return_mode,
                    max_results=max_results,
                    max_response_chars=self.max_response_chars,
                    include_values=include_values,
                )
        except asyncio.CancelledError:
            raise
        except re.error as exc:
            return _error_tool_response("invalid_regex", f"invalid regex: {exc}")
        except Exception as exc:
            return _error_tool_response("find_failed", str(exc))

        payload["file"] = str(relpath)
        if sheet_name is not None:
            sheet_scope = payload.get("sheet", sheet_name)
            payload.setdefault("sheet_scope", sheet_scope)
            payload.setdefault("searched_sheets", [sheet_scope])
            payload.setdefault("total_matches", payload.get("returned_matches", payload.get("returned", 0)))
            payload.setdefault("max_results", max_results)
            payload.setdefault("result_truncated", bool(payload.get("truncated")) and not payload.get("scan_truncated"))
            metrics.setdefault("sheet_scope", sheet_scope)
            metrics.setdefault("searched_sheets", 1)
            metrics.setdefault("total_matches", payload.get("total_matches", 0))
            metrics.setdefault("result_truncated", payload.get("result_truncated", False))
        metrics["file"] = str(relpath)
        payload = _fit_success_payload_to_max_chars(payload, self.max_response_chars)
        metrics["returned_matches"] = payload.get(
            "returned_matches",
            payload.get("returned", metrics.get("returned_matches", 0)),
        )
        metrics["truncated"] = bool(payload.get("truncated"))
        return ToolResponse(text=_json_dumps_compact(payload)), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
