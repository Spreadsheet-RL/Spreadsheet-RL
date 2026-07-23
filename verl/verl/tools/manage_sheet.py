# Copyright 2026 Bytedance Ltd. and/or its affiliates
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
from __future__ import annotations

import asyncio
import json
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .formula_fill import (
    _commit_recalc_workbook_under_lock_sync,
    _copy_workbook_to_temp_under_lock_sync,
    _get_max_response_chars,
    _normalize_sheet_name,
    _resolve_workspace_file,
    _sanitize_relpath,
    _scan_zip_metadata,
    _truncate_str,
)
from .schemas import OpenAIFunctionToolSchema, ToolResponse
from .workbook_formula_cache import (
    _collect_formula_cached_values_from_xlsx,
    _restore_formula_cached_values_in_xlsx,
)
from .worksheet_resolution import WorksheetResolutionError, resolve_worksheet

_INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")
_ACTIONS = {"create", "rename", "delete", "copy", "move", "hide", "unhide"}
_RENAME_WARNING = "formulas referencing the old sheet name are not rewritten"


@dataclass(frozen=True)
class _ManageResult:
    action: str
    sheet: str
    new_name: Optional[str]
    index: Optional[int]
    sheets: list[dict[str, Any]]
    warning: Optional[str] = None
    formula_cache_preserved: bool = True


class _ManageSheetError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _snippet(value: Any, *, max_chars: int = 160) -> str:
    try:
        text = json.dumps(value, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        text = repr(value)
    return _truncate_str(text, max_chars)


def _validation_response(code: str, message: str) -> tuple[ToolResponse, float, dict[str, Any]]:
    return ToolResponse(text=f"Error: {message}"), 0.0, {"status": "error", "error": code}


def _is_valid_excel_sheet_name(name: str) -> bool:
    return bool(name) and len(name) <= 31 and _INVALID_SHEET_CHARS_RE.search(name) is None


def _normalize_required_sheet_name(parameters: dict[str, Any], key: str) -> str:
    raw = parameters.get(key)
    if not isinstance(raw, str) or not raw.strip():
        raise _ManageSheetError(
            "invalid_sheet_name",
            f"{key} received {_snippet(raw)}; pass a non-empty worksheet name.",
        )
    name = _normalize_sheet_name(raw)
    if not _is_valid_excel_sheet_name(name):
        raise _ManageSheetError(
            "invalid_sheet_name",
            f"{key} received {_snippet(raw)}; pass a valid Excel sheet name (1-31 chars, no []:*?/\\).",
        )
    return name


def _parse_index(raw: Any, *, required: bool, max_index: int, parameter: str = "index") -> Optional[int]:
    if raw is None:
        if required:
            raise _ManageSheetError(
                "invalid_index",
                f"{parameter} received missing; pass a 0-based final sheet position.",
            )
        return None
    if isinstance(raw, bool):
        raise _ManageSheetError("invalid_index", f"{parameter} received {_snippet(raw)}; pass a 0-based integer.")
    if isinstance(raw, str):
        token = raw.strip()
        if not token:
            raise _ManageSheetError("invalid_index", f"{parameter} received empty; pass a 0-based integer.")
        try:
            value = int(token)
        except ValueError:
            raise _ManageSheetError(
                "invalid_index",
                f"{parameter} received {_snippet(raw)}; pass a 0-based integer.",
            ) from None
    else:
        try:
            value = int(raw)
        except (TypeError, ValueError):
            raise _ManageSheetError(
                "invalid_index",
                f"{parameter} received {_snippet(raw)}; pass a 0-based integer.",
            ) from None
    if value < 0 or value > max_index:
        raise _ManageSheetError(
            "invalid_index",
            f"{parameter} received {value}; pass a value between 0 and {max_index}.",
        )
    return value


def _sheet_by_casefold(wb: Any, requested: str):
    try:
        return resolve_worksheet(wb, requested)
    except WorksheetResolutionError as exc:
        if exc.code == "sheet_not_found":
            raise _ManageSheetError("sheet_not_found", str(exc)) from None
        if exc.code == "ambiguous_sheet_name":
            names = ", ".join(repr(name) for name in exc.matches[:10])
            raise _ManageSheetError(
                "invalid_sheet_name",
                f"sheet_name received {requested!r}: ambiguous sheet name matches {names}; pass the exact sheet title.",
            ) from None
        raise _ManageSheetError(
            "invalid_sheet_name",
            f"sheet_name received {requested!r}: sheet is not a worksheet.",
        ) from None


def _ensure_unique_sheet_name(
    wb: Any,
    name: str,
    *,
    parameter: str = "sheet_name",
    ignore_title: Optional[str] = None,
) -> None:
    for existing in getattr(wb, "sheetnames", []) or []:
        if ignore_title is not None and existing == ignore_title:
            continue
        if str(existing).casefold() == name.casefold():
            raise _ManageSheetError(
                "sheet_exists",
                f"{parameter} received {name!r} but conflicts with existing sheet {existing!r}; "
                "pass a unique sheet name.",
            )


def _unique_copy_name(wb: Any, source_name: str) -> str:
    existing = {str(name).casefold() for name in getattr(wb, "sheetnames", []) or []}
    base = f"{source_name} Copy"
    if len(base) > 31:
        base = f"{source_name[:26]} Copy"
    if base.casefold() not in existing and _is_valid_excel_sheet_name(base):
        return base
    idx = 2
    while True:
        suffix = f" {idx}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing and _is_valid_excel_sheet_name(candidate):
            return candidate
        idx += 1


def _visible_sheets(wb: Any) -> list[Any]:
    return [ws for ws in getattr(wb, "worksheets", []) if getattr(ws, "sheet_state", "visible") == "visible"]


def _sheet_list(wb: Any) -> list[dict[str, Any]]:
    sheets: list[dict[str, Any]] = []
    for idx, sheet in enumerate(getattr(wb, "_sheets", []) or []):
        item = {
            "name": str(getattr(sheet, "title", "")),
            "state": str(getattr(sheet, "sheet_state", "visible") or "visible"),
            "index": idx,
        }
        sheet_type = str(getattr(sheet, "_rel_type", "") or "").strip()
        if sheet_type and sheet_type != "worksheet":
            item["type"] = sheet_type
        sheets.append(item)
    return sheets


def _move_sheet_to_index(wb: Any, ws: Any, index: int) -> None:
    sheets = wb._sheets
    sheets.remove(ws)
    sheets.insert(index, ws)


def _mutate_workbook_sync(
    *,
    file_path: Path,
    action: str,
    sheet_name: str,
    new_name: Optional[str],
    index: Optional[int],
    max_sheets: int,
) -> _ManageResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    cached_formula_values = _collect_formula_cached_values_from_xlsx(file_path)
    wb = None
    title_map: dict[str, str] = {}
    warning: Optional[str] = None
    try:
        wb = load_workbook(filename=str(file_path), keep_vba=False, data_only=False, keep_links=True)
        worksheets = list(getattr(wb, "worksheets", None) or [])
        all_sheets = list(getattr(wb, "_sheets", None) or [])
        if not worksheets and action != "create":
            raise _ManageSheetError(
                "sheet_not_found",
                f"sheet_name received {sheet_name!r} but the workbook has no worksheets; use action='create' first.",
            )

        result_sheet = sheet_name
        result_new_name = new_name
        result_index = index

        if action == "create":
            _ensure_unique_sheet_name(wb, sheet_name)
            if max_sheets > 0 and len(all_sheets) >= max_sheets:
                raise _ManageSheetError(
                    "too_many_sheets",
                    f"action received 'create' but workbook already has {len(all_sheets)} sheets "
                    f">= {max_sheets}; delete a sheet first.",
                )
            final_index = _parse_index(index, required=False, max_index=len(all_sheets)) if index is not None else None
            ws = wb.create_sheet(title=sheet_name, index=final_index)
            result_index = wb._sheets.index(ws)

        elif action == "rename":
            ws = _sheet_by_casefold(wb, sheet_name)
            assert new_name is not None
            _ensure_unique_sheet_name(wb, new_name, parameter="new_name", ignore_title=str(ws.title))
            old_title = str(ws.title)
            ws.title = new_name
            title_map[old_title] = new_name
            result_sheet = old_title
            result_new_name = new_name
            result_index = wb._sheets.index(ws)
            warning = _RENAME_WARNING

        elif action == "delete":
            ws = _sheet_by_casefold(wb, sheet_name)
            if getattr(ws, "sheet_state", "visible") == "visible" and len(_visible_sheets(wb)) <= 1:
                raise _ManageSheetError(
                    "last_visible_sheet",
                    f"sheet_name received {sheet_name!r}; Excel requires at least one visible sheet.",
                )
            result_sheet = str(ws.title)
            result_index = wb._sheets.index(ws)
            wb.remove(ws)

        elif action == "copy":
            source = _sheet_by_casefold(wb, sheet_name)
            if max_sheets > 0 and len(all_sheets) >= max_sheets:
                raise _ManageSheetError(
                    "too_many_sheets",
                    f"action received 'copy' but workbook already has {len(all_sheets)} sheets "
                    f">= {max_sheets}; delete a sheet first.",
                )
            copy_name = new_name or _unique_copy_name(wb, str(source.title))
            _ensure_unique_sheet_name(wb, copy_name, parameter="new_name")
            copied = wb.copy_worksheet(source)
            copied.title = copy_name
            title_map[str(source.title)] = copy_name
            final_index = (
                _parse_index(index, required=False, max_index=len(wb._sheets) - 1) if index is not None else None
            )
            if final_index is not None:
                _move_sheet_to_index(wb, copied, final_index)
            result_sheet = str(source.title)
            result_new_name = copy_name
            result_index = wb._sheets.index(copied)

        elif action == "move":
            ws = _sheet_by_casefold(wb, sheet_name)
            final_index = _parse_index(index, required=True, max_index=len(wb._sheets) - 1)
            assert final_index is not None
            _move_sheet_to_index(wb, ws, final_index)
            result_sheet = str(ws.title)
            result_index = wb._sheets.index(ws)

        elif action == "hide":
            ws = _sheet_by_casefold(wb, sheet_name)
            if getattr(ws, "sheet_state", "visible") == "visible" and len(_visible_sheets(wb)) <= 1:
                raise _ManageSheetError(
                    "last_visible_sheet",
                    f"sheet_name received {sheet_name!r}; Excel requires at least one visible sheet.",
                )
            ws.sheet_state = "hidden"
            result_sheet = str(ws.title)
            result_index = wb._sheets.index(ws)

        elif action == "unhide":
            ws = _sheet_by_casefold(wb, sheet_name)
            if getattr(ws, "sheet_state", "visible") == "visible":
                raise _ManageSheetError(
                    "sheet_not_hidden",
                    f"sheet_name received {sheet_name!r} but the sheet is already visible; pass a hidden sheet name.",
                )
            ws.sheet_state = "visible"
            result_sheet = str(ws.title)
            result_index = wb._sheets.index(ws)

        else:
            raise _ManageSheetError(
                "invalid_action",
                f"action received {_snippet(action)}; pass one of create, rename, delete, copy, move, hide, unhide.",
            )

        wb.save(str(file_path))
        formula_cache_preserved = True
        try:
            _restore_formula_cached_values_in_xlsx(file_path, cached_formula_values, title_map=title_map)
        except Exception:
            formula_cache_preserved = False
        return _ManageResult(
            action=action,
            sheet=result_sheet,
            new_name=result_new_name,
            index=result_index,
            sheets=_sheet_list(wb),
            warning=warning,
            formula_cache_preserved=formula_cache_preserved,
        )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _truncate_payload_to_max_chars(payload: dict[str, Any], max_chars: int) -> str:
    text = _json_dumps_compact(payload)
    if len(text) <= max_chars:
        return text
    sheets = payload.get("sheets")
    if isinstance(sheets, list):
        original = list(sheets)
        sheet_count = len(original)
        low = 0
        high = len(original)
        best = 0
        while low <= high:
            mid = (low + high) // 2
            payload["sheets"] = original[:mid]
            payload["sheets_truncated"] = mid < sheet_count
            payload["sheet_count"] = sheet_count
            text = _json_dumps_compact(payload)
            if len(text) <= max_chars:
                best = mid
                low = mid + 1
            else:
                high = mid - 1
        payload["sheets"] = original[:best]
        payload["sheets_truncated"] = best < sheet_count
        payload["sheet_count"] = sheet_count
        text = _json_dumps_compact(payload)
        if len(text) <= max_chars:
            return text
    return _json_dumps_compact({"status": payload.get("status"), "action": payload.get("action"), "truncated": True})


class ManageSheetTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        try:
            self.lock_timeout_s = max(0.0, float(config.get("lock_timeout_s", 30)))
        except (TypeError, ValueError):
            self.lock_timeout_s = 30.0
        try:
            self.max_sheets = max(0, int(config.get("max_sheets", 200)))
        except (TypeError, ValueError):
            self.max_sheets = 200
        self.max_response_chars = _get_max_response_chars(config)

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return _validation_response(
                    "invalid_path",
                    f"path received {_snippet(raw_path)}; pass a relative .xlsx path such as data.xlsx.",
                )
        if relpath.suffix.lower() != ".xlsx":
            return _validation_response(
                "invalid_path",
                f"path received {str(relpath)!r}; pass a relative .xlsx workbook path such as data.xlsx.",
            )

        action_raw = parameters.get("action")
        if not isinstance(action_raw, str) or action_raw.strip().casefold() not in _ACTIONS:
            return _validation_response(
                "invalid_action",
                f"action received {_snippet(action_raw)}; "
                "pass one of create, rename, delete, copy, move, hide, unhide.",
            )
        action = action_raw.strip().casefold()
        try:
            sheet_name = _normalize_required_sheet_name(parameters, "sheet_name")
            new_name = None
            if action == "rename":
                new_name = _normalize_required_sheet_name(parameters, "new_name")
            elif action == "copy" and parameters.get("new_name") is not None:
                new_name = _normalize_required_sheet_name(parameters, "new_name")
            index = parameters.get("index")
            if action == "move":
                _parse_index(index, required=True, max_index=1_000_000)
        except _ManageSheetError as exc:
            return _validation_response(exc.code, str(exc))

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return _validation_response(
                "missing_workspace_id",
                "workspace_id received missing/invalid; pass a valid workspace_id for the workbook.",
            )
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return _validation_response(
                "file_not_found",
                f"path received {str(relpath)!r} but no workbook was found; create it in the workspace or fix path.",
            )
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return _validation_response(
                    "file_not_found",
                    f"path received {str(relpath)!r} could not be stat'ed ({exc}); verify the workbook exists.",
                )
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return _validation_response(
                    "file_too_large",
                    f"path received {str(relpath)!r} is {actual_mb:.1f}MB > {max_mb}MB; use a smaller workbook.",
                )

        tmp_path: Optional[Path] = None
        try:
            tmp_path, expected_sig = await asyncio.to_thread(
                _copy_workbook_to_temp_under_lock_sync,
                file_path=file_path,
                lock_timeout_s=self.lock_timeout_s,
            )
            result = await asyncio.to_thread(
                _mutate_workbook_sync,
                file_path=tmp_path,
                action=action,
                sheet_name=sheet_name,
                new_name=new_name,
                index=index,
                max_sheets=self.max_sheets,
            )
            zip_error = await asyncio.to_thread(
                _scan_zip_metadata,
                tmp_path,
                max_members=50_000,
                max_total_uncompressed_bytes=512 * 1024 * 1024,
                max_member_uncompressed_bytes=128 * 1024 * 1024,
                max_ratio=200.0,
            )
            if zip_error:
                raise RuntimeError(f"filled workbook rejected by zip safety checks: {zip_error}")
            content = await asyncio.to_thread(tmp_path.read_bytes)
            committed, writeback_error = await asyncio.to_thread(
                _commit_recalc_workbook_under_lock_sync,
                file_path=file_path,
                expected_sig=expected_sig,
                recalc_content=content,
                lock_timeout_s=self.lock_timeout_s,
            )
            if not committed:
                raise RuntimeError(writeback_error or "writeback failed")
            tmp_path.unlink(missing_ok=True)
            tmp_path = None
        except asyncio.CancelledError:
            raise
        except _ManageSheetError as exc:
            return _validation_response(exc.code, str(exc))
        except Exception as exc:
            payload = {
                "status": "error",
                "error": "manage_sheet_failed",
                "file": str(relpath),
                "rolled_back": True,
                "message": _truncate_str(str(exc), 500),
            }
            return (
                ToolResponse(text=_json_dumps_compact(payload)),
                0.0,
                {
                    "status": "error",
                    "error": "manage_sheet_failed",
                    "rolled_back": True,
                },
            )
        finally:
            if tmp_path is not None:
                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass

        payload: dict[str, Any] = {
            "status": "success",
            "action": result.action,
            "sheet": result.sheet,
            "index": result.index,
            "sheets": result.sheets,
        }
        if result.new_name is not None:
            payload["new_name"] = result.new_name
        if result.warning:
            payload["warning"] = result.warning
        if not result.formula_cache_preserved:
            payload["formula_cache_preserved"] = False
        response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
        metrics = {
            "status": "success",
            "action": result.action,
            "sheet": result.sheet,
            "sheet_count": len(result.sheets),
            "sheets_truncated": bool(payload.get("sheets_truncated")),
        }
        if result.new_name is not None:
            metrics["new_name"] = result.new_name
        if not result.formula_cache_preserved:
            metrics["formula_cache_preserved"] = False
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
