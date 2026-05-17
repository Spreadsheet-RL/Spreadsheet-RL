from __future__ import annotations

import asyncio
import datetime as dt
import errno
import json
import math
import os
import re
import stat
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, Optional

from verl.utils.paths import get_spreadsheet_rl_data_root, normalize_workspace_id
from verl.utils.rollout_trace import rollout_trace_op

from .base_tool import BaseTool
from .schemas import OpenAIFunctionToolSchema, ToolResponse

_A1_CELL_RE = r"[A-Z]{1,3}[0-9]{1,7}"
_A1_COL_RE = r"[A-Z]{1,3}"
_A1_ROW_RE = r"[0-9]{1,7}"
_A1_RANGE_RE = re.compile(
    rf"^(?:{_A1_CELL_RE}(?::{_A1_CELL_RE})?|{_A1_COL_RE}:{_A1_COL_RE}|{_A1_ROW_RE}:{_A1_ROW_RE})$"
)
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z0-9_]+$")

_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLS = 16_384


def _sanitize_relpath(value: Any) -> Optional[Path]:
    if not isinstance(value, str):
        return None
    raw = value.strip()
    if not raw:
        return None
    if "\0" in raw:
        return None
    p = Path(raw)
    if p.is_absolute():
        return None
    if any(part == ".." for part in p.parts):
        return None
    return p


def _resolve_workspace_file(*, workspace_id: Optional[str], relpath: Path) -> Optional[Path]:
    if not workspace_id:
        return None

    workspace_base = get_spreadsheet_rl_data_root()
    workspaces_base = workspace_base / "_workspaces"
    try:
        if workspace_base.is_symlink():
            return None
        workspaces_base_resolved = workspaces_base.resolve(strict=True)
    except (OSError, RuntimeError):
        return None

    workspace_root = workspaces_base / workspace_id
    try:
        if workspace_root.is_symlink():
            return None
        workspace_root_resolved = workspace_root.resolve(strict=True)
    except (OSError, RuntimeError):
        return None
    if not workspace_root_resolved.is_relative_to(workspaces_base_resolved):
        return None

    rel_candidates = [relpath]
    if relpath.parts and relpath.parts[0] == "data" and len(relpath.parts) > 1:
        rel_candidates.append(Path(*relpath.parts[1:]))

    for rel_candidate in rel_candidates:
        candidate = workspace_root / rel_candidate
        try:
            resolved = candidate.resolve(strict=True)
        except (OSError, RuntimeError):
            continue
        if not resolved.is_relative_to(workspace_root_resolved):
            continue

        if candidate.is_symlink():
            continue

        cursor = workspace_root
        safe = True
        for part in rel_candidate.parts:
            cursor = cursor / part
            try:
                if cursor.is_symlink():
                    safe = False
                    break
            except OSError:
                safe = False
                break
        if not safe:
            continue

        if resolved.is_file():
            return resolved
    return None


def _normalize_sheet_name(value: str) -> str:
    name = value.strip(" \t\n\r\f\v\"")
    if len(name) >= 2 and name[0] == "'" and name[-1] == "'":
        name = name[1:-1].replace("''", "'")
        return name.strip()

    if name.startswith("'") and not name.endswith("'"):
        name = name[1:]
    elif name.endswith("'") and not name.startswith("'"):
        name = name[:-1]
    return name.strip()


def _quote_sheet_name_for_a1(sheet_name: str) -> str:
    if _SAFE_SHEET_NAME_RE.fullmatch(sheet_name):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _normalize_cell_range(value: str) -> str:
    rng = value.strip(" \t\n\r\f\v\"'")
    rng = rng.replace(" ", "").replace("$", "")
    return rng.upper()


def _split_sheet_cell_range(token: str) -> tuple[Optional[str], str]:
    token = token.strip(" \t\n\r\f\v\"")
    if not token:
        raise ValueError("empty cell range")

    if "!" not in token:
        return None, token

    sheet_part, _, range_part = token.rpartition("!")
    sheet_name = _normalize_sheet_name(sheet_part)
    if not sheet_name:
        raise ValueError(f"empty sheet name: {token!r}")
    if not range_part.strip():
        raise ValueError(f"empty cell range: {token!r}")
    return sheet_name, range_part


def _fill_missing_boundaries(
    boundaries: tuple[Optional[int], Optional[int], Optional[int], Optional[int]],
) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = boundaries
    if min_col is None:
        min_col = 1
    if max_col is None:
        max_col = _EXCEL_MAX_COLS
    if min_row is None:
        min_row = 1
    if max_row is None:
        max_row = _EXCEL_MAX_ROWS
    return int(min_col), int(min_row), int(max_col), int(max_row)


def _get_max_string_chars() -> int:
    raw = os.environ.get("SPREADSHEET_RL_TOOL_MAX_STRING_CHARS", "200").strip()
    try:
        n = int(raw)
        return min(max(16, n), 100_000)
    except ValueError:
        return 200


def _get_max_response_chars(config: Optional[dict[str, Any]] = None) -> int:
    config_value = None
    if isinstance(config, dict):
        config_value = config.get("max_response_chars")
    for raw in (config_value, os.environ.get("SPREADSHEET_RL_TOOL_MAX_RESPONSE_CHARS", "").strip()):
        if raw is None:
            continue
        try:
            n = int(raw)
            return min(max(128, n), 1000)
        except (TypeError, ValueError):
            continue
    return 900


def _truncate_str(value: str, max_chars: int) -> str:
    if max_chars <= 0:
        return ""
    if len(value) <= max_chars:
        return value
    if max_chars <= 3:
        return value[:max_chars]
    return value[: max_chars - 3] + "..."


def _to_jsonable_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, str)):
        if isinstance(value, str):
            return _truncate_str(value, _get_max_string_chars())
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return str(value)


def _json_dumps_compact(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _scan_zip_metadata(
    file_path: Path,
    *,
    max_members: int,
    max_total_uncompressed_bytes: int,
    max_member_uncompressed_bytes: int,
    max_ratio: float,
) -> Optional[str]:
    try:
        import zipfile
    except ImportError:
        return None

    try:
        with zipfile.ZipFile(file_path) as zf:
            infos = zf.infolist()
            if len(infos) > max_members:
                return f"too many zip members (members={len(infos)}, max={max_members})"

            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                uncompressed = int(getattr(info, "file_size", 0) or 0)
                compressed = int(getattr(info, "compress_size", 0) or 0)

                total_uncompressed += uncompressed
                total_compressed += compressed

                if uncompressed > max_member_uncompressed_bytes:
                    return (
                        "zip member too large "
                        f"(name={getattr(info, 'filename', '?')!r}, bytes={uncompressed}, max={max_member_uncompressed_bytes})"
                    )
                if compressed > 0 and max_ratio > 0 and (uncompressed / compressed) > max_ratio:
                    return (
                        "zip member compression ratio too large "
                        f"(name={getattr(info, 'filename', '?')!r}, ratio={uncompressed / compressed:.1f}, max={max_ratio})"
                    )

                if total_uncompressed > max_total_uncompressed_bytes:
                    return (
                        "zip contents too large "
                        f"(total_bytes={total_uncompressed}, max={max_total_uncompressed_bytes})"
                    )

            if total_compressed > 0 and max_ratio > 0 and (total_uncompressed / total_compressed) > max_ratio:
                return (
                    "zip total compression ratio too large "
                    f"(ratio={total_uncompressed / total_compressed:.1f}, max={max_ratio})"
                )
    except zipfile.BadZipFile:
        return "invalid zip file"
    except Exception as exc:
        return f"failed to read zip metadata: {exc}"

    return None


def _acquire_lockfile(lock_path: Path, timeout_s: float):
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    flags = os.O_RDWR | os.O_CREAT
    nofollow = getattr(os, "O_NOFOLLOW", 0)
    if nofollow:
        flags |= nofollow
    else:
        try:
            if lock_path.is_symlink():
                raise RuntimeError(f"lock file path is a symlink: {lock_path}")
        except OSError:
            pass

    try:
        fd = os.open(lock_path, flags, 0o600)
    except OSError as exc:
        if exc.errno == errno.ELOOP:
            raise RuntimeError(f"lock file path is a symlink: {lock_path}") from None
        raise RuntimeError(f"failed to open lock file: {exc}") from None

    try:
        st = os.fstat(fd)
    except OSError as exc:
        os.close(fd)
        raise RuntimeError(f"failed to stat lock file: {exc}") from None
    if not stat.S_ISREG(st.st_mode):
        os.close(fd)
        raise RuntimeError(f"lock file is not a regular file: {lock_path}") from None

    lock_file = os.fdopen(fd, "r+", encoding="utf-8", errors="replace")
    try:
        import fcntl
    except ImportError:
        lock_file.close()
        raise RuntimeError("file locking requires fcntl (Unix-only)") from None

    try:
        deadline = time.monotonic() + max(0.0, timeout_s)
        while True:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                return lock_file
            except BlockingIOError:
                if time.monotonic() >= deadline:
                    raise TimeoutError(f"timed out waiting for lock: {lock_path}") from None
                time.sleep(0.1)
    except BaseException:
        lock_file.close()
        raise


def _resolve_target_worksheet(wb, requested_name: str):
    worksheets = getattr(wb, "worksheets", None) or []
    for ws in worksheets:
        if getattr(ws, "title", None) == requested_name:
            return ws

    requested_cf = requested_name.casefold()
    matches = [ws for ws in worksheets if getattr(ws, "title", "").casefold() == requested_cf]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        candidates = ", ".join(repr(getattr(ws, "title", "")) for ws in matches[:5])
        raise RuntimeError(f"ambiguous sheet name: {requested_name!r} matches {candidates}")

    sheetnames = getattr(wb, "sheetnames", None) or []
    if any(isinstance(name, str) and name.casefold() == requested_cf for name in sheetnames):
        raise RuntimeError(f"sheet is not a worksheet: {requested_name!r}")
    raise RuntimeError(f"sheet not found: {requested_name!r}")


def _format_a1_address(*, sheet_name: str, row: int, col: int) -> str:
    try:
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        col_letter = str(col)
    else:
        col_letter = get_column_letter(col)
    sheet_ref = _quote_sheet_name_for_a1(sheet_name)
    return f"{sheet_ref}!{col_letter}{row}"


def _sample_indices(total: int, *, head: int, tail: int) -> list[int]:
    if total <= 0:
        return []
    head_n = max(0, head)
    tail_n = max(0, tail)
    if total <= head_n + tail_n:
        return list(range(total))
    indices = list(range(head_n))
    indices.extend(range(total - tail_n, total))
    return indices


def _truncate_payload_to_max_chars(payload: dict[str, Any], max_chars: int) -> str:
    response_text = _json_dumps_compact(payload)
    if len(response_text) <= max_chars:
        return response_text

    payload["truncated"] = True
    samples_out = payload.get("samples")
    if not isinstance(samples_out, list):
        return _json_dumps_compact(payload)

    for head, tail in ((2, 2), (1, 1)):
        if len(samples_out) > head + tail:
            payload["samples"] = samples_out[:head] + samples_out[-tail:]
            samples_out = payload.get("samples")
        response_text = _json_dumps_compact(payload)
        if len(response_text) <= max_chars:
            return response_text

    if len(response_text) > max_chars and isinstance(samples_out, list) and len(samples_out) > 1:
        payload["samples"] = samples_out[:1]
        response_text = _json_dumps_compact(payload)

    return response_text


def _clear_range_in_workbook(
    *,
    file_path: Path,
    sheet_name: Optional[str],
    range_token: str,
    lock_timeout_s: float,
    max_iter_cells: int,
    max_scan_cells: int,
) -> tuple[str, str, int, list[dict[str, Any]], list[str]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required to edit workbooks; install openpyxl>=3.1.5") from None

    lock_file = _acquire_lockfile(file_path.with_suffix(file_path.suffix + ".lock"), timeout_s=lock_timeout_s)
    wb = None
    tmp_path: Optional[Path] = None
    try:
        try:
            orig_mode = stat.S_IMODE(file_path.stat().st_mode)
        except OSError as exc:
            raise RuntimeError(f"failed to stat workbook: {exc}") from None

        zip_error = _scan_zip_metadata(
            file_path,
            max_members=50_000,
            max_total_uncompressed_bytes=512 * 1024 * 1024,
            max_member_uncompressed_bytes=128 * 1024 * 1024,
            max_ratio=200.0,
        )
        if zip_error:
            raise RuntimeError(f"workbook rejected by zip safety checks: {zip_error}")

        try:
            wb = load_workbook(filename=str(file_path), keep_vba=False, data_only=False, keep_links=False)
        except Exception as exc:
            raise RuntimeError(f"failed to load workbook: {exc}") from None

        worksheets = getattr(wb, "worksheets", None) or []
        if not worksheets:
            raise RuntimeError("workbook has no worksheets")

        default_sheet = getattr(worksheets[0], "title", "") or ""
        parsed_sheet, range_part = _split_sheet_cell_range(range_token)
        if parsed_sheet is None:
            requested_sheet = sheet_name or default_sheet
        else:
            if sheet_name is not None and parsed_sheet.casefold() != sheet_name.casefold():
                raise RuntimeError(f"sheet mismatch: {sheet_name!r} vs {parsed_sheet!r}")
            requested_sheet = parsed_sheet

        ws = _resolve_target_worksheet(wb, requested_sheet) if requested_sheet else worksheets[0]
        resolved_sheet_name = getattr(ws, "title", requested_sheet) or requested_sheet or ""
        sheet_ref = _quote_sheet_name_for_a1(resolved_sheet_name)

        cell_range = _normalize_cell_range(range_part)
        if not _A1_RANGE_RE.fullmatch(cell_range):
            raise RuntimeError(f"invalid A1 cell range: {range_part!r}")

        try:
            boundaries = range_boundaries(cell_range if ":" in cell_range else f"{cell_range}:{cell_range}")
        except Exception as exc:
            raise RuntimeError(f"failed to parse range boundaries: {exc}") from None
        min_col, min_row, max_col, max_row = _fill_missing_boundaries(boundaries)
        if min_col < 1 or min_row < 1 or max_col > _EXCEL_MAX_COLS or max_row > _EXCEL_MAX_ROWS:
            raise RuntimeError("invalid range boundaries")

        if min_col > max_col:
            min_col, max_col = max_col, min_col
        if min_row > max_row:
            min_row, max_row = max_row, min_row

        if ":" not in cell_range:
            normalized_range = cell_range
        elif re.fullmatch(rf"{_A1_COL_RE}:{_A1_COL_RE}", cell_range):
            normalized_range = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
        elif re.fullmatch(rf"{_A1_ROW_RE}:{_A1_ROW_RE}", cell_range):
            normalized_range = f"{min_row}:{max_row}"
        else:
            normalized_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

        total_rows = max_row - min_row + 1
        total_cols = max_col - min_col + 1
        area = total_rows * total_cols
        scan_threshold = max(1, int(max_iter_cells))

        anchors_to_clear: set[tuple[int, int]] = set()
        merged_ranges = getattr(getattr(ws, "merged_cells", None), "ranges", None) or []
        for merged_range in merged_ranges:
            try:
                m_min_col = int(getattr(merged_range, "min_col"))
                m_min_row = int(getattr(merged_range, "min_row"))
                m_max_col = int(getattr(merged_range, "max_col"))
                m_max_row = int(getattr(merged_range, "max_row"))
            except Exception:
                continue
            if m_max_col < min_col or m_min_col > max_col or m_max_row < min_row or m_min_row > max_row:
                continue
            anchors_to_clear.add((m_min_row, m_min_col))

        cleared_merged_anchor_coords: set[tuple[int, int]] = set()
        cleared_cells = 0
        cells = getattr(ws, "_cells", None)
        if isinstance(cells, dict):
            max_scan_cells = max(0, int(max_scan_cells))
            use_grid_scan = area <= scan_threshold
            if not use_grid_scan:
                grid_ok = max_scan_cells == 0 or area <= max_scan_cells
                if grid_ok and area < len(cells):
                    use_grid_scan = True

            if use_grid_scan:
                for row_num in range(min_row, max_row + 1):
                    for col_num in range(min_col, max_col + 1):
                        cell = cells.get((row_num, col_num))
                        if cell is None:
                            continue
                        try:
                            current = getattr(cell, "value", None)
                        except Exception:
                            continue
                        if current is None:
                            continue
                        try:
                            cell.value = None
                        except Exception:
                            continue
                        cleared_cells += 1
            else:
                if max_scan_cells and len(cells) > max_scan_cells:
                    raise RuntimeError(f"worksheet too large to scan (cells={len(cells)}, max={max_scan_cells})")
                for (row_num, col_num), cell in cells.items():
                    if row_num < min_row or row_num > max_row or col_num < min_col or col_num > max_col:
                        continue
                    try:
                        current = getattr(cell, "value", None)
                    except Exception:
                        continue
                    if current is None:
                        continue
                    try:
                        cell.value = None
                    except Exception:
                        continue
                    cleared_cells += 1
        else:
            if area <= scan_threshold:
                for row_num in range(min_row, max_row + 1):
                    for col_num in range(min_col, max_col + 1):
                        try:
                            cell = ws.cell(row=row_num, column=col_num)
                        except Exception:
                            continue
                        try:
                            current = getattr(cell, "value", None)
                        except Exception:
                            continue
                        if current is None:
                            continue
                        try:
                            cell.value = None
                        except Exception:
                            continue
                        cleared_cells += 1
            else:
                raise RuntimeError("worksheet cell storage is unsupported")

        for row_num, col_num in anchors_to_clear:
            side_effect = row_num < min_row or row_num > max_row or col_num < min_col or col_num > max_col
            cell = None
            if isinstance(cells, dict):
                cell = cells.get((row_num, col_num))
            else:
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                except Exception:
                    continue
            if cell is None:
                continue
            try:
                current = getattr(cell, "value", None)
            except Exception:
                continue
            if current is None:
                continue
            try:
                cell.value = None
            except Exception:
                continue
            cleared_cells += 1
            if side_effect:
                cleared_merged_anchor_coords.add((row_num, col_num))

        if total_rows <= 1 and total_cols <= 1:
            coords = [(min_row, min_col)]
        elif total_rows == 1:
            sample_cols = [min_col + idx for idx in _sample_indices(total_cols, head=3, tail=3)]
            coords = [(min_row, col_num) for col_num in sample_cols]
        elif total_cols == 1:
            sample_rows = [min_row + idx for idx in _sample_indices(total_rows, head=3, tail=3)]
            coords = [(row_num, min_col) for row_num in sample_rows]
        else:
            sample_rows = [min_row + idx for idx in _sample_indices(total_rows, head=2, tail=2)]
            sample_cols = [min_col + idx for idx in _sample_indices(total_cols, head=2, tail=2)]
            coords = [(row_num, col_num) for row_num in sample_rows for col_num in sample_cols]

        if cleared_merged_anchor_coords:
            for anchor_coord in sorted(cleared_merged_anchor_coords):
                if anchor_coord not in coords:
                    coords.append(anchor_coord)

        samples: list[dict[str, Any]] = []
        for row_num, col_num in coords:
            cell = None
            if isinstance(cells, dict):
                cell = cells.get((row_num, col_num))
            elif hasattr(cells, "get"):
                try:
                    cell = cells.get((row_num, col_num))
                except Exception:
                    cell = None
            raw = None
            data_type = None
            if cell is not None:
                try:
                    raw = getattr(cell, "value", None)
                except Exception:
                    raw = None
                try:
                    data_type = getattr(cell, "data_type", None)
                except Exception:
                    data_type = None
            formula = None
            if data_type == "f" and isinstance(raw, str) and raw:
                formula = raw if raw.startswith("=") else f"={raw}"
            if isinstance(formula, str):
                formula = _truncate_str(formula, _get_max_string_chars())

            samples.append(
                {
                    "address": _format_a1_address(sheet_name=resolved_sheet_name, row=row_num, col=col_num),
                    "formula": formula,
                    "value": _to_jsonable_excel_value(raw if formula is None else None),
                }
            )

        cleared_merged_anchors = [
            _format_a1_address(sheet_name=resolved_sheet_name, row=row_num, col=col_num)
            for row_num, col_num in sorted(cleared_merged_anchor_coords)
        ]

        if cleared_cells == 0:
            return resolved_sheet_name, f"{sheet_ref}!{normalized_range}", cleared_cells, samples, cleared_merged_anchors

        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{file_path.name}.",
            suffix=file_path.suffix,
            dir=str(file_path.parent),
        )
        tmp_path = Path(tmp_name)
        try:
            with os.fdopen(fd, "w+b") as f:
                fd = -1
                wb.save(f)
        finally:
            if fd != -1:
                try:
                    os.close(fd)
                except OSError:
                    pass
        try:
            os.chmod(tmp_path, orig_mode)
        except OSError:
            pass
        os.replace(tmp_path, file_path)
        tmp_path = None
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        try:
            lock_file.close()
        except Exception:
            pass

    return resolved_sheet_name, f"{sheet_ref}!{normalized_range}", cleared_cells, samples, cleared_merged_anchors


class ClearRangeTool(BaseTool):
    def __init__(self, config: dict, tool_schema: OpenAIFunctionToolSchema):
        super().__init__(config, tool_schema)
        self._instance_dict: dict[str, dict[str, Any]] = {}
        max_file_size_mb_raw = config.get("max_file_size_mb", 100)
        if max_file_size_mb_raw is None:
            self.max_file_size_bytes = None
        else:
            try:
                max_file_size_mb = int(max_file_size_mb_raw)
            except (TypeError, ValueError):
                max_file_size_mb = 100
            self.max_file_size_bytes = max_file_size_mb * 1024 * 1024 if max_file_size_mb > 0 else None

        lock_timeout_s_raw = config.get("lock_timeout_s", 30)
        try:
            lock_timeout_s = float(lock_timeout_s_raw)
        except (TypeError, ValueError):
            lock_timeout_s = 30.0
        self.lock_timeout_s = max(0.0, lock_timeout_s)
        self.max_response_chars = _get_max_response_chars(config)

        max_iter_cells_raw = config.get("max_iter_cells", os.environ.get("SPREADSHEET_RL_CLEAR_RANGE_MAX_ITER_CELLS", "50000"))
        try:
            self.max_iter_cells = max(1, int(max_iter_cells_raw))
        except (TypeError, ValueError):
            self.max_iter_cells = 50_000

        max_scan_cells_raw = config.get("max_scan_cells", os.environ.get("SPREADSHEET_RL_CLEAR_RANGE_MAX_SCAN_CELLS", "500000"))
        try:
            self.max_scan_cells = max(0, int(max_scan_cells_raw))
        except (TypeError, ValueError):
            self.max_scan_cells = 500_000

    async def create(self, instance_id: Optional[str] = None, **kwargs) -> tuple[str, ToolResponse]:
        if instance_id is None:
            instance_id = uuid.uuid4().hex
        self._instance_dict[instance_id] = {}
        return instance_id, ToolResponse()

    @rollout_trace_op
    async def execute(self, instance_id: str, parameters: dict[str, Any], **kwargs) -> tuple[ToolResponse, float, dict]:
        raw_path = parameters.get("path")
        if raw_path is None or (isinstance(raw_path, str) and not raw_path.strip()):
            relpath = Path("data.xlsx")
        else:
            relpath = _sanitize_relpath(raw_path)
            if relpath is None:
                return ToolResponse(text="Error: invalid path."), 0.0, {"status": "error", "error": "invalid_path"}
        if relpath.suffix.lower() != ".xlsx":
            return ToolResponse(text="Error: only .xlsx workbooks are supported."), 0.0, {
                "status": "error",
                "error": "invalid_path",
            }

        range_raw = parameters.get("range")
        if not isinstance(range_raw, str) or not range_raw.strip():
            return ToolResponse(text="Error: range must be a non-empty A1 range (e.g. A1:C3 or Sheet1!B2:D10)."), 0.0, {
                "status": "error",
                "error": "invalid_range",
            }
        range_token = range_raw.strip()
        sheet_name_raw = parameters.get("sheet_name")
        sheet_name = None
        if sheet_name_raw is not None:
            if not isinstance(sheet_name_raw, str) or not sheet_name_raw.strip():
                return ToolResponse(text="Error: sheet_name must be a non-empty string."), 0.0, {
                    "status": "error",
                    "error": "invalid_sheet_name",
                }
            sheet_name = _normalize_sheet_name(sheet_name_raw)
            if not sheet_name:
                return ToolResponse(text="Error: sheet_name is empty after normalization."), 0.0, {
                    "status": "error",
                    "error": "invalid_sheet_name",
                }

        workspace_id = normalize_workspace_id(kwargs.get("workspace_id"))
        if workspace_id is None:
            return ToolResponse(text="Error: workspace_id is missing/invalid."), 0.0, {
                "status": "error",
                "error": "missing_workspace_id",
            }
        file_path = _resolve_workspace_file(workspace_id=workspace_id, relpath=relpath)
        if file_path is None:
            return ToolResponse(text=f"Error: file not found: {relpath}"), 0.0, {
                "status": "error",
                "error": "file_not_found",
            }
        if self.max_file_size_bytes is not None:
            try:
                file_size = file_path.stat().st_size
            except OSError as exc:
                return ToolResponse(text=f"Error: failed to stat file: {exc}"), 0.0, {
                    "status": "error",
                    "error": "stat_failed",
                }
            if file_size > self.max_file_size_bytes:
                max_mb = self.max_file_size_bytes // (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                return ToolResponse(text=f"Error: workbook is too large ({actual_mb:.1f}MB > {max_mb}MB)."), 0.0, {
                    "status": "error",
                    "error": "file_too_large",
                }

        try:
            resolved_sheet_name, resolved_range, cleared_cells, samples, cleared_merged_anchors = await asyncio.to_thread(
                _clear_range_in_workbook,
                file_path=file_path,
                sheet_name=sheet_name,
                range_token=range_token,
                lock_timeout_s=self.lock_timeout_s,
                max_iter_cells=self.max_iter_cells,
                max_scan_cells=self.max_scan_cells,
            )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            return ToolResponse(text=f"Error: failed to clear range: {exc}"), 0.0, {
                "status": "error",
                "error": "clear_failed",
            }

        payload: dict[str, Any] = {
            "status": "success",
            "file": str(relpath),
            "sheet": resolved_sheet_name,
            "range": resolved_range,
            "cleared_cells": cleared_cells,
            "samples": samples,
        }
        if cleared_merged_anchors:
            payload["cleared_merged_anchors"] = cleared_merged_anchors
        response_text = _truncate_payload_to_max_chars(payload, self.max_response_chars)
        samples_final = payload.get("samples")
        sample_count = len(samples_final) if isinstance(samples_final, list) else 0
        metrics = {
            "status": "success",
            "file": str(relpath),
            "cleared_cells": cleared_cells,
            "sample_cells": sample_count,
        }
        return ToolResponse(text=response_text), 0.0, metrics

    async def calc_reward(self, instance_id: str, **kwargs) -> float:
        return 0.0

    async def release(self, instance_id: str, **kwargs) -> None:
        self._instance_dict.pop(instance_id, None)
