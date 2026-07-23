from __future__ import annotations

import datetime
import threading
from pathlib import Path

import openpyxl
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from _tempdir import temporary_directory

from async_reward_api import eval as eval_mod
from async_reward_api.eval import (
    AnswerPositionError,
    _estimate_answer_position_cells,
    _get_gt_prepared_max_cells,
    compare_cell_value,
    compare_workbooks,
    parse_answer_position,
)


def _assert_equal(label: str, got: object, expected: object) -> None:
    if got != expected:
        raise AssertionError(f"{label}:\n  got={got!r}\n  expected={expected!r}")


def _assert_not_contains(label: str, text: str, unexpected: str) -> None:
    if unexpected in text:
        raise AssertionError(f"{label}: {text!r} contains {unexpected!r}")


def _expect_error(label: str, fn, /, *args, **kwargs) -> None:
    try:
        fn(*args, **kwargs)
    except AnswerPositionError:
        return
    raise AssertionError(f"Expected AnswerPositionError for {label!r}")


def _write_column_range_workbook(
    path: Path,
    *,
    include_extra_proc_row: bool,
    sheet_title: str = "Sheet1",
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws["A1"] = 1
    if include_extra_proc_row:
        ws["A2"] = 999
    wb.save(path)
    wb.close()


def _write_single_cell_workbook(path: Path, value: object, *, sheet_title: str = "Sheet1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws["A1"] = value
    wb.save(path)
    wb.close()


def _write_time_workbook(path: Path, value: datetime.time | datetime.datetime) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = value
    wb.save(path)
    wb.close()


def _compare_with_scan_count(
    gt_file: Path,
    proc_file: Path,
    answer_position: str,
) -> tuple[tuple[bool, str], list[tuple[str, int, int, int, int]]]:
    calls: list[tuple[str, int, int, int, int]] = []
    original_iter_rows = ReadOnlyWorksheet.iter_rows

    def counting_iter_rows(self, *args, **kwargs):
        calls.append(
            (
                self.title,
                int(kwargs["min_row"]),
                int(kwargs["max_row"]),
                int(kwargs["min_col"]),
                int(kwargs["max_col"]),
            )
        )
        return original_iter_rows(self, *args, **kwargs)

    ReadOnlyWorksheet.iter_rows = counting_iter_rows
    try:
        result = compare_workbooks(gt_file, proc_file, answer_position)
    finally:
        ReadOnlyWorksheet.iter_rows = original_iter_rows
    return result, calls


def main() -> int:
    default_sheet = "__DEFAULT__"
    cases: list[tuple[str, list[tuple[str, str]]]] = [
        (
            # Intentionally includes a dangling apostrophe in the sheet name.
            "Pendency'!D7:D11,Pendency'!G7:G11",
            [("Pendency", "D7:D11"), ("Pendency", "G7:G11")],
        ),
        (
            "Sheet1'!C21:F30\"",
            [("Sheet1", "C21:F30")],
        ),
        (
            "'STATS 2025'!C6:C7,'STATS 2025'!E6:E7,'STATS 2025'!G6:G7,'STATS 2025'!I6:I7",
            [
                ("STATS 2025", "C6:C7"),
                ("STATS 2025", "E6:E7"),
                ("STATS 2025", "G6:G7"),
                ("STATS 2025", "I6:I7"),
            ],
        ),
        (
            "'Sheet'!B17:D17,'Sheet'!B30:D30,'Sheet'!B43:D43,'Sheet'!B56:D56",
            [
                ("Sheet", "B17:D17"),
                ("Sheet", "B30:D30"),
                ("Sheet", "B43:D43"),
                ("Sheet", "B56:D56"),
            ],
        ),
        (
            "'My, Sheet'!A1:B2,'Other'!C3:D4",
            [("My, Sheet", "A1:B2"), ("Other", "C3:D4")],
        ),
        (
            "  Sheet1 !  A1:B2  ,  Sheet2  ! C3 : D4  ",
            [("Sheet1", "A1:B2"), ("Sheet2", "C3:D4")],
        ),
        (
            "'My, Sheet''s Data'!A1:B2",
            [("My, Sheet's Data", "A1:B2")],
        ),
        (
            "'Test''s'!A1",
            [("Test's", "A1")],
        ),
        (
            "A1:A1",
            [(default_sheet, "A1:A1")],
        ),
        (
            # Full-width colon should be normalized.
            "G12：J15",
            [(default_sheet, "G12:J15")],
        ),
        (
            # Actual Unicode full-width colon should also be normalized.
            "G12\uff1aJ15",
            [(default_sheet, "G12:J15")],
        ),
        (
            # Malformed quote placement around sheet tokens.
            "'RAWDATA!'A1:P6,'OUTPUT!'A1:P6'",
            [("RAWDATA", "A1:P6"), ("OUTPUT", "A1:P6")],
        ),
        (
            # Malformed duplicate-sheet fragment in range part.
            "'Received'!'Received!A1:G16'",
            [("Received", "A1:G16")],
        ),
        (
            # Column-only range should be accepted by parser.
            "Sheet3'!A:G,'Sheet4'!A:G",
            [("Sheet3", "A:G"), ("Sheet4", "A:G")],
        ),
        (
            "XFD1048576",
            [(default_sheet, "XFD1048576")],
        ),
        (
            "'Sheet'!XFD:XFD",
            [("Sheet", "XFD:XFD")],
        ),
        (
            "' Sheet '!A1",
            [(" Sheet ", "A1")],
        ),
        (
            # Malformed shorthand end-row form.
            "'Sheet1'!BD2:308",
            [("Sheet1", "BD2:BD308")],
        ),
        (
            # Curly quote variant should be normalized.
            "'Invoice List‘!A1:G10,'Invoice Items'!A1:G10",
            [("Invoice List", "A1:G10"), ("Invoice Items", "A1:G10")],
        ),
        (
            # Actual Unicode curly quotes should also be normalized.
            "'Invoice List\u2018!A1:G10,\u2018Invoice Items\u2019!A1:G10",
            [("Invoice List", "A1:G10"), ("Invoice Items", "A1:G10")],
        ),
    ]

    for raw, expected in cases:
        got = parse_answer_position(raw, default_sheet_name=default_sheet)
        _assert_equal(raw, got, expected)

    _expect_error(
        "Sheet1!A1:B2:C3",
        parse_answer_position,
        "Sheet1!A1:B2:C3",
        default_sheet_name=default_sheet,
    )

    for raw in [
        "''!A1",
        "Sheet1!$A$1:$B$2",
        "  $a$1 : $b$2  ",
        "Sheet1!A1!B2",
        "'My Sheet',Sheet2!A1",
        "XFE1",
        "A1048577",
        "A0",
        "XFE:XFE",
        "ZZZ9999999",
    ]:
        _expect_error(raw, parse_answer_position, raw, default_sheet_name=default_sheet)

    _assert_equal(
        "column-only range should bypass prepared GT cache",
        _estimate_answer_position_cells("Sheet1!A:G"),
        _get_gt_prepared_max_cells() + 1,
    )
    _assert_equal("boolean TRUE must not equal numeric 1", compare_cell_value(True, 1), False)
    _assert_equal("boolean FALSE must not equal numeric 0", compare_cell_value(False, 0), False)
    _assert_equal("matching booleans should still compare equal", compare_cell_value(True, True), True)
    _assert_equal("literal nan text should compare equal", compare_cell_value("nan", "nan"), True)
    _assert_equal("literal nan text should remain case-sensitive", compare_cell_value("nan", "NaN"), False)
    _assert_equal("literal inf variants should not compare equal", compare_cell_value("inf", "Infinity"), False)
    _assert_equal("overflowing numeric text should stay text", compare_cell_value("1e309", "inf"), False)
    _assert_equal(
        "time seconds must affect comparison",
        compare_cell_value(datetime.time(12, 34, 0), datetime.time(12, 34, 59)),
        False,
    )
    _assert_equal(
        "datetime time-of-day must affect comparison",
        compare_cell_value(
            datetime.datetime(2024, 1, 1, 0, 0, 0),
            datetime.datetime(2024, 1, 1, 10, 0, 0),
        ),
        False,
    )

    with temporary_directory(prefix="async_reward_api_answer_position_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_column_range_workbook(gt_file, include_extra_proc_row=False, sheet_title="sheet1")
        _write_column_range_workbook(proc_file, include_extra_proc_row=True, sheet_title="SHEET1")
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!A:A")
        if not ok:
            raise AssertionError(f"column-only streaming changed GT-row-bound semantics: {msg}")

    with temporary_directory(prefix="async_reward_api_sheet_case_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_single_cell_workbook(gt_file, 7, sheet_title="sheet1")
        _write_single_cell_workbook(proc_file, 7, sheet_title="SHEET1")
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!A1")
        if not ok:
            raise AssertionError(f"case-insensitive worksheet lookup failed: {msg}")

    with temporary_directory(prefix="async_reward_api_sheet_spaces_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_single_cell_workbook(gt_file, 7, sheet_title=" Sheet ")
        _write_single_cell_workbook(proc_file, 7, sheet_title=" Sheet ")
        ok, msg = compare_workbooks(gt_file, proc_file, "' Sheet '!A1")
        if not ok:
            raise AssertionError(f"quoted sheet name spaces were not preserved: {msg}")

    with temporary_directory(prefix="async_reward_api_non_finite_text_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_single_cell_workbook(gt_file, "nan")
        _write_single_cell_workbook(proc_file, "nan")
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!A1")
        if not ok:
            raise AssertionError(f"non-finite text workbook comparison failed: {msg}")

    with temporary_directory(prefix="async_reward_api_gt_cache_inflight_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        _write_single_cell_workbook(gt_file, 1)
        answer_position = "Sheet1!A1"
        original_prepare = eval_mod._prepare_ground_truth
        calls = {"count": 0}
        calls_lock = threading.Lock()
        barrier = threading.Barrier(8)
        errors: list[BaseException] = []

        def slow_prepare(gt_path: Path, answer: str):
            with calls_lock:
                calls["count"] += 1
            threading.Event().wait(0.05)
            return original_prepare(gt_path, answer)

        def load_prepared() -> None:
            try:
                barrier.wait(timeout=5.0)
                entry = eval_mod._get_or_prepare_ground_truth(gt_file, answer_position)
                if entry.error is not None or entry.prepared is None:
                    raise AssertionError(f"unexpected cache entry: {entry}")
            except BaseException as exc:
                errors.append(exc)

        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()
        eval_mod._prepare_ground_truth = slow_prepare
        try:
            threads = [threading.Thread(target=load_prepared) for _ in range(8)]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join(timeout=5.0)
            if errors:
                raise AssertionError(f"concurrent GT cache load failed: {errors!r}")
            _assert_equal("concurrent GT cache miss should prepare once", calls["count"], 1)
        finally:
            eval_mod._prepare_ground_truth = original_prepare
            with eval_mod._GT_CACHE_LOCK:
                eval_mod._GT_CACHE.clear()
                eval_mod._GT_INFLIGHT.clear()

    cache_key = ("gt.xlsx", "Sheet1!A1", 1, 1)
    fallback_entry = eval_mod._GroundTruthCacheEntry(prepared=None, error="fallback")
    original_gt_wait_s = eval_mod._GT_INFLIGHT_WAIT_S
    original_cache_key = eval_mod._ground_truth_cache_key
    original_prepare = eval_mod._prepare_ground_truth
    prepare_calls = {"count": 0}
    eval_mod._GT_INFLIGHT_WAIT_S = 0.01
    eval_mod._ground_truth_cache_key = lambda gt_path, answer: cache_key

    def fallback_prepare(gt_path: Path, answer: str):
        prepare_calls["count"] += 1
        return fallback_entry

    eval_mod._prepare_ground_truth = fallback_prepare
    with eval_mod._GT_CACHE_LOCK:
        eval_mod._GT_CACHE.clear()
        eval_mod._GT_INFLIGHT.clear()
        eval_mod._GT_INFLIGHT[cache_key] = eval_mod._GroundTruthInflight(event=threading.Event())
    try:
        entry = eval_mod._get_or_prepare_ground_truth(Path("gt.xlsx"), "Sheet1!A1")
        _assert_equal("stale GT inflight should fall back to local prepare", entry, fallback_entry)
        _assert_equal("stale GT inflight fallback prepare count", prepare_calls["count"], 1)
        with eval_mod._GT_CACHE_LOCK:
            _assert_equal(
                "stale GT inflight should be removed after timeout fallback",
                cache_key in eval_mod._GT_INFLIGHT,
                False,
            )
    finally:
        eval_mod._GT_INFLIGHT_WAIT_S = original_gt_wait_s
        eval_mod._ground_truth_cache_key = original_cache_key
        eval_mod._prepare_ground_truth = original_prepare
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()

    replacement_entry = eval_mod._GroundTruthCacheEntry(prepared=None, error="replacement")
    replacement_inflight = eval_mod._GroundTruthInflight(event=threading.Event())
    original_cache_key = eval_mod._ground_truth_cache_key
    original_prepare = eval_mod._prepare_ground_truth
    eval_mod._ground_truth_cache_key = lambda gt_path, answer: cache_key

    def replacing_prepare(gt_path: Path, answer: str):
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_INFLIGHT[cache_key] = replacement_inflight
        return replacement_entry

    eval_mod._prepare_ground_truth = replacing_prepare
    with eval_mod._GT_CACHE_LOCK:
        eval_mod._GT_CACHE.clear()
        eval_mod._GT_INFLIGHT.clear()
    try:
        entry = eval_mod._get_or_prepare_ground_truth(Path("gt.xlsx"), "Sheet1!A1")
        _assert_equal("GT inflight identity success result", entry, replacement_entry)
        with eval_mod._GT_CACHE_LOCK:
            _assert_equal(
                "GT inflight identity success preserved newer inflight",
                eval_mod._GT_INFLIGHT.get(cache_key),
                replacement_inflight,
            )
    finally:
        eval_mod._ground_truth_cache_key = original_cache_key
        eval_mod._prepare_ground_truth = original_prepare
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()

    replacement_error_inflight = eval_mod._GroundTruthInflight(event=threading.Event())
    expected_prepare_error = RuntimeError("prepare failed")
    original_cache_key = eval_mod._ground_truth_cache_key
    original_prepare = eval_mod._prepare_ground_truth
    eval_mod._ground_truth_cache_key = lambda gt_path, answer: cache_key

    def replacing_failing_prepare(gt_path: Path, answer: str):
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_INFLIGHT[cache_key] = replacement_error_inflight
        raise expected_prepare_error

    eval_mod._prepare_ground_truth = replacing_failing_prepare
    with eval_mod._GT_CACHE_LOCK:
        eval_mod._GT_CACHE.clear()
        eval_mod._GT_INFLIGHT.clear()
    try:
        try:
            eval_mod._get_or_prepare_ground_truth(Path("gt.xlsx"), "Sheet1!A1")
        except RuntimeError as exc:
            if exc is not expected_prepare_error:
                raise AssertionError(f"unexpected GT prepare error: {exc!r}") from exc
        else:
            raise AssertionError("GT prepare error did not propagate")
        with eval_mod._GT_CACHE_LOCK:
            _assert_equal(
                "GT inflight identity exception preserved newer inflight",
                eval_mod._GT_INFLIGHT.get(cache_key),
                replacement_error_inflight,
            )
    finally:
        eval_mod._ground_truth_cache_key = original_cache_key
        eval_mod._prepare_ground_truth = original_prepare
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()

    with temporary_directory(prefix="async_reward_api_bool_compare_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_single_cell_workbook(gt_file, True)
        _write_single_cell_workbook(proc_file, 1)
        ok, _ = compare_workbooks(gt_file, proc_file, "Sheet1!A1")
        _assert_equal("workbook boolean TRUE must not equal numeric 1", ok, False)

    with temporary_directory(prefix="async_reward_api_exact_match_fast_path_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append([1.234, "unchanged", True])
        wb.save(gt_file)
        wb.save(proc_file)
        wb.close()
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!A1:C1")
        if not ok:
            raise AssertionError(f"initial exact-value comparison failed: {msg}")

        original_transform = eval_mod.transform_value
        calls = {"count": 0}

        def counting_transform(value):
            calls["count"] += 1
            return original_transform(value)

        eval_mod.transform_value = counting_transform
        try:
            ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!A1:C1")
            if not ok:
                raise AssertionError(f"cached exact-value comparison failed: {msg}")
            _assert_equal("exact primitive matches should skip normalization", calls["count"], 0)
        finally:
            eval_mod.transform_value = original_transform
            with eval_mod._GT_CACHE_LOCK:
                eval_mod._GT_CACHE.clear()
                eval_mod._GT_INFLIGHT.clear()

    with temporary_directory(prefix="async_reward_api_grouped_range_scan_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B4"] = 11
        ws["B6"] = 12
        ws["D10"] = "left"
        ws["E10"] = "right"
        wb.save(gt_file)
        wb.save(proc_file)
        wb.close()
        answer_position = "Sheet1!B4,Sheet1!B6,Sheet1!D10:E10"
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()
        ok, msg = compare_workbooks(gt_file, proc_file, answer_position)
        if not ok:
            raise AssertionError(f"initial grouped-range comparison failed: {msg}")

        (ok, msg), scans = _compare_with_scan_count(gt_file, proc_file, answer_position)
        if not ok:
            raise AssertionError(f"warm grouped-range comparison failed: {msg}")
        _assert_equal(
            "same-sheet ranges should share one bounded worksheet scan",
            scans,
            [("Sheet1", 4, 10, 2, 5)],
        )

        wb = openpyxl.load_workbook(proc_file)
        ws = wb["Sheet1"]
        ws["B4"] = -1
        ws["E10"] = -2
        wb.save(proc_file)
        wb.close()
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!E10,Sheet1!B4")
        _assert_equal("grouped ranges should still fail", ok, False)
        _assert_equal(
            "grouped ranges should preserve answer_position mismatch order",
            msg,
            "Value mismatch at E10: your workbook has -2",
        )
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!E10,'Missing'!A1")
        _assert_equal("earlier mismatch should still fail", ok, False)
        _assert_equal(
            "later missing sheet should not replace an earlier mismatch",
            msg,
            "Value mismatch at E10: your workbook has -2",
        )
        ok, msg = compare_workbooks(gt_file, proc_file, "'Missing'!A1,Sheet1!E10")
        _assert_equal("earlier missing sheet should still fail", ok, False)
        _assert_equal(
            "earlier missing sheet should precede a later mismatch",
            msg,
            "Worksheet was not found in either workbook.",
        )
        ok, msg = compare_workbooks(
            gt_file,
            proc_file,
            "Sheet1!B6,'Missing'!A1,Sheet1!E10",
        )
        _assert_equal("intervening missing sheet should still fail", ok, False)
        _assert_equal(
            "intervening missing sheet should precede a later grouped mismatch",
            msg,
            "Worksheet was not found in either workbook.",
        )
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()

    with temporary_directory(prefix="async_reward_api_sparse_grouped_range_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 1
        ws["XFD1"] = 2
        wb.save(gt_file)
        wb.save(proc_file)
        wb.close()
        answer_position = "Sheet1!A1,Sheet1!XFD1"
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()
        ok, msg = compare_workbooks(gt_file, proc_file, answer_position)
        if not ok:
            raise AssertionError(f"initial sparse-range comparison failed: {msg}")

        (ok, msg), scans = _compare_with_scan_count(gt_file, proc_file, answer_position)
        if not ok:
            raise AssertionError(f"warm sparse-range comparison failed: {msg}")
        _assert_equal(
            "far-apart sparse ranges should not create a huge bounding scan",
            scans,
            [("Sheet1", 1, 1, 1, 1), ("Sheet1", 1, 1, 16384, 16384)],
        )
        with eval_mod._GT_CACHE_LOCK:
            eval_mod._GT_CACHE.clear()
            eval_mod._GT_INFLIGHT.clear()

    with temporary_directory(prefix="async_reward_api_large_prepared_gt_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 0
        ws["A50001"] = 1
        wb.save(gt_file)
        wb.save(proc_file)
        wb.close()
        answer_position = "Sheet1!A1:A50001"
        original_prepared_max_cells = eval_mod._get_gt_prepared_max_cells
        original_cache_size_getter = eval_mod._GT_LRU_CACHE._max_size_getter
        original_cache_weight_getter = eval_mod._GT_LRU_CACHE._max_weight_getter
        eval_mod._GT_LRU_CACHE._max_size_getter = lambda: 128
        eval_mod._GT_LRU_CACHE._max_weight_getter = lambda: 500000
        try:
            with eval_mod._GT_CACHE_LOCK:
                eval_mod._GT_CACHE.clear()
                eval_mod._GT_INFLIGHT.clear()
            eval_mod._get_gt_prepared_max_cells = lambda: 500000
            prepared_success = compare_workbooks(gt_file, proc_file, answer_position)
            with eval_mod._GT_CACHE_LOCK:
                entries = list(eval_mod._GT_CACHE.values())
            _assert_equal("large prepared GT cache entry count", len(entries), 1)
            _assert_equal(
                "large prepared GT cache weight",
                entries[0].prepared_cell_count,
                50001,
            )

            eval_mod._get_gt_prepared_max_cells = lambda: 50000
            streaming_success = compare_workbooks(gt_file, proc_file, answer_position)
            _assert_equal(
                ">50k prepared and streaming success semantics",
                prepared_success,
                streaming_success,
            )

            with eval_mod._GT_CACHE_LOCK:
                eval_mod._GT_CACHE.clear()
                eval_mod._GT_INFLIGHT.clear()
            eval_mod._GT_LRU_CACHE._max_weight_getter = lambda: 50000
            eval_mod._get_gt_prepared_max_cells = lambda: 500000
            overweight_success = compare_workbooks(gt_file, proc_file, answer_position)
            _assert_equal(">50k overweight result semantics", overweight_success, prepared_success)
            with eval_mod._GT_CACHE_LOCK:
                _assert_equal("overweight GT entry should not be cached", len(eval_mod._GT_CACHE), 0)

            wb = openpyxl.load_workbook(proc_file)
            wb["Sheet1"]["A50001"] = 2
            wb.save(proc_file)
            wb.close()
            eval_mod._GT_LRU_CACHE._max_weight_getter = lambda: 500000
            eval_mod._get_gt_prepared_max_cells = lambda: 500000
            prepared_mismatch = compare_workbooks(gt_file, proc_file, answer_position)
            eval_mod._get_gt_prepared_max_cells = lambda: 50000
            streaming_mismatch = compare_workbooks(gt_file, proc_file, answer_position)
            _assert_equal(
                ">50k prepared and streaming mismatch semantics",
                prepared_mismatch,
                streaming_mismatch,
            )
            _assert_equal(
                ">50k mismatch coordinate",
                prepared_mismatch,
                (False, "Value mismatch at A50001: your workbook has 2"),
            )
        finally:
            eval_mod._get_gt_prepared_max_cells = original_prepared_max_cells
            eval_mod._GT_LRU_CACHE._max_size_getter = original_cache_size_getter
            eval_mod._GT_LRU_CACHE._max_weight_getter = original_cache_weight_getter
            with eval_mod._GT_CACHE_LOCK:
                eval_mod._GT_CACHE.clear()
                eval_mod._GT_INFLIGHT.clear()

    duplicate_item = eval_mod._PreparedRange(
        sheet_name="Sheet1",
        cell_range="A1",
        min_col=1,
        min_row=1,
        max_col=1,
        max_row=1,
        gt_missing=False,
        expected_rows=((1,),),
    )
    overlapping_item = eval_mod._PreparedRange(
        sheet_name="Sheet1",
        cell_range="A1:B1",
        min_col=1,
        min_row=1,
        max_col=2,
        max_row=1,
        gt_missing=False,
        expected_rows=((1, 2),),
    )
    indexed_ranges = [
        (index, duplicate_item if index % 2 == 0 else overlapping_item)
        for index in range(5_000)
    ]
    original_range_cell_count = eval_mod._range_cell_count
    original_grouped_scan_bounds = eval_mod._grouped_scan_bounds
    planning_calls = {"range_cell_count": 0, "grouped_scan_bounds": 0}

    def counting_range_cell_count(item):
        planning_calls["range_cell_count"] += 1
        return original_range_cell_count(item)

    def counting_grouped_scan_bounds(items):
        planning_calls["grouped_scan_bounds"] += 1
        return original_grouped_scan_bounds(items)

    eval_mod._range_cell_count = counting_range_cell_count
    eval_mod._grouped_scan_bounds = counting_grouped_scan_bounds
    try:
        scans = eval_mod._plan_grouped_scans(indexed_ranges)
        _assert_equal("duplicate/overlap scan group count", len(scans), 1)
        _assert_equal("duplicate/overlap scan range count", len(scans[0]), 5_000)
        _assert_equal(
            "group planning should count each requested range once",
            planning_calls["range_cell_count"],
            5_000,
        )
        _assert_equal(
            "group planning should reuse maintained bounds",
            planning_calls["grouped_scan_bounds"],
            0,
        )
    finally:
        eval_mod._range_cell_count = original_range_cell_count
        eval_mod._grouped_scan_bounds = original_grouped_scan_bounds

    class _EarlyMismatchWorksheet:
        rows_read = 0

        def iter_rows(self, **kwargs):
            self.rows_read += 1
            yield (1,)
            raise AssertionError("singleton prepared comparison read past the first mismatch")

    class _EarlyMismatchWorkbook:
        sheetnames = ["Sheet1"]

        def __init__(self):
            self.worksheet = _EarlyMismatchWorksheet()

        def __getitem__(self, sheet_name):
            _assert_equal("singleton prepared sheet", sheet_name, "Sheet1")
            return self.worksheet

    large_singleton_range = eval_mod._PreparedRange(
        sheet_name="Sheet1",
        cell_range="A1:A500000",
        min_col=1,
        min_row=1,
        max_col=1,
        max_row=500_000,
        gt_missing=False,
        expected_rows=((0,),),
    )
    early_mismatch_workbook = _EarlyMismatchWorkbook()
    early_mismatch = eval_mod._compare_proc_with_prepared(
        eval_mod._PreparedGroundTruth(ranges=(large_singleton_range,)),
        early_mismatch_workbook,
    )
    _assert_equal(
        "singleton prepared range should preserve first-mismatch semantics",
        early_mismatch,
        (False, "Value mismatch at A1: your workbook has 1"),
    )
    _assert_equal(
        "singleton prepared range should stop after first mismatch",
        early_mismatch_workbook.worksheet.rows_read,
        1,
    )

    grouped_early_mismatch_workbook = _EarlyMismatchWorkbook()
    grouped_early_mismatch = eval_mod._compare_proc_with_prepared(
        eval_mod._PreparedGroundTruth(
            ranges=(
                eval_mod._PreparedRange(
                    sheet_name="Sheet1",
                    cell_range="A1",
                    min_col=1,
                    min_row=1,
                    max_col=1,
                    max_row=1,
                    gt_missing=False,
                    expected_rows=((0,),),
                ),
                eval_mod._PreparedRange(
                    sheet_name="Sheet1",
                    cell_range="A2:A100000",
                    min_col=1,
                    min_row=2,
                    max_col=1,
                    max_row=100_000,
                    gt_missing=False,
                    expected_rows=((None,),) * 99_999,
                ),
            )
        ),
        grouped_early_mismatch_workbook,
    )
    _assert_equal(
        "grouped prepared range should preserve first-mismatch semantics",
        grouped_early_mismatch,
        (False, "Value mismatch at A1: your workbook has 1"),
    )
    _assert_equal(
        "grouped prepared range should stop after the first requested-cell mismatch",
        grouped_early_mismatch_workbook.worksheet.rows_read,
        1,
    )

    class _GuardedWorksheet:
        def __init__(self, value, *, max_rows: int | None = None):
            self.value = value
            self.max_rows = max_rows
            self.rows_read = 0

        def iter_rows(self, **kwargs):
            for _ in range(kwargs["min_row"], kwargs["max_row"] + 1):
                self.rows_read += 1
                if self.max_rows is not None and self.rows_read > self.max_rows:
                    raise AssertionError("grouped comparison read across an answer-position gap")
                yield (self.value,)

    class _InterveningMismatchWorkbook:
        sheetnames = ["S1", "S2"]

        def __init__(self):
            self.s1 = _GuardedWorksheet(1, max_rows=1)
            self.s2 = _GuardedWorksheet(1)

        def __getitem__(self, sheet_name):
            return {"S1": self.s1, "S2": self.s2}[sheet_name]

    intervening_mismatch_workbook = _InterveningMismatchWorkbook()
    intervening_mismatch = eval_mod._compare_proc_with_prepared(
        eval_mod._PreparedGroundTruth(
            ranges=(
                eval_mod._PreparedRange("S1", "A1", 1, 1, 1, 1, False, ((1,),)),
                eval_mod._PreparedRange("S2", "A1", 1, 1, 1, 1, False, ((0,),)),
                eval_mod._PreparedRange("S1", "A2:A100000", 1, 2, 1, 100_000, False, ()),
            )
        ),
        intervening_mismatch_workbook,
    )
    _assert_equal(
        "intervening range should preserve answer-position mismatch order",
        intervening_mismatch,
        (False, "Value mismatch at A1: your workbook has 1"),
    )
    _assert_equal(
        "nonconsecutive grouped range should not be read early",
        intervening_mismatch_workbook.s1.rows_read,
        1,
    )

    sparse_ranges: list[tuple[int, eval_mod._PreparedRange]] = []
    for index in range(5_000):
        row = (index // 16) * 1_025 + 1
        col = (index % 16) * 1_025 + 1
        sparse_ranges.append(
            (
                index,
                eval_mod._PreparedRange(
                    sheet_name="Sheet1",
                    cell_range=f"{eval_mod.col_num2name(col)}{row}",
                    min_col=col,
                    min_row=row,
                    max_col=col,
                    max_row=row,
                    gt_missing=False,
                    expected_rows=((1,),),
                ),
            )
        )
    original_can_merge_grouped_scan = eval_mod._can_merge_grouped_scan
    merge_checks = {"count": 0}

    def counting_can_merge_grouped_scan(*, bounding_cells, requested_cells):
        merge_checks["count"] += 1
        return original_can_merge_grouped_scan(
            bounding_cells=bounding_cells,
            requested_cells=requested_cells,
        )

    eval_mod._can_merge_grouped_scan = counting_can_merge_grouped_scan
    try:
        sparse_scans = eval_mod._plan_grouped_scans(sparse_ranges)
    finally:
        eval_mod._can_merge_grouped_scan = original_can_merge_grouped_scan
    _assert_equal("pairwise sparse scan group count", len(sparse_scans), 5_000)
    _assert_equal(
        "sequential sparse scan merge checks",
        merge_checks["count"],
        len(sparse_ranges) - 1,
    )

    with temporary_directory(prefix="async_reward_api_time_compare_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_time_workbook(gt_file, datetime.time(12, 34, 0))
        _write_time_workbook(proc_file, datetime.time(12, 34, 59))
        ok, _ = compare_workbooks(gt_file, proc_file, "Sheet1!A1")
        _assert_equal("workbook time seconds must not compare equal", ok, False)

        _write_time_workbook(gt_file, datetime.datetime(2024, 1, 1, 0, 0, 0))
        _write_time_workbook(proc_file, datetime.datetime(2024, 1, 1, 10, 0, 0))
        ok, _ = compare_workbooks(gt_file, proc_file, "Sheet1!A1")
        _assert_equal("workbook datetime time-of-day must not compare equal", ok, False)

        aware_datetime = datetime.datetime(2024, 1, 1, 10, 30, 0, tzinfo=datetime.timezone.utc)
        _assert_equal(
            "timezone-aware datetime should compare as local Excel serial value",
            eval_mod.datetime_to_float(aware_datetime),
            eval_mod.datetime_to_float(aware_datetime.replace(tzinfo=None)),
        )

    with temporary_directory(prefix="async_reward_api_public_errors_") as tmp:
        tmp_path = Path(tmp)
        missing_gt = tmp_path / "private" / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_single_cell_workbook(proc_file, 1)
        ok, msg = compare_workbooks(missing_gt, proc_file, "Sheet1!A1")
        _assert_equal("missing GT should fail", ok, False)
        _assert_not_contains("missing GT message should not expose path", msg, str(missing_gt))

        gt_file = tmp_path / "gt.xlsx"
        bad_proc_file = tmp_path / "bad_proc.xlsx"
        _write_single_cell_workbook(gt_file, 1)
        bad_proc_file.write_text("not an xlsx file", encoding="utf-8")
        ok, msg = compare_workbooks(gt_file, bad_proc_file, "Sheet1!A1")
        _assert_equal("invalid proc workbook should fail", ok, False)
        _assert_not_contains("invalid proc message should not expose path", msg, str(bad_proc_file))

    with temporary_directory(prefix="async_reward_api_unicode_positions_") as tmp:
        tmp_path = Path(tmp)
        gt_file = tmp_path / "gt.xlsx"
        proc_file = tmp_path / "proc.xlsx"
        _write_single_cell_workbook(gt_file, 1)
        _write_single_cell_workbook(proc_file, 1)
        ok, msg = compare_workbooks(gt_file, proc_file, "Sheet1!A1\uff1aA1")
        if not ok:
            raise AssertionError(f"Unicode full-width colon comparison failed: {msg}")

    print("OK: answer_position parsing looks good")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
