from __future__ import annotations

import datetime
import logging
import math
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from .config import _get_gt_cache_max_cells, _get_gt_cache_size, _get_gt_prepared_max_cells
from .inflight_cache import Inflight
from .inflight_cache import InflightLruCache


logger = logging.getLogger(__name__)
_GroundTruthInflight = Inflight  # compat alias for tests

def datetime_to_float(dt: datetime.datetime | datetime.date) -> float:
    if isinstance(dt, datetime.datetime):
        dt_value = dt
    else:
        dt_value = datetime.datetime.combine(dt, datetime.time())
    if dt_value.tzinfo is not None:
        dt_value = dt_value.replace(tzinfo=None)
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt_value - excel_start_date
    return round(delta.total_seconds() / 86400.0, 8)


def transform_value(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = (v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1_000_000) / 86400.0
        v = round(v, 8)
    elif isinstance(v, datetime.datetime):
        v = datetime_to_float(v)
    elif isinstance(v, datetime.date):
        v = datetime_to_float(v)
    elif isinstance(v, str):
        try:
            parsed = float(v)
        except ValueError:
            pass
        else:
            if math.isfinite(parsed):
                v = round(parsed, 2)
    return v


_PRIMITIVE_EXACT_MATCH_TYPES = {bool, float, int, str}


def compare_cell_value(v1, v2) -> bool:
    if type(v1) is type(v2) and type(v1) in _PRIMITIVE_EXACT_MATCH_TYPES and v1 == v2:
        return True
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2


def col_num2name(n: int) -> str:
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name: str) -> int:
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


_EXCEL_MAX_COLUMN = 16384
_EXCEL_MAX_ROW = 1048576
_A1_CELL_RE = r"[A-Z]{1,3}[0-9]{1,7}"
_A1_RANGE_RE = re.compile(rf"^{_A1_CELL_RE}(?::{_A1_CELL_RE})?$")
_A1_CELL_COORD_RE = re.compile(r"^([A-Z]{1,3})([0-9]{1,7})$")
_A1_COLUMN_RANGE_RE = re.compile(r"^[A-Z]{1,3}:[A-Z]{1,3}$")
_A1_RANGE_START_RE = re.compile(rf"^{_A1_CELL_RE}")
_GROUPED_SCAN_MAX_CELLS = 100_000
_GROUPED_SCAN_MIN_CELLS = 1_024
_GROUPED_SCAN_MAX_EXPANSION = 4


class AnswerPositionError(ValueError):
    pass


@dataclass(frozen=True)
class _PreparedRange:
    sheet_name: str
    cell_range: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int
    gt_missing: bool
    expected_rows: tuple[tuple[object | None, ...], ...]


@dataclass(frozen=True)
class _PreparedGroundTruth:
    ranges: tuple[_PreparedRange, ...]


@dataclass(frozen=True)
class _GroundTruthCacheEntry:
    prepared: _PreparedGroundTruth | None
    error: str | None
    prepared_cell_count: int = 0


_GT_INFLIGHT_WAIT_S = 30.0
_GT_LRU_CACHE: InflightLruCache[tuple[str, str, int, int], _GroundTruthCacheEntry] = InflightLruCache(
    max_size_getter=_get_gt_cache_size,
    inflight_wait_s_getter=lambda: _GT_INFLIGHT_WAIT_S,
    max_weight_getter=_get_gt_cache_max_cells,
    weight_getter=lambda entry: entry.prepared_cell_count,
)
_GT_CACHE_LOCK = _GT_LRU_CACHE.lock
_GT_CACHE = _GT_LRU_CACHE.cache
_GT_INFLIGHT = _GT_LRU_CACHE.inflight


def _estimate_answer_position_cells(answer_position: str) -> int | None:
    try:
        parsed_ranges = parse_answer_position(answer_position, default_sheet_name="Sheet1")
    except AnswerPositionError:
        return None

    total_cells = 0
    for _, cell_range in parsed_ranges:
        if _A1_COLUMN_RANGE_RE.fullmatch(cell_range):
            return _get_gt_prepared_max_cells() + 1
        try:
            min_col, min_row, max_col, max_row = _cell_range_boundaries(cell_range)
        except AnswerPositionError:
            return None
        total_cells += (max_col - min_col + 1) * (max_row - min_row + 1)
    return total_cells


def _extract_range_values(
    worksheet,
    *,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
) -> tuple[tuple[object | None, ...], ...]:
    iter_rows = worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )

    expected_cols = max_col - min_col + 1
    empty_row = (None,) * expected_cols
    rows: list[tuple[object | None, ...]] = []
    for _ in range(min_row, max_row + 1):
        try:
            row = next(iter_rows)
        except StopIteration:
            row = empty_row
        rows.append(tuple(row[offset] if offset < len(row) else None for offset in range(expected_cols)))
    return tuple(rows)


def _resolve_sheet_name(workbook, sheet_name: str) -> str | None:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    folded = sheet_name.casefold()
    for candidate in workbook.sheetnames:
        if candidate.casefold() == folded:
            return candidate
    stripped = sheet_name.strip()
    if stripped and stripped != sheet_name:
        if stripped in workbook.sheetnames:
            return stripped
        stripped_folded = stripped.casefold()
        for candidate in workbook.sheetnames:
            if candidate.casefold() == stripped_folded:
                return candidate
    return None


def _ground_truth_cache_key(gt_path: Path, answer_position: str) -> tuple[str, str, int, int]:
    stat = gt_path.stat()
    resolved = gt_path.resolve()
    return str(resolved), answer_position, int(stat.st_mtime_ns), int(stat.st_size)


def _prepare_ground_truth(gt_path: Path, answer_position: str) -> _GroundTruthCacheEntry:
    wb_gt = None
    try:
        try:
            wb_gt = openpyxl.load_workbook(
                filename=str(gt_path),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except (OSError, zipfile.BadZipFile, InvalidFileException) as exc:
            logger.warning(f"[eval] Failed to load ground truth workbook ({gt_path}): {type(exc).__name__}: {exc}")
            return _GroundTruthCacheEntry(
                prepared=None,
                error="Failed to load ground truth workbook",
            )

        if not wb_gt.sheetnames:
            logger.warning(f"[eval] Ground truth workbook has no worksheets: {gt_path}")
            return _GroundTruthCacheEntry(
                prepared=None,
                error="Ground truth workbook has no worksheets",
            )

        try:
            parsed_ranges = parse_answer_position(
                answer_position, default_sheet_name=wb_gt.sheetnames[0]
            )
        except AnswerPositionError as exc:
            return _GroundTruthCacheEntry(prepared=None, error=str(exc))

        prepared_ranges: list[_PreparedRange] = []
        prepared_cell_count = 0
        for sheet_name, cell_range in parsed_ranges:
            gt_sheet_name = _resolve_sheet_name(wb_gt, sheet_name)
            ws_gt = wb_gt[gt_sheet_name] if gt_sheet_name is not None else None
            max_row_hint = max(1, int(ws_gt.max_row or 1)) if ws_gt is not None else 1
            min_col, min_row, max_col, max_row = _cell_range_boundaries(
                cell_range,
                max_row_hint=max_row_hint,
            )
            if gt_sheet_name is None:
                prepared_ranges.append(
                    _PreparedRange(
                        sheet_name=sheet_name,
                        cell_range=cell_range,
                        min_col=min_col,
                        min_row=min_row,
                        max_col=max_col,
                        max_row=max_row,
                        gt_missing=True,
                        expected_rows=(),
                    )
                )
                continue

            expected_rows = _extract_range_values(
                ws_gt,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            prepared_cell_count += (max_col - min_col + 1) * (max_row - min_row + 1)
            prepared_ranges.append(
                _PreparedRange(
                    sheet_name=sheet_name,
                    cell_range=cell_range,
                    min_col=min_col,
                    min_row=min_row,
                    max_col=max_col,
                    max_row=max_row,
                    gt_missing=False,
                    expected_rows=expected_rows,
                )
            )

        return _GroundTruthCacheEntry(
            prepared=_PreparedGroundTruth(ranges=tuple(prepared_ranges)),
            error=None,
            prepared_cell_count=prepared_cell_count,
        )
    finally:
        if wb_gt is not None:
            try:
                wb_gt.close()
            except Exception as exc:  # noqa: BLE001
                logger.warning(f"[eval] wb_gt.close failed: {type(exc).__name__}: {exc}")


def _get_or_prepare_ground_truth(gt_path: Path, answer_position: str) -> _GroundTruthCacheEntry:
    try:
        cache_key = _ground_truth_cache_key(gt_path, answer_position)
    except OSError as exc:
        logger.warning(f"[eval] Failed to stat ground truth workbook ({gt_path}): {type(exc).__name__}: {exc}")
        return _GroundTruthCacheEntry(
            prepared=None,
            error="Failed to stat ground truth workbook",
        )
    return _GT_LRU_CACHE.get_or_compute(
        cache_key,
        lambda: _prepare_ground_truth(gt_path, answer_position),
        should_cache=lambda entry: entry.error is None,
    )


def _split_answer_position(answer_position: str) -> list[str]:
    """
    Split an answer_position string into individual sheet/range tokens.

    We split on commas, but allow commas inside a quoted sheet name, e.g.:
      'My, Sheet'!A1:B2

    We intentionally only treat a leading quote (after optional whitespace) as the
    start of a quoted sheet name. This keeps us resilient to malformed inputs
    like: Sheet1'!A1:B2 (dangling apostrophe).
    """
    tokens: list[str] = []
    buf: list[str] = []
    in_quoted_sheet_name = False
    at_token_start = True

    def _flush_token() -> None:
        token = "".join(buf).strip()
        if token:
            tokens.append(token)

    i = 0
    while i < len(answer_position):
        ch = answer_position[i]

        # Allow benign wrappers (whitespace, double-quotes) before the token proper.
        # This helps with inputs like:  "'My, Sheet'!A1:B2,'Other'!C3:D4"
        if at_token_start and (ch.isspace() or ch == '"'):
            buf.append(ch)
            i += 1
            continue

        if at_token_start and ch == "'":
            in_quoted_sheet_name = True
            at_token_start = False
            buf.append(ch)
            i += 1
            continue

        at_token_start = False

        if in_quoted_sheet_name and ch == "'":
            # Excel escapes an apostrophe inside a quoted sheet name as ''.
            if i + 1 < len(answer_position) and answer_position[i + 1] == "'":
                buf.append("'")
                buf.append("'")
                i += 2
                continue

            # End quoted sheet name if this quote is followed by:
            #   - a sheet separator (!), or
            #   - a token separator (,), or
            #   - a direct range start (malformed form like: 'Sheet!'A1:B2), or
            #   - the end of the string
            lookahead = answer_position[i + 1 :].lstrip()
            if (
                not lookahead
                or lookahead.startswith(("!", ","))
                or _A1_RANGE_START_RE.match(lookahead) is not None
            ):
                in_quoted_sheet_name = False

            buf.append(ch)
            i += 1
            continue

        if ch == "," and not in_quoted_sheet_name:
            _flush_token()
            buf = []
            at_token_start = True
            i += 1
            continue

        buf.append(ch)
        i += 1

    _flush_token()
    return tokens


def _normalize_sheet_name(sheet_name: str) -> str:
    # Drop obvious wrapper quotes first (common when passing through JSON/CLI).
    name = sheet_name.strip(" \t\n\r\f\v\"")

    if len(name) >= 2 and name[0] == "'" and name[-1] == "'":
        # Unwrap and unescape Excel-style '' -> '
        name = name[1:-1].replace("''", "'")
        return name

    # Handle malformed/dangling apostrophes like: Sheet1'!A1
    if name.startswith("'") and not name.endswith("'"):
        name = name[1:]
    elif name.endswith("'") and not name.startswith("'"):
        name = name[:-1]
    return name.strip()


def _normalize_cell_range(cell_range: str) -> str:
    # Drop obvious wrapper quotes first (common when passing through JSON/CLI).
    rng = cell_range.strip(" \t\n\r\f\v\"'")
    rng = rng.replace(" ", "")
    rng = rng.replace("\uff1a", ":")
    rng = rng.replace("：", ":")
    rng = rng.upper()

    # Recover malformed shorthand like "BD2:308" -> "BD2:BD308".
    match = re.fullmatch(r"^([A-Z]{1,3})([0-9]{1,7}):([0-9]{1,7})$", rng)
    if match is not None:
        col = match.group(1)
        start_row = match.group(2)
        end_row = match.group(3)
        return f"{col}{start_row}:{col}{end_row}"

    return rng


def _normalize_answer_position_text(answer_position: str) -> str:
    text = answer_position
    # Normalize common Unicode punctuation variants seen in dataset sources.
    text = text.replace("\uff1a", ":")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("：", ":")
    text = text.replace("‘", "'").replace("’", "'")
    return text


def _is_supported_cell_range(cell_range: str) -> bool:
    if not (_A1_RANGE_RE.fullmatch(cell_range) or _A1_COLUMN_RANGE_RE.fullmatch(cell_range)):
        return False
    try:
        _cell_range_boundaries(
            cell_range,
            max_row_hint=1 if _A1_COLUMN_RANGE_RE.fullmatch(cell_range) else None,
        )
    except AnswerPositionError:
        return False
    return True


def parse_answer_position(answer_position: str, *, default_sheet_name: str) -> list[tuple[str, str]]:
    """
    Parse answer_position into a list of (sheet_name, cell_range) pairs.

    Supported forms:
      - A1
      - A1:B2
      - Sheet1!A1
      - Sheet1!A1:B2
      - 'Sheet With Spaces'!A1:B2

    Multiple tokens can be separated by commas.
    """
    if not isinstance(answer_position, str) or not answer_position.strip():
        raise AnswerPositionError("answer_position is empty")

    answer_position = _normalize_answer_position_text(answer_position)
    tokens = _split_answer_position(answer_position)

    parsed: list[tuple[str, str]] = []
    for token in tokens:
        token = token.strip(" \t\n\r\f\v\"")
        if not token:
            continue

        if "!" in token:
            if token.count("!") > 1:
                sheet_part, range_part = token.split("!", 1)
                # Recover malformed pattern like "'Received'!'Received!A1:G16'".
                # We only salvage when extra "!" appears inside a quoted range fragment.
                if "!" in range_part and range_part.lstrip().startswith("'"):
                    range_part = range_part.rsplit("!", 1)[-1]
                else:
                    raise AnswerPositionError(
                        f"Malformed sheet reference (multiple '!' characters): {token!r}"
                    )
            else:
                sheet_part, range_part = token.split("!", 1)
            sheet_name = _normalize_sheet_name(sheet_part)
            if not sheet_name:
                raise AnswerPositionError(f"Empty sheet name in token: {token!r}")
        else:
            sheet_name = default_sheet_name
            range_part = token

        cell_range = _normalize_cell_range(range_part)
        if not _is_supported_cell_range(cell_range):
            raise AnswerPositionError(
                f"Invalid cell range: {range_part!r} (normalized: {cell_range!r})"
            )

        parsed.append((sheet_name, cell_range))

    if not parsed:
        raise AnswerPositionError("answer_position is empty")

    return parsed


def _cell_range_boundaries(
    cell_range: str,
    *,
    max_row_hint: int | None = None,
) -> tuple[int, int, int, int]:
    if _A1_COLUMN_RANGE_RE.fullmatch(cell_range):
        start_col_name, end_col_name = cell_range.split(":", 1)
        start_col = col_name2num(start_col_name)
        end_col = col_name2num(end_col_name)
        min_col, max_col = sorted((start_col, end_col))
        if min_col < 1 or max_col > _EXCEL_MAX_COLUMN:
            raise AnswerPositionError(f"Column out of Excel bounds: {cell_range!r}")

        if max_row_hint is None:
            raise AnswerPositionError(
                f"Column-only range requires max_row_hint: {cell_range!r}"
            )
        max_row = max(1, int(max_row_hint))
        return min_col, 1, max_col, max_row

    if ":" in cell_range:
        start_cell, end_cell = cell_range.split(":", 1)
    else:
        start_cell, end_cell = cell_range, cell_range

    match_start = _A1_CELL_COORD_RE.fullmatch(start_cell)
    match_end = _A1_CELL_COORD_RE.fullmatch(end_cell)
    if match_start is None or match_end is None:
        raise AnswerPositionError(f"Invalid cell range: {cell_range!r}")

    start_col = col_name2num(match_start.group(1))
    start_row = int(match_start.group(2))
    end_col = col_name2num(match_end.group(1))
    end_row = int(match_end.group(2))

    min_col, max_col = sorted((start_col, end_col))
    min_row, max_row = sorted((start_row, end_row))
    if min_col < 1 or max_col > _EXCEL_MAX_COLUMN:
        raise AnswerPositionError(f"Column out of Excel bounds: {cell_range!r}")
    if min_row < 1 or max_row > _EXCEL_MAX_ROW:
        raise AnswerPositionError(f"Row out of Excel bounds: {cell_range!r}")
    return min_col, min_row, max_col, max_row


def _range_cell_count(item: _PreparedRange) -> int:
    return (item.max_col - item.min_col + 1) * (item.max_row - item.min_row + 1)


def _grouped_scan_bounds(
    indexed_ranges: list[tuple[int, _PreparedRange]],
) -> tuple[int, int, int, int, int]:
    min_col = min(item.min_col for _, item in indexed_ranges)
    min_row = min(item.min_row for _, item in indexed_ranges)
    max_col = max(item.max_col for _, item in indexed_ranges)
    max_row = max(item.max_row for _, item in indexed_ranges)
    bounding_cells = (max_col - min_col + 1) * (max_row - min_row + 1)
    return min_col, min_row, max_col, max_row, bounding_cells


def _can_merge_grouped_scan(*, bounding_cells: int, requested_cells: int) -> bool:
    return bounding_cells <= _GROUPED_SCAN_MAX_CELLS and bounding_cells <= max(
        _GROUPED_SCAN_MIN_CELLS,
        requested_cells * _GROUPED_SCAN_MAX_EXPANSION,
    )


@dataclass
class _GroupedScanState:
    indexed_ranges: list[tuple[int, _PreparedRange]]
    min_col: int
    min_row: int
    max_col: int
    max_row: int
    requested_cells: int


def _plan_grouped_scans(
    indexed_ranges: list[tuple[int, _PreparedRange]],
) -> list[list[tuple[int, _PreparedRange]]]:
    groups: list[_GroupedScanState] = []
    current_group: _GroupedScanState | None = None
    previous_index: int | None = None
    for indexed_range in indexed_ranges:
        index, item = indexed_range
        item_cells = _range_cell_count(item)
        if current_group is not None and previous_index is not None and index == previous_index + 1:
            min_col = min(current_group.min_col, item.min_col)
            min_row = min(current_group.min_row, item.min_row)
            max_col = max(current_group.max_col, item.max_col)
            max_row = max(current_group.max_row, item.max_row)
            bounding_cells = (max_col - min_col + 1) * (max_row - min_row + 1)
            requested_cells = current_group.requested_cells + item_cells
            if _can_merge_grouped_scan(
                bounding_cells=bounding_cells,
                requested_cells=requested_cells,
            ):
                current_group.indexed_ranges.append(indexed_range)
                current_group.min_col = min_col
                current_group.min_row = min_row
                current_group.max_col = max_col
                current_group.max_row = max_row
                current_group.requested_cells = requested_cells
                previous_index = index
                continue

        current_group = _GroupedScanState(
            indexed_ranges=[indexed_range],
            min_col=item.min_col,
            min_row=item.min_row,
            max_col=item.max_col,
            max_row=item.max_row,
            requested_cells=item_cells,
        )
        groups.append(current_group)
        previous_index = index
    return [group.indexed_ranges for group in groups]


def _compare_grouped_range_values(
    worksheet,
    indexed_ranges: list[tuple[int, _PreparedRange]],
) -> tuple[str | None, dict[int, str | None]]:
    min_col, min_row, max_col, max_row, _ = _grouped_scan_bounds(indexed_ranges)
    iter_rows = worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    starts: dict[int, list[tuple[int, _PreparedRange]]] = {}
    stops: dict[int, list[int]] = {}
    mismatches: dict[int, str] = {}
    for index, item in indexed_ranges:
        starts.setdefault(item.min_row, []).append((index, item))
        stops.setdefault(item.max_row + 1, []).append(index)

    ordered_indices = sorted(index for index, _ in indexed_ranges)
    early_indices = [ordered_indices[0]]
    for index in ordered_indices[1:]:
        if index != early_indices[-1] + 1:
            break
        early_indices.append(index)

    active: dict[int, _PreparedRange] = {}
    completed: set[int] = set()
    early_offset = 0
    for row_idx in range(min_row, max_row + 1):
        for index in stops.get(row_idx, ()):
            active.pop(index, None)
        for index, item in starts.get(row_idx, ()):
            active[index] = item
        row = next(iter_rows, ())
        for index, item in active.items():
            if index in mismatches:
                continue
            start = item.min_col - min_col
            expected_cols = item.max_col - item.min_col + 1
            values = (
                tuple(row[start : start + expected_cols]) if start < len(row) else ()
            )
            if len(values) < expected_cols:
                values += (None,) * (expected_cols - len(values))
            expected_offset = row_idx - item.min_row
            expected_row = (
                item.expected_rows[expected_offset]
                if expected_offset < len(item.expected_rows)
                else (None,) * expected_cols
            )
            for offset in range(expected_cols):
                v_gt = expected_row[offset] if offset < len(expected_row) else None
                v_proc = values[offset]
                if compare_cell_value(v_gt, v_proc):
                    continue
                coord = f"{col_num2name(item.min_col + offset)}{row_idx}"
                mismatches[index] = (
                    f"Value mismatch at {coord}: your workbook has {v_proc!r}"
                )
                break

        for index, item in active.items():
            if row_idx == item.max_row:
                completed.add(index)

        while early_offset < len(early_indices):
            index = early_indices[early_offset]
            mismatch = mismatches.get(index)
            if mismatch is not None:
                return mismatch, {}
            if index not in completed:
                break
            early_offset += 1

    return None, {index: mismatches.get(index) for index, _ in indexed_ranges}


def _range_mismatch_message(item: _PreparedRange, proc_rows) -> str | None:
    proc_iter = iter(proc_rows)
    expected_cols = item.max_col - item.min_col + 1
    empty_row = (None,) * expected_cols

    for row_offset, row_idx in enumerate(range(item.min_row, item.max_row + 1)):
        expected_row = item.expected_rows[row_offset] if row_offset < len(item.expected_rows) else empty_row
        proc_row = next(proc_iter, empty_row)

        for offset in range(expected_cols):
            v_gt = expected_row[offset]
            v_proc = proc_row[offset] if offset < len(proc_row) else None
            if compare_cell_value(v_gt, v_proc):
                continue
            col_idx = item.min_col + offset
            coord = f"{col_num2name(col_idx)}{row_idx}"
            return f"Value mismatch at {coord}: your workbook has {v_proc!r}"
    return None


def _compare_proc_with_prepared(
    prepared: _PreparedGroundTruth,
    wb_proc,
) -> tuple[bool, str]:
    proc_sheet_names = [
        _resolve_sheet_name(wb_proc, item.sheet_name)
        for item in prepared.ranges
    ]
    ranges_by_sheet: dict[str, list[tuple[int, _PreparedRange]]] = {}
    for index, (item, proc_sheet_name) in enumerate(zip(prepared.ranges, proc_sheet_names)):
        if item.gt_missing or proc_sheet_name is None:
            continue
        ranges_by_sheet.setdefault(proc_sheet_name, []).append((index, item))

    scan_by_index: dict[int, list[tuple[int, _PreparedRange]]] = {}
    for indexed_ranges in ranges_by_sheet.values():
        for scan in _plan_grouped_scans(indexed_ranges):
            for index, _ in scan:
                scan_by_index[index] = scan

    comparison_results: dict[int, str | None] = {}
    for index, item in enumerate(prepared.ranges):
        missing_in_gt = item.gt_missing
        proc_sheet_name = proc_sheet_names[index]
        missing_in_proc = proc_sheet_name is None
        if missing_in_gt or missing_in_proc:
            if missing_in_gt and missing_in_proc:
                return False, "Worksheet was not found in either workbook."
            if missing_in_gt:
                return False, "Worksheet was not found in the reference workbook."
            return False, "Worksheet was not found in your workbook."

        if index not in comparison_results:
            scan = scan_by_index[index]
            ws_proc = wb_proc[proc_sheet_name]
            if len(scan) == 1:
                _, only_item = scan[0]
                proc_rows = ws_proc.iter_rows(
                    min_col=only_item.min_col,
                    min_row=only_item.min_row,
                    max_col=only_item.max_col,
                    max_row=only_item.max_row,
                    values_only=True,
                )
                mismatch = _range_mismatch_message(item, proc_rows)
            else:
                mismatch, scan_results = _compare_grouped_range_values(ws_proc, scan)
                if mismatch is not None:
                    return False, mismatch
                comparison_results.update(scan_results)
                mismatch = comparison_results.pop(index)
        else:
            mismatch = comparison_results.pop(index)

        if mismatch is not None:
            return False, mismatch
    return True, ""


def _compare_workbooks_streaming(gt_path: Path, proc_path: Path, answer_position: str) -> tuple[bool, str]:
    wb_gt = None
    wb_proc = None
    try:
        try:
            wb_gt = openpyxl.load_workbook(
                filename=str(gt_path),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except (OSError, zipfile.BadZipFile, InvalidFileException) as exc:
            logger.warning(f"[eval] Failed to load ground truth workbook ({gt_path}): {type(exc).__name__}: {exc}")
            return False, "Failed to load ground truth workbook"
        try:
            wb_proc = openpyxl.load_workbook(
                filename=str(proc_path),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except (OSError, zipfile.BadZipFile, InvalidFileException) as exc:
            logger.warning(f"[eval] Failed to load processed workbook ({proc_path}): {type(exc).__name__}: {exc}")
            return False, "Failed to load your processed workbook"

        if not wb_gt.sheetnames:
            logger.warning(f"[eval] Ground truth workbook has no worksheets: {gt_path}")
            return False, "Ground truth workbook has no worksheets"

        try:
            parsed_ranges = parse_answer_position(answer_position, default_sheet_name=wb_gt.sheetnames[0])
        except AnswerPositionError as exc:
            return False, str(exc)

        for sheet_name, cell_range in parsed_ranges:
            gt_sheet_name = _resolve_sheet_name(wb_gt, sheet_name)
            proc_sheet_name = _resolve_sheet_name(wb_proc, sheet_name)
            missing_in_gt = gt_sheet_name is None
            missing_in_proc = proc_sheet_name is None
            if missing_in_gt or missing_in_proc:
                if missing_in_gt and missing_in_proc:
                    return False, "Worksheet was not found in either workbook."
                if missing_in_gt:
                    return False, "Worksheet was not found in the reference workbook."
                return False, "Worksheet was not found in your workbook."

            ws_gt = wb_gt[gt_sheet_name]
            ws_proc = wb_proc[proc_sheet_name]
            max_row_hint = max(1, int(ws_gt.max_row or 1))
            if not _A1_COLUMN_RANGE_RE.fullmatch(cell_range):
                max_row_hint = max(max_row_hint, int(ws_proc.max_row or 1))
            min_col, min_row, max_col, max_row = _cell_range_boundaries(
                cell_range,
                max_row_hint=max_row_hint,
            )
            iter_gt = ws_gt.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
            iter_proc = ws_proc.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
            expected_cols = max_col - min_col + 1
            empty_row = (None,) * expected_cols

            for row_idx in range(min_row, max_row + 1):
                try:
                    gt_row = next(iter_gt)
                except StopIteration:
                    gt_row = empty_row
                try:
                    proc_row = next(iter_proc)
                except StopIteration:
                    proc_row = empty_row

                for offset in range(expected_cols):
                    v_gt = gt_row[offset] if offset < len(gt_row) else None
                    v_proc = proc_row[offset] if offset < len(proc_row) else None
                    if compare_cell_value(v_gt, v_proc):
                        continue
                    col_idx = min_col + offset
                    coord = f"{col_num2name(col_idx)}{row_idx}"
                    msg = f"Value mismatch at {coord}: your workbook has {v_proc!r}"
                    return False, msg

        return True, ""
    finally:
        if wb_proc is not None:
            try:
                wb_proc.close()
            except Exception as exc:  # noqa: BLE001
                logger.warning(f"[eval] wb_proc.close failed: {type(exc).__name__}: {exc}")
        if wb_gt is not None:
            try:
                wb_gt.close()
            except Exception as exc:  # noqa: BLE001
                logger.warning(f"[eval] wb_gt.close failed: {type(exc).__name__}: {exc}")


def compare_workbooks(gt_file: str | Path, proc_file: str | Path, answer_position: str):
    gt_path = Path(gt_file)
    proc_path = Path(proc_file)

    if not gt_path.exists():
        logger.warning(f"[eval] Ground truth file does not exist: {gt_path}")
        return False, "Ground truth file does not exist"
    if not proc_path.exists():
        logger.warning(f"[eval] Processed file does not exist: {proc_path}")
        return False, "Your processed file does not exist"

    estimated_cells = _estimate_answer_position_cells(answer_position)
    max_prepared_cells = _get_gt_prepared_max_cells()
    if estimated_cells is not None and estimated_cells > max_prepared_cells:
        return _compare_workbooks_streaming(gt_path, proc_path, answer_position)

    gt_cache_entry = _get_or_prepare_ground_truth(gt_path, answer_position)
    if gt_cache_entry.error is not None:
        return False, gt_cache_entry.error
    if gt_cache_entry.prepared is None:
        return False, "ground truth preparation failed"

    wb_proc = None
    try:
        try:
            wb_proc = openpyxl.load_workbook(
                filename=str(proc_path),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except (OSError, zipfile.BadZipFile, InvalidFileException) as exc:
            logger.warning(f"[eval] Failed to load processed workbook ({proc_path}): {type(exc).__name__}: {exc}")
            return False, "Failed to load your processed workbook"

        return _compare_proc_with_prepared(gt_cache_entry.prepared, wb_proc)
    finally:
        if wb_proc is not None:
            try:
                wb_proc.close()
            except Exception as exc:  # noqa: BLE001
                logger.warning(f"[eval] wb_proc.close failed: {type(exc).__name__}: {exc}")


def compute_reward(gt_file: Path, proc_file: Path, answer_position: str) -> tuple[float, str]:
    ok, msg = compare_workbooks(gt_file, proc_file, answer_position)
    reward = 1.0 if ok else 0.0
    return reward, msg
